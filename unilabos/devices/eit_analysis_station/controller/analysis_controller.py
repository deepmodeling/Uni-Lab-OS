#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
功能:
    分析站上层控制器, 读取合成任务 xlsx 中的实验信息和仪器方法配置,
    自动生成分析任务 CSV 并通过 ZhidaClient 提交至对应仪器.
    当前已实现 GC_MS 和 UPLC_QTOF 提交, HPLC 自动提交流程预留占位.
参数:
    无(通过 Settings 传入配置).
返回:
    无.
"""

import csv
import json
import logging
import io
import re
import shutil
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from ..config.setting import Settings, configure_logging
from ..driver.zhida_driver import ZhidaClient
from ..processor.chromatogram_plotter import ChromatogramPlotter
from ..processor.data_reader import GCMSDataReader
from ..processor.molecular_mass_predictor import PIMPredictor
from ..processor.mspepsearch_predictor import MSPepSearchPredictor
from ..processor.nist_library_reader import NistLibraryReader
from ..processor.peak_integrator import PeakIntegrator, PeakResult
from ..processor.nist_matcher import NISTMatcher
from ..processor.report_generator import (
    PIMPrediction, SSHMPrediction, iHSHMPrediction,
    ReportGenerator, SampleResult,
)
from ..processor.structure_fetcher import NistLocalStructureFetcher
from ..processor.yield_calculator import YieldCalculator, YIELD_CONFIG_SHEET_NAME


def _natural_sort_key(path: Path) -> list:
    """
    功能:
        自然排序键函数, 将路径名中的连续数字段转换为 int 排序,
        非数字段按小写字符串排序, 实现 725-1 < 725-2 < 725-10 的效果.
    参数:
        path: 文件或目录路径.
    返回:
        list: 混合类型排序键列表.
    """
    parts = re.split(r'(\d+)', path.stem)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


class AnalysisStationController:
    """
    功能:
        检测站上层控制器, 提供以下核心流程:
        1. 定位合成任务目录(按 task_id 或取最新任务).
        2. 检查任务状态(非 COMPLETED 时发出 warning 但继续).
        3. 解析任务 xlsx, 提取实验数量及各仪器方法名称.
        4. 生成分析任务 CSV, 双路保存.
        5. 通过 ZhidaClient 提交 CSV 给 GC_MS 和 UPLC_QTOF 仪器.
    参数:
        settings: Settings 实例, 为 None 时从环境变量读取.
    返回:
        无.
    """

    # CSV 列头(按仪器协议固定顺序)
    _CSV_HEADERS: List[str] = [
        "SampleName", "AcqMethod", "RackCode", "VialPos", "SmplInjVol", "OutputFile"
    ]

    # 默认 Rack 编号
    _DEFAULT_RACK_CODE: str = "Rack 6"

    # 默认进样量
    _DEFAULT_INJ_VOL: int = 1

    def __init__(self, settings: Optional[Settings] = None) -> None:
        self._settings = settings or Settings.from_env()
        configure_logging(self._settings.log_level)
        self._logger = logging.getLogger(self.__class__.__name__)
        self._logger.info("分析站控制器初始化完成, 合成任务目录: %s", self._settings.synthesis_tasks_dir)

    # ------------------------------------------------------------------
    # 任务定位与状态检查
    # ------------------------------------------------------------------

    def _build_task_file_name_map(self, task_id: str) -> Dict[str, str]:
        """
        功能:
            构建任务目录中的旧文件名到新文件名映射, 用于历史文件自动迁移.
        参数:
            task_id: 任务 ID.
        返回:
            Dict[str, str], 键为旧文件名, 值为新文件名.
        """
        return {
            f"{task_id}.xlsx": f"{task_id}_experiment_plan.xlsx",
            f"integration_report_{task_id}.xlsx": f"{task_id}_integration_report.xlsx",
            f"yield_report_{task_id}.xlsx": f"{task_id}_yield_report.xlsx",
            f"task_report_{task_id}.xlsx": f"{task_id}_task_report.xlsx",
            f"task_report_{task_id}.csv": f"{task_id}_task_report.csv",
            f"task_report_{task_id}.pdf": f"{task_id}_task_report.pdf",
        }

    def _migrate_task_file_names(self, task_dir: Path, task_id: str) -> None:
        """
        功能:
            在任务目录中执行旧命名到新命名的自动迁移.
            若新文件已存在, 保留新文件并记录 warning.
        参数:
            task_dir: 任务目录.
            task_id: 任务 ID.
        返回:
            无.
        """
        rename_map = self._build_task_file_name_map(task_id)
        for old_name, new_name in rename_map.items():
            old_path = task_dir / old_name
            new_path = task_dir / new_name

            if old_path.exists() is False:
                continue

            if new_path.exists() is True:
                self._logger.warning("新命名文件已存在, 跳过迁移: %s -> %s", old_path, new_path)
                continue

            try:
                old_path.rename(new_path)
                self._logger.info("历史文件重命名完成: %s -> %s", old_path.name, new_path.name)
            except Exception as exc:
                self._logger.warning("历史文件重命名失败: %s -> %s, 错误: %s", old_path, new_path, exc)

    def _find_task_dir(self, task_id: Optional[str] = None) -> Tuple[Path, str]:
        """
        功能:
            定位合成任务目录.
            若指定 task_id 则直接定位, 否则取编号最大(最新)的任务目录.
        参数:
            task_id: 任务编号字符串, None 表示自动选取最新任务.
        返回:
            Tuple[Path, str]: (任务目录 Path, 任务 ID 字符串).
        """
        tasks_root = self._settings.synthesis_tasks_dir

        if not tasks_root.exists():
            raise FileNotFoundError(f"合成任务根目录不存在: {tasks_root}")

        if task_id is not None:
            # 按指定 ID 定位
            task_dir = tasks_root / str(task_id)
            if not task_dir.is_dir():
                raise FileNotFoundError(f"指定的任务目录不存在: {task_dir}")
            self._logger.info("使用指定任务目录: %s", task_dir)
            self._migrate_task_file_names(task_dir, str(task_id))
            return task_dir, str(task_id)

        # 自动选取编号最大的子目录
        sub_dirs = [d for d in tasks_root.iterdir() if d.is_dir()]
        if not sub_dirs:
            raise FileNotFoundError(f"合成任务根目录下没有任务: {tasks_root}")

        # 尝试将目录名解析为整数排序, 取最大值
        def _dir_key(d: Path) -> int:
            try:
                return int(d.name)
            except ValueError:
                return -1

        latest_dir = max(sub_dirs, key=_dir_key)
        self._migrate_task_file_names(latest_dir, latest_dir.name)
        self._logger.info("自动选取最新任务目录: %s", latest_dir)
        return latest_dir, latest_dir.name

    def _check_task_status(self, task_dir: Path) -> str:
        """
        功能:
            读取 task_info.json 中的任务状态, 非 COMPLETED 时发出 warning.
        参数:
            task_dir: 任务目录 Path.
        返回:
            str: 任务状态字符串(如 "COMPLETED").
        """
        info_path = task_dir / "task_info.json"
        if not info_path.exists():
            self._logger.warning("未找到 task_info.json: %s, 跳过状态检查", info_path)
            return "UNKNOWN"

        with info_path.open("r", encoding="utf-8") as f:
            info = json.load(f)

        status = info.get("status", "UNKNOWN")

        if status != "COMPLETED":
            self._logger.warning(
                "任务 %s 状态为 [%s], 并非 COMPLETED, 将继续生成分析CSV.",
                info.get("task_id", "?"), status
            )
        else:
            self._logger.info("任务状态: %s", status)

        return status

    # ------------------------------------------------------------------
    # xlsx 解析
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_sheet_text(value: Any) -> str:
        """
        功能:
            规范化工作表文本, 用于表头和关键字匹配.
        参数:
            value: 任意类型单元格值.
        返回:
            str, 去空白并转小写后的文本.
        """
        text = "" if value is None else str(value)
        return (
            text.replace(" ", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .strip()
            .lower()
        )

    def _find_header_in_sheet(
        self,
        worksheet: Any,
        header_keyword: str,
        *,
        max_scan_rows: int = 80,
        max_scan_cols: int = 40,
    ) -> Tuple[Optional[int], Optional[int]]:
        """
        功能:
            在单个工作表中查找目标表头位置.
        参数:
            worksheet: openpyxl 工作表对象.
            header_keyword: 目标表头关键字, 例如 "实验编号".
            max_scan_rows: 最大扫描行数.
            max_scan_cols: 最大扫描列数.
        返回:
            Tuple[Optional[int], Optional[int]], 命中时返回(行号, 列号), 未命中返回(None, None).
        """
        normalized_keyword = self._normalize_sheet_text(header_keyword)
        scan_rows = min(worksheet.max_row, max_scan_rows)
        scan_cols = min(worksheet.max_column, max_scan_cols)

        for row_index in range(1, scan_rows + 1):
            for col_index in range(1, scan_cols + 1):
                cell_text = self._normalize_sheet_text(worksheet.cell(row_index, col_index).value)
                if normalized_keyword in cell_text and cell_text != "":
                    return row_index, col_index

        return None, None

    def _sheet_contains_instrument_anchor(self, worksheet: Any) -> bool:
        """
        功能:
            判断工作表是否包含分析仪器方法锚点字段.
        参数:
            worksheet: openpyxl 工作表对象.
        返回:
            bool, 是否命中 GC_MS/UPLC_QTOF/HPLC 任一锚点.
        """
        anchors = {"GC_MS", "UPLC_QTOF", "HPLC"}
        scan_rows = min(worksheet.max_row, 300)

        for row_index in range(1, scan_rows + 1):
            cell_value = worksheet.cell(row_index, 1).value
            if cell_value is None:
                continue
            if str(cell_value).strip() in anchors:
                return True

        return False

    def _select_task_parse_sheet(self, workbook: Any) -> Tuple[Any, int, int]:
        """
        功能:
            为任务解析选择最合适的工作表, 并返回实验编号表头位置.
            选择顺序: 实验方案设定 > 当前 active > 其它工作表.
            命中优先级: 同时命中实验编号表头和仪器锚点 > 仅命中实验编号表头.
        参数:
            workbook: openpyxl Workbook 对象.
        返回:
            Tuple[Any, int, int], (工作表对象, 实验编号表头行号, 实验编号列号).
        """
        candidate_sheet_names: List[str] = []
        preferred_sheet_name = "实验方案设定"
        if preferred_sheet_name in workbook.sheetnames:
            candidate_sheet_names.append(preferred_sheet_name)

        active_sheet_name = workbook.active.title
        if active_sheet_name not in candidate_sheet_names:
            candidate_sheet_names.append(active_sheet_name)

        for sheet_name in workbook.sheetnames:
            if sheet_name not in candidate_sheet_names:
                candidate_sheet_names.append(sheet_name)

        anchor_candidates: List[Tuple[Any, int, int]] = []
        header_only_candidates: List[Tuple[Any, int, int]] = []
        for sheet_name in candidate_sheet_names:
            worksheet = workbook[sheet_name]
            header_row, exp_no_col = self._find_header_in_sheet(worksheet, "实验编号")
            if header_row is None or exp_no_col is None:
                continue

            has_anchor = self._sheet_contains_instrument_anchor(worksheet)
            if has_anchor:
                anchor_candidates.append((worksheet, header_row, exp_no_col))
                continue

            header_only_candidates.append((worksheet, header_row, exp_no_col))

        if len(anchor_candidates) > 0:
            worksheet, header_row, exp_no_col = anchor_candidates[0]
            if len(anchor_candidates) > 1:
                candidate_names = [item[0].title for item in anchor_candidates]
                self._logger.warning(
                    "任务解析命中多个候选工作表, 将按优先顺序使用 [%s], 其余候选: %s",
                    worksheet.title,
                    candidate_names[1:],
                )
            return worksheet, header_row, exp_no_col

        if len(header_only_candidates) > 0:
            worksheet, header_row, exp_no_col = header_only_candidates[0]
            if len(header_only_candidates) > 1:
                candidate_names = [item[0].title for item in header_only_candidates]
                self._logger.warning(
                    "任务解析命中多个仅含实验编号表头的工作表, 将按优先顺序使用 [%s], 其余候选: %s",
                    worksheet.title,
                    candidate_names[1:],
                )
            self._logger.warning(
                "任务解析工作表 [%s] 命中实验编号表头, 但未检测到仪器锚点, 将按兼容模式继续解析.",
                worksheet.title,
            )
            return worksheet, header_row, exp_no_col

        raise ValueError(
            f"未找到包含实验编号表头的工作表, 可用工作表: {workbook.sheetnames}"
        )

    def _parse_task_xlsx_legacy(self, task_dir: Path, task_id: str) -> Dict:
        """
        功能:
            解析合成任务 xlsx 文件, 提取实验数量及各仪器方法名称.
            扫描 col A 定位 GC_MS/UPLC_QTOF/HPLC 字段(ASCII 可靠锚点),
            扫描 col C 统计实验数量(连续整数字符串).
        参数:
            task_dir: 任务目录 Path.
            task_id: 任务 ID 字符串.
        返回:
            Dict, 包含以下键:
                task_id (str): 任务 ID.
                exp_count (int): 实验数量.
                gc_ms_method (str|None): GC_MS 方法名, None 表示不使用.
                gc_ms_exp_nums (List[int]|None): GC_MS 实验编号过滤列表.
                uplc_qtof_method (str|None): UPLC_QTOF 方法名.
                uplc_qtof_exp_nums (List[int]|None): UPLC_QTOF 实验编号过滤列表.
                hplc_method (str|None): HPLC 方法名.
                hplc_exp_nums (List[int]|None): HPLC 实验编号过滤列表.
        """
        # 优先查找新命名 xlsx, 兼容旧命名和 .csv
        xlsx_path = task_dir / f"{task_id}_experiment_plan.xlsx"
        legacy_xlsx_path = task_dir / f"{task_id}.xlsx"
        csv_path = task_dir / f"{task_id}.csv"

        if xlsx_path.exists():
            file_path = xlsx_path
        elif legacy_xlsx_path.exists():
            file_path = legacy_xlsx_path
        elif csv_path.exists():
            file_path = csv_path
        else:
            raise FileNotFoundError(
                f"未找到任务文件 {task_id}_experiment_plan.xlsx, {task_id}.xlsx 或 {task_id}.csv 于: {task_dir}"
            )

        self._logger.info("解析任务文件: %s", file_path)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        exp_count = 0
        gc_ms_method_raw: Optional[str] = None
        uplc_qtof_method_raw: Optional[str] = None
        hplc_method_raw: Optional[str] = None

        for row in ws.iter_rows(values_only=True):
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            col_c = row[2] if len(row) > 2 else None

            # 统计 col C 中连续整数形式的实验编号
            if col_c is not None:
                try:
                    exp_num = int(str(col_c).strip())
                    if exp_num > exp_count:
                        exp_count = exp_num  # 取最大值即为总数
                except (ValueError, TypeError):
                    pass

            # 定位仪器方法行(col A 为 ASCII 关键字)
            if col_a is None:
                continue
            col_a_str = str(col_a).strip()

            if col_a_str == "GC_MS":
                # col B 为方法名, 空值则跳过该仪器
                gc_ms_method_raw = str(col_b).strip() if col_b is not None else None
            elif col_a_str == "UPLC_QTOF":
                uplc_qtof_method_raw = str(col_b).strip() if col_b is not None else None
            elif col_a_str == "HPLC":
                hplc_method_raw = str(col_b).strip() if col_b is not None else None

        if exp_count == 0:
            raise ValueError(f"未能从 {file_path} 中读取到有效实验编号, 请检查 col C 数据.")

        # 支持 Method(1-8,9) 写法, 提取方法名与实验编号过滤列表.
        gc_ms_method, gc_ms_exp_nums = self._parse_method_and_exp_filter(
            gc_ms_method_raw, exp_count, "GC_MS"
        )
        uplc_qtof_method, uplc_qtof_exp_nums = self._parse_method_and_exp_filter(
            uplc_qtof_method_raw, exp_count, "UPLC_QTOF"
        )
        hplc_method, hplc_exp_nums = self._parse_method_and_exp_filter(
            hplc_method_raw, exp_count, "HPLC"
        )

        self._logger.info(
            "任务解析完成: 实验数=%d, GC_MS方法=%s, GC_MS实验=%s, UPLC_QTOF方法=%s, "
            "UPLC_QTOF实验=%s, HPLC方法=%s, HPLC实验=%s",
            exp_count,
            gc_ms_method,
            gc_ms_exp_nums if gc_ms_exp_nums is not None else "ALL",
            uplc_qtof_method,
            uplc_qtof_exp_nums if uplc_qtof_exp_nums is not None else "ALL",
            hplc_method,
            hplc_exp_nums if hplc_exp_nums is not None else "ALL",
        )

        return {
            "task_id": task_id,
            "exp_count": exp_count,
            "gc_ms_method": gc_ms_method,
            "gc_ms_exp_nums": gc_ms_exp_nums,
            "uplc_qtof_method": uplc_qtof_method,
            "uplc_qtof_exp_nums": uplc_qtof_exp_nums,
            "hplc_method": hplc_method,
            "hplc_exp_nums": hplc_exp_nums,
        }

    def _parse_task_xlsx(self, task_dir: Path, task_id: str) -> Dict:
        """
        功能:
            解析合成任务 xlsx 文件, 提取实验数量及各仪器方法名称.
            工作表选择优先级: 实验方案设定 > 当前 active > 其它工作表.
            实验数量统计基于“实验编号”表头定位到的列, 避免误读其它表格数值列.
        参数:
            task_dir: 任务目录 Path.
            task_id: 任务 ID 字符串.
        返回:
            Dict, 包含以下键:
                task_id (str): 任务 ID.
                exp_count (int): 实验数量.
                gc_ms_method (str|None): GC_MS 方法名, None 表示不使用.
                gc_ms_exp_nums (List[int]|None): GC_MS 实验编号过滤列表.
                uplc_qtof_method (str|None): UPLC_QTOF 方法名.
                uplc_qtof_exp_nums (List[int]|None): UPLC_QTOF 实验编号过滤列表.
                hplc_method (str|None): HPLC 方法名.
                hplc_exp_nums (List[int]|None): HPLC 实验编号过滤列表.
        """
        # 优先查找新命名 xlsx, 兼容旧命名和 .csv.
        xlsx_path = task_dir / f"{task_id}_experiment_plan.xlsx"
        legacy_xlsx_path = task_dir / f"{task_id}.xlsx"
        csv_path = task_dir / f"{task_id}.csv"

        if xlsx_path.exists():
            file_path = xlsx_path
        elif legacy_xlsx_path.exists():
            file_path = legacy_xlsx_path
        elif csv_path.exists():
            file_path = csv_path
        else:
            raise FileNotFoundError(
                f"未找到任务文件 {task_id}_experiment_plan.xlsx, {task_id}.xlsx 或 {task_id}.csv 于: {task_dir}"
            )

        self._logger.info("解析任务文件: %s", file_path)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws, header_row, exp_no_col = self._select_task_parse_sheet(wb)
            if ws.title != wb.active.title:
                self._logger.info(
                    "任务解析使用工作表: %s, 当前活动工作表: %s",
                    ws.title,
                    wb.active.title,
                )

            exp_count = 0
            gc_ms_method_raw: Optional[str] = None
            uplc_qtof_method_raw: Optional[str] = None
            hplc_method_raw: Optional[str] = None

            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                col_a = row[0] if len(row) > 0 else None
                col_b = row[1] if len(row) > 1 else None
                col_exp = row[exp_no_col - 1] if len(row) >= exp_no_col else None

                # 按实验编号表头列统计实验总数, 避免误读其它表格中的数值列.
                if row_index > header_row and col_exp is not None:
                    try:
                        exp_num = int(str(col_exp).strip())
                        if exp_num > exp_count:
                            exp_count = exp_num
                    except (ValueError, TypeError):
                        pass

                # 定位仪器方法行: col A 中 ASCII 关键字.
                if col_a is None:
                    continue
                col_a_str = str(col_a).strip()

                if col_a_str == "GC_MS":
                    # col B 为方法名, 空值则跳过该仪器.
                    gc_ms_method_raw = str(col_b).strip() if col_b is not None else None
                elif col_a_str == "UPLC_QTOF":
                    uplc_qtof_method_raw = str(col_b).strip() if col_b is not None else None
                elif col_a_str == "HPLC":
                    hplc_method_raw = str(col_b).strip() if col_b is not None else None

            if exp_count == 0:
                raise ValueError(
                    f"未能从工作表 [{ws.title}] 的实验编号列读取有效实验编号, "
                    f"表头位置: row={header_row}, col={exp_no_col}, 文件: {file_path}"
                )

            # 支持 Method(1-8,9) 写法, 提取方法名与实验编号过滤列表.
            gc_ms_method, gc_ms_exp_nums = self._parse_method_and_exp_filter(
                gc_ms_method_raw, exp_count, "GC_MS"
            )
            uplc_qtof_method, uplc_qtof_exp_nums = self._parse_method_and_exp_filter(
                uplc_qtof_method_raw, exp_count, "UPLC_QTOF"
            )
            hplc_method, hplc_exp_nums = self._parse_method_and_exp_filter(
                hplc_method_raw, exp_count, "HPLC"
            )

            self._logger.info(
                "任务解析完成: 实验数=%d, GC_MS方法=%s, GC_MS实验=%s, UPLC_QTOF方法=%s, "
                "UPLC_QTOF实验=%s, HPLC方法=%s, HPLC实验=%s",
                exp_count,
                gc_ms_method,
                gc_ms_exp_nums if gc_ms_exp_nums is not None else "ALL",
                uplc_qtof_method,
                uplc_qtof_exp_nums if uplc_qtof_exp_nums is not None else "ALL",
                hplc_method,
                hplc_exp_nums if hplc_exp_nums is not None else "ALL",
            )

            return {
                "task_id": task_id,
                "exp_count": exp_count,
                "gc_ms_method": gc_ms_method,
                "gc_ms_exp_nums": gc_ms_exp_nums,
                "uplc_qtof_method": uplc_qtof_method,
                "uplc_qtof_exp_nums": uplc_qtof_exp_nums,
                "hplc_method": hplc_method,
                "hplc_exp_nums": hplc_exp_nums,
            }
        finally:
            wb.close()

    def _parse_method_and_exp_filter(
        self,
        raw_method: Optional[str],
        exp_count: int,
        instrument: str,
    ) -> Tuple[Optional[str], Optional[List[int]]]:
        """
        功能:
            解析方法字符串中的实验编号过滤表达式.
            支持 `Generic_15min(1-8,9)` 和 `Generic_15min（1-8，9）` 等写法.
            仅当括号内容为数字范围表达式时启用过滤, 否则保持原方法名不变.
        参数:
            raw_method: 原始方法字符串, None 或空字符串表示未配置.
            exp_count: 实验总数, 用于校验编号范围.
            instrument: 仪器名称, 用于日志和错误信息.
        返回:
            Tuple[Optional[str], Optional[List[int]]]:
                第一个值为去掉过滤后的方法名, 第二个值为过滤实验编号列表.
                若未配置过滤, 第二个值为 None.
        """
        if raw_method is None:
            return None, None

        method_text = raw_method.strip()
        if method_text == "":
            return None, None

        # 仅匹配结尾一对括号, 避免干扰方法名中间内容.
        match = re.match(r"^(.*?)[\(（]\s*(.*?)\s*[\)）]\s*$", method_text)
        if match is None:
            return method_text, None

        method_name = match.group(1).strip()
        selector_text = match.group(2).strip()

        if selector_text == "":
            return method_text, None

        # 仅当括号中是编号表达式时才按过滤处理, 其余情况保持原方法名.
        if re.fullmatch(r"[0-9０-９\s,，、;；\-－—–~～到至]+", selector_text) is None:
            return method_text, None

        if method_name == "":
            raise ValueError(f"{instrument} 方法配置无效, 括号前方法名不能为空: {raw_method}")

        exp_nums = self._parse_exp_selector(selector_text, exp_count, instrument)
        return method_name, exp_nums

    def _parse_exp_selector(self, selector_text: str, exp_count: int, instrument: str) -> List[int]:
        """
        功能:
            将实验编号过滤表达式解析为有序去重的实验编号列表.
            支持中英文括号内的中文标点, 例如 `1-8,9` `1～8，9` `1到8、9`.
        参数:
            selector_text: 括号内原始表达式文本.
            exp_count: 实验总数, 用于范围校验.
            instrument: 仪器名称, 用于错误提示.
        返回:
            List[int], 有序去重后的实验编号列表.
        """
        full_width_digits = str.maketrans("０１２３４５６７８９", "0123456789")
        normalized = selector_text.translate(full_width_digits)

        # 统一分隔符与范围连接符, 简化后续解析.
        normalized = normalized.replace("，", ",").replace("、", ",").replace("；", ",").replace(";", ",")
        for range_sep in ("－", "—", "–", "~", "～", "到", "至"):
            normalized = normalized.replace(range_sep, "-")
        normalized = normalized.replace(" ", "")

        if normalized == "":
            raise ValueError(f"{instrument} 方法实验编号为空, 请检查括号内容: {selector_text}")

        exp_num_set = set()
        for token in normalized.split(","):
            if token == "":
                continue

            if "-" in token:
                parts = token.split("-")
                if len(parts) != 2 or parts[0] == "" or parts[1] == "":
                    raise ValueError(f"{instrument} 方法实验编号格式无效: {selector_text}")

                try:
                    start_num = int(parts[0])
                    end_num = int(parts[1])
                except ValueError as exc:
                    raise ValueError(f"{instrument} 方法实验编号格式无效: {selector_text}") from exc

                if start_num > end_num:
                    raise ValueError(f"{instrument} 方法实验编号范围无效: {selector_text}")

                for exp_num in range(start_num, end_num + 1):
                    exp_num_set.add(exp_num)
            else:
                try:
                    exp_num_set.add(int(token))
                except ValueError as exc:
                    raise ValueError(f"{instrument} 方法实验编号格式无效: {selector_text}") from exc

        if len(exp_num_set) == 0:
            raise ValueError(f"{instrument} 方法实验编号为空, 请检查括号内容: {selector_text}")

        exp_nums = sorted(exp_num_set)
        invalid_nums = [num for num in exp_nums if num < 1 or num > exp_count]
        if len(invalid_nums) > 0:
            raise ValueError(
                f"{instrument} 方法实验编号超出范围(1-{exp_count}): {invalid_nums}, 原始表达式: {selector_text}"
            )

        return exp_nums

    # ------------------------------------------------------------------
    # VialPos 计算
    # ------------------------------------------------------------------

    def _calc_vial_pos(self, exp_num: int) -> int:
        """
        功能:
            根据实验编号计算 GC-MS 进样位置 VialPos.

            样品托盘(闪滤瓶外瓶托盘) 规格: 6行(A-F) × 8列(1-8).
            装样遵循蛇形规则:
                偶数行(A/C/E): 从左到右 col 1→8.
                奇数行(B/D/F): 从右到左 col 8→1.
            GC-MS 进样位置从 F8=1 向上递增, 直至 A1=48.
        参数:
            exp_num: 实验编号, 范围 1-48.
        返回:
            int, 对应的 VialPos(1-48).
        """
        exp_0 = exp_num - 1              # 转为 0-indexed
        row_0 = exp_0 // 8              # 行索引: 0=A, 1=B, ..., 5=F
        col_within = exp_0 % 8          # 该行内第几个样品(0-indexed)

        # 蛇形: 奇数行(B/D/F)列方向翻转
        col_0 = col_within if row_0 % 2 == 0 else 7 - col_within

        row = row_0 + 1                 # 1=A, ..., 6=F
        col = col_0 + 1                 # 1-8

        # VialPos: F8=1, F7=2, ..., A1=48
        vial_pos = (6 - row) * 8 + (9 - col)
        return vial_pos

    # ------------------------------------------------------------------
    # CSV 生成
    # ------------------------------------------------------------------

    def _generate_gc_ms_csv(
        self,
        task_id: str,
        exp_count: int,
        method: str,
        exp_nums: Optional[List[int]] = None,
    ) -> str:
        """
        功能:
            生成 GC_MS 分析任务 CSV 字符串.
            每一行对应一个实验样品, VialPos 由蛇形映射公式计算.
            当传入 exp_nums 时, 仅为指定实验编号生成提交行.
        参数:
            task_id: 任务 ID, 用于拼接 SampleName/OutputFile.
            exp_count: 实验数量.
            method: GC_MS 方法名称.
            exp_nums: 可选实验编号列表, None 表示 1..exp_count 全量提交.
        返回:
            str: CSV 文本内容(含列头).
        """
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")

        # 写入列头
        writer.writerow(self._CSV_HEADERS)

        target_exp_nums = exp_nums if exp_nums is not None else list(range(1, exp_count + 1))

        for exp_num in target_exp_nums:
            sample_name = f"{task_id}-{exp_num}"          # 如 "719-1"
            vial_pos = self._calc_vial_pos(exp_num)        # VialPos 蛇形映射
            writer.writerow([
                sample_name,                               # SampleName
                method,                                    # AcqMethod
                self._DEFAULT_RACK_CODE,                   # RackCode
                vial_pos,                                  # VialPos
                self._DEFAULT_INJ_VOL,                     # SmplInjVol
                sample_name,                               # OutputFile(与 SampleName 相同)
            ])

        return output.getvalue()

    # ------------------------------------------------------------------
    # CSV 保存
    # ------------------------------------------------------------------

    def _save_csv(self, content: str, task_id: str, instrument: str) -> List[Path]:
        """
        功能:
            将 CSV 内容双路保存:
            1. 分析站本地数据目录: eit_analysis_station/data/<task_id>/<instrument>.csv.
            2. 合成任务目录: eit_synthesis_station/data/tasks/<task_id>/<instrument>.csv.
        参数:
            content: CSV 文本内容.
            task_id: 任务 ID.
            instrument: 仪器名称(如 "gc_ms", "hplc", "uplc_qtof").
        返回:
            List[Path]: 实际保存的文件路径列表.
        """
        filename = f"{instrument}.csv"
        saved_paths: List[Path] = []

        # 路径 1: 分析站本地 data/<task_id>/
        local_dir = self._settings.data_dir / task_id
        local_dir.mkdir(parents=True, exist_ok=True)
        local_path = local_dir / filename
        local_path.write_text(content, encoding="utf-8")
        self._logger.info("CSV 已保存至本地: %s", local_path)
        saved_paths.append(local_path)

        # 路径 2: 合成任务目录 synthesis_tasks/<task_id>/
        syn_dir = self._settings.synthesis_tasks_dir / task_id
        if syn_dir.is_dir():
            syn_path = syn_dir / filename
            syn_path.write_text(content, encoding="utf-8")
            self._logger.info("CSV 已同步至合成任务目录: %s", syn_path)
            saved_paths.append(syn_path)
        else:
            self._logger.warning("合成任务目录不存在, 跳过同步: %s", syn_dir)

        return saved_paths

    # ------------------------------------------------------------------
    # GC_MS 提交流程
    # ------------------------------------------------------------------

    def _do_submit_gc_ms(self, resolved_id: str, task_info: Dict) -> Dict:
        """
        功能:
            GC_MS 提交核心逻辑(内部方法), 接收已解析的任务信息直接执行,
            避免 run_analysis 统一调度时重复解析 xlsx.
        参数:
            resolved_id: 任务 ID 字符串.
            task_info: _parse_task_xlsx 返回的任务信息字典.
        返回:
            Dict: {"success": bool, "return_info": str}.
        """
        gc_ms_method = task_info["gc_ms_method"]

        if gc_ms_method is None:
            msg = f"任务 {resolved_id} 未配置 GC_MS 方法, 跳过 GC_MS 提交."
            self._logger.info(msg)
            return {"success": True, "return_info": msg}

        gc_ms_exp_nums = task_info.get("gc_ms_exp_nums")

        # 步骤1: 生成并保存 CSV
        csv_content = self._generate_gc_ms_csv(
            resolved_id, task_info["exp_count"], gc_ms_method, gc_ms_exp_nums
        )
        saved_paths = self._save_csv(csv_content, resolved_id, "gc_ms")

        # 步骤2: 通过 ZhidaClient 提交至 GC_MS 仪器
        client = ZhidaClient(
            host=self._settings.gc_ms_host,
            port=self._settings.gc_ms_port,
            timeout=self._settings.gc_ms_timeout,
        )
        self._logger.info(
            "连接 GC_MS: %s:%d", self._settings.gc_ms_host, self._settings.gc_ms_port
        )
        if gc_ms_exp_nums is not None:
            self._logger.info("GC_MS 按实验编号过滤提交: %s", gc_ms_exp_nums)

        try:
            client.connect()
            # 使用本地保存的第一份 CSV 文件提交
            submit_path = str(saved_paths[0])
            result = client.start_with_csv_file(string=submit_path)
        finally:
            client.close()  # 确保连接关闭

        self._logger.info("GC_MS 提交结果: %s", result)
        return result

    def submit_gc_ms(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            GC_MS 分析任务完整提交流程(公开入口, 独立调用时使用):
            1. 定位任务目录(task_id 或最新).
            2. 检查 task_info.json 状态(非 COMPLETED 则 warning 后继续).
            3. 解析 xlsx, 若无 gc_ms_method 则跳过提交.
            4. 生成分析 CSV 并双路保存.
            5. ZhidaClient 连接 GC_MS 并调用 start_with_csv_file 提交.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
        返回:
            Dict: {"success": bool, "return_info": str}.
        """
        try:
            # 定位任务目录
            task_dir, resolved_id = self._find_task_dir(task_id)
            # 检查任务状态(仅 warning, 不阻断)
            self._check_task_status(task_dir)
            # 解析 xlsx
            task_info = self._parse_task_xlsx(task_dir, resolved_id)
            # 调用核心提交逻辑
            return self._do_submit_gc_ms(resolved_id, task_info)

        except Exception as exc:
            msg = f"GC_MS 提交失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

    # ------------------------------------------------------------------
    # UPLC_QTOF 提交流程
    # ------------------------------------------------------------------

    def _do_submit_uplc_qtof(self, resolved_id: str, task_info: Dict) -> Dict:
        """
        功能:
            UPLC_QTOF 提交核心逻辑(内部方法), 接收已解析的任务信息直接执行.
            通过复用统一 CSV 协议格式, 避免 run_analysis 统一调度时重复解析 xlsx.
        参数:
            resolved_id: 任务 ID 字符串.
            task_info: _parse_task_xlsx 返回的任务信息字典.
        返回:
            Dict: {"success": bool, "return_info": str}.
        """
        uplc_qtof_method = task_info["uplc_qtof_method"]

        if uplc_qtof_method is None:
            msg = f"任务 {resolved_id} 未配置 UPLC_QTOF 方法, 跳过 UPLC_QTOF 提交."
            self._logger.info(msg)
            return {"success": True, "return_info": msg}

        uplc_qtof_exp_nums = task_info.get("uplc_qtof_exp_nums")

        # UPLC_QTOF 与 GC_MS 使用一致的 CSV 协议格式.
        csv_content = self._generate_gc_ms_csv(
            resolved_id, task_info["exp_count"], uplc_qtof_method, uplc_qtof_exp_nums
        )

        if self._settings.uplc_qtof_append_wash_stop is True:
            # 末尾追加 wash_stop 行, 仅写方法名, 不填进样信息.
            wash_buf = io.StringIO()
            wash_writer = csv.writer(wash_buf, lineterminator="\n")
            wash_writer.writerow(["", "wash_stop", "", "", "", ""])
            csv_content += wash_buf.getvalue()
            self._logger.info("UPLC_QTOF 已追加停止方法行: wash_stop")
        else:
            self._logger.info("UPLC_QTOF 已关闭追加停止方法")

        saved_paths = self._save_csv(csv_content, resolved_id, "uplc_qtof")

        client = ZhidaClient(
            host=self._settings.uplc_qtof_host,
            port=self._settings.uplc_qtof_port,
            timeout=self._settings.uplc_qtof_timeout,
        )
        self._logger.info(
            "连接 UPLC_QTOF: %s:%d",
            self._settings.uplc_qtof_host,
            self._settings.uplc_qtof_port,
        )
        if uplc_qtof_exp_nums is not None:
            self._logger.info("UPLC_QTOF 按实验编号过滤提交: %s", uplc_qtof_exp_nums)

        try:
            client.connect()
            submit_path = str(saved_paths[0])
            result = client.start_with_csv_file(string=submit_path)
        finally:
            client.close()

        self._logger.info("UPLC_QTOF 提交结果: %s", result)
        return result

    def submit_uplc_qtof(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            UPLC_QTOF 分析任务完整提交流程(公开入口, 独立调用时使用):
            1. 定位任务目录(task_id 或最新).
            2. 检查 task_info.json 状态, 非 COMPLETED 时 warning 后继续.
            3. 解析 xlsx, 若无 uplc_qtof_method 则跳过提交.
            4. 生成分析 CSV 并双路保存.
            5. ZhidaClient 连接 UPLC_QTOF 并调用 start_with_csv_file 提交.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
        返回:
            Dict: {"success": bool, "return_info": str}.
        """
        try:
            task_dir, resolved_id = self._find_task_dir(task_id)
            self._check_task_status(task_dir)
            task_info = self._parse_task_xlsx(task_dir, resolved_id)
            return self._do_submit_uplc_qtof(resolved_id, task_info)
        except Exception as exc:
            msg = f"UPLC_QTOF 提交失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

    def submit_by_csv_path(self, instrument: str, csv_file_path: str) -> Dict:
        """
        功能:
            按指定仪器和 CSV 文件路径直接提交分析任务.
            支持 gc_ms, uplc_qtof, hplc 三种仪器.
        参数:
            instrument: 仪器标识, 可选值为 gc_ms/uplc_qtof/hplc.
            csv_file_path: CSV 文件路径.
        返回:
            Dict: {"success": bool, "return_info": str}.
        """
        instrument_key = instrument.strip().lower()
        instrument_configs = {
            "gc_ms": ("GC_MS", self._settings.gc_ms_host, self._settings.gc_ms_port, self._settings.gc_ms_timeout),
            "uplc_qtof": (
                "UPLC_QTOF",
                self._settings.uplc_qtof_host,
                self._settings.uplc_qtof_port,
                self._settings.uplc_qtof_timeout,
            ),
            "hplc": ("HPLC", self._settings.hplc_host, self._settings.hplc_port, self._settings.hplc_timeout),
        }

        if instrument_key not in instrument_configs:
            msg = "仪器参数无效, 可选值: gc_ms, uplc_qtof, hplc."
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

        csv_path = Path(csv_file_path)
        if csv_path.exists() is False or csv_path.is_file() is False:
            msg = f"CSV 文件不存在或不是文件: {csv_file_path}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

        instrument_name, host, port, timeout = instrument_configs[instrument_key]
        client = ZhidaClient(host=host, port=port, timeout=timeout)
        self._logger.info(
            "按路径提交 CSV, 仪器=%s, 地址=%s:%d, 文件=%s",
            instrument_name, host, port, csv_path,
        )

        try:
            client.connect()
            result = client.start_with_csv_file(string=str(csv_path))
            self._logger.info("%s 手工 CSV 提交结果: %s", instrument_name, result)
            return result
        except Exception as exc:
            msg = f"{instrument_name} CSV 提交失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}
        finally:
            client.close()

    # ------------------------------------------------------------------
    # 统一分析入口
    # ------------------------------------------------------------------

    def run_analysis(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            统一分析入口, 依据 xlsx 中各仪器方法配置依次处理.
            xlsx 仅解析一次, 各仪器提交直接调用内部核心方法避免重复解析.
            支持方法写法 `Method(1-8,9)`, 仅提交括号中实验编号对应样品.
            当前已实现: GC_MS, UPLC_QTOF.
            预留未实现: HPLC 自动提交流程(方法存在时发出 warning).
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
        返回:
            Dict: {"gc_ms": result_dict, "uplc_qtof": result_dict, "hplc": result_dict}.
        """
        results: Dict = {}

        try:
            # 定位任务目录并解析 xlsx(仅执行一次)
            task_dir, resolved_id = self._find_task_dir(task_id)
            self._check_task_status(task_dir)
            task_info = self._parse_task_xlsx(task_dir, resolved_id)
        except Exception as exc:
            msg = f"任务初始化失败: {exc}"
            self._logger.error(msg)
            return {"error": msg}

        # ---------- GC_MS ----------
        if task_info["gc_ms_method"] is not None:
            self._logger.info("开始提交 GC_MS 分析任务...")
            try:
                # 直接调用核心方法，跳过重复的定位+解析步骤
                results["gc_ms"] = self._do_submit_gc_ms(resolved_id, task_info)
            except Exception as exc:
                results["gc_ms"] = {"success": False, "return_info": f"GC_MS 提交失败: {exc}"}
        else:
            results["gc_ms"] = {"success": True, "return_info": "未配置 GC_MS 方法, 已跳过."}

        # ---------- UPLC_QTOF ----------
        if task_info["uplc_qtof_method"] is not None:
            self._logger.info("开始提交 UPLC_QTOF 分析任务...")
            try:
                results["uplc_qtof"] = self._do_submit_uplc_qtof(resolved_id, task_info)
            except Exception as exc:
                results["uplc_qtof"] = {"success": False, "return_info": f"UPLC_QTOF 提交失败: {exc}"}
        else:
            results["uplc_qtof"] = {"success": True, "return_info": "未配置 UPLC_QTOF 方法, 已跳过."}

        # ---------- HPLC(预留) ----------
        if task_info["hplc_method"] is not None:
            self._logger.warning(
                "任务 %s 配置了 HPLC 方法 [%s], 但 HPLC 接入尚未实现, 已跳过.",
                resolved_id, task_info["hplc_method"]
            )
            results["hplc"] = {"success": False, "return_info": "HPLC 接入尚未实现."}
        else:
            results["hplc"] = {"success": True, "return_info": "未配置 HPLC 方法, 已跳过."}

        self._logger.info("分析任务提交完毕, 结果: %s", results)
        return results

    # ------------------------------------------------------------------
    # GC_MS 结果处理(积分 + 定性 + 报告)
    # ------------------------------------------------------------------

    def _enumerate_d_dirs(self, task_id: str) -> List[Path]:
        """
        功能:
            枚举指定任务下所有 .D 结果目录.
            先在仪器侧网络目录查找, 再在本地 data 目录查找.
        参数:
            task_id: 任务 ID.
        返回:
            List[Path]: 找到的 .D 目录列表, 按样品编号排序.
        """
        d_dirs: List[Path] = []

        # 优先查找仪器侧网络目录
        remote_data_dir = self._settings.gc_ms_data_dir
        if remote_data_dir.exists():
            # 匹配 <task_id>-<num>.D 格式
            for d_dir in sorted(remote_data_dir.glob(f"{task_id}-*.D"), key=_natural_sort_key):
                if d_dir.is_dir():
                    d_dirs.append(d_dir)

        if d_dirs:
            self._logger.info("从仪器侧目录找到 %d 个 .D 文件: %s", len(d_dirs), remote_data_dir)
            return d_dirs

        # 备选: 在本地 data/<task_id>/ 目录下查找
        local_data_dir = self._settings.data_dir / task_id
        if local_data_dir.exists():
            for d_dir in sorted(local_data_dir.glob("*.D"), key=_natural_sort_key):
                if d_dir.is_dir():
                    d_dirs.append(d_dir)

        if d_dirs:
            self._logger.info("从本地目录找到 %d 个 .D 文件: %s", len(d_dirs), local_data_dir)
        else:
            self._logger.warning("未找到任务 %s 的 .D 结果目录", task_id)

        return d_dirs

    def _load_expected_samples(self, task_id: str) -> List[str]:
        """
        功能:
            从本地数据目录的 gc_ms.csv 读取预期样品列表, 按 CSV 行顺序返回.
            CSV 由 _save_csv 生成, 格式为:
            SampleName,AcqMethod,RackCode,VialPos,SmplInjVol,OutputFile.
        参数:
            task_id: 任务 ID 字符串.
        返回:
            List[str]: 样品名称列表, 如 ["725-1", "725-2", ..., "725-12"].
        """
        csv_path = self._settings.data_dir / task_id / "gc_ms.csv"
        if not csv_path.exists():
            raise FileNotFoundError(f"未找到样品列表文件: {csv_path}")

        with csv_path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            samples = [row["SampleName"] for row in reader]

        if not samples:
            raise ValueError(f"样品列表为空: {csv_path}")

        self._logger.info("从 gc_ms.csv 加载 %d 个预期样品", len(samples))
        return samples

    def _read_run_completed_flag(self, d_dir: Path) -> Optional[bool]:
        """
        功能:
            读取 .D 目录下 AcqData/sample_info.xml 中的 RunCompletedFlag 字段.
            解析 XML 中所有 <Field> 元素, 找到 Name 为 "RunCompletedFlag" 的条目,
            返回其 Value 的布尔解析结果.
        参数:
            d_dir: .D 目录路径.
        返回:
            Optional[bool]: True 表示采集完成, False 表示采集中,
                            None 表示文件不存在或解析失败.
        """
        info_path = d_dir / "AcqData" / "sample_info.xml"
        if not info_path.exists():
            return None

        try:
            tree = ET.parse(str(info_path))
            root = tree.getroot()
            for field_elem in root.findall("Field"):
                name = field_elem.findtext("Name", "")
                if name == "RunCompletedFlag":
                    value = field_elem.findtext("Value", "").strip()
                    return value.lower() == "true"
            # RunCompletedFlag 字段不存在, 视为未完成
            return False
        except (ET.ParseError, OSError) as exc:
            # 文件可能正在被仪器写入, 视为采集中
            self._logger.debug("解析 %s 失败: %s, 视为采集中", info_path, exc)
            return None

    def _filter_peaks(
        self,
        peaks: List[PeakResult],
        area_min: Optional[float] = None,
        area_max: Optional[float] = None,
    ) -> List[PeakResult]:
        """
        功能:
            按保留时间范围和面积阈值过滤峰列表, 过滤后重新计算面积百分比.
        参数:
            peaks: 积分后的峰列表.
            area_min: 峰面积下限, None 表示不过滤.
            area_max: 峰面积上限, None 表示不过滤.
        返回:
            List[PeakResult]: 过滤后的峰列表.
        """
        filtered = peaks

        # 保留时间范围过滤 (TIC/FID 共用)
        if self._settings.peak_rt_min is not None:
            filtered = [p for p in filtered if p.retention_time >= self._settings.peak_rt_min]
        if self._settings.peak_rt_max is not None:
            filtered = [p for p in filtered if p.retention_time <= self._settings.peak_rt_max]

        # 面积范围过滤 (TIC/FID 分别传入不同阈值)
        if area_min is not None:
            filtered = [p for p in filtered if p.area >= area_min]
        if area_max is not None:
            filtered = [p for p in filtered if p.area <= area_max]

        # 重新计算面积百分比
        if len(filtered) < len(peaks):
            total_area = sum(p.area for p in filtered)
            if total_area > 0:
                for p in filtered:
                    p.area_percent = (p.area / total_area) * 100.0
            self._logger.info("峰过滤: %d -> %d 个峰", len(peaks), len(filtered))

        return filtered

    def _process_single_sample(
        self, d_dir: Path, nist: NISTMatcher, report_dir: Optional[Path] = None,
        mspepsearch_predictor: Optional[MSPepSearchPredictor] = None,
    ) -> SampleResult:
        """
        功能:
            处理单个 .D 目录: 读取 TIC/FID, 积分, NIST 匹配, PIM/SS-HM/iHS-HM 预测, 生成色谱图.
        参数:
            d_dir: .D 目录路径.
            nist: NISTMatcher 实例 (由外部传入, 保持状态复用).
            report_dir: 报告输出目录, 用于保存色谱图. None 则不生成图.
            mspepsearch_predictor: MSPepSearchPredictor 实例, None 则跳过 SS-HM/iHS-HM 预测.
        返回:
            SampleResult: 该样品的完整积分结果.
        """
        reader = GCMSDataReader()

        # 读取样品元数据
        sample_info = reader.read_sample_info(d_dir)
        sample_name = sample_info.get("sample_name") or d_dir.stem  # 空值回退到目录名
        acq_time = sample_info.get("acq_time", "")

        result = SampleResult(
            sample_name=sample_name,
            d_dir=d_dir,
            acq_time=acq_time,
        )

        # 缓存色谱数据, 供后续绘图复用
        tic_times = tic_intensities = None
        fid_times = fid_intensities = None
        tic_baseline = fid_baseline = None  # 积分基线, 供绘图使用
        peak_ms_cache: Dict[int, Tuple] = {}

        pim_predictor: Optional[PIMPredictor] = None
        if self._settings.pim_enable is True:
            try:
                pim_predictor = PIMPredictor(
                    ab_m=self._settings.pim_ab_m,
                    beta=self._settings.pim_beta,
                    epsilon_f=self._settings.pim_epsilon_f,
                )
            except Exception as e:
                self._logger.error("样品 %s PIM 预测器初始化失败: %s", sample_name, e)

        # TIC 积分
        try:
            tic_times, tic_intensities = reader.read_tic(d_dir)
            tic_integrator = PeakIntegrator(
                smoothing_window=self._settings.peak_smoothing_window,
                prominence=self._settings.peak_prominence,
                min_distance=self._settings.peak_min_distance,
                width_rel_height=self._settings.peak_width_rel_height,
                use_als_baseline=self._settings.use_als_baseline,
                als_lambda=self._settings.als_lambda,
                als_p=self._settings.als_p,
                use_valley_boundary=self._settings.use_valley_boundary,
                integration_mode=self._settings.integration_mode,
                baseline_method=self._settings.baseline_method,
                baseline_quantile=self._settings.baseline_quantile,
                baseline_window_min=self._settings.baseline_window_min,
                boundary_sigma_factor=self._settings.boundary_sigma_factor,
                boundary_edge_ratio=self._settings.boundary_edge_ratio,
                boundary_expand_factor=self._settings.boundary_expand_factor,
                boundary_min_span_min=self._settings.boundary_min_span_min,
                boundary_max_span_min=self._settings.boundary_max_span_min,
                shoulder_filter_enable=self._settings.robust_v3_shoulder_filter_enable,
                shoulder_filter_width_max_min=self._settings.robust_v3_shoulder_width_max_min,
                shoulder_filter_gap_max_min=self._settings.robust_v3_shoulder_gap_max_min,
                shoulder_filter_relative_prominence_max=self._settings.robust_v3_shoulder_relative_prominence_max,
                tail_artifact_filter_enable=self._settings.robust_v3_tail_artifact_filter_enable,
                tail_artifact_gap_max_min=self._settings.robust_v3_tail_artifact_gap_max_min,
                tail_artifact_relative_prominence_max=self._settings.robust_v3_tail_artifact_relative_prominence_max,
                tail_artifact_half_width_asymmetry_min=self._settings.robust_v3_tail_artifact_half_width_asymmetry_min,
                tail_monotonic_filter_enable=self._settings.robust_v3_tail_monotonic_filter_enable,
                tail_monotonic_ratio_max=self._settings.robust_v3_tail_monotonic_ratio_max,
                max_peak_width_min=self._settings.robust_v3_max_peak_width_min,
                leading_edge_filter_enable=self._settings.robust_v3_leading_edge_filter_enable,
                leading_edge_relative_prominence_max=self._settings.robust_v3_leading_edge_relative_prominence_max,
                leading_edge_monotonic_ratio_min=self._settings.robust_v3_leading_edge_monotonic_ratio_min,
                gcpy_whittaker_lmbd=self._settings.gcpy_whittaker_lmbd,
                use_cwt_detection=self._settings.use_cwt_detection,
                cwt_min_width_min=self._settings.cwt_min_width_min,
                cwt_max_width_min=self._settings.cwt_max_width_min,
                cwt_min_snr=self._settings.cwt_min_snr,
                cwt_noise_perc=self._settings.cwt_noise_perc,
            )
            result.tic_peaks = tic_integrator.integrate(tic_times, tic_intensities)
            tic_baseline = tic_integrator.last_baseline
            result.tic_peaks = self._filter_peaks(
                result.tic_peaks,
                area_min=self._settings.tic_area_min,
                area_max=self._settings.tic_area_max,
            )
            self._logger.info("样品 %s TIC 积分: %d 个峰", sample_name, len(result.tic_peaks))
        except Exception as e:
            self._logger.error("样品 %s TIC 积分失败: %s", sample_name, e)

        # FID 积分
        try:
            fid_times, fid_intensities = reader.read_fid(d_dir)
            fid_integrator = PeakIntegrator(
                smoothing_window=self._settings.peak_smoothing_window,
                prominence=self._settings.fid_peak_prominence,
                min_distance=self._settings.fid_peak_min_distance,
                width_rel_height=self._settings.peak_width_rel_height,
                use_als_baseline=self._settings.use_als_baseline,
                als_lambda=self._settings.als_lambda,
                als_p=self._settings.als_p,
                use_valley_boundary=self._settings.use_valley_boundary,
                integration_mode=self._settings.integration_mode,
                baseline_method=self._settings.baseline_method,
                baseline_quantile=self._settings.baseline_quantile,
                baseline_window_min=self._settings.baseline_window_min,
                boundary_sigma_factor=self._settings.boundary_sigma_factor,
                boundary_edge_ratio=self._settings.boundary_edge_ratio,
                boundary_expand_factor=self._settings.boundary_expand_factor,
                boundary_min_span_min=self._settings.boundary_min_span_min,
                boundary_max_span_min=self._settings.boundary_max_span_min,
                shoulder_filter_enable=self._settings.robust_v3_shoulder_filter_enable,
                shoulder_filter_width_max_min=self._settings.robust_v3_shoulder_width_max_min,
                shoulder_filter_gap_max_min=self._settings.robust_v3_shoulder_gap_max_min,
                shoulder_filter_relative_prominence_max=self._settings.robust_v3_shoulder_relative_prominence_max,
                tail_artifact_filter_enable=self._settings.robust_v3_tail_artifact_filter_enable,
                tail_artifact_gap_max_min=self._settings.robust_v3_tail_artifact_gap_max_min,
                tail_artifact_relative_prominence_max=self._settings.robust_v3_tail_artifact_relative_prominence_max,
                tail_artifact_half_width_asymmetry_min=self._settings.robust_v3_tail_artifact_half_width_asymmetry_min,
                tail_monotonic_filter_enable=self._settings.robust_v3_tail_monotonic_filter_enable,
                tail_monotonic_ratio_max=self._settings.robust_v3_tail_monotonic_ratio_max,
                max_peak_width_min=self._settings.robust_v3_max_peak_width_min,
                leading_edge_filter_enable=self._settings.robust_v3_leading_edge_filter_enable,
                leading_edge_relative_prominence_max=self._settings.robust_v3_leading_edge_relative_prominence_max,
                leading_edge_monotonic_ratio_min=self._settings.robust_v3_leading_edge_monotonic_ratio_min,
                gcpy_whittaker_lmbd=self._settings.gcpy_whittaker_lmbd,
                use_cwt_detection=self._settings.use_cwt_detection,
                cwt_min_width_min=self._settings.cwt_min_width_min,
                cwt_max_width_min=self._settings.cwt_max_width_min,
                cwt_min_snr=self._settings.cwt_min_snr,
                cwt_noise_perc=self._settings.cwt_noise_perc,
            )
            result.fid_peaks = fid_integrator.integrate(fid_times, fid_intensities)
            fid_baseline = fid_integrator.last_baseline
            result.fid_peaks = self._filter_peaks(
                result.fid_peaks,
                area_min=self._settings.fid_area_min,
                area_max=self._settings.fid_area_max,
            )
            self._logger.info("样品 %s FID 积分: %d 个峰", sample_name, len(result.fid_peaks))
        except Exception as e:
            self._logger.error("样品 %s FID 积分失败: %s", sample_name, e)

        # NIST 化合物匹配 (优先使用 NIST MS Search 自动化)
        try:
            if nist.nist_available and result.tic_peaks:
                # 提取各 TIC 峰的质谱并通过 NIST 搜索匹配
                result.compound_matches = nist.match_peaks_with_nist(
                    d_dir, result.tic_peaks, reader,
                    avg_scans=self._settings.nist_avg_scans,
                )
                if result.compound_matches:
                    self._logger.info(
                        "样品 %s NIST 自动匹配: %d 个峰有匹配结果",
                        sample_name, len(result.compound_matches)
                    )
                # 记录 NIST 结果文件副本路径
                if hasattr(nist, '_saved_srcreslt') and nist._saved_srcreslt is not None:
                    result.nist_result_path = nist._saved_srcreslt
            else:
                # 降级: 从 MassHunter 已有报告中提取
                result.compound_matches = nist.match_from_qual_results(d_dir)

            if not result.compound_matches:
                self._logger.info("样品 %s 无可用的 NIST 定性结果", sample_name)
        except Exception as e:
            self._logger.error("样品 %s NIST 匹配失败: %s", sample_name, e)

        # 提取 TIC 峰质谱并执行 PIM 预测
        if len(result.tic_peaks) > 0:
            for peak_num, peak in enumerate(result.tic_peaks, start=1):
                try:
                    mz_values, ms_intensities = reader.read_ms_spectra_at_peak(
                        d_dir, peak.start_time, peak.end_time,
                        avg_scans=self._settings.nist_avg_scans,
                    )
                    peak_ms_cache[peak_num] = (mz_values, ms_intensities)
                except Exception as e:
                    self._logger.error("样品 %s 峰%d 质谱提取失败: %s", sample_name, peak_num, e)
                    if pim_predictor is not None:
                        result.pim_predictions[peak.retention_time] = PIMPrediction(
                            status="error",
                            message=f"质谱提取失败: {e}",
                        )
                    continue

                if pim_predictor is None:
                    continue

                try:
                    prediction = pim_predictor.predict(mz_values, ms_intensities)
                    result.pim_predictions[peak.retention_time] = prediction
                except Exception as e:
                    self._logger.error("样品 %s 峰%d PIM 预测失败: %s", sample_name, peak_num, e)
                    result.pim_predictions[peak.retention_time] = PIMPrediction(
                        status="error",
                        message=f"PIM 预测失败: {e}",
                    )

        # SS-HM / iHS-HM 预测 (复用已缓存的峰质谱)
        if mspepsearch_predictor is not None and mspepsearch_predictor.available:
            enable_sshm_search = self._settings.process_gc_ms_enable_sshm_search
            enable_ihshm_search = self._settings.process_gc_ms_enable_ihshm_search
            for peak_num, peak in enumerate(result.tic_peaks, start=1):
                cached_spectrum = peak_ms_cache.get(peak_num)
                if cached_spectrum is None:
                    continue

                mz_vals, ms_ints = cached_spectrum

                # SS-HM
                if enable_sshm_search is True:
                    try:
                        sshm_pred = mspepsearch_predictor.predict_sshm(mz_vals, ms_ints)
                        result.sshm_predictions[peak.retention_time] = sshm_pred
                    except Exception as e:
                        self._logger.error("样品 %s 峰%d SS-HM 预测失败: %s", sample_name, peak_num, e)
                        result.sshm_predictions[peak.retention_time] = SSHMPrediction(
                            status="error", message=f"SS-HM 预测失败: {e}",
                        )

                # iHS-HM
                if enable_ihshm_search is True:
                    try:
                        ihshm_pred = mspepsearch_predictor.predict_ihshm(mz_vals, ms_ints)
                        result.ihshm_predictions[peak.retention_time] = ihshm_pred
                    except Exception as e:
                        self._logger.error("样品 %s 峰%d iHS-HM 预测失败: %s", sample_name, peak_num, e)
                        result.ihshm_predictions[peak.retention_time] = iHSHMPrediction(
                            status="error", message=f"iHS-HM 预测失败: {e}",
                        )

        # 生成色谱图 (TIC + FID)
        if report_dir is not None:
            plotter = ChromatogramPlotter(
                chromatogram_ppi=self._settings.chromatogram_plot_ppi,
                ms_spectrum_ppi=self._settings.ms_spectrum_plot_ppi,
            )
            plot_dir = report_dir / "plots"

            # 根据积分模式决定填充基线方式, 使绘图区域与实际积分一致
            mode = self._settings.integration_mode.strip().lower()
            if mode == "gcpy" or mode == "robust_v3":
                fill_mode = "global"
            elif mode == "legacy" and self._settings.use_als_baseline is True:
                fill_mode = "global"
            else:
                # legacy 无 ALS 时, 积分使用局部端点连线.
                fill_mode = "local"

            # TIC 色谱图
            if tic_times is not None and result.tic_peaks:
                try:
                    tic_plot = plot_dir / f"{sample_name}_tic.png"
                    result.tic_plot_path = plotter.plot_chromatogram(
                        tic_times, tic_intensities, result.tic_peaks,
                        compound_matches=(result.compound_matches or None) if self._settings.tic_plot_show_compound else None,
                        title=f"TIC Chromatogram - {sample_name}",
                        ylabel="TIC Intensity",
                        output_path=tic_plot,
                        rt_min=self._settings.peak_rt_min,
                        rt_max=self._settings.peak_rt_max,
                        baseline=tic_baseline,
                        fill_baseline_mode=fill_mode,
                    )
                except Exception as e:
                    self._logger.error("样品 %s TIC 色谱图生成失败: %s", sample_name, e)

            # FID 色谱图
            if fid_times is not None and result.fid_peaks:
                try:
                    fid_plot = plot_dir / f"{sample_name}_fid.png"
                    result.fid_plot_path = plotter.plot_chromatogram(
                        fid_times, fid_intensities, result.fid_peaks,
                        compound_matches=None,  # FID 不标注化合物
                        title=f"FID Chromatogram - {sample_name}",
                        ylabel="FID Signal",
                        output_path=fid_plot,
                        rt_min=self._settings.peak_rt_min,
                        rt_max=self._settings.peak_rt_max,
                        y_range_min=100,  # FID 信号较小, 确保 Y 轴最小范围
                        baseline=fid_baseline,
                        fill_baseline_mode=fill_mode,
                    )
                except Exception as e:
                    self._logger.error("样品 %s FID 色谱图生成失败: %s", sample_name, e)

            # 各 TIC 峰的质谱图
            if result.tic_peaks:
                ms_plot_dir = report_dir / "ms_plots"
                for peak_num, peak in enumerate(result.tic_peaks, start=1):
                    try:
                        cached_spectrum = peak_ms_cache.get(peak_num)
                        if cached_spectrum is None:
                            mz, ms_intensities = reader.read_ms_spectra_at_peak(
                                d_dir, peak.start_time, peak.end_time,
                                avg_scans=self._settings.nist_avg_scans,
                            )
                        else:
                            mz, ms_intensities = cached_spectrum
                        if len(mz) > 0:
                            ms_plot_path = ms_plot_dir / f"{sample_name}_peak{peak_num}_ms.png"
                            plotter.plot_ms_spectrum(
                                mz, ms_intensities,
                                title=f"Mass Spectrum - {sample_name} Peak {peak_num} (RT {peak.retention_time:.2f} min)",
                                output_path=ms_plot_path,
                            )
                            result.ms_plot_paths[peak_num] = ms_plot_path
                    except Exception as e:
                        self._logger.error(
                            "样品 %s 峰 %d 质谱图生成失败: %s", sample_name, peak_num, e
                        )

        return result

    def process_gc_ms_results(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            GC-MS 运行完成后的结果处理入口:
            1. 定位任务目录, 找到所有 .D 结果文件.
            2. 逐个读取 TIC/FID 数据并积分.
            3. 读取/匹配 NIST 定性结果.
            4. 按任务汇总生成 Excel 报告.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
        返回:
            Dict: {"success": bool, "return_info": str, "report_path": str}.
        """
        try:
            # 定位任务(取 resolved_id, task_dir 不直接使用)
            _, resolved_id = self._find_task_dir(task_id)
            self._logger.info("开始处理任务 %s 的 GC-MS 结果", resolved_id)

            # 枚举 .D 目录
            d_dirs = self._enumerate_d_dirs(resolved_id)
            if not d_dirs:
                return {
                    "success": False,
                    "return_info": f"任务 {resolved_id} 未找到 .D 结果目录",
                }

            # 初始化 NIST 匹配器 (复用同一实例)
            nist = NISTMatcher(
                nist_path=self._settings.nist_path,
                max_hits=self._settings.nist_max_hits,
                search_timeout=self._settings.nist_search_timeout,
            )

            # 本地报告目录 (色谱图也保存在此)
            local_report_dir = self._settings.report_dir / resolved_id

            # 初始化 MSPepSearch 预测器 (SS-HM / iHS-HM)
            mspepsearch_predictor: Optional[MSPepSearchPredictor] = None
            if (
                self._settings.mspepsearch_enable is True
                and (
                    self._settings.process_gc_ms_enable_sshm_search is True
                    or self._settings.process_gc_ms_enable_ihshm_search is True
                )
            ):
                try:
                    # PIMPredictor 供 MSPepSearchPredictor 内部复用
                    pim_for_mspep = PIMPredictor(
                        ab_m=self._settings.pim_ab_m,
                        beta=self._settings.pim_beta,
                        epsilon_f=self._settings.pim_epsilon_f,
                    )
                    # 库谱读取器 (SS-HM 完整版需要, 不可用时自动降级)
                    nist_lib_reader = NistLibraryReader(self._settings.nist_mainlib_msp)

                    mspepsearch_predictor = MSPepSearchPredictor(
                        mspepsearch_exe=self._settings.mspepsearch_exe,
                        lib_path=self._settings.mspepsearch_lib_path,
                        lib_type=self._settings.mspepsearch_lib_type,
                        pim_predictor=pim_for_mspep,
                        nist_lib_reader=nist_lib_reader,
                        work_dir=local_report_dir,
                        sshm_hits=self._settings.sshm_hits,
                        sshm_b_ss=self._settings.sshm_b_ss,
                        ihshm_hits=self._settings.ihshm_hits,
                        ihshm_mEMF=self._settings.ihshm_mEMF,
                        timeout=self._settings.mspepsearch_timeout,
                    )
                    if mspepsearch_predictor.available:
                        self._logger.info("MSPepSearch 预测器初始化成功")
                    else:
                        self._logger.warning("MSPepSearch 预测器不可用, 将跳过 SS-HM/iHS-HM 预测")
                except Exception as exc:
                    self._logger.warning("MSPepSearch 预测器初始化失败, 将跳过 SS-HM/iHS-HM: %s", exc)
                    mspepsearch_predictor = None
            else:
                self._logger.info(
                    "已跳过 SS-HM/iHS-HM 预测, mspepsearch_enable=%s, process_gc_ms_enable_sshm_search=%s, process_gc_ms_enable_ihshm_search=%s",
                    self._settings.mspepsearch_enable,
                    self._settings.process_gc_ms_enable_sshm_search,
                    self._settings.process_gc_ms_enable_ihshm_search,
                )

            # 逐样品处理
            sample_results: List[SampleResult] = []
            for d_dir in d_dirs:
                self._logger.info("处理样品: %s", d_dir.name)
                sr = self._process_single_sample(
                    d_dir, nist, report_dir=local_report_dir,
                    mspepsearch_predictor=mspepsearch_predictor,
                )
                sample_results.append(sr)

            # 收集所有命中项并按本地索引批量获取结构图.
            structure_images = {}
            all_matches = []
            for sr in sample_results:
                for match_list in sr.compound_matches.values():
                    for m in match_list:
                        all_matches.append(m)

            if all_matches:
                try:
                    fetcher = NistLocalStructureFetcher(
                        task_cache_dir=local_report_dir / "structures",
                        seed_msp_path=self._settings.nist_structure_seed_msp,
                        seed_mol_dir=self._settings.nist_structure_seed_mol_dir,
                        runtime_cache_path=self._settings.nist_structure_runtime_cache_path,
                        offline_only=self._settings.structure_offline_only,
                        global_cache_dir=self._settings.structure_cache_dir,
                        image_ppi=self._settings.structure_image_ppi,
                        image_size=self._settings.structure_image_size,
                    )
                    structure_images = fetcher.fetch_batch_from_matches(all_matches)
                except Exception as e:
                    self._logger.warning("化合物结构图获取失败, 报告将不含结构图: %s", e)

            # 生成 Excel 报告
            generator = ReportGenerator()

            # 确定各预测方法是否启用, 控制报告中对应列的显示
            pim_enabled = self._settings.pim_enable
            sshm_enabled = (
                self._settings.mspepsearch_enable is True
                and self._settings.process_gc_ms_enable_sshm_search is True
            )
            ihshm_enabled = (
                self._settings.mspepsearch_enable is True
                and self._settings.process_gc_ms_enable_ihshm_search is True
            )

            # 保存到本地数据目录
            report_path = generator.generate_task_report(
                resolved_id, sample_results, local_report_dir,
                structure_images=structure_images,
                nist_max_hits=self._settings.nist_max_hits,
                alignment_tolerance=self._settings.alignment_tolerance,
                include_tic_only=self._settings.alignment_include_tic_only,
                include_fid_only=self._settings.alignment_include_fid_only,
                pim_enabled=pim_enabled,
                sshm_enabled=sshm_enabled,
                ihshm_enabled=ihshm_enabled,
            )

            # 同步到合成任务目录
            syn_dir = self._settings.synthesis_tasks_dir / resolved_id
            if syn_dir.is_dir():
                syn_report = generator.generate_task_report(
                    resolved_id, sample_results, syn_dir,
                    structure_images=structure_images,
                    nist_max_hits=self._settings.nist_max_hits,
                    alignment_tolerance=self._settings.alignment_tolerance,
                    include_tic_only=self._settings.alignment_include_tic_only,
                    include_fid_only=self._settings.alignment_include_fid_only,
                    pim_enabled=pim_enabled,
                    sshm_enabled=sshm_enabled,
                    ihshm_enabled=ihshm_enabled,
                )
                self._logger.info("报告已同步至合成任务目录: %s", syn_report)

            # 统计摘要
            total_tic_peaks = sum(len(sr.tic_peaks) for sr in sample_results)
            total_fid_peaks = sum(len(sr.fid_peaks) for sr in sample_results)
            msg = (
                f"任务 {resolved_id} 结果处理完成: "
                f"{len(sample_results)} 个样品, "
                f"TIC 共 {total_tic_peaks} 个峰, "
                f"FID 共 {total_fid_peaks} 个峰, "
                f"报告: {report_path}"
            )
            self._logger.info(msg)

            # 尝试自动触发产率计算
            self._try_auto_yield_calculation(resolved_id)

            return {
                "success": True,
                "return_info": msg,
                "report_path": str(report_path),
            }

        except Exception as exc:
            msg = f"结果处理失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

    def calculate_yields(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            产率计算入口, 定位实验方案/积分报告/化学品清单后调用 YieldCalculator.
            实验方案中需包含 "GC产率计算" Sheet, 否则跳过.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
        返回:
            Dict: {"success": bool, "return_info": str, "report_path": str}.
        """
        try:
            _, resolved_id = self._find_task_dir(task_id)
            self._logger.info("开始产率计算, 任务: %s", resolved_id)

            # 定位文件
            syn_dir = self._settings.synthesis_tasks_dir / resolved_id
            plan_path = syn_dir / f"{resolved_id}_experiment_plan.xlsx"
            report_path = syn_dir / f"{resolved_id}_integration_report.xlsx"
            chemical_list_path = self._settings.chemical_list_path

            if not plan_path.exists():
                return {"success": False, "return_info": f"未找到实验方案: {plan_path}"}
            if not report_path.exists():
                # 尝试本地报告目录
                local_report = self._settings.report_dir / resolved_id / f"{resolved_id}_integration_report.xlsx"
                legacy_local_report = self._settings.report_dir / resolved_id / f"integration_report_{resolved_id}.xlsx"
                if local_report.exists():
                    report_path = local_report
                elif legacy_local_report.exists():
                    report_path = legacy_local_report
                else:
                    return {"success": False, "return_info": f"未找到积分报告: {report_path}"}
            if not chemical_list_path.exists():
                return {"success": False, "return_info": f"未找到化学品清单: {chemical_list_path}"}

            # 检查实验方案是否包含 "GC产率计算" Sheet
            wb_check = openpyxl.load_workbook(str(plan_path), data_only=True)
            has_yield_sheet = YIELD_CONFIG_SHEET_NAME in wb_check.sheetnames
            wb_check.close()
            if not has_yield_sheet:
                return {
                    "success": False,
                    "return_info": f"实验方案中未包含 '{YIELD_CONFIG_SHEET_NAME}' Sheet, 跳过产率计算",
                }

            # 执行产率计算
            sshm_enabled = (
                self._settings.mspepsearch_enable is True
                and self._settings.process_gc_ms_enable_sshm_search is True
            )
            ihshm_enabled = (
                self._settings.mspepsearch_enable is True
                and self._settings.process_gc_ms_enable_ihshm_search is True
            )
            calc = YieldCalculator(
                rt_tolerance=self._settings.yield_rt_tolerance,
                nist_mainlib_msp_path=self._settings.nist_mainlib_msp,
                pim_enabled=self._settings.pim_enable,
                sshm_enabled=sshm_enabled,
                ihshm_enabled=ihshm_enabled,
            )
            config, results = calc.process_task(plan_path, report_path, chemical_list_path)

            # 按自然顺序排序 (729-1, 729-2, ..., 729-10 而非字典序 729-1, 729-10, 729-2)
            results.sort(key=lambda r: [
                int(p) if p.isdigit() else p.lower()
                for p in re.split(r'(\d+)', r.sample_name)
            ])

            # 保存到合成任务目录
            yield_path = calc.generate_yield_report(resolved_id, config, results, syn_dir)

            # 同步到本地分析数据目录
            local_report_dir = self._settings.report_dir / resolved_id
            if local_report_dir != syn_dir:
                calc.generate_yield_report(resolved_id, config, results, local_report_dir)

            # 统计
            valid_count = sum(1 for r in results if r.yield_percent is not None)
            msg = f"任务 {resolved_id} 产率计算完成: {len(results)} 条结果, 其中 {valid_count} 条有效产率"
            self._logger.info(msg)

            return {
                "success": True,
                "return_info": msg,
                "report_path": str(yield_path),
            }

        except Exception as exc:
            msg = f"产率计算失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

    def _try_auto_yield_calculation(self, resolved_id: str) -> None:
        """
        功能:
            在积分报告生成后尝试自动触发产率计算.
            仅当实验方案包含 "GC产率计算" Sheet 时执行, 失败不影响积分报告.
        参数:
            resolved_id: 任务 ID 字符串.
        返回:
            无.
        """
        try:
            result = self.calculate_yields(resolved_id)
            if result["success"]:
                self._logger.info("自动产率计算成功: %s", result["return_info"])
            else:
                self._logger.info("自动产率计算跳过: %s", result["return_info"])
        except Exception as exc:
            self._logger.warning("自动产率计算失败(不影响积分报告): %s", exc)

    def poll_analysis_run(
        self, task_id: Optional[str] = None, poll_interval: float = 30.0
    ) -> Dict:
        """
        功能:
            基于文件监控的 GC-MS 分析任务轮询.
            1. 从 gc_ms.csv 读取预期样品列表, 确定总样品数和采集顺序.
            2. 循环检查数据目录中 .D 目录的出现情况.
            3. 对每个 .D 目录, 解析 AcqData/sample_info.xml 中的
               RunCompletedFlag 判断采集状态:
               - .D 目录不存在 -> "等待进样"
               - .D 存在但 RunCompletedFlag 非 True -> "采集中"
               - RunCompletedFlag 为 True -> "采集结束"
            4. 实时反馈每个样品的状态变更和整体进度.
            5. 当所有样品均为 "采集结束" 时, 调用 process_gc_ms_results 处理结果.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
            poll_interval: 轮询间隔(秒), 默认 30 秒.
        返回:
            Dict: process_gc_ms_results 的返回值, 包含 success/return_info/report_path.
        """
        # 解析任务 ID
        _, resolved_id = self._find_task_dir(task_id)

        # 从 gc_ms.csv 加载预期样品列表(按表格顺序)
        try:
            expected_samples = self._load_expected_samples(resolved_id)
        except (FileNotFoundError, ValueError) as exc:
            msg = f"加载预期样品列表失败: {exc}"
            self._logger.error(msg)
            return {"success": False, "return_info": msg}

        total = len(expected_samples)

        # 确定 .D 目录搜索路径: 优先远程仪器目录, 回退到本地数据目录
        remote_data_dir = self._settings.gc_ms_data_dir
        local_data_dir = self._settings.data_dir / resolved_id
        if remote_data_dir.exists():
            search_dir = remote_data_dir
        else:
            self._logger.warning(
                "远程数据目录不可达: %s, 回退到本地目录: %s",
                remote_data_dir, local_data_dir
            )
            search_dir = local_data_dir

        self._logger.info(
            "开始文件监控轮询, 任务 %s, 共 %d 个样品, 间隔 %.0f 秒, 监控目录: %s",
            resolved_id, total, poll_interval, search_dir
        )

        # 记录每个样品的上次状态, 用于检测状态变更
        prev_status: Dict[str, str] = {name: "" for name in expected_samples}
        prev_completed = -1  # -1 保证第一轮必然输出进度

        try:
            while True:
                completed_count = 0

                for idx, sample_name in enumerate(expected_samples, start=1):
                    d_dir = search_dir / f"{sample_name}.D"

                    # 判断当前样品采集状态
                    if not d_dir.exists():
                        status = "等待进样"
                    else:
                        flag = self._read_run_completed_flag(d_dir)
                        if flag is True:
                            status = "采集结束"
                        else:
                            status = "采集中"

                    # 状态变更时输出日志
                    if status != prev_status[sample_name]:
                        self._logger.info(
                            "[%d/%d] 样品 %s: %s -> %s",
                            idx, total, sample_name,
                            prev_status[sample_name] or "(初始)", status
                        )
                        prev_status[sample_name] = status

                    if status == "采集结束":
                        completed_count += 1

                # 进度有变化时才输出, 避免无变化时重复刷屏
                if completed_count != prev_completed:
                    self._logger.info(
                        "轮询进度: %d/%d 样品已完成采集", completed_count, total
                    )
                    prev_completed = completed_count

                # 全部采集完成, 进入结果处理
                if completed_count == total:
                    self._logger.info(
                        "任务 %s 全部 %d 个样品采集完成, 开始处理结果...",
                        resolved_id, total
                    )
                    return self.process_gc_ms_results(resolved_id)

                time.sleep(poll_interval)

        except KeyboardInterrupt:
            self._logger.info("轮询被用户中断")
            return {"success": False, "return_info": "轮询被用户中断"}

    # ------------------------------------------------------------------
    # 实验数据归档
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_copy(
        src: Path, dest: Path, logger: logging.Logger
    ) -> bool:
        """
        功能:
            安全复制单个文件, 自动创建目标父目录.
            失败时记录 warning 而非抛出异常, 保证归档流程不因单个文件中断.
        参数:
            src: 源文件路径.
            dest: 目标文件路径.
            logger: 日志记录器.
        返回:
            bool: 复制是否成功.
        """
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dest)
            return True
        except (OSError, shutil.Error) as exc:
            logger.warning("复制文件失败: %s -> %s, 原因: %s", src, dest, exc)
            return False

    @staticmethod
    def _select_plan_sheet(wb: openpyxl.Workbook):
        """
        功能:
            从实验方案工作簿中选取包含 '实验编号' 表头的工作表.
            优先尝试名为 '实验方案设定' 的工作表, 其次激活表, 最后遍历其余表.
        参数:
            wb: openpyxl Workbook 对象.
        返回:
            命中的 Worksheet; 未命中则回退到激活表.
        """
        preferred = "实验方案设定"
        candidates = []
        if preferred in wb.sheetnames:
            candidates.append(preferred)
        active_name = wb.active.title
        if active_name not in candidates:
            candidates.append(active_name)
        for name in wb.sheetnames:
            if name not in candidates:
                candidates.append(name)

        for name in candidates:
            ws = wb[name]
            # 在前 5 行中查找 '实验编号' 关键词
            for row in range(1, min(ws.max_row + 1, 6)):
                for col in range(1, min(ws.max_column + 1, 20)):
                    val = ws.cell(row, col).value
                    if val is not None and "实验编号" in str(val):
                        return ws
        # 未命中则回退到激活表
        return wb.active

    def _collect_experiment_plan(
        self,
        task_id: str,
        syn_task_dir: Path,
        dest_dir: Path,
    ) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
        """
        功能:
            从合成站任务目录复制实验方案 Excel 到归档目录的 experiment_plan/ 子目录.
        参数:
            task_id: 任务 ID.
            syn_task_dir: 合成站任务目录路径.
            dest_dir: 归档目标根目录 ({task_id}/).
        返回:
            Tuple[Dict, List]:
                第一个值为成功复制的文件清单 {"experiment_plan": "relative/path"}.
                第二个值为缺失文件记录列表.
        """
        copied: Dict[str, str] = {}
        missing: List[Dict[str, str]] = []
        plan_dir = dest_dir / "experiment_plan"
        plan_dir.mkdir(parents=True, exist_ok=True)

        # 实验方案 Excel: 优先 {task_id}_experiment_plan.xlsx, 后备 {task_id}.xlsx
        plan_name = f"{task_id}_experiment_plan.xlsx"
        plan_src = syn_task_dir / plan_name
        if not plan_src.exists():
            plan_name = f"{task_id}.xlsx"
            plan_src = syn_task_dir / plan_name
        if plan_src.exists():
            rel = f"experiment_plan/{plan_name}"
            if self._safe_copy(plan_src, plan_dir / plan_name, self._logger):
                copied["experiment_plan"] = rel
                # 从实验方案 sheet 中提取实验名称
                try:
                    wb = openpyxl.load_workbook(plan_src, data_only=True)
                    ws = self._select_plan_sheet(wb)
                    # 扫描 A 列, 找到 "实验名称" 标签对应的 B 列值
                    exp_name = None
                    for r in range(1, min(ws.max_row + 1, 10)):
                        label = ws.cell(r, 1).value
                        if label is not None and "实验名称" in str(label):
                            exp_name = ws.cell(r, 2).value
                            break
                    wb.close()
                    if exp_name is not None and str(exp_name).strip() != "":
                        copied["experiment_name"] = str(exp_name).strip()
                except Exception:
                    self._logger.debug("无法从实验方案中提取实验名称")
        else:
            missing.append({
                "expected": f"{task_id}_experiment_plan.xlsx",
                "reason": "合成站任务目录中未找到实验方案 Excel",
            })

        return copied, missing

    def _detect_instruments(
        self,
        analysis_data_dir: Path,
    ) -> List[str]:
        """
        功能:
            检测分析站数据目录中存在哪些仪器的数据,
            通过检查对应 CSV 文件是否存在来判断.
        参数:
            analysis_data_dir: 分析站本地数据目录 (data/{task_id}/).
        返回:
            List[str]: 有数据的仪器标识列表, 如 ["gc_ms", "uplc_qtof"].
        """
        instruments: List[str] = []
        # 按 CSV 文件名判断仪器是否有数据
        instrument_csv_map = {
            "gc_ms": "gc_ms.csv",
            "uplc_qtof": "uplc_qtof.csv",
            "hplc": "hplc.csv",
        }
        for instrument, csv_name in instrument_csv_map.items():
            if (analysis_data_dir / csv_name).exists():
                instruments.append(instrument)
                self._logger.info("检测到 %s 数据: %s", instrument, csv_name)

        if not instruments:
            # 若无 CSV, 通过 plots/ 目录推断是否有 GC-MS 数据
            plots_dir = analysis_data_dir / "plots"
            if plots_dir.exists() and list(plots_dir.glob("*_tic.png")):
                instruments.append("gc_ms")
                self._logger.info("通过 plots/ 目录推断存在 GC-MS 数据")

        return instruments

    def _collect_analysis_data(
        self,
        task_id: str,
        analysis_data_dir: Path,
        dest_dir: Path,
        instruments: List[str],
    ) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
        """
        功能:
            按仪器收集分析数据到 analysis_data/{instrument}/ 子目录,
            包括 CSV 清单, 积分报告, 色谱图, 质谱图和结构图.
        参数:
            task_id: 任务 ID.
            analysis_data_dir: 分析站本地数据目录.
            dest_dir: 归档目标根目录 ({task_id}/).
            instruments: 有数据的仪器标识列表.
        返回:
            Tuple[Dict, List]:
                第一个值为按仪器组织的归档信息, 如 {"gc_ms": {"csv": ..., "samples": [...], ...}}.
                第二个值为缺失文件记录列表.
        """
        result: Dict[str, Any] = {}
        missing: List[Dict[str, str]] = []

        for instrument in instruments:
            inst_dir = dest_dir / "analysis_data" / instrument
            inst_dir.mkdir(parents=True, exist_ok=True)
            inst_info: Dict[str, Any] = {}

            if instrument == "gc_ms":
                inst_info, inst_missing = self._collect_gc_ms_analysis(
                    task_id, analysis_data_dir, inst_dir
                )
                missing.extend(inst_missing)
            elif instrument == "uplc_qtof":
                inst_info, inst_missing = self._collect_uplc_qtof_analysis(
                    task_id, analysis_data_dir, inst_dir
                )
                missing.extend(inst_missing)
            elif instrument == "hplc":
                inst_info, inst_missing = self._collect_hplc_analysis(
                    task_id, analysis_data_dir, inst_dir
                )
                missing.extend(inst_missing)

            result[instrument] = inst_info

        return result, missing

    def _collect_gc_ms_analysis(
        self,
        task_id: str,
        analysis_data_dir: Path,
        inst_dir: Path,
    ) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
        """
        功能:
            收集 GC-MS 仪器的分析数据: CSV 清单, 积分报告, 色谱图, 质谱图, 结构图.
        参数:
            task_id: 任务 ID.
            analysis_data_dir: 分析站本地数据目录.
            inst_dir: 目标仪器子目录 (analysis_data/gc_ms/).
        返回:
            Tuple[Dict, List]:
                第一个值为 GC-MS 归档信息字典.
                第二个值为缺失文件记录列表.
        """
        info: Dict[str, Any] = {}
        missing: List[Dict[str, str]] = []

        # gc_ms.csv
        csv_src = analysis_data_dir / "gc_ms.csv"
        if csv_src.exists():
            if self._safe_copy(csv_src, inst_dir / "gc_ms.csv", self._logger):
                info["csv"] = "analysis_data/gc_ms/gc_ms.csv"
        else:
            missing.append({"expected": "gc_ms.csv", "reason": "GC-MS 样品清单缺失"})

        # 积分报告
        report_name = f"{task_id}_integration_report.xlsx"
        report_src = analysis_data_dir / report_name
        if report_src.exists():
            if self._safe_copy(report_src, inst_dir / report_name, self._logger):
                info["integration_report"] = f"analysis_data/gc_ms/{report_name}"
        else:
            missing.append({"expected": report_name, "reason": "GC-MS 积分报告缺失"})

        # 解析样品列表
        sample_names: List[str] = []
        if csv_src.exists():
            with csv_src.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                sample_names = [row["SampleName"] for row in reader]
        if not sample_names:
            # 从 plots/ 目录推断
            plots_dir = analysis_data_dir / "plots"
            if plots_dir.exists():
                for png in sorted(plots_dir.glob("*_tic.png"), key=_natural_sort_key):
                    name = png.stem.replace("_tic", "")
                    if name not in sample_names:
                        sample_names.append(name)
        info["sample_names"] = sample_names

        # 色谱图
        chrom_dir = inst_dir / "chromatograms"
        chrom_dir.mkdir(parents=True, exist_ok=True)
        tic_count = 0
        fid_count = 0
        for sample_name in sample_names:
            # TIC
            tic_name = f"{sample_name}_tic.png"
            tic_src = analysis_data_dir / "plots" / tic_name
            if tic_src.exists():
                if self._safe_copy(tic_src, chrom_dir / tic_name, self._logger):
                    tic_count += 1
            else:
                missing.append({
                    "expected": tic_name,
                    "reason": f"样品 {sample_name} 的 TIC 色谱图缺失",
                })
            # FID
            fid_name = f"{sample_name}_fid.png"
            fid_src = analysis_data_dir / "plots" / fid_name
            if fid_src.exists():
                if self._safe_copy(fid_src, chrom_dir / fid_name, self._logger):
                    fid_count += 1
            else:
                missing.append({
                    "expected": fid_name,
                    "reason": f"样品 {sample_name} 的 FID 色谱图缺失",
                })
        info["tic_count"] = tic_count
        info["fid_count"] = fid_count

        # 质谱图
        ms_dir = inst_dir / "ms_spectra"
        ms_dir.mkdir(parents=True, exist_ok=True)
        ms_count = 0
        ms_plots_dir = analysis_data_dir / "ms_plots"
        if ms_plots_dir.exists():
            for sample_name in sample_names:
                ms_pattern = f"{sample_name}_peak*_ms.png"
                for ms_src in sorted(ms_plots_dir.glob(ms_pattern), key=_natural_sort_key):
                    if self._safe_copy(ms_src, ms_dir / ms_src.name, self._logger):
                        ms_count += 1
        info["ms_count"] = ms_count

        # 结构图
        struct_dir = inst_dir / "structures"
        struct_dir.mkdir(parents=True, exist_ok=True)
        struct_count = 0
        src_struct_dir = analysis_data_dir / "structures"
        if src_struct_dir.exists():
            for struct_src in sorted(src_struct_dir.glob("*.png")):
                if self._safe_copy(struct_src, struct_dir / struct_src.name, self._logger):
                    struct_count += 1
        info["struct_count"] = struct_count

        return info, missing

    def _collect_uplc_qtof_analysis(
        self,
        task_id: str,
        analysis_data_dir: Path,
        inst_dir: Path,
    ) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
        """
        功能:
            收集 UPLC-QTOF 仪器的分析数据.
            当前仅收集 CSV 清单, 图片相关功能预留.
        参数:
            task_id: 任务 ID.
            analysis_data_dir: 分析站本地数据目录.
            inst_dir: 目标仪器子目录 (analysis_data/uplc_qtof/).
        返回:
            Tuple[Dict, List]:
                第一个值为 UPLC-QTOF 归档信息字典.
                第二个值为缺失文件记录列表.
        """
        info: Dict[str, Any] = {}
        missing: List[Dict[str, str]] = []

        # uplc_qtof.csv
        csv_src = analysis_data_dir / "uplc_qtof.csv"
        if csv_src.exists():
            if self._safe_copy(csv_src, inst_dir / "uplc_qtof.csv", self._logger):
                info["csv"] = "analysis_data/uplc_qtof/uplc_qtof.csv"
        else:
            missing.append({"expected": "uplc_qtof.csv", "reason": "UPLC-QTOF 样品清单缺失"})

        # 解析样品列表
        sample_names: List[str] = []
        if csv_src.exists():
            with csv_src.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                sample_names = [row["SampleName"] for row in reader]
        info["sample_names"] = sample_names

        # 预留: 色谱图, 质谱图, 结构图子目录
        for sub in ("chromatograms", "ms_spectra", "structures"):
            (inst_dir / sub).mkdir(parents=True, exist_ok=True)

        return info, missing

    def _collect_hplc_analysis(
        self,
        task_id: str,
        analysis_data_dir: Path,
        inst_dir: Path,
    ) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
        """
        功能:
            收集 HPLC 仪器的分析数据.
            当前仅收集 CSV 清单, 图片相关功能预留.
        参数:
            task_id: 任务 ID.
            analysis_data_dir: 分析站本地数据目录.
            inst_dir: 目标仪器子目录 (analysis_data/hplc/).
        返回:
            Tuple[Dict, List]:
                第一个值为 HPLC 归档信息字典.
                第二个值为缺失文件记录列表.
        """
        info: Dict[str, Any] = {}
        missing: List[Dict[str, str]] = []

        # hplc.csv
        csv_src = analysis_data_dir / "hplc.csv"
        if csv_src.exists():
            if self._safe_copy(csv_src, inst_dir / "hplc.csv", self._logger):
                info["csv"] = "analysis_data/hplc/hplc.csv"
        else:
            missing.append({"expected": "hplc.csv", "reason": "HPLC 样品清单缺失"})

        # 解析样品列表
        sample_names: List[str] = []
        if csv_src.exists():
            with csv_src.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                sample_names = [row["SampleName"] for row in reader]
        info["sample_names"] = sample_names

        return info, missing

    def _collect_results(
        self,
        task_id: str,
        syn_task_dir: Path,
        analysis_data_dir: Path,
        dest_dir: Path,
    ) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
        """
        功能:
            收集结果报告到 results/ 子目录,
            包括任务报告 (合成站) 和产率报告 (分析站).
        参数:
            task_id: 任务 ID.
            syn_task_dir: 合成站任务目录路径.
            analysis_data_dir: 分析站本地数据目录.
            dest_dir: 归档目标根目录 ({task_id}/).
        返回:
            Tuple[Dict, List]:
                第一个值为成功复制的文件清单.
                第二个值为缺失文件记录列表.
        """
        copied: Dict[str, str] = {}
        missing: List[Dict[str, str]] = []
        results_dir = dest_dir / "results"
        results_dir.mkdir(parents=True, exist_ok=True)

        # 任务报告 (来自合成站)
        report_name = f"{task_id}_task_report.xlsx"
        report_src = syn_task_dir / report_name
        if report_src.exists():
            if self._safe_copy(report_src, results_dir / report_name, self._logger):
                copied["task_report"] = f"results/{report_name}"
        else:
            missing.append({
                "expected": report_name,
                "reason": "合成站任务目录中未找到任务报告",
            })

        # 产率报告 (来自分析站)
        yield_name = f"{task_id}_yield_report.xlsx"
        yield_src = analysis_data_dir / yield_name
        if yield_src.exists():
            if self._safe_copy(yield_src, results_dir / yield_name, self._logger):
                copied["yield_report"] = f"results/{yield_name}"
        else:
            missing.append({
                "expected": yield_name,
                "reason": "分析站数据目录中未找到产率报告",
            })

        return copied, missing

    def _collect_raw_data(
        self,
        task_id: str,
        dest_dir: Path,
        instrument_samples: Dict[str, List[str]],
    ) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """
        功能:
            从各仪器数据目录复制 .D 原始数据到 raw_data/{instrument}/ 子目录.
        参数:
            task_id: 任务 ID.
            dest_dir: 归档目标根目录 ({task_id}/).
            instrument_samples: 按仪器分组的样品名列表,
                如 {"gc_ms": ["771-1", "771-2"], "uplc_qtof": ["771-3"]}.
        返回:
            Tuple[List, List]:
                第一个值为成功复制的原始数据记录列表.
                第二个值为缺失文件记录列表.
        """
        copied: List[Dict[str, str]] = []
        missing: List[Dict[str, str]] = []

        # 仪器标识 -> 仪器数据目录的映射
        instrument_data_dirs = {
            "gc_ms": self._settings.gc_ms_data_dir,
            "uplc_qtof": self._settings.uplc_qtof_data_dir,
            "hplc": self._settings.hplc_data_dir,
        }

        for instrument, sample_names in instrument_samples.items():
            data_dir = instrument_data_dirs.get(instrument)
            if data_dir is None:
                self._logger.warning("未知仪器类型: %s, 跳过原始数据收集", instrument)
                continue

            raw_dir = dest_dir / "raw_data" / instrument
            raw_dir.mkdir(parents=True, exist_ok=True)

            for sample_name in sample_names:
                d_dir_name = f"{sample_name}.D"
                d_src = data_dir / d_dir_name
                d_dest = raw_dir / d_dir_name
                if d_src.exists() and d_src.is_dir():
                    try:
                        if d_dest.exists():
                            shutil.rmtree(d_dest)
                        shutil.copytree(d_src, d_dest)
                        copied.append({
                            "instrument": instrument,
                            "sample_name": sample_name,
                            "path": f"raw_data/{instrument}/{d_dir_name}",
                        })
                        self._logger.info("已复制原始数据: %s/%s", instrument, d_dir_name)
                    except (OSError, shutil.Error) as exc:
                        self._logger.warning(
                            "复制原始数据失败: %s -> %s, 原因: %s", d_src, d_dest, exc
                        )
                        missing.append({
                            "expected": d_dir_name,
                            "reason": f"{instrument} 原始数据复制失败: {exc}",
                        })
                else:
                    missing.append({
                        "expected": d_dir_name,
                        "reason": f"{instrument} 仪器目录 {data_dir} 中未找到",
                    })

        return copied, missing

    def _build_raw_data_references(
        self,
        instrument_samples: Dict[str, List[str]],
    ) -> Dict[str, List[Dict[str, str]]]:
        """
        功能:
            构建各仪器原始数据路径引用, 记录各样品 .D 目录的原始位置.
            不论是否复制原始数据, 均记录以便溯源.
        参数:
            instrument_samples: 按仪器分组的样品名列表.
        返回:
            Dict[str, List]: 按仪器分组的引用列表.
        """
        instrument_data_dirs = {
            "gc_ms": self._settings.gc_ms_data_dir,
            "uplc_qtof": self._settings.uplc_qtof_data_dir,
            "hplc": self._settings.hplc_data_dir,
        }

        refs: Dict[str, List[Dict[str, str]]] = {}
        for instrument, sample_names in instrument_samples.items():
            data_dir = instrument_data_dirs.get(instrument)
            if data_dir is None:
                continue
            inst_refs: List[Dict[str, str]] = []
            for sample_name in sample_names:
                d_dir_name = f"{sample_name}.D"
                inst_refs.append({
                    "sample_name": sample_name,
                    "d_dir": str(data_dir / d_dir_name),
                })
            refs[instrument] = inst_refs

        return refs

    def _write_archive_summary_txt(
        self,
        task_id: str,
        dest_dir: Path,
        experiment_plan: Dict[str, str],
        analysis_info: Dict[str, Any],
        results_files: Dict[str, str],
        raw_refs: Dict[str, List[Dict[str, str]]],
        missing_files: List[Dict[str, str]],
        raw_copied: Optional[List[Dict[str, str]]] = None,
    ) -> Path:
        """
        功能:
            在任务归档目录内生成 archive_summary.txt,
            按四大类 (实验计划/分析数据/结果报告/原始数据) 汇总.
        参数:
            task_id: 任务 ID.
            dest_dir: 归档目标目录 ({task_id}/).
            experiment_plan: 实验计划已归档文件清单.
            analysis_info: 按仪器组织的分析数据信息.
            results_files: 结果报告已归档文件清单.
            raw_refs: 按仪器分组的原始数据引用.
            missing_files: 缺失文件记录列表.
            raw_copied: 已复制的原始数据记录列表, None 表示未执行复制.
        返回:
            Path: 生成的 TXT 文件路径.
        """
        txt_path = dest_dir / "archive_summary.txt"
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        instrument_labels = {
            "gc_ms": "GC-MS",
            "uplc_qtof": "UPLC-QTOF",
            "hplc": "HPLC",
        }

        lines: List[str] = []
        lines.append(f"实验数据归档清单 - 任务 {task_id}")
        lines.append(f"归档时间: {now_str}")
        lines.append(f"归档目录: {dest_dir}")
        # 实验名称 (从实验方案 Excel B1 单元格提取)
        exp_name = experiment_plan.get("experiment_name")
        if exp_name:
            lines.append(f"实验名称: {exp_name}")
        lines.append("")

        # 一. 实验计划
        lines.append("=" * 40)
        lines.append("实验计划")
        lines.append("=" * 40)
        if "experiment_plan" in experiment_plan:
            lines.append(f"[实验方案] {experiment_plan['experiment_plan']}")
        else:
            lines.append("(无)")
        lines.append("")

        # 二. 分析数据
        lines.append("=" * 40)
        lines.append("分析数据")
        lines.append("=" * 40)
        if analysis_info:
            for instrument, inst_data in analysis_info.items():
                label = instrument_labels.get(instrument, instrument)
                lines.append(f"--- {label} ---")
                if "csv" in inst_data:
                    lines.append(f"[样品清单] {inst_data['csv']}")
                if "integration_report" in inst_data:
                    lines.append(f"[积分报告] {inst_data['integration_report']}")
                tic_count = inst_data.get("tic_count", 0)
                fid_count = inst_data.get("fid_count", 0)
                ms_count = inst_data.get("ms_count", 0)
                struct_count = inst_data.get("struct_count", 0)
                if tic_count > 0:
                    lines.append(
                        f"[TIC色谱图] analysis_data/{instrument}/chromatograms/ "
                        f"... 共 {tic_count} 个"
                    )
                if fid_count > 0:
                    lines.append(
                        f"[FID色谱图] analysis_data/{instrument}/chromatograms/ "
                        f"... 共 {fid_count} 个"
                    )
                if ms_count > 0:
                    lines.append(
                        f"[质谱图] analysis_data/{instrument}/ms_spectra/ "
                        f"... 共 {ms_count} 个"
                    )
                if struct_count > 0:
                    lines.append(
                        f"[结构图] analysis_data/{instrument}/structures/ "
                        f"... 共 {struct_count} 个"
                    )
                # 仅有 CSV 无图片数据的仪器, 显示样品数
                sample_count = len(inst_data.get("sample_names", []))
                if sample_count > 0 and tic_count == 0 and ms_count == 0:
                    lines.append(f"[样品数] {sample_count} 个")
                lines.append("")
        else:
            lines.append("(无)")
            lines.append("")

        # 三. 结果报告
        lines.append("=" * 40)
        lines.append("结果报告")
        lines.append("=" * 40)
        if "task_report" in results_files:
            lines.append(f"[任务报告] {results_files['task_report']}")
        if "yield_report" in results_files:
            lines.append(f"[产率报告] {results_files['yield_report']}")
        if not results_files:
            lines.append("(无)")
        lines.append("")

        # 四. 原始数据
        lines.append("=" * 40)
        lines.append("原始数据")
        lines.append("=" * 40)
        has_raw_info = False
        # 已复制的原始数据
        if raw_copied is not None and len(raw_copied) > 0:
            # 按仪器分组统计
            inst_counts: Dict[str, int] = {}
            for r in raw_copied:
                inst = r["instrument"]
                inst_counts[inst] = inst_counts.get(inst, 0) + 1
            for inst, count in inst_counts.items():
                label = instrument_labels.get(inst, inst)
                lines.append(
                    f"[{label}] raw_data/{inst}/ ... 共 {count} 个 .D 目录 (已复制)"
                )
            has_raw_info = True
        # 原始数据引用
        if raw_refs:
            for instrument, inst_refs in raw_refs.items():
                label = instrument_labels.get(instrument, instrument)
                if inst_refs:
                    first_ref = inst_refs[0]["d_dir"]
                    lines.append(f"[{label}] {first_ref} ... 共 {len(inst_refs)} 个")
            has_raw_info = True
        if not has_raw_info:
            lines.append("(无)")
        lines.append("")

        # 缺失文件
        lines.append("=" * 40)
        lines.append("缺失文件")
        lines.append("=" * 40)
        if missing_files:
            for item in missing_files:
                lines.append(f"  {item['expected']} - {item['reason']}")
        else:
            lines.append("(无)")
        lines.append("")

        txt_path.parent.mkdir(parents=True, exist_ok=True)
        txt_path.write_text("\n".join(lines), encoding="utf-8")
        self._logger.info("归档清单已写入: %s", txt_path)
        return txt_path

    def aggregate_task_data(
        self,
        task_id: Optional[str] = None,
        archive_dir: Optional[Path] = None,
        copy_raw_data: Optional[bool] = None,
    ) -> Dict:
        """
        功能:
            将指定任务的全部实验数据从合成站和分析站汇聚到统一归档目录.
            按四大类组织: 实验计划, 分析数据 (按仪器划分), 结果报告, 原始数据.
        参数:
            task_id: 任务 ID 字符串, None 表示自动选取最新任务.
            archive_dir: 归档输出根目录, None 则使用 settings.archive_dir.
            copy_raw_data: 是否将 .D 原始数据目录复制到归档,
                None 则使用 settings.archive_copy_raw_data.
        返回:
            Dict: {"success": bool, "return_info": str, "archive_path": str}.
        """
        try:
            # 1. 定位任务目录
            syn_task_dir, resolved_id = self._find_task_dir(task_id)
            self._logger.info("开始归档任务 %s ...", resolved_id)

            # 2. 确定源目录
            analysis_data_dir = self._settings.data_dir / resolved_id

            # 3. 确定归档目标目录
            target_archive_dir = archive_dir if archive_dir is not None else self._settings.archive_dir
            dest_dir = target_archive_dir / resolved_id
            dest_dir.mkdir(parents=True, exist_ok=True)
            self._logger.info("归档目标目录: %s", dest_dir)

            all_missing: List[Dict[str, str]] = []

            # 4. 验证源目录
            if not syn_task_dir.exists():
                self._logger.warning("合成站任务目录不存在: %s", syn_task_dir)
            if not analysis_data_dir.exists():
                self._logger.warning("分析站数据目录不存在: %s", analysis_data_dir)

            # 5. 检测有数据的仪器
            instruments: List[str] = []
            if analysis_data_dir.exists():
                instruments = self._detect_instruments(analysis_data_dir)
            self._logger.info("检测到仪器: %s", instruments)

            # 6. 收集实验计划
            experiment_plan: Dict[str, str] = {}
            if syn_task_dir.exists():
                experiment_plan, plan_missing = self._collect_experiment_plan(
                    resolved_id, syn_task_dir, dest_dir
                )
                all_missing.extend(plan_missing)

            # 7. 收集分析数据 (按仪器划分)
            analysis_info: Dict[str, Any] = {}
            if analysis_data_dir.exists() and instruments:
                analysis_info, ana_missing = self._collect_analysis_data(
                    resolved_id, analysis_data_dir, dest_dir, instruments
                )
                all_missing.extend(ana_missing)

            # 8. 收集结果报告
            results_files: Dict[str, str] = {}
            if syn_task_dir.exists() or analysis_data_dir.exists():
                results_files, res_missing = self._collect_results(
                    resolved_id, syn_task_dir, analysis_data_dir, dest_dir
                )
                all_missing.extend(res_missing)

            # 9. 构建按仪器分组的样品名映射
            instrument_samples: Dict[str, List[str]] = {}
            for instrument, inst_data in analysis_info.items():
                sample_names = inst_data.get("sample_names", [])
                if sample_names:
                    instrument_samples[instrument] = sample_names

            # 10. 解析是否复制原始数据: 参数优先, 否则取配置
            should_copy_raw = copy_raw_data if copy_raw_data is not None else self._settings.archive_copy_raw_data

            # 11. 可选复制原始数据
            raw_copied: Optional[List[Dict[str, str]]] = None
            if should_copy_raw is True and instrument_samples:
                raw_copied, raw_missing = self._collect_raw_data(
                    resolved_id, dest_dir, instrument_samples
                )
                all_missing.extend(raw_missing)

            # 12. 构建原始数据引用
            raw_refs = self._build_raw_data_references(instrument_samples)

            # 13. 写入归档清单 TXT (位于 {task_id}/ 目录内)
            txt_path = self._write_archive_summary_txt(
                task_id=resolved_id,
                dest_dir=dest_dir,
                experiment_plan=experiment_plan,
                analysis_info=analysis_info,
                results_files=results_files,
                raw_refs=raw_refs,
                missing_files=all_missing,
                raw_copied=raw_copied,
            )

            # 14. 统计归档文件数量
            file_count = len(experiment_plan) + len(results_files)
            for inst_data in analysis_info.values():
                file_count += (1 if "csv" in inst_data else 0)
                file_count += (1 if "integration_report" in inst_data else 0)
                file_count += inst_data.get("tic_count", 0)
                file_count += inst_data.get("fid_count", 0)
                file_count += inst_data.get("ms_count", 0)
                file_count += inst_data.get("struct_count", 0)
            if raw_copied is not None:
                file_count += len(raw_copied)

            info = (
                f"任务 {resolved_id} 归档完成: "
                f"已归档 {file_count} 项, "
                f"缺失 {len(all_missing)} 项. "
                f"归档目录: {dest_dir}, "
                f"清单文件: {txt_path}"
            )
            self._logger.info(info)
            return {
                "success": True,
                "return_info": info,
                "archive_path": str(dest_dir),
            }

        except Exception as exc:
            self._logger.error("归档失败: %s", exc, exc_info=True)
            return {"success": False, "return_info": f"归档失败: {exc}"}


# ------------------------------------------------------------------
# 交互式测试入口
# ------------------------------------------------------------------

def _print_result(result: Dict) -> None:
    """
    功能:
        格式化打印函数返回结果.
    参数:
        result: 函数返回的字典.
    返回:
        无.
    """
    print("\n========== 执行结果 ==========")
    for key, value in result.items():
        print(f"  {key}: {value}")
    print("==============================\n")


def main() -> None:
    """
    功能:
        交互式菜单, 用于手动测试 run_analysis / process_gc_ms_results /
        poll_analysis_run / get_status / get_methods / calculate_yields /
        submit_by_csv_path / aggregate_task_data.
        用户可选择功能并输入 task_id, 输入 q 退出.
    参数:
        无.
    返回:
        无.
    """
    configure_logging("DEBUG")
    logger = logging.getLogger("main")
    logger.info("初始化分析站控制器...")

    controller = AnalysisStationController()

    menu = (
        "\n===== 分析站交互式测试菜单 =====\n"
        "  1. run_analysis          - 统一分析入口(生成CSV并提交至仪器)\n"
        "  2. process_gc_ms_results - GC-MS结果处理(积分+定性+报告)\n"
        "  3. poll_analysis_run     - 轮询GC-MS分析任务状态并自动处理结果\n"
        "  4. get_status            - 获取GC-MS设备当前状态\n"
        "  5. get_methods           - 获取当前Project的方法列表\n"
        "  6. calculate_yields      - 产率计算\n"
        "  7. submit_by_csv_path    - 选择仪器并按CSV路径直接提交任务\n"
        "  8. aggregate_task_data   - 实验数据归档汇总\n"
        "  9. transfer_to_shelf   - 分析完成样品→货架转运(等待空闲后自动执行)\n"
        "  0. 退出\n"
        "================================"
    )

    while True:
        print(menu)
        choice = input("请选择功能编号: ").strip()

        if choice == "0":
            print("已退出测试.")
            break

        if choice not in ("1", "2", "3", "4", "5", "6", "7", "8", "9"):
            print("无效选择, 请输入 0/1/2/3/4/5/6/7/8/9.")
            continue

        # 选项 4/5 直接操作设备驱动, 不需要 task_id
        if choice in ("4", "5"):
            settings = controller._settings
            client = ZhidaClient(
                host=settings.gc_ms_host,
                port=settings.gc_ms_port,
                timeout=settings.gc_ms_timeout,
            )
            try:
                client.connect()
                if choice == "4":
                    print("\n>>> 调用 ZhidaClient.get_status_detail()")
                    status_detail = client.get_status_detail()
                    raw_status = status_detail["raw_status"] or "(空)"
                    sub_status = status_detail["sub_status"] or "(无)"
                    print(
                        "\n"
                        f"  原始状态: {raw_status}\n"
                        f"  主状态: {status_detail['base_status']}\n"
                        f"  子状态: {sub_status}\n"
                    )
                else:
                    print("\n>>> 调用 ZhidaClient.get_methods()")
                    methods = client.get_methods()
                    _print_result(methods)
            except Exception as exc:
                logger.error("设备操作失败: %s", exc)
                print(f"\n  操作失败: {exc}\n")
            finally:
                client.close()
            continue

        if choice == "7":
            instrument_options = {
                "1": ("gc_ms", "GC-MS"),
                "2": ("uplc_qtof", "UPLC_QTOF"),
                "3": ("hplc", "HPLC"),
            }
            print("\n请选择仪器:")
            for option, (_, instrument_name) in instrument_options.items():
                print(f"  {option}. {instrument_name}")

            instrument_choice = input("请输入仪器编号(1/2/3): ").strip()
            if instrument_choice not in instrument_options:
                print("无效选择, 请输入 1/2/3.")
                continue

            instrument = instrument_options[instrument_choice][0]
            csv_file_path = input("请输入CSV文件路径: ").strip()
            print(
                f"\n>>> 调用 submit_by_csv_path("
                f"instrument={instrument!r}, csv_file_path={csv_file_path!r})"
            )
            result = controller.submit_by_csv_path(
                instrument=instrument, csv_file_path=csv_file_path
            )
            _print_result(result)
            continue

        # 获取 task_id, 空字符串视为 None(自动选取最新任务)
        task_id_input = input("请输入 task_id (留空则自动选取最新任务): ").strip()
        task_id = task_id_input if task_id_input else None

        if choice == "1":
            print(f"\n>>> 调用 run_analysis(task_id={task_id!r})")
            result = controller.run_analysis(task_id=task_id)
            _print_result(result)

        elif choice == "2":
            print(f"\n>>> 调用 process_gc_ms_results(task_id={task_id!r})")
            result = controller.process_gc_ms_results(task_id=task_id)
            _print_result(result)

        elif choice == "3":
            # poll_analysis_run 额外支持配置轮询间隔
            interval_input = input("请输入轮询间隔秒数 (留空默认30): ").strip()
            try:
                interval = float(interval_input) if interval_input else 30.0
            except ValueError:
                print("无效数值, 使用默认30秒.")
                interval = 30.0

            print(
                f"\n>>> 调用 poll_analysis_run(task_id={task_id!r}, "
                f"poll_interval={interval})"
            )
            result = controller.poll_analysis_run(
                task_id=task_id, poll_interval=interval
            )
            _print_result(result)

        elif choice == "6":
            print(f"\n>>> 调用 calculate_yields(task_id={task_id!r})")
            result = controller.calculate_yields(task_id=task_id)
            _print_result(result)

        elif choice == "8":
            default_copy = controller._settings.archive_copy_raw_data
            hint = "Y/n" if default_copy is True else "y/N"
            copy_raw_input = input(
                f"是否复制原始数据(.D目录)? ({hint}, 留空使用配置默认值): "
            ).strip().lower()
            if copy_raw_input == "":
                copy_raw: Optional[bool] = None  # 使用配置默认值
            else:
                copy_raw = copy_raw_input in ("y", "yes")
            print(
                f"\n>>> 调用 aggregate_task_data("
                f"task_id={task_id!r}, copy_raw_data={copy_raw})"
            )
            result = controller.aggregate_task_data(
                task_id=task_id, copy_raw_data=copy_raw
            )
            _print_result(result)

        elif choice == "9":
            # 分析完成样品→货架转运
            print("\n>>> 分析完成样品→货架转运")
            print("说明: 轮询智达进样设备状态, 等待空闲后将样品从分析站转运到货架空位\n")

            try:
                from unilabos.devices.eit_agv.controller.agv_controller import AGVController

                agv = AGVController(timeout=180000)

                # 显示当前货架状态
                agv.shelf_manager.print_status()

                # 询问源托盘
                print("默认源托盘: analysis_station_tray_1-2")
                source_input = input(
                    "请输入源托盘(多个用逗号分隔, 直接回车使用默认): "
                ).strip()

                if source_input == "":
                    source_trays = ["analysis_station_tray_1-2"]
                else:
                    source_trays = [
                        s.strip() for s in source_input.split(",") if s.strip() != ""
                    ]

                # 询问轮询间隔
                interval_input = input("请输入轮询间隔秒数 (留空默认30): ").strip()
                try:
                    interval = float(interval_input) if interval_input else 30.0
                except ValueError:
                    print("无效数值, 使用默认30秒.")
                    interval = 30.0

                print(f"\n源托盘: {source_trays}")
                print(f"轮询间隔: {interval} 秒")

                # 执行转运
                print("\n开始执行分析站→货架样品转运...")
                success = agv.transfer_analysis_to_shelf(
                    source_trays=source_trays,
                    poll_interval=interval,
                )

                _print_result({
                    "success": success,
                    "source_trays": source_trays,
                    "poll_interval": interval,
                })

                if success:
                    agv.shelf_manager.print_status()

            except Exception as exc:
                logger.error("分析站→货架转运失败: %s", exc)
                print(f"\n  操作失败: {exc}\n")


if __name__ == "__main__":
    main()

