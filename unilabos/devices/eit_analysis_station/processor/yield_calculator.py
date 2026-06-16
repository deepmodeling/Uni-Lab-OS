#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
功能:
    基于 GC-FID 积分报告和实验方案, 计算各样品的产率.
    支持 ECN 法, 标准曲线法和响应因子法三种计算方式.
    从 chemical_list.xlsx 自动推算内标摩尔量, 无需手动输入浓度.
参数:
    无.
返回:
    无.
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pysmiles import PTE
from pysmiles.read_smiles import read_smiles

from .ecn import smiles2carbontypes, ecn_dct, class_dct

try:
    from rdkit import Chem
    from rdkit.Chem import Descriptors
    from rdkit.Chem import inchi as RDKitInchi
except Exception:
    Chem = None
    Descriptors = None
    RDKitInchi = None

logger = logging.getLogger(__name__)
YIELD_CONFIG_SHEET_NAME = "GC产率计算"


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------

@dataclass
class TargetProduct:
    """
    功能:
        存储单个目标产物的配置信息.
    参数:
        name: 目标产物名称.
        smiles: SMILES 字符串.
        formula: 分子式 (从 SMILES 自动推导).
        ecn: 有效碳数 (从 SMILES 自动计算).
        expected_rt: 预期保留时间(min), None 表示使用分子式匹配.
        applicable_experiments: 适用实验编号列表, 如 [1,2,3].
        equivalent: 目标产物当量(eq), 默认1.0. 理论产物量 = 反应规模 * equivalent.
    返回:
        TargetProduct.
    """
    name: str = ""
    smiles: str = ""
    formula: str = ""
    molecular_weight: Optional[float] = None
    ecn: float = 0.0
    expected_rt: Optional[float] = None
    applicable_experiments: List[int] = field(default_factory=list)
    equivalent: float = 1.0
    nist_has_record: bool = False
    nist_query_mode: str = ""
    nist_reference_names: List[str] = field(default_factory=list)


@dataclass
class YieldCalcConfig:
    """
    功能:
        存储产率计算的完整配置, 从实验方案和 chemical_list 共同构建.
    参数:
        is_name: 内标名称 (从原参数 Sheet 的 "内标种类" 读取).
        is_smiles: 内标 SMILES (从 "GC产率计算" Sheet 读取).
        is_formula: 内标分子式 (自动推导).
        is_ecn: 内标 ECN (自动计算).
        is_expected_rt: 内标预期保留时间(min), 可选.
        is_amount: 内标用量原始值 (μL 或 mg).
        is_moles: 内标实际加入摩尔量(mol), 从 chemical_list 推算.
        reaction_scale_mmol: 反应规模(mmol).
        calc_method: 产率计算方法, "ECN" / "标准曲线" / "响应因子".
        curve_slope: 标准曲线斜率.
        curve_intercept: 标准曲线截距.
        response_factor: 响应因子.
        products: 目标产物列表, 各自有适用实验范围.
    返回:
        YieldCalcConfig.
    """
    # 内标信息
    is_name: str = ""
    is_smiles: str = ""
    is_formula: str = ""
    is_molecular_weight: Optional[float] = None
    is_ecn: float = 0.0
    is_expected_rt: Optional[float] = None
    is_amount: float = 0.0
    is_moles: float = 0.0
    is_nist_has_record: bool = False
    is_nist_query_mode: str = ""
    is_nist_reference_names: List[str] = field(default_factory=list)
    # 反应信息
    reaction_scale_mmol: float = 0.0
    # 计算方法
    calc_method: str = "ECN"
    curve_slope: Optional[float] = None
    curve_intercept: Optional[float] = None
    response_factor: Optional[float] = None
    # 目标产物
    products: List[TargetProduct] = field(default_factory=list)


@dataclass
class SampleYieldResult:
    """
    功能:
        存储单个 (样品 x 产物) 组合的产率计算结果.
    参数:
        sample_name: 样品名, 如 "729-1".
        product_name: 目标产物名称.
        product_fid_rt: 产物 FID 保留时间(min).
        product_fid_area: 产物 FID 峰面积.
        product_match_compound: 匹配到的 NIST 化合物名.
        is_fid_rt: 内标 FID 保留时间(min).
        is_fid_area: 内标 FID 峰面积.
        is_match_compound: 内标匹配到的 NIST 化合物名.
        area_ratio: FID 面积比 (产物/内标).
        ecn_product: 产物 ECN.
        ecn_is: 内标 ECN.
        molar_ratio: 摩尔比 (产物/内标).
        n_product_mol: 产物物质的量(mol).
        yield_percent: 产率(%).
        match_method: 峰匹配方式, 可选 NIST命中/RT命中/分子量命中.
        confidence_level: 峰判定置信度分数, 0-100.
        confidence_score: 峰判定置信度分数值, 0-100.
        confidence_reason: 置信度说明.
        hit_summary: 命中详情摘要.
        nist_matched_mw: NIST 命中化合物分子量(Da).
        pim_predicted_mw: PIM 预测分子量(Da).
        sshm_predicted_mw: SS-HM 预测分子量(Da).
        ihshm_predicted_mw: iHS-HM 预测分子量(Da).
        warnings: 警告信息列表.
    返回:
        SampleYieldResult.
    """
    sample_name: str = ""
    product_name: str = ""
    # 产物峰
    product_fid_rt: Optional[float] = None
    product_fid_area: Optional[float] = None
    product_match_compound: str = ""
    # 内标峰
    is_fid_rt: Optional[float] = None
    is_fid_area: Optional[float] = None
    is_match_compound: str = ""
    # 计算结果
    area_ratio: Optional[float] = None
    ecn_product: Optional[float] = None
    ecn_is: Optional[float] = None
    molar_ratio: Optional[float] = None
    n_product_mol: Optional[float] = None
    yield_percent: Optional[float] = None
    match_method: str = ""
    confidence_level: str = ""
    confidence_score: Optional[float] = None
    confidence_reason: str = ""
    hit_summary: str = ""
    remark: str = ""
    nist_matched_mw: Optional[float] = None
    pim_predicted_mw: Optional[float] = None
    sshm_predicted_mw: Optional[float] = None
    ihshm_predicted_mw: Optional[float] = None
    warnings: List[str] = field(default_factory=list)


@dataclass
class NISTLibraryQueryResult:
    """
    功能:
        存储 NIST 库收录查询结果.
    参数:
        has_record: 是否检索到对应记录.
        query_mode: 命中模式, 可选 inchikey_exact/smiles_exact/formula_mw_fallback/not_found.
        reference_names: 命中记录中的化合物名称集合.
        reference_formulas: 命中记录中的分子式集合.
        reference_mw: 命中记录中的分子量集合.
    返回:
        NISTLibraryQueryResult.
    """
    has_record: bool = False
    query_mode: str = "not_found"
    reference_names: List[str] = field(default_factory=list)
    reference_formulas: List[str] = field(default_factory=list)
    reference_mw: List[float] = field(default_factory=list)


@dataclass
class NISTHitInfo:
    """
    功能:
        存储单条峰行内的 NIST 命中信息.
    参数:
        rank: 命中序号, 1 表示化合物1, 2 表示化合物2, 依此类推.
        compound_name: 命中名称.
        formula: 命中分子式.
        molecular_weight: 命中分子量.
    返回:
        NISTHitInfo.
    """
    rank: int = 0
    compound_name: str = ""
    formula: str = ""
    molecular_weight: Optional[float] = None


@dataclass
class PeakDecision:
    """
    功能:
        存储单个候选峰的判定结果.
    参数:
        row: 原始对照表行数据.
        match_method: 峰匹配方式文本.
        fid_rt: FID 保留时间(min).
        fid_area: FID 峰面积.
        nist_hit: 选中的 NIST 命中信息.
        nist_target_matched: NIST 命中是否为目标化合物.
        nist_mw_matched: NIST 命中分子量是否与目标分子量匹配.
        nist_formula_matched: NIST 命中分子式是否与目标分子式匹配.
        pim_mw: PIM 预测分子量.
        sshm_mw: SS-HM 预测分子量.
        ihshm_mw: iHS-HM 预测分子量.
        mass_match_methods: 分子量命中的预测方法列表.
        confidence_level: 置信度分数文本, 0-100.
        confidence_score: 置信度分数值, 0-100.
        confidence_reason: 置信度说明.
        hit_summary: 命中详情摘要.
    返回:
        PeakDecision.
    """
    row: Dict[str, Any] = field(default_factory=dict)
    match_method: str = ""
    fid_rt: Optional[float] = None
    fid_area: Optional[float] = None
    nist_hit: Optional[NISTHitInfo] = None
    nist_target_matched: bool = False
    nist_mw_matched: bool = False
    nist_formula_matched: bool = False
    pim_mw: Optional[float] = None
    sshm_mw: Optional[float] = None
    ihshm_mw: Optional[float] = None
    mass_match_methods: List[str] = field(default_factory=list)
    confidence_level: str = ""
    confidence_score: float = 0.0
    confidence_reason: str = ""
    hit_summary: str = ""
    remark: str = ""


# ---------------------------------------------------------------------------
# 核心类
# ---------------------------------------------------------------------------

class YieldCalculator:
    """
    功能:
        基于 GC-FID 积分报告和实验方案计算各样品的产率.
        支持 ECN 法, 标准曲线法和响应因子法.
    参数:
        rt_tolerance: 保留时间匹配容差(min), 默认 0.1.
        nist_mainlib_msp_path: NIST 主库导出的 MSP 文件路径, 用于查询化合物是否收录.
        mw_tolerance_da: 分子量匹配容差(Da), 用于 NIST/预测分子量一致性判断.
        pim_enabled: 是否启用 PIM 预测列读取.
        sshm_enabled: 是否启用 SS-HM 预测列读取.
        ihshm_enabled: 是否启用 iHS-HM 预测列读取.
    返回:
        无.
    """

    # 报告表头样式
    _HEADER_FONT = Font(bold=True, size=11)
    _HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    _PREDICTION_WEIGHTS: Dict[str, float] = {
        "PIM": 20.0,
        "SS-HM": 15.0,
        "iHS-HM": 15.0,
    }

    _PREDICTION_MW_COLUMNS: Dict[str, str] = {
        "PIM": "PIM预测分子量(Da)",
        "SS-HM": "SS-HM预测分子量(Da)",
        "iHS-HM": "iHS-HM预测分子量(Da)",
    }

    _PREDICTION_CONF_COLUMNS: Dict[str, str] = {
        "PIM": "PIM置信指数",
        "SS-HM": "SS-HM置信度",
        "iHS-HM": "iHS-HM置信度",
    }

    _CONFIDENCE_BAND_RULES: Dict[str, Tuple[float, float, float]] = {
        "nist_strong_with_mw": (92.0, 100.0, 8.0),
        "nist_strong_without_mw": (85.0, 92.0, 7.0),
        "nist_weak": (55.0, 70.0, 15.0),
        "nist_none_absent_multi_match": (65.0, 80.0, 15.0),
        "nist_none_absent_single_match": (45.0, 60.0, 15.0),
        "nist_none_absent_none": (15.0, 15.0, 0.0),
        "nist_none_contradict_multi_match": (60.0, 75.0, 15.0),
        "nist_none_contradict_single_match": (40.0, 55.0, 15.0),
        "nist_none_contradict_none": (10.0, 10.0, 0.0),
    }

    def __init__(
        self,
        rt_tolerance: float = 0.1,
        nist_mainlib_msp_path: Optional[Path] = None,
        mw_tolerance_da: float = 1.0,
        pim_enabled: bool = True,
        sshm_enabled: bool = True,
        ihshm_enabled: bool = True,
    ) -> None:
        """
        功能:
            初始化产率计算器并加载峰判定策略相关配置.
        参数:
            rt_tolerance: 保留时间匹配容差(min).
            nist_mainlib_msp_path: NIST 主库 MSP 路径.
            mw_tolerance_da: 分子量匹配容差(Da).
            pim_enabled: 是否启用 PIM 分子量读取.
            sshm_enabled: 是否启用 SS-HM 分子量读取.
            ihshm_enabled: 是否启用 iHS-HM 分子量读取.
        返回:
            无.
        """
        self._rt_tolerance = rt_tolerance
        self._nist_mainlib_msp_path = nist_mainlib_msp_path
        self._mw_tolerance_da = mw_tolerance_da
        self._pim_enabled = pim_enabled
        self._sshm_enabled = sshm_enabled
        self._ihshm_enabled = ihshm_enabled

        self._nist_index_loaded = False
        self._nist_smiles_field_detected = False
        self._nist_inchikey_field_detected = False
        self._nist_index_by_inchikey: Dict[str, List[Dict[str, Any]]] = {}
        self._nist_index_by_smiles: Dict[str, List[Dict[str, Any]]] = {}
        self._nist_index_by_formula: Dict[str, List[Dict[str, Any]]] = {}

    def _get_enabled_prediction_methods(self) -> List[str]:
        """
        功能:
            返回当前启用的分子量预测方法列表.
        参数:
            无.
        返回:
            List[str]: 启用方法名列表, 顺序固定为 PIM, SS-HM, iHS-HM.
        """
        methods: List[str] = []
        if self._pim_enabled is True:
            methods.append("PIM")
        if self._sshm_enabled is True:
            methods.append("SS-HM")
        if self._ihshm_enabled is True:
            methods.append("iHS-HM")
        return methods

    def _build_yield_headers(self) -> List[str]:
        """
        功能:
            根据启用的预测方法动态构建产率结果表头.
        参数:
            无.
        返回:
            List[str]: 动态表头列表.
        """
        headers = [
            "样品名", "目标产物", "产物保留时间(min)", "产物FID面积",
            "产物匹配化合物", "内标保留时间(min)", "内标FID面积",
            "内标匹配化合物", "Ratio", "产物ECN", "内标ECN",
            "产率(%)", "匹配方式", "置信度",
            "NIST匹配分子量(Da)",
        ]

        enabled_methods = self._get_enabled_prediction_methods()
        for method_name in enabled_methods:
            headers.append(self._PREDICTION_MW_COLUMNS[method_name])

        headers.append("备注")
        return headers

    # ------------------------------------------------------------------
    # 静态/辅助方法
    # ------------------------------------------------------------------

    @staticmethod
    def calculate_ecn(smiles: str) -> float:
        """
        功能:
            根据 SMILES 计算化合物的有效碳数(ECN).
        参数:
            smiles: 化合物 SMILES 字符串.
        返回:
            float: ECN 值.
        """
        ecn = 0.0
        for ary, typ in smiles2carbontypes(smiles):
            if typ == 'alcohol':
                # 醇类需要区分伯/仲/叔
                key = (ary + ' ' + typ) if ary else typ
                ecn += ecn_dct[key]
            else:
                ecn += ecn_dct[class_dct[typ]]
        return ecn

    @staticmethod
    def smiles_to_formula(smiles: str) -> str:
        """
        功能:
            从 SMILES 推导分子式, 使用 Hill 排序 (C, H 在前, 其余按字母序).
        参数:
            smiles: 化合物 SMILES 字符串.
        返回:
            str: 分子式, 如 "C13H13N".
        """
        graph = read_smiles(smiles, explicit_hydrogen=True)
        atom_counts: Dict[str, int] = {}
        for node in graph.nodes:
            elem = graph.nodes[node].get('element', '')
            if elem:
                atom_counts[elem] = atom_counts.get(elem, 0) + 1

        # Hill 排序: C 在前, H 次之, 其余按字母序
        parts = []
        for elem in ['C', 'H']:
            if elem in atom_counts:
                count = atom_counts.pop(elem)
                parts.append(elem + (str(count) if count > 1 else ""))
        for elem in sorted(atom_counts.keys()):
            count = atom_counts[elem]
            parts.append(elem + (str(count) if count > 1 else ""))
        return "".join(parts)

    @staticmethod
    def smiles_to_molecular_weight(smiles: str) -> Optional[float]:
        """
        功能:
            从 SMILES 计算分子量.
            计算顺序:
            1. 优先使用 RDKit 的 MolWt, 覆盖元素更完整.
            2. RDKit 不可用或解析失败时, 回退到 pysmiles + PTE 原子量累加.
        参数:
            smiles: 化合物 SMILES 字符串.
        返回:
            Optional[float]: 分子量(Da), 失败时返回 None.
        """
        smiles_text = str(smiles).strip()
        if smiles_text == "":
            return None

        if Chem is not None and Descriptors is not None:
            try:
                mol = Chem.MolFromSmiles(smiles_text)
                if mol is not None:
                    return float(Descriptors.MolWt(mol))
                logger.warning("RDKit 无法解析 SMILES, 将回退 PTE 计算: %s", smiles_text)
            except Exception as exc:
                logger.warning("RDKit 计算分子量失败, 将回退 PTE 计算: %s, 错误=%s", smiles_text, exc)

        try:
            graph = read_smiles(smiles_text, explicit_hydrogen=True)
        except Exception as exc:
            logger.warning("SMILES 解析失败, 无法计算分子量: %s, 错误=%s", smiles_text, exc)
            return None

        total_weight = 0.0
        missing_elements: Set[str] = set()
        for node in graph.nodes:
            element_symbol = str(graph.nodes[node].get("element", "")).strip()
            if element_symbol == "":
                continue
            pte_entry = PTE.get(element_symbol)
            if pte_entry is None:
                missing_elements.add(element_symbol)
                continue
            atomic_mass = pte_entry.get("AtomicMass")
            if atomic_mass is None or atomic_mass == "":
                missing_elements.add(element_symbol)
                continue
            total_weight += float(atomic_mass)

        if len(missing_elements) > 0:
            logger.warning(
                "分子量计算失败, 缺少元素原子量: SMILES=%s, 元素=%s",
                smiles_text,
                sorted(missing_elements),
            )
            return None

        return total_weight

    @staticmethod
    def parse_experiment_range(range_str: str) -> List[int]:
        """
        功能:
            解析 "适用实验" 字段, 支持单个编号/范围/逗号分隔/混合格式.
            "all" 或空字符串返回空列表, 表示适用于所有实验.
        参数:
            range_str: 适用实验字符串, 如 "1-6,9" 或 "all".
        返回:
            List[int]: 实验编号列表, 空列表表示全部适用.
        """
        text = str(range_str).strip().lower()
        if text == "" or text == "all":
            return []

        result = []
        for part in text.split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                bounds = part.split("-", 1)
                try:
                    start = int(bounds[0].strip())
                    end = int(bounds[1].strip())
                    result.extend(range(start, end + 1))
                except ValueError:
                    logger.warning("无法解析实验范围: '%s'", part)
            else:
                try:
                    result.append(int(part))
                except ValueError:
                    logger.warning("无法解析实验编号: '%s'", part)
        return sorted(set(result))

    @staticmethod
    def _extract_experiment_number(sample_name: str) -> Optional[int]:
        """
        功能:
            从样品名提取实验编号, 如 "729-3" → 3.
        参数:
            sample_name: 样品名字符串.
        返回:
            Optional[int]: 实验编号, 提取失败返回 None.
        """
        match = re.search(r"-(\d+)$", sample_name.strip())
        if match is not None:
            return int(match.group(1))
        return None

    # ------------------------------------------------------------------
    # 内标摩尔量推算
    # ------------------------------------------------------------------

    def _calculate_is_moles(
        self, is_name: str, is_amount: float, chemical_list_path: Path
    ) -> float:
        """
        功能:
            从 chemical_list.xlsx 查询内标的物化属性,
            根据 physical_state / physical_form / active_content 推算实际加入摩尔量.
        参数:
            is_name: 内标种类名称 (含括号描述).
            is_amount: 内标用量(μL 或 mg).
            chemical_list_path: chemical_list.xlsx 文件路径.
        返回:
            float: 内标摩尔量(mol).
        """
        # 读取化学品库, 多工作表时按必需列优先选表.
        chem_sheets = pd.read_excel(chemical_list_path, sheet_name=None)
        chem_df = None
        required_columns = {"substance", "molecular_weight", "physical_state"}
        for sheet_name, sheet_df in chem_sheets.items():
            normalized_columns = {str(col).strip().lower() for col in sheet_df.columns}
            if required_columns.issubset(normalized_columns):
                chem_df = sheet_df
                logger.info("化学品库解析使用工作表: %s", sheet_name)
                break

        if chem_df is None:
            first_sheet_name = next(iter(chem_sheets.keys()))
            chem_df = chem_sheets[first_sheet_name]
            logger.warning(
                "chemical_list.xlsx 未命中必需列%s, 回退到第一张工作表: %s",
                sorted(required_columns),
                first_sheet_name,
            )
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row, *keys, default=None):
            for k in keys:
                if k in row and pd.notna(row[k]):
                    return row[k]
            return default

        # 在化学品库中精确匹配内标名称
        chem_info = None
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if name == is_name:
                chem_info = {
                    "molecular_weight": _pick(row, "molecular_weight", "mw"),
                    "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                    "density": _pick(row, "density (g/ml)", "density(g/ml)", "density_g_ml", "density", default=None),
                    "physical_form": str(_pick(row, "physical_form", default="") or "").strip().lower(),
                    "active_content": _pick(
                        row, "active_content",
                        "active_content(mmol/ml or wt%)",
                        "active_content(mol/l or wt%)",
                        default=""
                    ),
                }
                break

        if chem_info is None:
            logger.error("在 chemical_list.xlsx 中未找到内标: '%s' (精确匹配)", is_name)
            raise ValueError(f"在 chemical_list.xlsx 中未找到内标: '{is_name}' (需精确匹配)")

        mw = chem_info["molecular_weight"]
        state = chem_info["physical_state"]
        form = chem_info["physical_form"]
        density = chem_info["density"]
        active_content = chem_info["active_content"]

        if mw is None or float(mw) <= 0:
            raise ValueError(f"内标 '{is_name}' 缺少有效分子量")

        mw = float(mw)

        # 纯物质 (neat)
        if form == "neat" or form == "":
            if state == "liquid":
                # 用量单位为 μL, 需要密度
                if density is None or float(density) <= 0:
                    raise ValueError(f"液体内标 '{is_name}' 缺少密度")
                mass_g = is_amount / 1000.0 * float(density)
                n_mol = mass_g / mw
            else:
                # 固体, 用量单位为 mg
                n_mol = is_amount / 1000.0 / mw
            logger.info("内标(纯物质) '%s': %.4f mol", is_name, n_mol)
            return n_mol

        # 溶液 (solution)
        if form == "solution":
            content_type, content_value = self._parse_active_content(active_content, form)
            if content_type == "mmol_per_ml":
                # content_value 为 mmol/mL, is_amount 为 μL
                n_mol = content_value * (is_amount / 1000.0) / 1000.0
            elif content_type == "wt_percent":
                if density is None or float(density) <= 0:
                    raise ValueError(f"溶液内标 '{is_name}' 按 wt% 换算需要密度")
                mass_total_g = is_amount / 1000.0 * float(density)
                mass_active_g = mass_total_g * content_value / 100.0
                n_mol = mass_active_g / mw
            else:
                raise ValueError(f"溶液内标 '{is_name}' 的 active_content 无法解析")
            logger.info("内标(溶液) '%s': %.6f mol", is_name, n_mol)
            return n_mol

        # 负载型 (beads)
        if form == "beads":
            content_type, content_value = self._parse_active_content(active_content, form)
            if content_type == "wt_percent":
                # is_amount 为 mg
                mass_active_mg = is_amount * content_value / 100.0
                n_mol = mass_active_mg / 1000.0 / mw
            else:
                raise ValueError(f"负载型内标 '{is_name}' 的 active_content 应为 wt%")
            logger.info("内标(负载型) '%s': %.6f mol", is_name, n_mol)
            return n_mol

        raise ValueError(f"内标 '{is_name}' 的 physical_form '{form}' 未支持")

    @staticmethod
    def _parse_active_content(value: Any, physical_form: str) -> Tuple[str, float]:
        """
        功能:
            解析 active_content 字段, 结合 physical_form 区分 mmol/mL 与 wt%.
            逻辑与 station_controller._parse_active_content 一致.
        参数:
            value: active_content 原始值.
            physical_form: 物理形态 (solution/beads 等).
        返回:
            Tuple[str, float]: (类型标记, 数值), 无法解析返回 ("", 0.0).
        """
        form = (physical_form or "").lower().strip()
        if value is None:
            return "", 0.0
        text_raw = str(value).strip()
        if text_raw == "":
            return "", 0.0

        try:
            num_val = float(value)
        except (ValueError, TypeError):
            text_norm = text_raw.lower()
            numbers = re.findall(r"[0-9]+(?:\.[0-9]+)?", text_norm)
            num_val = float(numbers[0]) if len(numbers) > 0 else 0.0

        # 按 physical_form 优先判断
        if form == "solution":
            return "mmol_per_ml", num_val
        if form == "beads":
            return "wt_percent", num_val

        # 按文本关键词判断
        text = text_raw.lower()
        if "mmol/ml" in text or "mmol per ml" in text or "mmolml" in text:
            return "mmol_per_ml", num_val
        if "wt%" in text or "wt percent" in text or "wt" in text:
            return "wt_percent", num_val
        if num_val > 0:
            return "mmol_per_ml", num_val
        return "", 0.0

    # ------------------------------------------------------------------
    # 配置解析
    # ------------------------------------------------------------------

    def parse_yield_config_legacy(
        self, plan_path: Path, chemical_list_path: Path
    ) -> YieldCalcConfig:
        """
        功能:
            从实验方案 xlsx 解析产率计算配置:
            1. 读原参数 Sheet: 反应规模(mmol), 内标种类, 内标用量(μL/mg).
            2. 读 "GC产率计算" Sheet: 内标SMILES, 目标产物列表, 计算方法.
            3. 查 chemical_list.xlsx, 推算内标摩尔量.
            4. 自动计算各化合物的分子式和 ECN.
        参数:
            plan_path: 实验方案 xlsx 文件路径.
            chemical_list_path: chemical_list.xlsx 文件路径.
        返回:
            YieldCalcConfig: 产率计算配置.
        """
        wb = openpyxl.load_workbook(str(plan_path), data_only=True)

        # ---------- 1. 从原参数 Sheet 读取已有字段 ----------
        ws_main = wb.worksheets[0]
        params = self._read_kv_params(ws_main)

        reaction_scale = self._parse_float(params.get("反应规模(mmol)", 0))
        is_name = str(params.get("内标种类", "")).strip()
        is_amount = self._parse_float(params.get("内标用量(μL/mg)", 0))
        # 兼容可能的其他格式
        if is_amount == 0:
            is_amount = self._parse_float(params.get("内标用量(ul/mg)", 0))
            if is_amount == 0:
                is_amount = self._parse_float(params.get("内标用量", 0))

        # ---------- 2. 读 "GC产率计算" Sheet ----------
        if YIELD_CONFIG_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"实验方案中未找到 '{YIELD_CONFIG_SHEET_NAME}' Sheet")
        ws_yield = wb[YIELD_CONFIG_SHEET_NAME]

        yield_params = self._read_kv_params(ws_yield)

        is_smiles = str(yield_params.get("内标SMILES", "")).strip()
        is_expected_rt = self._parse_opt_float(yield_params.get("内标预期RT(min)"))
        calc_method = str(yield_params.get("产率计算方法", "ECN")).strip()
        curve_slope = self._parse_opt_float(yield_params.get("标准曲线斜率"))
        curve_intercept = self._parse_opt_float(yield_params.get("标准曲线截距"))
        response_factor = self._parse_opt_float(yield_params.get("响应因子"))

        if not is_smiles:
            raise ValueError(f"{YIELD_CONFIG_SHEET_NAME} Sheet 中未填写 '内标SMILES'")

        # ---------- 3. 读目标产物列表 ----------
        products = self._read_product_table(ws_yield)

        if len(products) == 0:
            raise ValueError(f"{YIELD_CONFIG_SHEET_NAME} Sheet 中未找到目标产物列表")

        # ---------- 4. 计算分子式和 ECN ----------
        is_formula = self.smiles_to_formula(is_smiles)
        is_molecular_weight = self.smiles_to_molecular_weight(is_smiles)
        is_ecn = self.calculate_ecn(is_smiles)
        logger.info(
            "内标 '%s': 分子式=%s, 分子量=%s, ECN=%.2f",
            is_name,
            is_formula,
            round(is_molecular_weight, 4) if is_molecular_weight is not None else "未知",
            is_ecn,
        )

        for p in products:
            p.formula = self.smiles_to_formula(p.smiles)
            p.molecular_weight = self.smiles_to_molecular_weight(p.smiles)
            p.ecn = self.calculate_ecn(p.smiles)
            logger.info(
                "产物 '%s': 分子式=%s, 分子量=%s, ECN=%.2f",
                p.name,
                p.formula,
                round(p.molecular_weight, 4) if p.molecular_weight is not None else "未知",
                p.ecn,
            )

        # ---------- 5. 推算内标摩尔量 ----------
        is_moles = self._calculate_is_moles(is_name, is_amount, chemical_list_path)

        wb.close()

        return YieldCalcConfig(
            is_name=is_name,
            is_smiles=is_smiles,
            is_formula=is_formula,
            is_molecular_weight=is_molecular_weight,
            is_ecn=is_ecn,
            is_expected_rt=is_expected_rt,
            is_amount=is_amount,
            is_moles=is_moles,
            reaction_scale_mmol=reaction_scale,
            calc_method=calc_method,
            curve_slope=curve_slope,
            curve_intercept=curve_intercept,
            response_factor=response_factor,
            products=products,
        )

    @staticmethod
    def _normalize_sheet_text(value: Any) -> str:
        """
        功能:
            规范化工作表文本, 用于关键字匹配.
        参数:
            value: 任意类型单元格值.
        返回:
            str, 去空白并转小写后的文本.
        """
        text = "" if value is None else str(value)
        return (
            text.replace(" ", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .strip()
            .lower()
        )

    def _score_main_config_sheet(self, worksheet: Any) -> int:
        """
        功能:
            计算工作表作为“实验方案主配置表”的匹配分数.
        参数:
            worksheet: openpyxl Worksheet.
        返回:
            int, 命中关键字段数量.
        """
        keyword_flags = {
            "反应规模(mmol)": False,
            "内标种类": False,
            "内标用量": False,
        }
        normalized_keywords = {
            key: self._normalize_sheet_text(key) for key in keyword_flags.keys()
        }

        scan_rows = min(worksheet.max_row, 120)
        for row_index in range(1, scan_rows + 1):
            cell_text = self._normalize_sheet_text(worksheet.cell(row_index, 1).value)
            if cell_text == "":
                continue
            for key, normalized_key in normalized_keywords.items():
                if normalized_key in cell_text:
                    keyword_flags[key] = True

        return sum(1 for matched in keyword_flags.values() if matched)

    def _select_main_config_sheet(self, workbook: Any) -> Any:
        """
        功能:
            选择用于读取主参数的工作表.
            选择顺序: 实验方案设定 > 当前 active > 其它工作表.
            命中规则: 反应规模/内标种类/内标用量 关键字段分数最高.
        参数:
            workbook: openpyxl Workbook 对象.
        返回:
            Any, 选中的工作表对象.
        """
        candidate_sheet_names: List[str] = []
        preferred_sheet_name = "实验方案设定"
        if preferred_sheet_name in workbook.sheetnames:
            candidate_sheet_names.append(preferred_sheet_name)

        active_sheet_name = workbook.active.title
        if active_sheet_name not in candidate_sheet_names:
            candidate_sheet_names.append(active_sheet_name)

        for sheet_name in workbook.sheetnames:
            if sheet_name not in candidate_sheet_names:
                candidate_sheet_names.append(sheet_name)

        best_candidates: List[Any] = []
        best_score = -1
        for sheet_name in candidate_sheet_names:
            worksheet = workbook[sheet_name]
            score = self._score_main_config_sheet(worksheet)
            if score > best_score:
                best_score = score
                best_candidates = [worksheet]
                continue
            if score == best_score:
                best_candidates.append(worksheet)

        if len(best_candidates) > 0 and best_score > 0:
            selected_sheet = best_candidates[0]
            if len(best_candidates) > 1:
                candidate_names = [sheet.title for sheet in best_candidates]
                logger.warning(
                    "产率配置命中多个主配置候选工作表, 将按优先顺序使用 [%s], 其余候选: %s",
                    selected_sheet.title,
                    candidate_names[1:],
                )
            return selected_sheet

        # 兼容历史文件: 若未命中关键字段, 回退到第一张表并记录告警.
        fallback_sheet = workbook.worksheets[0]
        logger.warning(
            "未命中主配置关键字段, 回退到第一张工作表: %s, 可用工作表: %s",
            fallback_sheet.title,
            workbook.sheetnames,
        )
        return fallback_sheet

    def parse_yield_config(
        self, plan_path: Path, chemical_list_path: Path
    ) -> YieldCalcConfig:
        """
        功能:
            从实验方案 xlsx 解析产率计算配置:
            1. 读主参数 Sheet: 反应规模(mmol), 内标种类, 内标用量(μL/mg).
            2. 读 "GC产率计算" Sheet: 内标SMILES, 目标产物列表, 计算方法.
            3. 查 chemical_list.xlsx, 推算内标摩尔量.
            4. 自动计算各化合物的分子式和 ECN.
        参数:
            plan_path: 实验方案 xlsx 文件路径.
            chemical_list_path: chemical_list.xlsx 文件路径.
        返回:
            YieldCalcConfig: 产率计算配置.
        """
        wb = openpyxl.load_workbook(str(plan_path), data_only=True)
        try:
            # ---------- 1. 从主参数 Sheet 读取已有字段 ----------
            ws_main = self._select_main_config_sheet(wb)
            if ws_main.title != wb.active.title:
                logger.info(
                    "产率配置解析使用工作表: %s, 当前活动工作表: %s",
                    ws_main.title,
                    wb.active.title,
                )
            params = self._read_kv_params(ws_main)

            reaction_scale = self._parse_float(params.get("反应规模(mmol)", 0))
            is_name = str(params.get("内标种类", "")).strip()
            is_amount = self._parse_float(params.get("内标用量(μL/mg)", 0))
            # 兼容可能的其它格式.
            if is_amount == 0:
                is_amount = self._parse_float(params.get("内标用量(ul/mg)", 0))
                if is_amount == 0:
                    is_amount = self._parse_float(params.get("内标用量", 0))

            # ---------- 2. 读 "GC产率计算" Sheet ----------
            if YIELD_CONFIG_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"实验方案中未找到 '{YIELD_CONFIG_SHEET_NAME}' Sheet")
            ws_yield = wb[YIELD_CONFIG_SHEET_NAME]

            yield_params = self._read_kv_params(ws_yield)

            is_smiles = str(yield_params.get("内标SMILES", "")).strip()
            is_expected_rt = self._parse_opt_float(yield_params.get("内标预期RT(min)"))
            calc_method = str(yield_params.get("产率计算方法", "ECN")).strip()
            curve_slope = self._parse_opt_float(yield_params.get("标准曲线斜率"))
            curve_intercept = self._parse_opt_float(yield_params.get("标准曲线截距"))
            response_factor = self._parse_opt_float(yield_params.get("响应因子"))

            if not is_smiles:
                raise ValueError(f"{YIELD_CONFIG_SHEET_NAME} Sheet 中未填写 '内标SMILES'")

            # ---------- 3. 读目标产物列表 ----------
            products = self._read_product_table(ws_yield)

            if len(products) == 0:
                raise ValueError(f"{YIELD_CONFIG_SHEET_NAME} Sheet 中未找到目标产物列表")

            # ---------- 4. 计算分子式和 ECN ----------
            is_formula = self.smiles_to_formula(is_smiles)
            is_molecular_weight = self.smiles_to_molecular_weight(is_smiles)
            is_ecn = self.calculate_ecn(is_smiles)
            logger.info(
                "内标 '%s': 分子式=%s, 分子量=%s, ECN=%.2f",
                is_name,
                is_formula,
                round(is_molecular_weight, 4) if is_molecular_weight is not None else "未知",
                is_ecn,
            )

            for product in products:
                product.formula = self.smiles_to_formula(product.smiles)
                product.molecular_weight = self.smiles_to_molecular_weight(product.smiles)
                product.ecn = self.calculate_ecn(product.smiles)
                logger.info(
                    "产物 '%s': 分子式=%s, 分子量=%s, ECN=%.2f",
                    product.name,
                    product.formula,
                    round(product.molecular_weight, 4) if product.molecular_weight is not None else "未知",
                    product.ecn,
                )

            # ---------- 5. 推算内标摩尔量 ----------
            is_moles = self._calculate_is_moles(is_name, is_amount, chemical_list_path)

            return YieldCalcConfig(
                is_name=is_name,
                is_smiles=is_smiles,
                is_formula=is_formula,
                is_molecular_weight=is_molecular_weight,
                is_ecn=is_ecn,
                is_expected_rt=is_expected_rt,
                is_amount=is_amount,
                is_moles=is_moles,
                reaction_scale_mmol=reaction_scale,
                calc_method=calc_method,
                curve_slope=curve_slope,
                curve_intercept=curve_intercept,
                response_factor=response_factor,
                products=products,
            )
        finally:
            wb.close()

    def _read_kv_params(self, ws) -> Dict[str, Any]:
        """
        功能:
            从 worksheet 的 A/B 列读取 key-value 参数.
            遇到空行或表格表头行时停止 (表头行特征: A 列为 "适用实验" 等).
        参数:
            ws: openpyxl Worksheet.
        返回:
            Dict[str, Any]: 参数字典.
        """
        params: Dict[str, Any] = {}
        # 已知的产物表表头关键词, 遇到则停止 KV 读取
        table_header_keywords = {"适用实验", "目标产物名称", "目标产物"}
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            key = str(row[0] or "").strip()
            if not key:
                continue
            if key in table_header_keywords:
                break
            value = row[1]
            if value is not None:
                params[key] = value
        return params

    def _read_product_table(self, ws) -> List[TargetProduct]:
        """
        功能:
            从 "GC产率计算" Sheet 中读取目标产物列表.
            查找表头行 (含 "适用实验" / "目标产物名称"), 然后逐行读取数据.
        参数:
            ws: openpyxl Worksheet ("GC产率计算" Sheet).
        返回:
            List[TargetProduct]: 目标产物列表.
        """
        products = []
        header_row = None
        col_map: Dict[str, int] = {}

        # 定位表头行
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), start=1):
            values = [str(cell.value or "").strip() for cell in row]
            # 检测是否为产物表头行
            lower_values = [v.lower() for v in values]
            if any("适用实验" in v for v in values) or any("目标产物" in v for v in values):
                header_row = row_idx
                for col_idx, val in enumerate(values):
                    if val:
                        col_map[val] = col_idx
                break

        if header_row is None:
            return products

        # 从表头行下一行开始读取数据
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = [v for v in row]
            # 跳过空行
            if all(v is None or str(v).strip() == "" for v in values):
                continue

            def _get(header_name: str, default=""):
                for name, idx in col_map.items():
                    if header_name in name:
                        if idx < len(values) and values[idx] is not None:
                            return values[idx]
                return default

            product_name = str(_get("目标产物名称", _get("目标产物"))).strip()
            smiles = str(_get("SMILES")).strip()
            rt_val = _get("预期RT", _get("RT"))
            range_str = str(_get("适用实验")).strip()
            eq_val = _get("当量", _get("eq", 1.0))

            if not product_name or not smiles:
                continue

            expected_rt = self._parse_opt_float(rt_val)
            applicable = self.parse_experiment_range(range_str)
            equivalent = self._parse_float(eq_val, default=1.0)

            products.append(TargetProduct(
                name=product_name,
                smiles=smiles,
                expected_rt=expected_rt,
                applicable_experiments=applicable,
                equivalent=equivalent,
            ))

        return products

    # ------------------------------------------------------------------
    # 数据加载
    # ------------------------------------------------------------------

    def load_alignment_data(self, report_path: Path) -> Dict[str, List[Dict]]:
        """
        功能:
            从积分报告的 "TIC-FID对照表" Sheet 读取对齐数据, 按样品名分组.
        参数:
            report_path: 积分报告 xlsx 文件路径.
        返回:
            Dict[str, List[Dict]]: 样品名 → 对齐行字典列表.
        """
        wb = openpyxl.load_workbook(str(report_path), data_only=True)
        if "TIC-FID对照表" not in wb.sheetnames:
            wb.close()
            raise ValueError("积分报告中未找到 'TIC-FID对照表' Sheet")

        ws = wb["TIC-FID对照表"]

        # 读表头
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip())

        # 按行读取数据
        data: Dict[str, List[Dict]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict: Dict[str, Any] = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val

            sample_name = str(row_dict.get("样品名", "")).strip()
            if not sample_name:
                continue

            if sample_name not in data:
                data[sample_name] = []
            data[sample_name].append(row_dict)

        wb.close()
        return data

    # ------------------------------------------------------------------
    # 峰匹配
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_smiles(smiles: str) -> str:
        """
        功能:
            规范化 SMILES 字符串, 仅做空白清理用于库查询键.
        参数:
            smiles: 原始 SMILES 字符串.
        返回:
            str, 规范化后的 SMILES.
        """
        return re.sub(r"\s+", "", str(smiles).strip())

    @staticmethod
    def _normalize_inchikey(inchikey: str) -> str:
        """
        功能:
            规范化 InChIKey 字符串, 统一去空白并转为大写.
        参数:
            inchikey: 原始 InChIKey 字符串.
        返回:
            str, 规范化后的 InChIKey.
        """
        return re.sub(r"\s+", "", str(inchikey).strip()).upper()

    def _smiles_to_inchikey(self, smiles: str) -> str:
        """
        功能:
            根据 SMILES 生成 InChIKey, 失败时返回空字符串.
        参数:
            smiles: 目标 SMILES 字符串.
        返回:
            str, 生成成功时返回规范化 InChIKey, 失败返回空字符串.
        """
        smiles_text = self._normalize_smiles(smiles)
        if smiles_text == "":
            return ""

        if Chem is None or RDKitInchi is None:
            logger.warning("RDKit 不可用, 无法生成 InChIKey: %s", smiles_text)
            return ""

        try:
            molecule = Chem.MolFromSmiles(smiles_text)
            if molecule is None:
                logger.warning("SMILES 解析失败, 无法生成 InChIKey: %s", smiles_text)
                return ""
            inchikey = RDKitInchi.MolToInchiKey(molecule)
        except Exception as exc:
            logger.warning("生成 InChIKey 失败: SMILES=%s, 错误=%s", smiles_text, exc)
            return ""

        normalized_inchikey = self._normalize_inchikey(inchikey)
        if normalized_inchikey == "":
            logger.warning("InChIKey 生成结果为空: %s", smiles_text)
        return normalized_inchikey

    @staticmethod
    def _normalize_name(name: str) -> str:
        """
        功能:
            规范化化合物名称, 用于松弛匹配.
        参数:
            name: 原始名称.
        返回:
            str, 去符号后的低噪声名称键.
        """
        return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "", str(name).strip().lower())

    def _add_nist_entry_to_index(self, entry: Dict[str, Any]) -> None:
        """
        功能:
            将单条 NIST MSP 记录写入运行时索引.
        参数:
            entry: 单条 MSP 记录字典.
        返回:
            无.
        """
        name_text = str(entry.get("name", "")).strip()
        if name_text == "":
            return

        inchikey_text = self._normalize_inchikey(str(entry.get("inchikey", "")))
        smiles_text = self._normalize_smiles(str(entry.get("smiles", "")))
        formula_text = str(entry.get("formula", "")).strip()

        if inchikey_text != "":
            if inchikey_text not in self._nist_index_by_inchikey:
                self._nist_index_by_inchikey[inchikey_text] = []
            self._nist_index_by_inchikey[inchikey_text].append(entry.copy())
            self._nist_inchikey_field_detected = True

        if smiles_text != "":
            if smiles_text not in self._nist_index_by_smiles:
                self._nist_index_by_smiles[smiles_text] = []
            self._nist_index_by_smiles[smiles_text].append(entry.copy())
            self._nist_smiles_field_detected = True

        if formula_text != "":
            if formula_text not in self._nist_index_by_formula:
                self._nist_index_by_formula[formula_text] = []
            self._nist_index_by_formula[formula_text].append(entry.copy())

    def _ensure_nist_library_index(self) -> None:
        """
        功能:
            懒加载 NIST MSP 索引, 支持按 InChIKey/SMILES/Formula+MW 查询收录状态.
        参数:
            无.
        返回:
            无.
        """
        if self._nist_index_loaded is True:
            return

        self._nist_index_loaded = True
        self._nist_index_by_inchikey.clear()
        self._nist_index_by_smiles.clear()
        self._nist_index_by_formula.clear()
        self._nist_inchikey_field_detected = False
        self._nist_smiles_field_detected = False

        if self._nist_mainlib_msp_path is None:
            logger.warning("未配置 NIST MSP 路径, 跳过 NIST 收录查询")
            return

        if self._nist_mainlib_msp_path.exists() is not True:
            logger.warning("NIST MSP 文件不存在: %s", self._nist_mainlib_msp_path)
            return

        logger.info("开始加载 NIST MSP 索引: %s", self._nist_mainlib_msp_path)
        current_entry: Dict[str, Any] = {}
        record_count = 0
        try:
            with self._nist_mainlib_msp_path.open("r", encoding="utf-8", errors="replace") as handle:
                for raw_line in handle:
                    line = raw_line.strip()
                    if line == "":
                        if len(current_entry) > 0:
                            self._add_nist_entry_to_index(current_entry)
                            record_count += 1
                            current_entry = {}
                        continue

                    if ":" not in line:
                        continue

                    key_text, value_text = line.split(":", 1)
                    key_name = key_text.strip().upper()
                    value = value_text.strip()

                    if key_name == "NAME":
                        current_entry["name"] = value
                        continue
                    if key_name == "INCHIKEY":
                        current_entry["inchikey"] = value
                        continue
                    if key_name == "SMILES":
                        current_entry["smiles"] = value
                        continue
                    if key_name == "FORMULA":
                        current_entry["formula"] = value
                        continue
                    if key_name == "MW":
                        current_entry["mw"] = self._parse_opt_float(value)
                        continue
                    if key_name == "EXACTMASS":
                        if current_entry.get("mw") is None:
                            current_entry["mw"] = self._parse_opt_float(value)
                        continue

            if len(current_entry) > 0:
                self._add_nist_entry_to_index(current_entry)
                record_count += 1

            logger.info(
                "NIST MSP 索引加载完成: 记录=%d, InChIKey键=%d, SMILES键=%d, FORMULA键=%d, 含InChIKey字段=%s, 含SMILES字段=%s",
                record_count,
                len(self._nist_index_by_inchikey),
                len(self._nist_index_by_smiles),
                len(self._nist_index_by_formula),
                self._nist_inchikey_field_detected,
                self._nist_smiles_field_detected,
            )
        except Exception as exc:
            logger.error("加载 NIST MSP 索引失败: %s", exc)
            self._nist_index_by_inchikey.clear()
            self._nist_index_by_smiles.clear()
            self._nist_index_by_formula.clear()
            self._nist_inchikey_field_detected = False
            self._nist_smiles_field_detected = False

    def _query_nist_library_by_smiles(
        self,
        smiles: str,
        formula: str,
        target_mw: Optional[float],
    ) -> NISTLibraryQueryResult:
        """
        功能:
            查询目标化合物是否在 NIST 库中收录.
            查询顺序: InChIKey 精确命中 -> SMILES 精确命中 -> Formula+MW 回退命中.
        参数:
            smiles: 目标 SMILES.
            formula: 目标分子式.
            target_mw: 目标分子量.
        返回:
            NISTLibraryQueryResult, 查询结果对象.
        """
        self._ensure_nist_library_index()
        normalized_smiles = self._normalize_smiles(smiles)
        target_inchikey = self._smiles_to_inchikey(normalized_smiles)

        if target_inchikey != "":
            inchikey_entries = self._nist_index_by_inchikey.get(target_inchikey, [])
            if len(inchikey_entries) > 0:
                names = sorted({str(item.get("name", "")).strip() for item in inchikey_entries if str(item.get("name", "")).strip() != ""})
                formulas = sorted({str(item.get("formula", "")).strip() for item in inchikey_entries if str(item.get("formula", "")).strip() != ""})
                mw_values = sorted({float(item.get("mw")) for item in inchikey_entries if item.get("mw") is not None})
                return NISTLibraryQueryResult(
                    has_record=True,
                    query_mode="inchikey_exact",
                    reference_names=names,
                    reference_formulas=formulas,
                    reference_mw=mw_values,
                )

        if normalized_smiles != "":
            hit_entries = self._nist_index_by_smiles.get(normalized_smiles, [])
            if len(hit_entries) > 0:
                names = sorted({str(item.get("name", "")).strip() for item in hit_entries if str(item.get("name", "")).strip() != ""})
                formulas = sorted({str(item.get("formula", "")).strip() for item in hit_entries if str(item.get("formula", "")).strip() != ""})
                mw_values = sorted({float(item.get("mw")) for item in hit_entries if item.get("mw") is not None})
                return NISTLibraryQueryResult(
                    has_record=True,
                    query_mode="smiles_exact",
                    reference_names=names,
                    reference_formulas=formulas,
                    reference_mw=mw_values,
                )

        formula_text = str(formula).strip()
        if formula_text != "":
            fallback_entries = self._nist_index_by_formula.get(formula_text, [])
            matched_entries: List[Dict[str, Any]] = []
            for item in fallback_entries:
                item_mw = item.get("mw")
                if item_mw is None:
                    continue
                if self._is_mass_match(float(item_mw), target_mw) is True:
                    matched_entries.append(item)

            if len(matched_entries) > 0:
                names = sorted({str(item.get("name", "")).strip() for item in matched_entries if str(item.get("name", "")).strip() != ""})
                formulas = sorted({str(item.get("formula", "")).strip() for item in matched_entries if str(item.get("formula", "")).strip() != ""})
                mw_values = sorted({float(item.get("mw")) for item in matched_entries if item.get("mw") is not None})
                return NISTLibraryQueryResult(
                    has_record=True,
                    query_mode="formula_mw_fallback",
                    reference_names=names,
                    reference_formulas=formulas,
                    reference_mw=mw_values,
                )

        return NISTLibraryQueryResult(has_record=False, query_mode="not_found")

    def _extract_rt_for_row(self, row: Dict[str, Any]) -> Optional[float]:
        """
        功能:
            从对照表行中提取用于匹配的保留时间.
            优先 FID 保留时间, 缺失时回退 TIC 保留时间.
        参数:
            row: 对照表行字典.
        返回:
            Optional[float]: 保留时间(min).
        """
        fid_rt = self._parse_opt_float(row.get("FID保留时间(min)"))
        if fid_rt is not None:
            return fid_rt
        tic_rt = self._parse_opt_float(row.get("TIC保留时间(min)"))
        return tic_rt

    def _extract_nist_hits(self, row: Dict[str, Any]) -> List[NISTHitInfo]:
        """
        功能:
            从对照表行动态提取 NIST 命中信息.
            支持化合物1, 化合物2, 化合物3 等任意数量候选.
        参数:
            row: 对照表行字典.
        返回:
            List[NISTHitInfo]: 命中信息列表.
        """
        hits: List[NISTHitInfo] = []
        candidate_ranks: Set[int] = set()

        for key in row.keys():
            if isinstance(key, str) is False:
                continue
            key_text = key.strip()
            match = re.match(r"^化合物(\d+)\((名称|分子式|分子量)\)$", key_text)
            if match is None:
                continue
            candidate_ranks.add(int(match.group(1)))

        for rank in sorted(candidate_ranks):
            name_text = str(row.get(f"化合物{rank}(名称)", "")).strip()
            formula_text = str(row.get(f"化合物{rank}(分子式)", "")).strip()
            mw_value = self._parse_opt_float(row.get(f"化合物{rank}(分子量)"))
            if name_text != "" or formula_text != "" or mw_value is not None:
                hits.append(
                    NISTHitInfo(
                        rank=rank,
                        compound_name=name_text,
                        formula=formula_text,
                        molecular_weight=mw_value,
                    )
                )
        return hits

    def _extract_mass_predictions(
        self, row: Dict[str, Any]
    ) -> Dict[str, Dict[str, Optional[float]]]:
        """
        功能:
            从对照表行中提取已启用算法的分子量预测值与置信度.
        参数:
            row: 对照表行字典.
        返回:
            Dict[str, Dict[str, Optional[float]]]:
                方法名 -> {"mw": 分子量, "confidence": 置信度}.
        """
        predictions: Dict[str, Dict[str, Optional[float]]] = {}
        enabled_methods = self._get_enabled_prediction_methods()
        for method_name in enabled_methods:
            mw_column = self._PREDICTION_MW_COLUMNS[method_name]
            conf_column = self._PREDICTION_CONF_COLUMNS[method_name]
            mw_value = self._parse_opt_float(row.get(mw_column))
            conf_value = self._parse_opt_float(row.get(conf_column))
            predictions[method_name] = {
                "mw": mw_value,
                "confidence": conf_value,
            }
        return predictions

    @staticmethod
    def _clip_zero_to_one(value: float) -> float:
        """
        功能:
            将数值截断到 [0, 1] 区间.
        参数:
            value: 输入数值.
        返回:
            float: 截断后的数值.
        """
        if value < 0.0:
            return 0.0
        if value > 1.0:
            return 1.0
        return value

    def _normalize_prediction_confidence(
        self,
        method_name: str,
        raw_confidence: Optional[float],
        mass_matched: bool,
    ) -> float:
        """
        功能:
            归一化不同预测方法的置信度到 [0, 1] 区间.
        参数:
            method_name: 方法名, PIM/SS-HM/iHS-HM.
            raw_confidence: 原始置信度.
            mass_matched: 分子量是否命中.
        返回:
            float: 归一化置信度.
        """
        if mass_matched is False:
            return 0.0
        if raw_confidence is None:
            return 0.5
        if method_name == "PIM":
            return self._clip_zero_to_one(raw_confidence / 2.0)
        return self._clip_zero_to_one(raw_confidence)

    def _is_mass_match(self, observed_mw: Optional[float], target_mw: Optional[float]) -> bool:
        """
        功能:
            判断观测分子量与目标分子量是否命中.
            命中规则:
            1. 两者都存在时, 分别四舍五入为整数.
            2. 仅当整数完全相等时判定命中.
        参数:
            observed_mw: 观测分子量.
            target_mw: 目标分子量.
        返回:
            bool: 命中返回 True.
        """
        if observed_mw is None:
            return False
        if target_mw is None:
            return False
        observed_int = int(round(observed_mw))
        target_int = int(round(target_mw))
        return observed_int == target_int

    def _is_name_matched(
        self,
        hit_name: str,
        target_name: str,
        reference_names: List[str],
    ) -> bool:
        """
        功能:
            判断 NIST 命中名称是否与目标化合物一致.
        参数:
            hit_name: NIST 命中名称.
            target_name: 目标化合物名称.
            reference_names: 由 NIST 收录查询得到的参考名称集合.
        返回:
            bool: 名称一致返回 True.
        """
        normalized_hit = self._normalize_name(hit_name)
        if normalized_hit == "":
            return False

        candidates: Set[str] = set()
        normalized_target = self._normalize_name(target_name)
        if normalized_target != "":
            candidates.add(normalized_target)
        for item in reference_names:
            normalized_item = self._normalize_name(item)
            if normalized_item != "":
                candidates.add(normalized_item)

        if len(candidates) == 0:
            return False
        return normalized_hit in candidates

    def _classify_nist_evidence(
        self,
        nist_query: NISTLibraryQueryResult,
        nist_target_matched: bool,
        nist_mw_matched: bool,
    ) -> str:
        """
        功能:
            将当前峰的 NIST 证据归类为强支持, 弱支持, 缺失或矛盾.
        参数:
            nist_query: NIST 收录查询结果.
            nist_target_matched: 当前峰的 NIST 候选是否直接支持目标.
            nist_mw_matched: 当前峰的 NIST 候选分子量是否命中目标.
        返回:
            str: 证据层级, 可选 nist_strong/nist_weak/nist_none_absent/nist_none_contradict.
        """
        if nist_target_matched is True:
            return "nist_strong"
        if nist_mw_matched is True:
            return "nist_weak"
        if nist_query.has_record is True:
            return "nist_none_contradict"
        return "nist_none_absent"

    def _summarize_prediction_evidence(
        self,
        mass_match_methods: List[str],
        prediction_details: Dict[str, Dict[str, Optional[float]]],
    ) -> Dict[str, Any]:
        """
        功能:
            汇总各预测方法的命中情况, 归一化置信度和平均预测质量.
        参数:
            mass_match_methods: 分子量预测命中的方法列表.
            prediction_details: 分子量预测值与置信度明细.
        返回:
            Dict[str, Any]: 包含预测层级, 命中方法和逐方法评分明细.
        """
        enabled_methods = self._get_enabled_prediction_methods()
        matched_method_names: List[str] = []
        matched_confidences: List[float] = []
        prediction_scores: Dict[str, Dict[str, Any]] = {}

        for method_name in enabled_methods:
            method_weight = self._PREDICTION_WEIGHTS[method_name]
            method_data = prediction_details.get(method_name, {})
            method_mw = method_data.get("mw")
            method_confidence = method_data.get("confidence")
            method_matched = method_name in mass_match_methods
            method_conf_norm = self._normalize_prediction_confidence(
                method_name=method_name,
                raw_confidence=method_confidence,
                mass_matched=method_matched,
            )
            if method_matched is True:
                matched_method_names.append(method_name)
                matched_confidences.append(method_conf_norm)

            # 保留逐方法明细, 便于解释最终评分来源.
            prediction_scores[method_name] = {
                "weight": method_weight,
                "mw": method_mw,
                "confidence_raw": method_confidence,
                "matched": method_matched,
                "confidence_norm": round(method_conf_norm, 4),
                "score": round(method_weight * method_conf_norm, 2),
            }

        matched_method_count = len(matched_method_names)
        if matched_method_count >= 2:
            prediction_evidence_level = "multi_match"
        elif matched_method_count == 1:
            prediction_evidence_level = "single_match"
        else:
            prediction_evidence_level = "none"

        predictor_quality = 0.0
        if matched_method_count > 0:
            predictor_quality = round(sum(matched_confidences) / matched_method_count, 4)

        return {
            "enabled_methods": enabled_methods,
            "matched_method_names": matched_method_names,
            "matched_method_count": matched_method_count,
            "prediction_evidence_level": prediction_evidence_level,
            "predictor_quality": predictor_quality,
            "prediction_scores": prediction_scores,
        }

    def _resolve_confidence_band(
        self,
        nist_evidence_level: str,
        prediction_evidence_level: str,
        nist_mw_matched: bool,
    ) -> Tuple[str, float, float, float]:
        """
        功能:
            根据 NIST 与预测证据层级选出固定评分分段和倍率.
        参数:
            nist_evidence_level: NIST 证据层级.
            prediction_evidence_level: 预测证据层级.
            nist_mw_matched: 当前峰的 NIST 分子量是否命中目标.
        返回:
            Tuple[str, float, float, float]:
                (评分层级, 分段下界, 分段上界, 预测质量倍率).
        """
        if nist_evidence_level == "nist_strong":
            if nist_mw_matched is True:
                scoring_tier = "nist_strong_with_mw"
            else:
                scoring_tier = "nist_strong_without_mw"
        elif nist_evidence_level == "nist_weak":
            scoring_tier = "nist_weak"
        else:
            scoring_tier = f"{nist_evidence_level}_{prediction_evidence_level}"

        band_min, band_max, quality_multiplier = self._CONFIDENCE_BAND_RULES[scoring_tier]
        return scoring_tier, band_min, band_max, quality_multiplier

    @staticmethod
    def _format_score_band_value(value: float) -> str:
        """
        功能:
            将评分分段端点格式化为紧凑文本.
        参数:
            value: 分段端点值.
        返回:
            str: 适合展示的分值文本.
        """
        if float(value).is_integer():
            return str(int(value))
        return f"{value:.2f}"

    def _describe_nist_evidence(
        self,
        nist_evidence_level: str,
        nist_mw_matched: bool,
    ) -> str:
        """
        功能:
            生成 NIST 证据层级的人类可读描述.
        参数:
            nist_evidence_level: NIST 证据层级.
            nist_mw_matched: 当前峰的 NIST 分子量是否命中目标.
        返回:
            str: NIST 证据描述.
        """
        if nist_evidence_level == "nist_strong":
            if nist_mw_matched is True:
                return "NIST命中目标化合物, 且分子量一致"
            return "NIST命中目标化合物, 但分子量未一致"
        if nist_evidence_level == "nist_weak":
            return "NIST仅支持分子量一致"
        if nist_evidence_level == "nist_none_absent":
            return "NIST无目标记录, 当前峰未获得NIST支持"
        return "NIST有目标记录, 但当前峰候选未支持目标"

    def _describe_prediction_evidence(
        self,
        enabled_methods: List[str],
        matched_method_names: List[str],
    ) -> str:
        """
        功能:
            生成预测证据层级的人类可读描述.
        参数:
            enabled_methods: 当前启用的预测方法列表.
            matched_method_names: 分子量命中的预测方法列表.
        返回:
            str: 预测证据描述.
        """
        if len(enabled_methods) == 0:
            return "未启用预测方法"
        if len(matched_method_names) == 0:
            return "预测未命中"
        return f"预测命中{len(matched_method_names)}项({','.join(matched_method_names)})"

    def _describe_current_peak_nist_support(
        self,
        nist_target_matched: bool,
        nist_formula_matched: bool,
        nist_mw_matched: bool,
    ) -> str:
        """
        功能:
            生成“当前峰是否被 NIST 支持”的文本描述.
        参数:
            nist_target_matched: 当前峰的 NIST 候选是否直接支持目标.
            nist_formula_matched: 当前峰的 NIST 候选分子式是否命中目标.
            nist_mw_matched: 当前峰的 NIST 候选分子量是否命中目标.
        返回:
            str: 支持说明文本.
        """
        if nist_target_matched is True and nist_formula_matched is True and nist_mw_matched is True:
            return "是(目标+分子式+分子量)"
        if nist_target_matched is True and nist_formula_matched is True:
            return "是(目标+分子式)"
        if nist_target_matched is True and nist_mw_matched is True:
            return "是(目标+分子量)"
        if nist_target_matched is True:
            return "是(目标)"
        if nist_mw_matched is True:
            return "是(分子量)"
        return "否"

    @staticmethod
    def _is_full_nist_match(
        nist_target_matched: bool,
        nist_mw_matched: bool,
    ) -> bool:
        """
        功能:
            判断当前峰是否满足“完整 NIST 命中”.
        参数:
            nist_target_matched: 当前峰的 NIST 候选是否支持目标.
            nist_mw_matched: 当前峰的 NIST 候选分子量是否命中目标.
        返回:
            bool: 目标与分子量同时命中时返回 True.
        """
        return nist_target_matched is True and nist_mw_matched is True

    def _resolve_match_method_label(
        self,
        base_match_method: str,
        nist_target_matched: bool,
        nist_mw_matched: bool,
    ) -> str:
        """
        功能:
            将内部匹配路径转换为结果表使用的匹配方式文本.
        参数:
            base_match_method: 内部匹配路径标记.
            nist_target_matched: 当前峰的 NIST 候选是否支持目标.
            nist_mw_matched: 当前峰的 NIST 候选分子量是否命中目标.
        返回:
            str: 匹配方式文本.
        """
        if self._is_full_nist_match(
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
        ) is True:
            return "NIST命中"
        if base_match_method == "rt":
            return "RT命中"
        return "分子量命中"

    def _build_confidence(
        self,
        nist_query: NISTLibraryQueryResult,
        nist_target_matched: bool,
        nist_mw_matched: bool,
        nist_formula_matched: bool,
        mass_match_methods: List[str],
        prediction_details: Dict[str, Dict[str, Optional[float]]],
    ) -> Tuple[float, str, Dict[str, Any]]:
        """
        功能:
            根据 NIST 与分子量预测证据输出 0-100 置信度分数与说明.
            分数使用“证据分层 + 固定分段”规则, 不再按满分归一化.
        参数:
            nist_query: NIST 收录查询结果.
            nist_target_matched: NIST 命中是否为目标化合物.
            nist_mw_matched: NIST 命中分子量是否匹配目标.
            nist_formula_matched: NIST 命中分子式是否匹配目标.
            mass_match_methods: 分子量预测命中的方法列表.
            prediction_details: 分子量预测值与置信度明细.
        返回:
            Tuple[float, str, Dict[str, Any]]:
                (置信度分数, 置信说明, 评分明细字典).
        """
        nist_evidence_level = self._classify_nist_evidence(
            nist_query=nist_query,
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
        )
        prediction_summary = self._summarize_prediction_evidence(
            mass_match_methods=mass_match_methods,
            prediction_details=prediction_details,
        )
        scoring_tier, band_min, band_max, quality_multiplier = self._resolve_confidence_band(
            nist_evidence_level=nist_evidence_level,
            prediction_evidence_level=prediction_summary["prediction_evidence_level"],
            nist_mw_matched=nist_mw_matched,
        )

        predictor_quality = float(prediction_summary["predictor_quality"])
        confidence_score = round(band_min + quality_multiplier * predictor_quality, 2)
        if confidence_score < band_min:
            confidence_score = band_min
        if confidence_score > band_max:
            confidence_score = band_max

        band_min_text = self._format_score_band_value(band_min)
        band_max_text = self._format_score_band_value(band_max)
        reason_parts = [
            self._describe_nist_evidence(
                nist_evidence_level=nist_evidence_level,
                nist_mw_matched=nist_mw_matched,
            ),
            self._describe_prediction_evidence(
                enabled_methods=prediction_summary["enabled_methods"],
                matched_method_names=prediction_summary["matched_method_names"],
            ),
            f"平均预测置信度={predictor_quality:.4f}",
            f"评分分段={band_min_text}-{band_max_text}",
            f"综合得分={int(round(confidence_score))}",
        ]
        detail = {
            "scoring_tier": scoring_tier,
            "nist_evidence_level": nist_evidence_level,
            "prediction_evidence_level": prediction_summary["prediction_evidence_level"],
            "score_band_min": band_min,
            "score_band_max": band_max,
            "predictor_quality": predictor_quality,
            "matched_method_count": prediction_summary["matched_method_count"],
            "matched_method_names": prediction_summary["matched_method_names"],
            "prediction_scores": prediction_summary["prediction_scores"],
        }
        return confidence_score, "; ".join(reason_parts), detail

    def _build_hit_summary(
        self,
        row: Dict[str, Any],
        expected_rt: Optional[float],
        nist_query: NISTLibraryQueryResult,
        nist_target_matched: bool,
        nist_mw_matched: bool,
        nist_formula_matched: bool,
        confidence_score: float,
        confidence_detail: Dict[str, Any],
    ) -> str:
        """
        功能:
            生成单峰命中详情文本, 便于审计所有证据来源.
        参数:
            row: 峰行数据.
            expected_rt: 目标预期 RT.
            nist_query: NIST 收录查询结果.
            nist_target_matched: NIST 名称/分子式命中目标.
            nist_mw_matched: NIST 分子量命中目标.
            nist_formula_matched: NIST 分子式命中目标.
            confidence_score: 综合分数.
            confidence_detail: 评分明细.
        返回:
            str: 命中详情摘要.
        """
        summary_parts: List[str] = []
        row_rt = self._extract_rt_for_row(row)
        if expected_rt is not None and row_rt is not None:
            rt_diff = abs(row_rt - expected_rt)
            summary_parts.append(
                f"RT={row_rt:.3f}, 预期={expected_rt:.3f}, 偏差={rt_diff:.3f}, 容差内={'是' if rt_diff <= self._rt_tolerance else '否'}"
            )
        elif row_rt is not None:
            summary_parts.append(f"RT={row_rt:.3f}, 无预期RT")
        else:
            summary_parts.append("RT缺失")

        summary_parts.append(
            "NIST收录="
            + ("是" if nist_query.has_record is True else "否")
            + f", 查询模式={nist_query.query_mode}, 名称/目标命中={'是' if nist_target_matched is True else '否'}"
            + f", 分子式命中={'是' if nist_formula_matched is True else '否'}"
            + f", 分子量命中={'是' if nist_mw_matched is True else '否'}"
        )

        prediction_scores = confidence_detail.get("prediction_scores", {})
        enabled_methods = self._get_enabled_prediction_methods()
        for method_name in enabled_methods:
            method_data = prediction_scores.get(method_name, {})
            method_mw = method_data.get("mw")
            method_conf_raw = method_data.get("confidence_raw")
            method_matched = method_data.get("matched")
            method_score = method_data.get("score")
            method_conf_norm = method_data.get("confidence_norm")
            if method_mw is None:
                mw_text = "无"
            else:
                mw_text = f"{float(method_mw):.4f}"
            if method_conf_raw is None:
                conf_text = "无"
            else:
                conf_text = f"{float(method_conf_raw):.4f}"
            if method_score is None:
                score_text = "0.00"
            else:
                score_text = f"{float(method_score):.2f}"
            if method_conf_norm is None:
                conf_norm_text = "0.0000"
            else:
                conf_norm_text = f"{float(method_conf_norm):.4f}"
            summary_parts.append(
                f"{method_name}(MW={mw_text}, 置信度={conf_text}, 命中={'是' if method_matched is True else '否'}, 归一化={conf_norm_text}, 参考分={score_text})"
            )

        band_min = float(confidence_detail.get("score_band_min", 0.0))
        band_max = float(confidence_detail.get("score_band_max", 0.0))
        predictor_quality = float(confidence_detail.get("predictor_quality", 0.0))
        scoring_tier = str(confidence_detail.get("scoring_tier", ""))
        summary_parts.append(
            f"评分层级={scoring_tier}, 分段={self._format_score_band_value(band_min)}-{self._format_score_band_value(band_max)}, 平均预测置信度={predictor_quality:.4f}, score={int(round(confidence_score))}"
        )
        return "; ".join(summary_parts)

    def _build_remark(
        self,
        nist_query: NISTLibraryQueryResult,
        nist_target_matched: bool,
        nist_mw_matched: bool,
        nist_formula_matched: bool,
        mass_match_methods: List[str],
        prediction_details: Dict[str, Dict[str, Optional[float]]],
        nist_hit: Optional[NISTHitInfo],
    ) -> str:
        """
        功能:
            生成精简的备注文本, 明确区分目标收录状态与当前峰证据状态.
        参数:
            nist_query: NIST 收录查询结果.
            nist_target_matched: NIST 命中是否为目标化合物.
            nist_mw_matched: NIST 命中分子量是否匹配目标.
            nist_formula_matched: NIST 命中分子式是否匹配目标.
            mass_match_methods: 分子量预测命中的方法列表.
            prediction_details: 分子量预测值与置信度明细.
            nist_hit: 当前峰选中的最佳 NIST 候选.
        返回:
            str: 备注文本.
        """
        parts: List[str] = []
        if self._is_full_nist_match(
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
        ) is True:
            parts.append(f"NIST目标记录=是({nist_query.query_mode})")
        else:
            parts.append("NIST目标记录=否")
        parts.append(
            "当前峰NIST支持="
            + self._describe_current_peak_nist_support(
                nist_target_matched=nist_target_matched,
                nist_formula_matched=nist_formula_matched,
                nist_mw_matched=nist_mw_matched,
            )
        )
        if nist_hit is None or nist_hit.molecular_weight is None:
            parts.append("当前峰NIST最佳候选MW=无")
        else:
            parts.append(f"当前峰NIST最佳候选MW={int(round(float(nist_hit.molecular_weight)))}")

        # 各预测方法的分子量和置信度 (MW 保留整数)
        enabled_methods = self._get_enabled_prediction_methods()
        for method_name in enabled_methods:
            method_data = prediction_details.get(method_name, {})
            mw = method_data.get("mw")
            conf = method_data.get("confidence")
            mw_text = str(int(round(float(mw)))) if mw is not None else "无"
            conf_text = f"{float(conf):.4f}" if conf is not None else "无"
            method_matched = "命中" if method_name in mass_match_methods else "未命中"
            parts.append(f"{method_name}: MW={mw_text}, {method_matched}, 置信度={conf_text}")

        return "; ".join(parts)

    @staticmethod
    def _confidence_rank(level: str) -> int:
        """
        功能:
            将置信度文本映射为排序分值.
            优先支持 0-100 分数字符串, 兼容旧的高/中/低文本.
        参数:
            level: 置信度文本.
        返回:
            int: 排序分值.
        """
        level_text = str(level).strip()
        try:
            numeric_score = float(level_text)
            return int(round(numeric_score))
        except ValueError:
            pass

        if level == "高":
            return 3
        if level == "中":
            return 2
        if level == "低":
            return 1
        return 0

    def _build_peak_decision(
        self,
        row: Dict[str, Any],
        match_method: str,
        target_name: str,
        target_formula: str,
        target_mw: Optional[float],
        nist_query: NISTLibraryQueryResult,
        expected_rt: Optional[float] = None,
    ) -> PeakDecision:
        """
        功能:
            对单行峰数据生成结构化判定结果.
        参数:
            row: 对照表行.
            match_method: 匹配路径标记.
            target_name: 目标化合物名称.
            target_formula: 目标分子式.
            target_mw: 目标分子量.
            nist_query: NIST 收录查询结果.
        返回:
            PeakDecision: 候选峰判定对象.
        """
        hits = self._extract_nist_hits(row)
        predictions = self._extract_mass_predictions(row)
        matched_methods: List[str] = []
        for method_name, method_data in predictions.items():
            method_mw = method_data.get("mw")
            if self._is_mass_match(method_mw, target_mw) is True:
                matched_methods.append(method_name)

        best_hit: Optional[NISTHitInfo] = None
        nist_target_matched = False
        nist_mw_matched = False
        nist_formula_matched = False
        nist_reference_names = nist_query.reference_names
        best_hit_priority = -1
        best_hit_rank = 99

        for hit in hits:
            name_matched = self._is_name_matched(hit.compound_name, target_name, nist_reference_names)
            formula_matched = False
            if target_formula != "" and hit.formula != "":
                formula_matched = (hit.formula == target_formula)
            mw_matched = self._is_mass_match(hit.molecular_weight, target_mw)

            if name_matched is True or formula_matched is True:
                nist_target_matched = True
            if formula_matched is True:
                nist_formula_matched = True
            if mw_matched is True:
                nist_mw_matched = True

            if name_matched is True or formula_matched is True:
                current_priority = 4
            elif mw_matched is True and formula_matched is True:
                current_priority = 3
            elif mw_matched is True:
                current_priority = 2
            else:
                current_priority = 1

            if current_priority > best_hit_priority:
                best_hit_priority = current_priority
                best_hit_rank = hit.rank
                best_hit = hit
            elif current_priority == best_hit_priority and hit.rank < best_hit_rank:
                best_hit_rank = hit.rank
                best_hit = hit

        confidence_score, confidence_reason, confidence_detail = self._build_confidence(
            nist_query=nist_query,
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
            nist_formula_matched=nist_formula_matched,
            mass_match_methods=matched_methods,
            prediction_details=predictions,
        )
        hit_summary = self._build_hit_summary(
            row=row,
            expected_rt=expected_rt,
            nist_query=nist_query,
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
            nist_formula_matched=nist_formula_matched,
            confidence_score=confidence_score,
            confidence_detail=confidence_detail,
        )
        remark = self._build_remark(
            nist_query=nist_query,
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
            nist_formula_matched=nist_formula_matched,
            mass_match_methods=matched_methods,
            prediction_details=predictions,
            nist_hit=best_hit,
        )

        return PeakDecision(
            row=row,
            match_method=self._resolve_match_method_label(
                base_match_method=match_method,
                nist_target_matched=nist_target_matched,
                nist_mw_matched=nist_mw_matched,
            ),
            fid_rt=self._parse_opt_float(row.get("FID保留时间(min)")),
            fid_area=self._parse_opt_float(row.get("FID峰面积")),
            nist_hit=best_hit,
            nist_target_matched=nist_target_matched,
            nist_mw_matched=nist_mw_matched,
            nist_formula_matched=nist_formula_matched,
            pim_mw=predictions.get("PIM", {}).get("mw"),
            sshm_mw=predictions.get("SS-HM", {}).get("mw"),
            ihshm_mw=predictions.get("iHS-HM", {}).get("mw"),
            mass_match_methods=matched_methods,
            confidence_level=str(int(round(confidence_score))),
            confidence_score=confidence_score,
            confidence_reason=confidence_reason,
            hit_summary=hit_summary,
            remark=remark,
        )

    def _sort_peak_decisions(
        self,
        decisions: List[PeakDecision],
        expected_rt: Optional[float] = None,
    ) -> List[PeakDecision]:
        """
        功能:
            按综合分数、NIST命中、峰面积与 RT 偏差排序候选峰.
        参数:
            decisions: 候选峰列表.
            expected_rt: 预期 RT, RT 模式下用于偏差排序.
        返回:
            List[PeakDecision]: 排序后的候选峰列表.
        """
        def _rt_bias(item: PeakDecision) -> float:
            if expected_rt is None:
                return 0.0
            if item.fid_rt is None:
                return float("-inf")
            return -abs(item.fid_rt - expected_rt)

        return sorted(
            decisions,
            key=lambda item: (
                item.confidence_score,
                1 if item.nist_target_matched is True else 0,
                item.fid_area if item.fid_area is not None else -1.0,
                _rt_bias(item),
            ),
            reverse=True,
        )

    def _match_rows_by_rt(
        self,
        sample_rows: List[Dict],
        expected_rt: float,
    ) -> List[Dict[str, Any]]:
        """
        功能:
            按保留时间过滤, 返回 RT 容差范围内的全部峰行.
        参数:
            sample_rows: 对照表数据行列表.
            expected_rt: 预期保留时间(min).
        返回:
            List[Dict[str, Any]]: 容差范围内候选峰行列表.
        """
        matched_rows: List[Dict[str, Any]] = []
        for row in sample_rows:
            row_rt = self._extract_rt_for_row(row)
            if row_rt is None:
                continue
            diff = abs(row_rt - expected_rt)
            if diff <= self._rt_tolerance:
                matched_rows.append(row)
        return matched_rows

    def identify_peaks(
        self,
        sample_rows: List[Dict],
        expected_rt: Optional[float],
        target_name: str,
        target_formula: str,
        target_mw: Optional[float],
        nist_query: NISTLibraryQueryResult,
        allow_multiple: bool = False,
    ) -> List[PeakDecision]:
        """
        功能:
            根据 RT/NIST/分子量预测综合规则识别目标峰.
            规则:
            1. 有 expected_rt 时, 以 RT 匹配为主并输出单峰.
            2. 无 expected_rt 且 NIST 收录时, NIST 命中或分子量预测任一命中即保留.
            3. 无 expected_rt 且 NIST 未收录时, 以分子量预测为主.
        参数:
            sample_rows: 样品对照表行列表.
            expected_rt: 目标预期保留时间.
            target_name: 目标化合物名称.
            target_formula: 目标分子式.
            target_mw: 目标分子量.
            nist_query: NIST 收录查询结果.
            allow_multiple: 是否允许返回多个候选峰.
        返回:
            List[PeakDecision]: 候选峰判定列表.
        """
        decisions: List[PeakDecision] = []

        if expected_rt is not None:
            rt_rows = self._match_rows_by_rt(sample_rows, expected_rt)
            if len(rt_rows) == 0:
                return decisions
            for row in rt_rows:
                decision = self._build_peak_decision(
                    row=row,
                    match_method="rt",
                    target_name=target_name,
                    target_formula=target_formula,
                    target_mw=target_mw,
                    nist_query=nist_query,
                    expected_rt=expected_rt,
                )
                decisions.append(decision)
            decisions = self._sort_peak_decisions(decisions, expected_rt=expected_rt)
            if allow_multiple is False and len(decisions) > 1:
                return [decisions[0]]
            return decisions

        for row in sample_rows:
            decision = self._build_peak_decision(
                row=row,
                match_method="mass",
                target_name=target_name,
                target_formula=target_formula,
                target_mw=target_mw,
                nist_query=nist_query,
                expected_rt=None,
            )

            if nist_query.has_record is True:
                has_nist_evidence = (
                    decision.nist_target_matched is True
                    or decision.nist_mw_matched is True
                )
            else:
                has_nist_evidence = (
                    decision.nist_mw_matched is True
                    and decision.nist_formula_matched is True
                )

            has_mass_evidence = (len(decision.mass_match_methods) > 0)
            if has_nist_evidence is False and has_mass_evidence is False:
                continue

            decisions.append(decision)

        decisions = self._sort_peak_decisions(decisions, expected_rt=None)
        if allow_multiple is False and len(decisions) > 1:
            return [decisions[0]]
        return decisions

    # ------------------------------------------------------------------
    # 产率计算
    # ------------------------------------------------------------------

    def _calculate_yield(
        self,
        fid_area_product: float,
        fid_area_is: float,
        ecn_product: float,
        ecn_is: float,
        config: YieldCalcConfig,
        product_equivalent: float = 1.0,
    ) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
        """
        功能:
            根据计算方法分发到 ECN/标准曲线/响应因子法, 计算产率.
        参数:
            fid_area_product: 产物 FID 峰面积.
            fid_area_is: 内标 FID 峰面积.
            ecn_product: 产物 ECN.
            ecn_is: 内标 ECN.
            config: 产率计算配置.
            product_equivalent: 目标产物当量(eq), 理论产物量 = 反应规模 * equivalent.
        返回:
            Tuple: (area_ratio, molar_ratio, n_product_mol, yield_percent).
                   任何环节失败返回 None.
        """
        if fid_area_is <= 0:
            logger.warning("内标 FID 面积 <= 0, 无法计算面积比")
            return (None, None, None, None)

        ratio = fid_area_product / fid_area_is

        method = config.calc_method.strip().upper()

        # ECN 法
        if method == "ECN":
            if ecn_product <= 0:
                logger.warning("产物 ECN <= 0, 无法计算摩尔比")
                return (ratio, None, None, None)
            molar_ratio = ratio * (ecn_is / ecn_product)

        # 标准曲线法
        elif method in ("标准曲线", "CALIBRATION", "CURVE"):
            if config.curve_slope is None or config.curve_slope == 0:
                logger.warning("标准曲线斜率无效, 无法计算")
                return (ratio, None, None, None)
            intercept = config.curve_intercept if config.curve_intercept is not None else 0.0
            molar_ratio = (ratio - intercept) / config.curve_slope

        # 响应因子法
        elif method in ("响应因子", "RF", "RESPONSE_FACTOR"):
            if config.response_factor is None or config.response_factor == 0:
                logger.warning("响应因子无效, 无法计算")
                return (ratio, None, None, None)
            molar_ratio = ratio * config.response_factor

        else:
            logger.warning("未知计算方法: '%s'", config.calc_method)
            return (ratio, None, None, None)

        # 计算产物物质的量
        if config.is_moles <= 0:
            logger.warning("内标摩尔量 <= 0, 无法计算产物物质的量")
            return (ratio, molar_ratio, None, None)

        n_product = molar_ratio * config.is_moles

        # 计算产率
        if config.reaction_scale_mmol <= 0:
            logger.warning("反应规模 <= 0, 无法计算产率")
            return (ratio, molar_ratio, n_product, None)

        n_theoretical = (config.reaction_scale_mmol * product_equivalent) / 1000.0  # mmol * eq → mol
        yield_percent = (n_product / n_theoretical) * 100.0

        return (ratio, molar_ratio, n_product, yield_percent)

    # ------------------------------------------------------------------
    # 主流程
    # ------------------------------------------------------------------

    def process_task(
        self,
        plan_path: Path,
        report_path: Path,
        chemical_list_path: Path,
    ) -> Tuple[YieldCalcConfig, List[SampleYieldResult]]:
        """
        功能:
            产率计算完整流程:
            1. 从实验方案和 chemical_list 解析配置.
            2. 从积分报告加载 TIC-FID 对照表数据.
            3. 逐 (样品 x 产物) 计算产率.
        参数:
            plan_path: 实验方案 xlsx 路径.
            report_path: 积分报告 xlsx 路径.
            chemical_list_path: chemical_list.xlsx 路径.
        返回:
            Tuple[YieldCalcConfig, List[SampleYieldResult]]:
                配置和所有样品的计算结果列表.
        """
        # 1. 解析配置
        config = self.parse_yield_config(plan_path, chemical_list_path)
        logger.info(
            "产率计算配置: 方法=%s, 内标=%s(%.6f mol), 反应规模=%.2f mmol, 产物数=%d",
            config.calc_method, config.is_name, config.is_moles,
            config.reaction_scale_mmol, len(config.products),
        )

        # 2. 加载对齐数据
        alignment_data = self.load_alignment_data(report_path)
        logger.info("加载了 %d 个样品的 TIC-FID 对照数据", len(alignment_data))

        # 3. 预先查询 NIST 收录状态
        is_nist_query = self._query_nist_library_by_smiles(
            smiles=config.is_smiles,
            formula=config.is_formula,
            target_mw=config.is_molecular_weight,
        )
        config.is_nist_has_record = is_nist_query.has_record
        config.is_nist_query_mode = is_nist_query.query_mode
        config.is_nist_reference_names = is_nist_query.reference_names

        product_nist_query_map: Dict[int, NISTLibraryQueryResult] = {}
        for product in config.products:
            product_nist_query = self._query_nist_library_by_smiles(
                smiles=product.smiles,
                formula=product.formula,
                target_mw=product.molecular_weight,
            )
            product.nist_has_record = product_nist_query.has_record
            product.nist_query_mode = product_nist_query.query_mode
            product.nist_reference_names = product_nist_query.reference_names
            product_nist_query_map[id(product)] = product_nist_query
            logger.info(
                "产物 '%s' NIST收录=%s, 查询模式=%s",
                product.name,
                "是" if product.nist_has_record is True else "否",
                product.nist_query_mode,
            )

        # 4. 逐样品计算
        results: List[SampleYieldResult] = []
        for sample_name in sorted(alignment_data.keys()):
            sample_rows = alignment_data[sample_name]
            exp_num = self._extract_experiment_number(sample_name)

            # 先识别内标峰, 内标始终只取最佳单峰.
            is_decisions = self.identify_peaks(
                sample_rows=sample_rows,
                expected_rt=config.is_expected_rt,
                target_name=config.is_name,
                target_formula=config.is_formula,
                target_mw=config.is_molecular_weight,
                nist_query=is_nist_query,
                allow_multiple=False,
            )
            is_decision = is_decisions[0] if len(is_decisions) > 0 else None

            is_fid_rt: Optional[float] = None
            is_fid_area: Optional[float] = None
            is_match_compound = ""
            if is_decision is not None:
                is_fid_rt = is_decision.fid_rt
                if is_fid_rt is None:
                    is_fid_rt = self._extract_rt_for_row(is_decision.row)
                is_fid_area = is_decision.fid_area
                if is_decision.nist_hit is not None:
                    is_match_compound = is_decision.nist_hit.compound_name

            # 筛选适用于该实验的目标产物.
            for product in config.products:
                if (
                    len(product.applicable_experiments) > 0
                    and exp_num is not None
                    and exp_num not in product.applicable_experiments
                ):
                    continue

                base_result = SampleYieldResult(
                    sample_name=sample_name,
                    product_name=product.name,
                    is_fid_rt=is_fid_rt,
                    is_fid_area=is_fid_area,
                    is_match_compound=is_match_compound,
                    ecn_product=product.ecn,
                    ecn_is=config.is_ecn,
                )

                if is_decision is None:
                    base_result.warnings.append("未检测到内标")
                    results.append(base_result)
                    continue

                if is_fid_area is None or is_fid_area <= 0:
                    base_result.warnings.append("内标FID面积无效")
                    results.append(base_result)
                    continue

                product_nist_query = product_nist_query_map.get(
                    id(product),
                    NISTLibraryQueryResult(),
                )
                prod_decisions = self.identify_peaks(
                    sample_rows=sample_rows,
                    expected_rt=product.expected_rt,
                    target_name=product.name,
                    target_formula=product.formula,
                    target_mw=product.molecular_weight,
                    nist_query=product_nist_query,
                    allow_multiple=True,
                )

                if len(prod_decisions) == 0:
                    base_result.remark = "没有目标产物"
                    if product.expected_rt is not None:
                        base_result.warnings.append("未检测到目标产物(RT匹配失败)")
                    elif product_nist_query.has_record is True:
                        base_result.warnings.append("未检测到满足NIST或分子量条件的目标峰")
                    else:
                        base_result.warnings.append("未检测到分子量预测命中的目标峰")
                    results.append(base_result)
                    continue

                for prod_decision in prod_decisions:
                    result = SampleYieldResult(
                        sample_name=sample_name,
                        product_name=product.name,
                        is_fid_rt=is_fid_rt,
                        is_fid_area=is_fid_area,
                        is_match_compound=is_match_compound,
                        ecn_product=product.ecn,
                        ecn_is=config.is_ecn,
                    )

                    result.product_fid_rt = prod_decision.fid_rt
                    if result.product_fid_rt is None:
                        result.product_fid_rt = self._extract_rt_for_row(prod_decision.row)
                    result.product_fid_area = prod_decision.fid_area
                    if prod_decision.nist_hit is not None:
                        result.product_match_compound = prod_decision.nist_hit.compound_name
                        result.nist_matched_mw = prod_decision.nist_hit.molecular_weight
                    result.match_method = prod_decision.match_method
                    result.confidence_level = prod_decision.confidence_level
                    result.confidence_score = prod_decision.confidence_score
                    result.confidence_reason = prod_decision.confidence_reason
                    result.hit_summary = prod_decision.hit_summary
                    result.remark = prod_decision.remark
                    result.pim_predicted_mw = prod_decision.pim_mw
                    result.sshm_predicted_mw = prod_decision.sshm_mw
                    result.ihshm_predicted_mw = prod_decision.ihshm_mw

                    if result.product_fid_area is None or result.product_fid_area <= 0:
                        result.warnings.append("产物FID面积无效")
                        results.append(result)
                        continue

                    ratio, molar_ratio, n_product, yield_pct = self._calculate_yield(
                        result.product_fid_area,
                        is_fid_area,
                        product.ecn,
                        config.is_ecn,
                        config,
                        product_equivalent=product.equivalent,
                    )
                    result.area_ratio = ratio
                    result.molar_ratio = molar_ratio
                    result.n_product_mol = n_product
                    result.yield_percent = yield_pct
                    results.append(result)

        logger.info("产率计算完成, 共 %d 条结果", len(results))
        return config, results

    # ------------------------------------------------------------------
    # 报告生成
    # ------------------------------------------------------------------

    def generate_yield_report(
        self,
        task_id: str,
        config: YieldCalcConfig,
        results: List[SampleYieldResult],
        output_dir: Path,
    ) -> Path:
        """
        功能:
            生成产率报告 Excel 文件, 包含 "产率计算结果" 和 "计算参数" 两个 Sheet.
        参数:
            task_id: 任务 ID.
            config: 产率计算配置.
            results: 各样品计算结果列表.
            output_dir: 输出目录.
        返回:
            Path: 生成的 xlsx 文件路径.
        """
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{task_id}_yield_report.xlsx"

        wb = openpyxl.Workbook()

        # Sheet 1: 产率计算结果
        ws_result = wb.active
        ws_result.title = "产率计算结果"
        self._write_yield_sheet(ws_result, results)

        # Sheet 2: 计算参数
        ws_params = wb.create_sheet("计算参数")
        self._write_params_sheet(ws_params, config)

        wb.save(str(output_path))
        logger.info("产率报告已保存: %s", output_path)
        return output_path

    def _write_yield_sheet(
        self, ws, results: List[SampleYieldResult]
    ) -> None:
        """写入产率计算结果 Sheet."""
        headers = self._build_yield_headers()
        col_map = {name: idx for idx, name in enumerate(headers, start=1)}
        self._write_header(ws, headers)

        for row_idx, r in enumerate(results, start=2):
            ws.cell(row=row_idx, column=col_map["样品名"], value=r.sample_name)
            ws.cell(row=row_idx, column=col_map["目标产物"], value=r.product_name)
            ws.cell(row=row_idx, column=col_map["产物保留时间(min)"],
                    value=round(r.product_fid_rt, 3) if r.product_fid_rt is not None else "")
            ws.cell(row=row_idx, column=col_map["产物FID面积"],
                    value=round(r.product_fid_area, 6) if r.product_fid_area is not None else "")
            ws.cell(row=row_idx, column=col_map["产物匹配化合物"], value=r.product_match_compound)
            ws.cell(row=row_idx, column=col_map["内标保留时间(min)"],
                    value=round(r.is_fid_rt, 3) if r.is_fid_rt is not None else "")
            ws.cell(row=row_idx, column=col_map["内标FID面积"],
                    value=round(r.is_fid_area, 6) if r.is_fid_area is not None else "")
            ws.cell(row=row_idx, column=col_map["内标匹配化合物"], value=r.is_match_compound)
            ws.cell(row=row_idx, column=col_map["Ratio"],
                    value=round(r.area_ratio, 4) if r.area_ratio is not None else "")
            ws.cell(row=row_idx, column=col_map["产物ECN"],
                    value=round(r.ecn_product, 2) if r.ecn_product is not None else "")
            ws.cell(row=row_idx, column=col_map["内标ECN"],
                    value=round(r.ecn_is, 2) if r.ecn_is is not None else "")
            if r.yield_percent is not None:
                yield_display = "<1" if r.yield_percent < 1 else round(r.yield_percent)
            else:
                yield_display = ""
            ws.cell(row=row_idx, column=col_map["产率(%)"], value=yield_display)
            ws.cell(row=row_idx, column=col_map["匹配方式"], value=r.match_method)
            ws.cell(row=row_idx, column=col_map["置信度"], value=r.confidence_level)
            ws.cell(
                row=row_idx,
                column=col_map["NIST匹配分子量(Da)"],
                value=round(r.nist_matched_mw, 4) if r.nist_matched_mw is not None else "",
            )

            if "PIM预测分子量(Da)" in col_map:
                ws.cell(
                    row=row_idx,
                    column=col_map["PIM预测分子量(Da)"],
                    value=round(r.pim_predicted_mw, 4) if r.pim_predicted_mw is not None else "",
                )
            if "SS-HM预测分子量(Da)" in col_map:
                ws.cell(
                    row=row_idx,
                    column=col_map["SS-HM预测分子量(Da)"],
                    value=round(r.sshm_predicted_mw, 4) if r.sshm_predicted_mw is not None else "",
                )
            if "iHS-HM预测分子量(Da)" in col_map:
                ws.cell(
                    row=row_idx,
                    column=col_map["iHS-HM预测分子量(Da)"],
                    value=round(r.ihshm_predicted_mw, 4) if r.ihshm_predicted_mw is not None else "",
                )

            # 组装最终备注: remark + 额外警告
            remark_text = r.remark
            if len(r.warnings) > 0:
                extra = "; ".join(r.warnings)
                if remark_text != "":
                    remark_text = remark_text + "; " + extra
                else:
                    remark_text = extra
            ws.cell(row=row_idx, column=col_map["备注"], value=remark_text)

        self._auto_column_width(ws)

    def _write_params_sheet(self, ws, config: YieldCalcConfig) -> None:
        """写入计算参数 Sheet, 记录本次计算的完整配置."""
        self._write_header(ws, ["参数", "值"])

        rows = [
            ("产率计算方法", config.calc_method),
            ("反应规模(mmol)", config.reaction_scale_mmol),
            ("", ""),
            ("── 内标信息 ──", ""),
            ("内标名称", config.is_name),
            ("内标SMILES", config.is_smiles),
            ("内标分子式", config.is_formula),
            (
                "内标分子量(Da)",
                round(config.is_molecular_weight, 4) if config.is_molecular_weight is not None else "",
            ),
            ("内标ECN", round(config.is_ecn, 4)),
            ("内标NIST收录", "是" if config.is_nist_has_record is True else "否"),
            ("内标NIST查询模式", config.is_nist_query_mode),
            ("内标用量(μL/mg)", config.is_amount),
            ("内标摩尔量(mmol)", round(config.is_moles * 1000, 6)),
            ("内标预期RT(min)", config.is_expected_rt if config.is_expected_rt is not None else ""),
            ("", ""),
        ]

        # 标准曲线/响应因子参数
        if config.curve_slope is not None:
            rows.append(("标准曲线斜率", config.curve_slope))
        if config.curve_intercept is not None:
            rows.append(("标准曲线截距", config.curve_intercept))
        if config.response_factor is not None:
            rows.append(("响应因子", config.response_factor))

        rows.append(("", ""))
        rows.append(("── 目标产物 ──", ""))

        for i, p in enumerate(config.products, start=1):
            exp_str = (
                ",".join(str(e) for e in p.applicable_experiments)
                if p.applicable_experiments else "all"
            )
            rows.append((f"产物{i} 名称", p.name))
            rows.append((f"产物{i} SMILES", p.smiles))
            rows.append((f"产物{i} 分子式", p.formula))
            rows.append((f"产物{i} 分子量(Da)", round(p.molecular_weight, 4) if p.molecular_weight is not None else ""))
            rows.append((f"产物{i} ECN", round(p.ecn, 4)))
            rows.append((f"产物{i} NIST收录", "是" if p.nist_has_record is True else "否"))
            rows.append((f"产物{i} NIST查询模式", p.nist_query_mode))
            rows.append((f"产物{i} 预期RT(min)", p.expected_rt if p.expected_rt is not None else ""))
            rows.append((f"产物{i} 适用实验", exp_str))
            rows.append((f"产物{i} 当量(eq)", p.equivalent))
            rows.append(("", ""))

        left_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        for row_idx, (key, value) in enumerate(rows, start=2):
            cell_key = ws.cell(row=row_idx, column=1, value=key)
            cell_val = ws.cell(row=row_idx, column=2, value=value)
            cell_key.alignment = left_align
            cell_val.alignment = center_align

        self._auto_column_width(ws)

    # ------------------------------------------------------------------
    # 工具方法
    # ------------------------------------------------------------------

    def _write_header(self, ws, headers: List[str]) -> None:
        """写入表头行并设置样式."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._HEADER_ALIGN

    @staticmethod
    def _auto_column_width(ws) -> None:
        """根据内容自动调整列宽, 所有单元格居中对齐, 备注列列宽加大."""
        center_align = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            header_value = str(col_cells[0].value) if col_cells[0].value is not None else ""
            for cell in col_cells:
                if cell.value is not None:
                    val_str = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_length = max(max_length, char_len)
                # 居中对齐 (表头行保留原样式, 数据行设置居中)
                if cell.row > 1:
                    cell.alignment = center_align
            # 备注列使用更大的列宽上限, 确保容纳完整文本
            if header_value == "备注":
                ws.column_dimensions[col_letter].width = min(max_length + 3, 100)
            else:
                ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    @staticmethod
    def _parse_float(value: Any, default: float = 0.0) -> float:
        """安全解析浮点数."""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            text = str(value).strip()
            numbers = re.findall(r"[0-9]+(?:\.[0-9]+)?", text)
            if numbers:
                return float(numbers[0])
            return default

    @staticmethod
    def _parse_opt_float(value: Any) -> Optional[float]:
        """安全解析可选浮点数, 无效返回 None."""
        if value is None:
            return None
        try:
            result = float(value)
            return result
        except (ValueError, TypeError):
            text = str(value).strip()
            if text == "":
                return None
            numbers = re.findall(r"[0-9]+(?:\.[0-9]+)?", text)
            if numbers:
                return float(numbers[0])
            return None

