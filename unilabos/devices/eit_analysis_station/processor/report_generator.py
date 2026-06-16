#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
功能:
    将积分结果汇总为 Excel (.xlsx) 表格.
    按任务汇总, 每个任务生成一个 xlsx 文件,
    包含 TIC 峰表, FID 峰表, TIC-FID 对照表, 样品汇总四个 Sheet.
参数:
    无.
返回:
    无.
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .nist_matcher import CompoundMatch
from .peak_integrator import PeakResult

logger = logging.getLogger(__name__)


@dataclass
class PIMPrediction:
    """
    功能:
        存储单个峰的 PIM 分子量预测结果.
    参数:
        predicted_mz: 预测分子离子峰 m/z.
        predicted_mw: 预测分子量(Da).
        confidence_index: PIM 置信指数.
        status: 结果状态, 可选 ok/no_spectrum/error.
        message: 状态说明.
    返回:
        PIMPrediction.
    """
    predicted_mz: Optional[int] = None
    predicted_mw: Optional[int] = None
    confidence_index: Optional[float] = None
    status: str = ""
    message: str = ""


@dataclass
class SSHMPrediction:
    """
    功能:
        存储单个峰的 SS-HM (Simple Search Hitlist Method) 分子量预测结果.
    参数:
        predicted_mw: SS-HM 预测分子量(Da).
        confidence: 概率置信度 (0-1).
        correction: 最佳修正值 (sigma = PIM + correction).
        status: 结果状态, 可选 ok/no_spectrum/error.
        message: 状态说明.
    返回:
        SSHMPrediction.
    """
    predicted_mw: Optional[int] = None
    confidence: Optional[float] = None
    correction: Optional[int] = None
    status: str = ""
    message: str = ""


@dataclass
class iHSHMPrediction:
    """
    功能:
        存储单个峰的 iHS-HM (iterative Hybrid Search Hitlist Method) 分子量预测结果.
    参数:
        predicted_mw: iHS-HM 预测分子量(Da).
        confidence: 置信度 (Omega1 - Omega2) / 999.
        status: 结果状态, 可选 ok/no_spectrum/error.
        message: 状态说明.
    返回:
        iHSHMPrediction.
    """
    predicted_mw: Optional[int] = None
    confidence: Optional[float] = None
    status: str = ""
    message: str = ""


@dataclass
class SampleResult:
    """
    功能:
        存储单个样品的积分结果(TIC + FID + 化合物匹配).
    参数:
        sample_name: 样品名称.
        d_dir: .D 目录路径.
        tic_peaks: TIC 峰检测与积分结果列表.
        fid_peaks: FID 峰检测与积分结果列表.
        compound_matches: 保留时间 -> 化合物匹配结果列表.
        acq_time: 采集时间字符串.
        nist_result_path: NIST SRCRESLT 结果文件副本路径.
        tic_plot_path: TIC 色谱图图片路径.
        fid_plot_path: FID 色谱图图片路径.
        pim_predictions: 保留时间 -> PIM 预测结果字典.
    返回:
        SampleResult.
    """
    sample_name: str = ""
    d_dir: Path = field(default_factory=Path)
    tic_peaks: List[PeakResult] = field(default_factory=list)
    fid_peaks: List[PeakResult] = field(default_factory=list)
    compound_matches: Dict[float, List[CompoundMatch]] = field(default_factory=dict)
    acq_time: str = ""
    nist_result_path: Optional[Path] = None
    tic_plot_path: Optional[Path] = None
    fid_plot_path: Optional[Path] = None
    ms_plot_paths: Dict[int, Path] = field(default_factory=dict)  # 峰号(1-based) -> 质谱图路径
    pim_predictions: Dict[float, PIMPrediction] = field(default_factory=dict)  # 保留时间 -> PIM 预测结果
    sshm_predictions: Dict[float, SSHMPrediction] = field(default_factory=dict)  # 保留时间 -> SS-HM 预测结果
    ihshm_predictions: Dict[float, iHSHMPrediction] = field(default_factory=dict)  # 保留时间 -> iHS-HM 预测结果

class ReportGenerator:
    """
    功能:
        将积分结果汇总为 Excel 表格.
        按任务汇总, 生成包含 TIC峰表/FID峰表/TIC-FID对照表/样品汇总 四个 Sheet 的 xlsx 文件.
    参数:
        无.
    返回:
        无.
    """

    # 表头样式
    _HEADER_FONT = Font(bold=True, size=11)
    _HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    # FID 峰表列定义
    _FID_HEADERS = [
        "样品名", "峰号", "保留时间(min)", "峰高", "峰面积",
        "面积%", "峰起始(min)", "峰结束(min)", "峰宽(min)",
    ]

    # 样品汇总列定义
    _SUMMARY_HEADERS = [
        "样品名", "TIC峰数", "FID峰数", "TIC总面积", "FID总面积", "采集时间",
        "TIC色谱图", "FID色谱图",
    ]

    def generate_task_report(
        self,
        task_id: str,
        sample_results: List[SampleResult],
        output_dir: Path,
        structure_images: Optional[Dict[str, Optional[Path]]] = None,
        nist_max_hits: int = 5,
        alignment_tolerance: float = 0.05,
        include_tic_only: bool = True,
        include_fid_only: bool = True,
        pim_enabled: bool = True,
        sshm_enabled: bool = True,
        ihshm_enabled: bool = True,
    ) -> Path:
        """
        功能:
            生成任务级汇总 Excel 报告.
        参数:
            task_id: 任务 ID.
            sample_results: 各样品的积分结果列表.
            output_dir: 输出目录.
            structure_images: 结构键 -> 结构图 PNG 路径映射, None 表示不嵌入结构图.
            nist_max_hits: NIST 匹配化合物数量上限, 与 setting.py 中配置保持一致.
            alignment_tolerance: FID 与 TIC 峰保留时间对齐容差(min).
            include_tic_only: 对照表是否输出 TIC 有峰但 FID 无峰的行.
            include_fid_only: 对照表是否输出 FID 有峰但 TIC 无峰的行.
            pim_enabled: 是否在报告中包含 PIM 预测列.
            sshm_enabled: 是否在报告中包含 SS-HM 预测列.
            ihshm_enabled: 是否在报告中包含 iHS-HM 预测列.
        返回:
            Path: 生成的 xlsx 文件路径.
        """
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{task_id}_integration_report.xlsx"

        wb = openpyxl.Workbook()

        # Sheet 1: TIC 峰表
        ws_tic = wb.active
        ws_tic.title = "TIC峰表"
        self._write_tic_sheet(
            ws_tic, sample_results, structure_images,
            nist_max_hits=nist_max_hits,
            pim_enabled=pim_enabled,
            sshm_enabled=sshm_enabled,
            ihshm_enabled=ihshm_enabled,
        )

        # Sheet 2: FID 峰表
        ws_fid = wb.create_sheet("FID峰表")
        self._write_fid_sheet(ws_fid, sample_results)

        # Sheet 3: TIC-FID 对照表
        ws_align = wb.create_sheet("TIC-FID对照表")
        self._write_alignment_sheet(
            ws_align,
            sample_results=sample_results,
            structure_images=structure_images,
            nist_max_hits=nist_max_hits,
            tolerance=alignment_tolerance,
            include_tic_only=include_tic_only,
            include_fid_only=include_fid_only,
            pim_enabled=pim_enabled,
            sshm_enabled=sshm_enabled,
            ihshm_enabled=ihshm_enabled,
        )

        # Sheet 4: 样品汇总
        ws_summary = wb.create_sheet("样品汇总")
        self._write_summary_sheet(ws_summary, sample_results)

        wb.save(str(output_path))
        logger.info("积分报告已保存: %s", output_path)
        return output_path

    def _write_header(self, ws, headers: List[str]) -> None:
        """写入表头行并设置样式."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._HEADER_ALIGN

    def _auto_column_width(self, ws) -> None:
        """根据内容自动调整列宽."""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    # 中文字符按2字节计算宽度
                    val_str = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_length = max(max_length, char_len)
            ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    @staticmethod
    def _build_tic_headers(
        compound_slot_count: int,
        pim_enabled: bool, sshm_enabled: bool, ihshm_enabled: bool,
    ) -> List[str]:
        """
        功能:
            根据启用的预测方法动态构建 TIC 峰表表头列表.
        参数:
            compound_slot_count: 化合物候选列组数量.
            pim_enabled: 是否启用 PIM 预测方法.
            sshm_enabled: 是否启用 SS-HM 预测方法.
            ihshm_enabled: 是否启用 iHS-HM 预测方法.
        返回:
            List[str], TIC 峰表表头列表.
        """
        headers = [
            "样品名", "峰号", "保留时间(min)", "峰高", "峰面积",
            "面积%", "峰起始(min)", "峰结束(min)", "峰宽(min)",
        ]
        headers.extend(
            ReportGenerator._build_compound_headers(compound_slot_count)
        )
        headers.append("质谱图")
        if pim_enabled is True:
            headers.extend(["PIM预测分子量(Da)", "PIM置信指数"])
        if sshm_enabled is True:
            headers.extend(["SS-HM预测分子量(Da)", "SS-HM置信度"])
        if ihshm_enabled is True:
            headers.extend(["iHS-HM预测分子量(Da)", "iHS-HM置信度"])
        return headers

    @staticmethod
    def _build_alignment_headers(
        compound_slot_count: int,
        pim_enabled: bool, sshm_enabled: bool, ihshm_enabled: bool,
    ) -> List[str]:
        """
        功能:
            根据启用的预测方法动态构建 TIC-FID 对照表表头列表.
        参数:
            compound_slot_count: 化合物候选列组数量.
            pim_enabled: 是否启用 PIM 预测方法.
            sshm_enabled: 是否启用 SS-HM 预测方法.
            ihshm_enabled: 是否启用 iHS-HM 预测方法.
        返回:
            List[str], TIC-FID 对照表表头列表.
        """
        headers = [
            "样品名", "FID峰号", "FID保留时间(min)",
            "TIC峰号", "TIC保留时间(min)", "FID峰面积",
        ]
        headers.extend(
            ReportGenerator._build_compound_headers(compound_slot_count)
        )
        if pim_enabled is True:
            headers.extend(["PIM预测分子量(Da)", "PIM置信指数"])
        if sshm_enabled is True:
            headers.extend(["SS-HM预测分子量(Da)", "SS-HM置信度"])
        if ihshm_enabled is True:
            headers.extend(["iHS-HM预测分子量(Da)", "iHS-HM置信度"])
        headers.append("质谱图")
        return headers

    @staticmethod
    def _build_col_map(headers: List[str]) -> Dict[str, int]:
        """
        功能:
            将表头列表转换为列名 -> 1-based 列号的映射字典.
        参数:
            headers: 表头名称列表.
        返回:
            Dict[str, int], 列名到列号的映射.
        """
        return {name: idx for idx, name in enumerate(headers, start=1)}

    @staticmethod
    def _normalize_nist_max_hits(nist_max_hits: int) -> int:
        """
        功能:
            归一化 NIST 匹配化合物数量上限, 确保至少为 1.
        参数:
            nist_max_hits: setting.py 中配置的命中数量上限.
        返回:
            int, 归一化后的命中数量上限.
        """
        if nist_max_hits < 1:
            return 1
        return nist_max_hits

    @staticmethod
    def _build_compound_headers(compound_slot_count: int) -> List[str]:
        """
        功能:
            根据化合物候选数量构建动态列组表头.
        参数:
            compound_slot_count: 化合物候选列组数量.
        返回:
            List[str], 动态化合物列组表头.
        """
        headers: List[str] = []
        for rank in range(1, compound_slot_count + 1):
            headers.extend([
                f"化合物{rank}(名称)",
                f"化合物{rank}(匹配度)",
                f"化合物{rank}(分子式)",
                f"化合物{rank}(分子量)",
            ])
        return headers

    @staticmethod
    def _build_compound_col_groups(
        col_map: Dict[str, int],
        compound_slot_count: int,
    ) -> List[Tuple[int, int, int, int]]:
        """
        功能:
            将动态化合物列组转换为列号元组列表.
        参数:
            col_map: 表头名称到列号的映射.
            compound_slot_count: 化合物候选列组数量.
        返回:
            List[Tuple[int, int, int, int]], 每组依次为名称/匹配度/分子式/分子量列号.
        """
        compound_cols: List[Tuple[int, int, int, int]] = []
        for rank in range(1, compound_slot_count + 1):
            compound_cols.append(
                (
                    col_map[f"化合物{rank}(名称)"],
                    col_map[f"化合物{rank}(匹配度)"],
                    col_map[f"化合物{rank}(分子式)"],
                    col_map[f"化合物{rank}(分子量)"],
                )
            )
        return compound_cols

    def _determine_compound_slot_count(
        self,
        sample_results: List[SampleResult],
        nist_max_hits: int,
    ) -> int:
        """
        功能:
            根据 setting.py 中的命中上限和当前任务实际结果, 确定报表中的化合物列组数量.
        参数:
            sample_results: 各样品积分结果列表.
            nist_max_hits: setting.py 中配置的命中数量上限.
        返回:
            int, 报表应输出的化合物列组数量.
        """
        normalized_hits = self._normalize_nist_max_hits(nist_max_hits)
        actual_max_hits = 0
        for sample_result in sample_results:
            for match_list in sample_result.compound_matches.values():
                actual_max_hits = max(actual_max_hits, len(match_list))

        if actual_max_hits < 1:
            return 1
        return min(normalized_hits, actual_max_hits)

    @staticmethod
    def _format_acq_time(raw_time: str) -> str:
        """
        功能:
            将 ISO 8601 采集时间字符串格式化为 YYYY-MM-DD HH:MM:SS.
            解析失败时原样返回, 空字符串直接返回.
        参数:
            raw_time: 原始时间字符串, 如 "2026-02-26T22:34:30.9070148+08:00".
        返回:
            str, 格式化后的时间字符串, 如 "2026-02-26 22:34:30".
        """
        if not raw_time:
            return raw_time
        try:
            # 截断子秒到6位, 确保 fromisoformat 兼容 Python < 3.11
            truncated = re.sub(r"(\.\d{1,6})\d*", r"\1", raw_time)
            dt = datetime.fromisoformat(truncated)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            return raw_time

    def _write_tic_sheet(
        self,
        ws,
        sample_results: List[SampleResult],
        structure_images: Optional[Dict[str, Optional[Path]]] = None,
        nist_max_hits: int = 5,
        pim_enabled: bool = True,
        sshm_enabled: bool = True,
        ihshm_enabled: bool = True,
    ) -> None:
        """
        功能:
            写入 TIC 峰表 Sheet, 包含所有样品的 TIC 峰检测积分结果,
            多候选化合物匹配结果, 并在化合物名称列写入结构图超链接.
            根据启用的预测方法动态构建表头, 未启用的方法不出现对应列.
        参数:
            ws: openpyxl Worksheet.
            sample_results: 各样品积分结果列表.
            structure_images: 结构键 -> 结构图 PNG 路径, None 表示无结构图.
            nist_max_hits: setting.py 中配置的命中数量上限.
            pim_enabled: 是否启用 PIM 预测方法.
            sshm_enabled: 是否启用 SS-HM 预测方法.
            ihshm_enabled: 是否启用 iHS-HM 预测方法.
        返回:
            无.
        """
        compound_slot_count = self._determine_compound_slot_count(
            sample_results=sample_results,
            nist_max_hits=nist_max_hits,
        )
        headers = self._build_tic_headers(
            compound_slot_count, pim_enabled, sshm_enabled, ihshm_enabled
        )
        col = self._build_col_map(headers)
        self._write_header(ws, headers)

        # 化合物列组: (名称列, 匹配度列, 分子式列, 分子量列)
        compound_cols = self._build_compound_col_groups(col, compound_slot_count)

        row = 2
        for sr in sample_results:
            for peak_num, peak in enumerate(sr.tic_peaks, start=1):
                # 按保留时间查找化合物匹配列表
                match_list = self._find_match_list(
                    peak.retention_time, sr.compound_matches
                )

                ws.cell(row=row, column=col["样品名"], value=sr.sample_name)
                ws.cell(row=row, column=col["峰号"], value=peak_num)
                ws.cell(row=row, column=col["保留时间(min)"], value=round(peak.retention_time, 3))
                ws.cell(row=row, column=col["峰高"], value=round(peak.height, 0))
                ws.cell(row=row, column=col["峰面积"], value=round(peak.area, 2))
                ws.cell(row=row, column=col["面积%"], value=round(peak.area_percent, 2))
                ws.cell(row=row, column=col["峰起始(min)"], value=round(peak.start_time, 3))
                ws.cell(row=row, column=col["峰结束(min)"], value=round(peak.end_time, 3))
                ws.cell(row=row, column=col["峰宽(min)"], value=round(peak.width, 3))

                # 填充化合物候选列组 (每个化合物占4列: 名称, 匹配度, 分子式, 分子量)
                for i, (c_name, c_score, c_formula, c_mw) in enumerate(compound_cols):
                    if match_list is not None and i < len(match_list):
                        m = match_list[i]
                        name_cell = ws.cell(row=row, column=c_name, value=m.compound_name)
                        ws.cell(row=row, column=c_score, value=round(m.match_score, 1))
                        # 名称单元格挂结构图链接, 保持表头简洁.
                        self._link_compound_name_cell(name_cell, m, structure_images)

                        # 分子式和分子量
                        ws.cell(row=row, column=c_formula, value=m.formula)
                        ws.cell(row=row, column=c_mw, value=round(m.mw, 2) if m.mw else "")
                    else:
                        ws.cell(row=row, column=c_name, value="")
                        ws.cell(row=row, column=c_score, value="")
                        ws.cell(row=row, column=c_formula, value="")
                        ws.cell(row=row, column=c_mw, value="")

                # 质谱图超链接
                if peak_num in sr.ms_plot_paths:
                    ms_path = sr.ms_plot_paths[peak_num]
                    if ms_path.exists():
                        cell = ws.cell(row=row, column=col["质谱图"], value="查看质谱")
                        cell.hyperlink = str(ms_path)
                        cell.font = Font(color="0563C1", underline="single")

                # PIM 预测结果
                if pim_enabled is True:
                    pim_prediction = self._find_pim_prediction(
                        peak.retention_time, sr.pim_predictions
                    )
                    if pim_prediction is not None:
                        if pim_prediction.predicted_mw is not None:
                            ws.cell(row=row, column=col["PIM预测分子量(Da)"], value=pim_prediction.predicted_mw)
                        if pim_prediction.confidence_index is not None:
                            ws.cell(row=row, column=col["PIM置信指数"], value=round(pim_prediction.confidence_index, 4))

                # SS-HM 预测结果
                if sshm_enabled is True:
                    sshm_prediction = self._find_prediction_by_rt(
                        peak.retention_time, sr.sshm_predictions
                    )
                    if sshm_prediction is not None:
                        if sshm_prediction.predicted_mw is not None:
                            ws.cell(row=row, column=col["SS-HM预测分子量(Da)"], value=sshm_prediction.predicted_mw)
                        if sshm_prediction.confidence is not None:
                            ws.cell(row=row, column=col["SS-HM置信度"], value=round(sshm_prediction.confidence, 4))

                # iHS-HM 预测结果
                if ihshm_enabled is True:
                    ihshm_prediction = self._find_prediction_by_rt(
                        peak.retention_time, sr.ihshm_predictions
                    )
                    if ihshm_prediction is not None:
                        if ihshm_prediction.predicted_mw is not None:
                            ws.cell(row=row, column=col["iHS-HM预测分子量(Da)"], value=ihshm_prediction.predicted_mw)
                        if ihshm_prediction.confidence is not None:
                            ws.cell(row=row, column=col["iHS-HM置信度"], value=round(ihshm_prediction.confidence, 6))

                row += 1

        self._auto_column_width(ws)

    def _write_fid_sheet(self, ws, sample_results: List[SampleResult]) -> None:
        """
        功能:
            写入 FID 峰表 Sheet, 包含所有样品的 FID 峰检测积分结果.
        """
        self._write_header(ws, self._FID_HEADERS)

        row = 2
        for sr in sample_results:
            for peak_num, peak in enumerate(sr.fid_peaks, start=1):
                ws.cell(row=row, column=1, value=sr.sample_name)
                ws.cell(row=row, column=2, value=peak_num)
                ws.cell(row=row, column=3, value=round(peak.retention_time, 3))
                ws.cell(row=row, column=4, value=round(peak.height, 4))
                ws.cell(row=row, column=5, value=round(peak.area, 6))
                ws.cell(row=row, column=6, value=round(peak.area_percent, 2))
                ws.cell(row=row, column=7, value=round(peak.start_time, 3))
                ws.cell(row=row, column=8, value=round(peak.end_time, 3))
                ws.cell(row=row, column=9, value=round(peak.width, 3))
                row += 1

        self._auto_column_width(ws)

    def _write_summary_sheet(self, ws, sample_results: List[SampleResult]) -> None:
        """
        功能:
            写入样品汇总 Sheet, 每个样品一行, 包含峰数、总面积和色谱图超链接.
        """
        self._write_header(ws, self._SUMMARY_HEADERS)

        for row_idx, sr in enumerate(sample_results, start=2):
            tic_total_area = sum(p.area for p in sr.tic_peaks)
            fid_total_area = sum(p.area for p in sr.fid_peaks)

            ws.cell(row=row_idx, column=1, value=sr.sample_name)
            ws.cell(row=row_idx, column=2, value=len(sr.tic_peaks))
            ws.cell(row=row_idx, column=3, value=len(sr.fid_peaks))
            ws.cell(row=row_idx, column=4, value=round(tic_total_area, 2))
            ws.cell(row=row_idx, column=5, value=round(fid_total_area, 6))
            ws.cell(row=row_idx, column=6, value=self._format_acq_time(sr.acq_time))

            # TIC 色谱图超链接
            if sr.tic_plot_path is not None and sr.tic_plot_path.exists():
                cell = ws.cell(row=row_idx, column=7, value="查看色谱图")
                cell.hyperlink = str(sr.tic_plot_path)
                cell.font = Font(color="0563C1", underline="single")

            # FID 色谱图超链接
            if sr.fid_plot_path is not None and sr.fid_plot_path.exists():
                cell = ws.cell(row=row_idx, column=8, value="查看色谱图")
                cell.hyperlink = str(sr.fid_plot_path)
                cell.font = Font(color="0563C1", underline="single")

        self._auto_column_width(ws)

    @staticmethod
    def _align_fid_tic_peaks(
        fid_peaks: List[PeakResult],
        tic_peaks: List[PeakResult],
        tolerance: float = 0.05,
    ) -> List[Tuple[Optional[PeakResult], Optional[int],
                     Optional[PeakResult], Optional[int]]]:
        """
        功能:
            按保留时间对齐 FID 峰和 TIC 峰.
            贪心匹配: 遍历 FID 峰, 对每个 FID 峰在未匹配的 TIC 峰中
            找保留时间最接近且在容差内的峰进行配对.
        参数:
            fid_peaks: FID 峰列表.
            tic_peaks: TIC 峰列表.
            tolerance: 保留时间容差(min), 在此范围内视为同一峰.
        返回:
            List of (fid_peak, fid_peak_num, tic_peak, tic_peak_num).
            未匹配的峰对应字段为 None.
        """
        matched_tic: set = set()  # 已匹配的 TIC 峰索引集合
        pairs: List[Tuple[Optional[PeakResult], Optional[int],
                          Optional[PeakResult], Optional[int]]] = []

        # 遍历 FID 峰, 贪心匹配最近的 TIC 峰
        for fi, fp in enumerate(fid_peaks):
            best_ti: Optional[int] = None
            best_diff = tolerance + 1.0
            for ti, tp in enumerate(tic_peaks):
                if ti in matched_tic:
                    continue
                diff = abs(fp.retention_time - tp.retention_time)
                if diff <= tolerance and diff < best_diff:
                    best_diff = diff
                    best_ti = ti
            if best_ti is not None:
                # FID-TIC 配对成功
                pairs.append((fp, fi + 1, tic_peaks[best_ti], best_ti + 1))
                matched_tic.add(best_ti)
            else:
                # FID 峰无对应 TIC 峰
                pairs.append((fp, fi + 1, None, None))

        # 收集未匹配的 TIC 峰
        for ti, tp in enumerate(tic_peaks):
            if ti not in matched_tic:
                pairs.append((None, None, tp, ti + 1))

        # 按保留时间排序 (取非 None 峰的 RT)
        pairs.sort(key=lambda x: (x[0] or x[2]).retention_time)
        return pairs

    def _write_alignment_sheet(
        self,
        ws,
        sample_results: List[SampleResult],
        structure_images: Optional[Dict[str, Optional[Path]]] = None,
        nist_max_hits: int = 5,
        tolerance: float = 0.05,
        include_tic_only: bool = True,
        include_fid_only: bool = True,
        pim_enabled: bool = True,
        sshm_enabled: bool = True,
        ihshm_enabled: bool = True,
    ) -> None:
        """
        功能:
            写入 TIC-FID 对照表 Sheet, 按保留时间对齐 FID 和 TIC 峰,
            并展示 TIC 峰对应的多候选化合物预测结果.
            根据启用的预测方法动态构建表头, 未启用的方法不出现对应列.
        参数:
            ws: openpyxl Worksheet.
            sample_results: 各样品积分结果列表.
            structure_images: 结构键 -> 结构图 PNG 路径, None 表示无结构图.
            nist_max_hits: setting.py 中配置的命中数量上限.
            tolerance: FID-TIC 峰保留时间对齐容差(min).
            include_tic_only: 是否输出 TIC 有峰但 FID 无峰的行.
            include_fid_only: 是否输出 FID 有峰但 TIC 无峰的行.
            pim_enabled: 是否启用 PIM 预测方法.
            sshm_enabled: 是否启用 SS-HM 预测方法.
            ihshm_enabled: 是否启用 iHS-HM 预测方法.
        返回:
            无.
        """
        compound_slot_count = self._determine_compound_slot_count(
            sample_results=sample_results,
            nist_max_hits=nist_max_hits,
        )
        headers = self._build_alignment_headers(
            compound_slot_count, pim_enabled, sshm_enabled, ihshm_enabled
        )
        col = self._build_col_map(headers)
        self._write_header(ws, headers)

        # 化合物列组: (名称列, 匹配度列, 分子式列, 分子量列)
        compound_cols = self._build_compound_col_groups(col, compound_slot_count)

        row = 2
        for sr in sample_results:
            aligned = self._align_fid_tic_peaks(
                sr.fid_peaks, sr.tic_peaks, tolerance
            )
            for fid_peak, fid_num, tic_peak, tic_num in aligned:
                # 根据配置过滤仅单侧有峰的行
                if fid_peak is None and not include_tic_only:
                    continue
                if tic_peak is None and not include_fid_only:
                    continue

                ws.cell(row=row, column=col["样品名"], value=sr.sample_name)

                # FID 峰信息
                if fid_peak is not None:
                    ws.cell(row=row, column=col["FID峰号"], value=fid_num)
                    ws.cell(row=row, column=col["FID保留时间(min)"],
                            value=round(fid_peak.retention_time, 3))
                    ws.cell(row=row, column=col["FID峰面积"],
                            value=round(fid_peak.area, 6))

                # TIC 峰信息
                if tic_peak is not None:
                    ws.cell(row=row, column=col["TIC峰号"], value=tic_num)
                    ws.cell(row=row, column=col["TIC保留时间(min)"],
                            value=round(tic_peak.retention_time, 3))

                    # 查找化合物匹配结果
                    match_list = self._find_match_list(
                        tic_peak.retention_time, sr.compound_matches
                    )
                    # 填充化合物候选列组 (每个化合物占4列: 名称/匹配度/分子式/分子量)
                    for i, (c_name, c_score, c_formula, c_mw) in enumerate(compound_cols):
                        if match_list is not None and i < len(match_list):
                            m = match_list[i]
                            name_cell = ws.cell(row=row, column=c_name, value=m.compound_name)
                            # 对照表同样在名称列挂结构图链接.
                            self._link_compound_name_cell(name_cell, m, structure_images)
                            ws.cell(row=row, column=c_score,
                                    value=round(m.match_score, 1))
                            ws.cell(row=row, column=c_formula,
                                    value=m.formula)
                            ws.cell(row=row, column=c_mw,
                                    value=round(m.mw, 2) if m.mw else "")
                        else:
                            ws.cell(row=row, column=c_name, value="")
                            ws.cell(row=row, column=c_score, value="")
                            ws.cell(row=row, column=c_formula, value="")
                            ws.cell(row=row, column=c_mw, value="")

                    # PIM 预测结果
                    if pim_enabled is True:
                        pim_prediction = self._find_pim_prediction(
                            tic_peak.retention_time, sr.pim_predictions
                        )
                        if pim_prediction is not None:
                            if pim_prediction.predicted_mw is not None:
                                ws.cell(row=row, column=col["PIM预测分子量(Da)"], value=pim_prediction.predicted_mw)
                            if pim_prediction.confidence_index is not None:
                                ws.cell(row=row, column=col["PIM置信指数"], value=round(pim_prediction.confidence_index, 4))

                    # SS-HM 预测结果
                    if sshm_enabled is True:
                        sshm_prediction = self._find_prediction_by_rt(
                            tic_peak.retention_time, sr.sshm_predictions
                        )
                        if sshm_prediction is not None:
                            if sshm_prediction.predicted_mw is not None:
                                ws.cell(row=row, column=col["SS-HM预测分子量(Da)"], value=sshm_prediction.predicted_mw)
                            if sshm_prediction.confidence is not None:
                                ws.cell(row=row, column=col["SS-HM置信度"], value=round(sshm_prediction.confidence, 4))

                    # iHS-HM 预测结果
                    if ihshm_enabled is True:
                        ihshm_prediction = self._find_prediction_by_rt(
                            tic_peak.retention_time, sr.ihshm_predictions
                        )
                        if ihshm_prediction is not None:
                            if ihshm_prediction.predicted_mw is not None:
                                ws.cell(row=row, column=col["iHS-HM预测分子量(Da)"], value=ihshm_prediction.predicted_mw)
                            if ihshm_prediction.confidence is not None:
                                ws.cell(row=row, column=col["iHS-HM置信度"], value=round(ihshm_prediction.confidence, 6))

                    # 质谱图超链接
                    if tic_num is not None:
                        if tic_num in sr.ms_plot_paths:
                            ms_plot_path = sr.ms_plot_paths[tic_num]
                            if ms_plot_path.exists() is True:
                                ms_cell = ws.cell(row=row, column=col["质谱图"], value="查看质谱图")
                                ms_cell.hyperlink = str(ms_plot_path)
                                ms_cell.font = Font(color="0563C1", underline="single")

                row += 1

        self._auto_column_width(ws)

    @staticmethod
    def _find_match_list(
        rt: float,
        matches: Dict[float, List[CompoundMatch]],
        tolerance: float = 0.05,
    ) -> Optional[List[CompoundMatch]]:
        """按保留时间在匹配字典中查找最接近的化合物匹配列表."""
        if not matches:
            return None
        closest_rt = min(matches.keys(), key=lambda r: abs(r - rt))
        if abs(closest_rt - rt) <= tolerance:
            return matches[closest_rt]
        return None

    @staticmethod
    def _build_structure_key(match: CompoundMatch) -> Optional[str]:
        """
        功能:
            根据命中信息构建结构图映射键.
            规则:
            1. 有效 nist_id 时使用 NIST:<id>.
            2. 否则使用 CAS:<digits>.
        参数:
            match: 化合物命中对象.
        返回:
            Optional[str], 结构键.
        """
        if match.nist_id is not None and match.nist_id > 0:
            return f"NIST:{match.nist_id}"

        cas_digits = re.sub(r"\D", "", match.cas_number.strip())
        if cas_digits:
            return f"CAS:{cas_digits}"
        return None

    @staticmethod
    def _link_compound_name_cell(
        cell,
        match: CompoundMatch,
        structure_images: Optional[Dict[str, Optional[Path]]] = None,
    ) -> None:
        """
        功能:
            若结构图存在, 将化合物名称单元格设置为结构图超链接.
        参数:
            cell: openpyxl 单元格对象.
            match: 化合物命中对象.
            structure_images: 结构键 -> 结构图 PNG 路径, None 表示不设置超链接.
        返回:
            无.
        """
        if structure_images is None:
            return

        structure_key = ReportGenerator._build_structure_key(match)
        if structure_key is None:
            return

        img_path = structure_images.get(structure_key)
        if img_path is None:
            return
        if img_path.exists() is not True:
            return

        cell.hyperlink = str(img_path)
        cell.font = Font(color="0563C1", underline="single")

    @staticmethod
    def _find_pim_prediction(
        rt: float,
        predictions: Dict[float, PIMPrediction],
        tolerance: float = 0.05,
    ) -> Optional[PIMPrediction]:
        """
        功能:
            按保留时间在 PIM 预测字典中查找最接近结果.
        参数:
            rt: 目标保留时间(min).
            predictions: 保留时间 -> PIM 预测结果.
            tolerance: 保留时间容差(min).
        返回:
            Optional[PIMPrediction], 容差内最近结果.
        """
        if len(predictions) == 0:
            return None
        closest_rt = min(predictions.keys(), key=lambda key_rt: abs(key_rt - rt))
        if abs(closest_rt - rt) <= tolerance:
            return predictions[closest_rt]
        return None

    @staticmethod
    def _find_prediction_by_rt(
        rt: float,
        predictions: Dict,
        tolerance: float = 0.05,
    ):
        """
        功能:
            按保留时间在预测字典中查找最接近结果 (通用版, 支持任意预测类型).
        参数:
            rt: 目标保留时间(min).
            predictions: 保留时间 -> 预测结果字典.
            tolerance: 保留时间容差(min).
        返回:
            容差内最近的预测结果, 或 None.
        """
        if len(predictions) == 0:
            return None
        closest_rt = min(predictions.keys(), key=lambda key_rt: abs(key_rt - rt))
        if abs(closest_rt - rt) <= tolerance:
            return predictions[closest_rt]
        return None


