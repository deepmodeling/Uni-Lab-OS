                          # -*- coding: utf-8 -*-
import csv
import re
import logging
import shutil
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, Alignment, NamedStyle 
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# 引入底层的控制器
from ..controller.station_controller import SynthesisStationController
from ..config.setting import Settings, configure_logging
from ..config.constants import (
    CONSUMABLE_ALIAS_TO_CODE,
    CONSUMABLE_CODE_DISPLAY_NAME,
    CONSUMABLE_CODE_TO_TRAY_CODE,
    ResourceCode,
    TRAY_CODE_DISPLAY_NAME,
    TraySpec,
)
from .synchronizer import EITSynthesisWorkstation

from ..driver.exceptions import ValidationError,ApiError
from ..utils.file_utils import safe_excel_write, safe_workbook_save
from ..chem_tools.chemical_append_utils import (
    build_append_row_data,
    build_append_row_data_for_smiles,
    build_prepared_chemical_row_data,
    build_duplicate_check_specs,
    collect_missing_append_headers,
    get_excel_write_value,
    save_chemicalbook_record,
)

logger = logging.getLogger("StationManager")

JsonDict = Dict[str, Any]

# 模块根目录, 用于构建相对路径
MODULE_ROOT = Path(__file__).resolve().parent.parent

class SynthesisStationManager(EITSynthesisWorkstation, SynthesisStationController):
    """
    功能:
        上层面向用户的管理器，继承自 SynthesisStationController。
        负责处理 CSV/Excel 文件读取、生成模板，将文件内容转换为中间格式(List/Dict)，
        然后调用父类方法执行具体的业务逻辑。
    """

    def __init__(
        self,
        settings: Optional[Settings] = None,
        config: Optional[Dict[str, Any]] = None,
        deck: Optional[Any] = None,
        **kwargs,
    ):
        settings = settings or Settings.from_env()
        configure_logging(settings.log_level)
        SynthesisStationController.__init__(self, settings)
        EITSynthesisWorkstation.__init__(
            self,
            config=config,
            deck=deck,
            controller=self,
            **kwargs,
        )
        self._reset_task_queue_state()

    def _reset_task_queue_state(self) -> None:
        """
        功能:
            重置“先入队后执行”的任务缓存状态。
        """
        self._queued_layout_list: List[JsonDict] = []
        self._queued_task_name: str = "Queued_Auto_Task"
        self._queued_task_setup: JsonDict = {
            "subtype": None,
            "experiment_num": 0,
            "vessel": "551000502",
            "added_slots": "",
        }
        self._queued_common_fields: JsonDict = {
            "layout_code": "",
            "src_layout_code": "",
            "resource_type": "551000502",
            "status": 0,
            "tray_QR_code": "",
            "QR_code": "",
        }
        self._workflow_context: JsonDict = {
            "chemical_db": None,
            "reaction_scale_mmol": 0.0,
            "weighing_error_pct": 1.0,
            "max_error_mg": 1.0,
            "experiment_num": None,
        }
        self._workflow_row_counters: Dict[int, int] = {}
        # begin_task 后启用：试剂/磁子先入队，add_task 时经 build_task_payload 与 Excel 路径一致编排
        self._workflow_deferred: bool = False
        self._workflow_build_params: JsonDict = {}
        self._workflow_reagent_queue: List[JsonDict] = []

    def _configure_queued_task(
        self,
        *,
        task_name: Optional[str] = None,
        task_setup: Optional[JsonDict] = None,
        common_fields: Optional[JsonDict] = None) -> JsonDict:
        """
        功能:
            配置入队任务的名称、task_setup 和 unit 公共字段。
        参数:
            task_name: 可选任务名称。
            task_setup: 可选 task_setup 覆盖字段。
            common_fields: 可选 unit 通用字段覆盖。
        返回:
            Dict, 当前队列配置摘要。
        """
        if task_name is not None and str(task_name).strip() != "":
            self._queued_task_name = str(task_name).strip()

        if isinstance(task_setup, dict):
            self._queued_task_setup.update(task_setup)

        if isinstance(common_fields, dict):
            self._queued_common_fields.update(common_fields)

        return self._get_queued_task_status()

    def _get_queued_task_status(self) -> JsonDict:
        """
        功能:
            获取当前入队任务的状态摘要。
        返回:
            Dict, 包含任务名、unit 数量、实验列统计与配置。
        """
        if getattr(self, "_workflow_deferred", False):
            return {
                "task_name": self._queued_task_name,
                "deferred_layout": True,
                "queued_reagents": len(self._workflow_reagent_queue),
                "workflow_build_params": dict(self._workflow_build_params),
                "task_setup": self._queued_task_setup.copy(),
                "common_fields": self._queued_common_fields.copy(),
            }

        used_columns: set[int] = set()
        used_rows: set[int] = set()
        for unit in self._queued_layout_list:
            col = unit.get("unit_column")
            row = unit.get("unit_row")
            if isinstance(col, int):
                used_columns.add(col)
            if isinstance(row, int):
                used_rows.add(row)

        return {
            "task_name": self._queued_task_name,
            "queued_units": len(self._queued_layout_list),
            "experiment_columns": sorted(used_columns),
            "unit_rows": sorted(used_rows),
            "task_setup": self._queued_task_setup.copy(),
            "common_fields": self._queued_common_fields.copy(),
        }

    def _preview_queued_task_payload(
        self,
        *,
        experiment_num: Optional[int] = None,
        is_audit_log: int = 1,
        is_copy: bool = False) -> JsonDict:
        """
        功能:
            预览当前入队单元将生成的任务 payload（不提交）。
        参数:
            experiment_num: 可选实验数量，未传时按 unit_column 自动推断。
            is_audit_log: 审计日志标记，默认 1。
            is_copy: 复制标记，默认 False。
        返回:
            Dict, 可直接用于 add_task 的 payload。
        """
        if getattr(self, "_workflow_deferred", False):
            return self._materialize_workflow_payload_from_excel_model(
                is_audit_log=is_audit_log,
                is_copy=is_copy,
            )

        if len(self._queued_layout_list) == 0:
            raise ValidationError("当前队列为空，请先调用 begin_task 和 add_* 接口入队")

        inferred_exp_num = 0
        for unit in self._queued_layout_list:
            col = unit.get("unit_column")
            if isinstance(col, int):
                inferred_exp_num = max(inferred_exp_num, col + 1)

        final_experiment_num = int(experiment_num) if experiment_num is not None else int(inferred_exp_num)
        if final_experiment_num <= 0:
            raise ValidationError("无法推断 experiment_num，请传入 experiment_num")

        task_setup = self._queued_task_setup.copy()
        task_setup["experiment_num"] = final_experiment_num

        return {
            "task_id": 0,
            "task_name": self._queued_task_name,
            "layout_list": [item.copy() for item in self._queued_layout_list],
            "task_setup": task_setup,
            "is_audit_log": int(is_audit_log),
            "is_copy": bool(is_copy),
        }

    def _execute_queued_task(
        self,
        *,
        experiment_num: Optional[int] = None,
        is_audit_log: int = 1,
        is_copy: bool = False,
        auto_start: bool = False,
        check_glovebox_env: bool = True,
        water_limit_ppm: float = 10.0,
        oxygen_limit_ppm: float = 10.0,
        clear_on_success: bool = True) -> JsonDict:
        """
        功能:
            将当前队列统一构建为 payload 并一次性提交到工站，可选自动启动任务。
        参数:
            experiment_num: 可选实验数量，未传时按 unit_column 自动推断。
            is_audit_log: 审计日志标记，默认 1。
            is_copy: 复制标记，默认 False。
            auto_start: 提交成功后是否自动启动。
            check_glovebox_env: 自动启动时是否校验手套箱环境。
            water_limit_ppm: 自动启动水含量阈值。
            oxygen_limit_ppm: 自动启动氧含量阈值。
            clear_on_success: 成功后是否自动清空队列。
        返回:
            Dict, 包含 payload、add_task 响应、task_id 以及可选 start_task 响应。
        """
        payload = self._preview_queued_task_payload(
            experiment_num=experiment_num,
            is_audit_log=is_audit_log,
            is_copy=is_copy,
        )

        add_resp = self._submit_task_payload(payload)
        task_id = add_resp.get("task_id") or add_resp.get("result", {}).get("task_id") or add_resp.get("data", {}).get("task_id")

        result: JsonDict = {
            "task_id": task_id,
            "add_task_response": add_resp,
            "payload": payload,
        }

        if auto_start is True:
            if task_id is None:
                raise ValidationError(f"任务创建成功但未解析到 task_id, add_resp={add_resp}")
            start_resp = SynthesisStationController.start_task(
                self,
                int(task_id),
                check_glovebox_env=check_glovebox_env,
                water_limit_ppm=float(water_limit_ppm),
                oxygen_limit_ppm=float(oxygen_limit_ppm),
            )
            result["start_task_response"] = start_resp

        if clear_on_success is True:
            self._reset_task_queue_state()

        return result

    def _submit_task_payload(self, payload: JsonDict) -> JsonDict:
        """
        功能:
            直接调用底层 AddTask 接口提交 payload。
        """
        return SynthesisStationController.add_task(self, payload)

    # ---------- 前端工作流接口 ----------
    def load_chemical_db_from_file(self, chemical_db_path: str) -> Dict[str, Dict[str, Any]]:
        """
        功能:
            从 Excel/CSV 化学品库文件加载化学品字典，供前端工作流接口复用。
        参数:
            chemical_db_path: 化学品库文件路径。
        返回:
            Dict[str, Dict[str, Any]], 以化学品名称为键的化学品信息表。
        """
        c_path = Path(chemical_db_path)
        if c_path.exists() is False:
            raise FileNotFoundError(f"未找到化学品库文件: {c_path}")

        chem_df = pd.read_excel(c_path) if c_path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(c_path)
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row: JsonDict, *keys: str, default: Any = None) -> Any:
            for key in keys:
                if key in row and pd.notna(row[key]):
                    return row[key]
            return default

        chemical_db: Dict[str, Dict[str, Any]] = {}
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if name == "":
                continue
            chemical_db[name] = {
                "chemical_id": _pick(row, "chemical_id"),
                "molecular_weight": _pick(row, "molecular_weight", "mw"),
                "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                "density (g/mL)": _pick(row, "density (g/ml)", "density(g/ml)", "density_g_ml", "density", default=None),
                "physical_form": str(_pick(row, "physical_form", default="") or "").strip().lower(),
                "active_content": _pick(row, "active_content", "active_content(mmol/ml or wt%)", "active_content(mol/l or wt%)", default=""),
            }
        return chemical_db

    def _resolve_workflow_chemical_db(
        self,
        *,
        chemical_db: Optional[Dict[str, Dict[str, Any]]] = None,
        chemical_db_path: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
        if isinstance(chemical_db, dict):
            return chemical_db
        if chemical_db_path is not None and str(chemical_db_path).strip() != "":
            return self.load_chemical_db_from_file(str(chemical_db_path))
        current_db = self._workflow_context.get("chemical_db")
        if isinstance(current_db, dict):
            return current_db
        raise ValidationError("必须提供 chemical_db 或 chemical_db_path，或先调用 begin_task 设置化学品库")

    def _get_workflow_chemical_info(
        self,
        name: str,
        *,
        chemical_db: Optional[Dict[str, Dict[str, Any]]] = None,
        chemical_db_path: Optional[str] = None) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
        name_text = str(name).strip()
        if name_text == "":
            raise ValidationError("化学品名称不能为空")
        db = self._resolve_workflow_chemical_db(chemical_db=chemical_db, chemical_db_path=chemical_db_path)
        if name_text not in db:
            raise ValidationError(f"化学品库中未找到 '{name_text}'")
        return db, db[name_text]

    def _experiment_to_col(self, experiment_no: int) -> int:
        exp_no = int(experiment_no)
        if exp_no <= 0:
            raise ValidationError("experiment_no 必须为大于 0 的整数")
        return exp_no - 1

    def _allocate_workflow_row(self, experiment_no: int) -> int:
        exp_no = int(experiment_no)
        current_row = self._workflow_row_counters.get(exp_no, 0)
        self._workflow_row_counters[exp_no] = current_row + 1
        return current_row

    def _default_experiment_no(self) -> int:
        """
        功能:
            前端编排式接口默认使用单实验模式，统一落在实验编号 1。
        """
        return 1

    @staticmethod
    def _format_reagent_amount_cell(amount: float, unit: str) -> str:
        """将数值与单位拼成与 Excel 单元格相近的字符串，供 _split_amount_unit 解析。"""
        u = str(unit).strip().replace("µ", "μ")
        try:
            x = float(amount)
            if x == int(x):
                num = str(int(x))
            else:
                num = str(x).rstrip("0").rstrip(".")
        except Exception:
            num = str(amount)
        return f"{num}{u}"

    def _build_headers_and_data_rows_from_reagent_queue(self) -> Tuple[List[str], List[List[Any]]]:
        """
        功能:
            将试剂/磁子入队记录还原为 build_task_payload 所需的表头与数据行（按实验编号连续展开行）。
        """
        per_exp: Dict[int, List[Tuple[str, str]]] = {}
        for op in self._workflow_reagent_queue:
            if op.get("op") == "reagent":
                exp = int(op["experiment_no"])
                cell = self._format_reagent_amount_cell(float(op["amount"]), str(op["unit"]))
                per_exp.setdefault(exp, []).append((str(op["name"]).strip(), cell))
            elif op.get("op") == "magnet":
                exp = int(op["experiment_no"])
                per_exp.setdefault(exp, []).append(("加磁子", ""))

        if not per_exp:
            raise ValidationError("请先使用 add_reagent / add_reagent_list / add_magnet 入队至少一条操作")

        max_exp_no = max(per_exp.keys())
        if max_exp_no < 1:
            raise ValidationError("experiment_no 必须为从 1 开始的整数")

        num_slots = max(len(per_exp[e]) for e in per_exp)
        if num_slots <= 0:
            raise ValidationError("试剂列为空")

        headers: List[str] = ["实验编号"]
        for i in range(1, num_slots + 1):
            headers.append(f"试剂名称_{i}")
            headers.append(f"试剂量_{i}")

        data_rows: List[List[Any]] = []
        for exp_no in range(1, max_exp_no + 1):
            slots = per_exp.get(exp_no, [])
            row: List[Any] = [str(exp_no)]
            for idx in range(num_slots):
                if idx < len(slots):
                    row.append(slots[idx][0])
                    row.append(slots[idx][1])
                else:
                    row.append("")
                    row.append("")
            data_rows.append(row)

        return headers, data_rows

    def _materialize_workflow_payload_from_excel_model(
        self,
        *,
        is_audit_log: int = 1,
        is_copy: bool = False,
    ) -> JsonDict:
        """
        功能:
            使用与 create_task_by_file 相同的 build_task_payload 生成最终 AddTask 请求体。
        """
        headers, data_rows = self._build_headers_and_data_rows_from_reagent_queue()
        db = self._workflow_context.get("chemical_db")
        if not isinstance(db, dict):
            raise ValidationError("化学品库未初始化，请先调用 begin_task")

        params = dict(self._workflow_build_params)
        params.setdefault("实验名称", self._queued_task_name)

        payload = self.build_task_payload(params, headers, data_rows, db)
        payload["is_audit_log"] = int(is_audit_log)
        payload["is_copy"] = bool(is_copy)
        return payload

    def _submit_workflow_payload_with_name_conflict_retry(self, task_payload: JsonDict) -> JsonDict:
        """与 create_task_by_file 一致：任务名冲突(409)时自动加时间戳重试。"""
        try:
            return self._submit_task_payload(task_payload)
        except ApiError as exc:
            if getattr(exc, "code", None) != 409:
                raise
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            task_name = task_payload.get("task_name") or self._workflow_build_params.get("实验名称")
            new_task_name = f"{task_name}_{timestamp}"
            task_payload["task_name"] = new_task_name
            self._workflow_build_params["实验名称"] = new_task_name
            logger.info("任务名称重复, 自动重命名为: %s", new_task_name)
            try:
                return self._submit_task_payload(task_payload)
            except ApiError as retry_exc:
                logger.error("重命名后任务提交仍失败: %s", retry_exc)
                raise

    def _execute_deferred_workflow(
        self,
        *,
        experiment_num: Optional[int] = None,
        is_audit_log: int = 1,
        is_copy: bool = False,
        auto_start: bool = False,
        check_glovebox_env: bool = True,
        water_limit_ppm: float = 10.0,
        oxygen_limit_ppm: float = 10.0,
        clear_on_success: bool = True,
    ) -> JsonDict:
        payload = self._materialize_workflow_payload_from_excel_model(
            is_audit_log=is_audit_log,
            is_copy=is_copy,
        )
        if experiment_num is not None:
            payload.setdefault("task_setup", {})["experiment_num"] = int(experiment_num)

        add_resp = self._submit_workflow_payload_with_name_conflict_retry(payload)
        task_id = add_resp.get("task_id") or add_resp.get("result", {}).get("task_id") or add_resp.get("data", {}).get("task_id")

        result: JsonDict = {
            "task_id": task_id,
            "add_task_response": add_resp,
            "payload": payload,
        }

        if auto_start is True:
            if task_id is None:
                raise ValidationError(f"任务创建成功但未解析到 task_id, add_resp={add_resp}")
            start_resp = SynthesisStationController.start_task(
                self,
                int(task_id),
                check_glovebox_env=check_glovebox_env,
                water_limit_ppm=float(water_limit_ppm),
                oxygen_limit_ppm=float(oxygen_limit_ppm),
            )
            result["start_task_response"] = start_resp

        if clear_on_success is True:
            self._reset_task_queue_state()

        return result

    def begin_task(
        self,
        *,
        task_name: str = "Auto_task",
        task_id: int = 0,
        chemical_db: Optional[Dict[str, Dict[str, Any]]] = None,
        chemical_db_path: Optional[str] = None,
        fixed_addition_order: bool = False,
        auto_magnet: bool = True,
        weighing_error_pct: float = 3.0,
        max_error_mg: float = 1.0) -> JsonDict:
        """
        功能:
            初始化编排式任务上下文。参数与 Excel 模板左侧"实验设定 / 称量设定 / 加料设定"一一对应。
        参数:
            task_name: 实验名称。
            task_id: 实验ID（默认 0 表示新建）。
            chemical_db/chemical_db_path: 化学品库对象或路径，至少提供一个。
            fixed_addition_order: 固定加料顺序（对应 Excel "是/否"），默认 False。
            auto_magnet: 自动加磁子（对应 Excel "是/否"），默认 True。
            weighing_error_pct: 称量误差(%)。
            max_error_mg: 最大称量误差(mg)。
        返回:
            Dict, 当前任务上下文摘要。
        """
        self._reset_task_queue_state()
        db = self._resolve_workflow_chemical_db(chemical_db=chemical_db, chemical_db_path=chemical_db_path)
        task_setup: JsonDict = {
            "subtype": None,
            "vessel": "551000502",
            "added_slots": "",
        }
        self._configure_queued_task(
            task_name=str(task_name).strip(),
            task_setup=task_setup,
        )
        self._workflow_context.update({
            "chemical_db": db,
            "reaction_scale_mmol": 0.0,
            "weighing_error_pct": float(weighing_error_pct),
            "max_error_mg": float(max_error_mg),
            "experiment_num": None,
        })
        self._workflow_row_counters = {}
        self._workflow_deferred = True
        self._workflow_reagent_queue = []
        self._workflow_build_params = {
            "实验名称": str(task_name).strip(),
            "实验ID": int(task_id),
            "固定加料顺序": "是" if fixed_addition_order else "否",
            "自动加磁子": "是" if auto_magnet else "否",
            "称量误差(%)": float(weighing_error_pct),
            "最大称量误差(mg)": float(max_error_mg),
        }
        return self._get_queued_task_status()

    def add_reagent(
        self,
        *,
        name: str,
        amount: float,
        unit: str,
        experiment_no: Optional[int] = None) -> JsonDict:
        """
        功能:
            以 Excel 里的“试剂 / 试剂量”语义入队（默认实验编号 1）；在 add_task 时与 create_task_by_file 相同经 build_task_payload 编排。
        参数:
            name: 试剂名称。
            amount: 用量数值。
            unit: 用量单位，如 eq/mmol/mg/g/μL/mL。
            experiment_no: 实验编号（从 1 开始）；缺省为 1。
            化学品库来源: 统一使用 begin_task 设置的 chemical_db/chemical_db_path。
        返回:
            Dict, 当前入队试剂/磁子条数与实验编号。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._get_workflow_chemical_info(name)
        exp_no = int(experiment_no) if experiment_no is not None else self._default_experiment_no()
        self._workflow_reagent_queue.append({
            "op": "reagent",
            "experiment_no": exp_no,
            "name": str(name).strip(),
            "amount": float(amount),
            "unit": str(unit).strip(),
        })
        return {
            "queued_reagents": len(self._workflow_reagent_queue),
            "experiment_no": exp_no,
        }

    def add_reagent_list(
        self,
        *,
        experiments: List[JsonDict],
    ) -> JsonDict:
        """
        功能:
            一次入队多个实验编号的配方（顺序与 Excel 中各实验试剂从左到右一致）。
        参数:
            experiments: 每项形如 {"experiment_no": 1, "reagents": [{"name": "...", "amount": 1.0, "unit": "mmol"}, ...]}。
        返回:
            Dict, 当前入队试剂/磁子总条数。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        for block in experiments:
            exp_no = int(block["experiment_no"])
            for r in block.get("reagents") or []:
                self._get_workflow_chemical_info(str(r["name"]).strip())
                self._workflow_reagent_queue.append({
                    "op": "reagent",
                    "experiment_no": exp_no,
                    "name": str(r["name"]).strip(),
                    "amount": float(r["amount"]),
                    "unit": str(r["unit"]).strip(),
                })
        return {"queued_reagents": len(self._workflow_reagent_queue)}

    def add_magnet(self, *, experiment_no: Optional[int] = None) -> JsonDict:
        """
        功能:
            以“加磁子”语义入队（默认实验编号 1）；编排规则与 Excel 一致。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        exp_no = int(experiment_no) if experiment_no is not None else self._default_experiment_no()
        self._workflow_reagent_queue.append({"op": "magnet", "experiment_no": exp_no})
        return {"queued_reagents": len(self._workflow_reagent_queue), "experiment_no": exp_no}

    def add_reaction(
        self,
        *,
        reaction_scale_mmol: float = 0.2,
        reactor_type: str = "heat",
        reaction_time_h: float = 8.0,
        reaction_temp_c: Optional[float] = 40.0,
        stir_speed_rpm: int = 500,
        target_temp_c: Optional[float] = None,
        wait_target_temp: bool = False) -> JsonDict:
        """
        功能:
            设置 Excel 模板中"反应设定"区域的全部参数。
        参数:
            reaction_scale_mmol: 反应规模(mmol)。
            reactor_type: 反应器类型（如 "heat"）。
            reaction_time_h: 反应时间，单位小时（写入 build_task_payload 时转为 "8h" 格式）。
            reaction_temp_c: 反应温度(°C)。
            stir_speed_rpm: 转速(rpm)。
            target_temp_c: 搅拌后目标温度(°C)；None 表示不设置。
            wait_target_temp: 等待目标温度（对应 Excel "是/否"）。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._workflow_build_params["反应规模(mmol)"] = float(reaction_scale_mmol)
        self._workflow_context["reaction_scale_mmol"] = float(reaction_scale_mmol)
        self._workflow_build_params["反应器类型"] = str(reactor_type).strip()
        self._workflow_build_params["反应时间(min/h)"] = f"{float(reaction_time_h)}h"
        self._workflow_build_params["转速(rpm)"] = int(stir_speed_rpm)
        self._workflow_build_params["等待目标温度"] = "是" if wait_target_temp else "否"
        if reaction_temp_c is not None:
            self._workflow_build_params["反应温度(°C)"] = float(reaction_temp_c)
        else:
            self._workflow_build_params.pop("反应温度(°C)", None)
        if target_temp_c is not None:
            self._workflow_build_params["搅拌后目标温度(°C)"] = float(target_temp_c)
        else:
            self._workflow_build_params.pop("搅拌后目标温度(°C)", None)
        return {"workflow_build_params": dict(self._workflow_build_params)}

    def add_internal_standard(
        self,
        *,
        name: str,
        amount_ul_or_mg: float) -> JsonDict:
        """
        功能:
            设置 Excel 模板中"内标设定"的内标种类与内标用量(μL/mg)。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._get_workflow_chemical_info(name)
        self._workflow_build_params["内标种类"] = str(name).strip()
        self._workflow_build_params["内标用量(μL/mg)"] = float(amount_ul_or_mg)
        return {"workflow_build_params": dict(self._workflow_build_params)}

    def add_stir(
        self,
        *,
        time_min: float) -> JsonDict:
        """
        功能:
            设置 Excel 模板中"加入内标后搅拌时间(min)"。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._workflow_build_params["加入内标后搅拌时间(min)"] = float(time_min)
        return {"workflow_build_params": dict(self._workflow_build_params)}

    def add_dilution(
        self,
        *,
        diluent_name: str,
        dilution_volume_ul: float) -> JsonDict:
        """
        功能:
            设置 Excel 模板中"稀释设定"（稀释液种类、稀释量(μL)）。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._get_workflow_chemical_info(diluent_name)
        self._workflow_build_params["稀释液种类"] = str(diluent_name).strip()
        self._workflow_build_params["稀释量(μL)"] = float(dilution_volume_ul)
        return {"workflow_build_params": dict(self._workflow_build_params)}

    def add_filter(
        self,
        *,
        filter_liquid_name: str,
        filter_volume_ul: float,
        sampling_volume_ul: float,
        filter_experiment_numbers: str = "全部") -> JsonDict:
        """
        功能:
            设置 Excel 模板中"闪滤设定"（闪滤液种类、闪滤液用量(μL)、取样量(μL)、闪滤实验编号）。
        参数:
            filter_liquid_name: 闪滤液种类。
            filter_volume_ul: 闪滤液用量(μL)。
            sampling_volume_ul: 取样量(μL)。
            filter_experiment_numbers: 闪滤实验编号，"全部" 或 "1-12,24,28" 格式。
        """
        if getattr(self, "_workflow_deferred", False) is False:
            raise ValidationError("请先调用 begin_task 初始化编排上下文")
        self._get_workflow_chemical_info(filter_liquid_name)
        self._workflow_build_params["闪滤液种类"] = str(filter_liquid_name).strip()
        self._workflow_build_params["闪滤液用量(μL)"] = float(filter_volume_ul)
        self._workflow_build_params["取样量(μL)"] = float(sampling_volume_ul)
        self._workflow_build_params["闪滤实验编号"] = str(filter_experiment_numbers).strip()
        return {"workflow_build_params": dict(self._workflow_build_params)}


    def add_task(
        self,
        *,
        payload: Optional[JsonDict] = None) -> JsonDict:
        """
        功能:
            作为前端工作流的“添加任务”节点，仅执行 add_task，不做 start_task。
            兼容旧接口：若传入 payload，则直连底层 AddTask 提交。
        参数:
            payload: 可选。传入时直接调用底层 AddTask。
        返回:
            Dict, 包含 payload、add_task 响应与 task_id。
        """
        if payload is not None:
            return self._submit_task_payload(payload)

        if getattr(self, "_workflow_deferred", False):
            experiment_num = self._workflow_context.get("experiment_num")
            return self._execute_deferred_workflow(
                experiment_num=int(experiment_num) if experiment_num is not None else None,
                auto_start=False,
                clear_on_success=True,
            )

        experiment_num = self._workflow_context.get("experiment_num")
        return self._execute_queued_task(
            experiment_num=int(experiment_num) if experiment_num is not None else None,
            auto_start=False,
            clear_on_success=True,
        )

    def execute_task(self, task_id: int | None = None):
        return super().start_task(task_id, check_glovebox_env=True, water_limit_ppm=10.0, oxygen_limit_ppm=10.0)

    def login(self) -> tuple:
        """
        功能:
            登录并缓存 token, 登录成功后根据配置自动启动异常通知监控.
        参数:
            无.
        返回:
            Tuple[str, str], (token_type, access_token).
        """
        result = super().login()
        # 登录成功后自动启动异常通知邮件监控 (CLI 启动路径)
        try:
            if self._settings.notification.enabled:
                self.start_notification_monitor()
                logger.info("异常通知邮件监控已启动")
        except Exception as e:
            logger.warning("异常通知监控启动失败, 不影响主流程: %s", e)
        return result

    def _read_table_file_with_required_columns(
        self,
        path: Path,
        *,
        required_columns: Optional[List[str]] = None,
        preferred_sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        功能:
            读取 CSV/Excel 文件.
            当为 Excel 且存在多工作表时, 根据必需列选择工作表.
        参数:
            path: 文件路径.
            required_columns: 必需列名列表, None 表示直接读取默认表.
            preferred_sheet_name: 优先工作表名.
        返回:
            DataFrame, 读取后的表格数据.
        """
        if path.suffix.lower() not in [".xlsx", ".xls"]:
            return pd.read_csv(path)

        if required_columns is None:
            return pd.read_excel(path)

        required = {str(col).strip().lower() for col in required_columns}
        all_sheets = pd.read_excel(path, sheet_name=None)

        candidate_sheet_names: List[str] = []
        if preferred_sheet_name is not None and preferred_sheet_name in all_sheets:
            candidate_sheet_names.append(preferred_sheet_name)
        for sheet_name in all_sheets.keys():
            if sheet_name not in candidate_sheet_names:
                candidate_sheet_names.append(sheet_name)

        matched_sheet_names: List[str] = []
        for sheet_name in candidate_sheet_names:
            df = all_sheets[sheet_name]
            normalized_columns = {str(col).strip().lower() for col in df.columns}
            if required.issubset(normalized_columns):
                matched_sheet_names.append(sheet_name)

        if len(matched_sheet_names) > 0:
            selected_sheet_name = matched_sheet_names[0]
            if len(matched_sheet_names) > 1:
                logger.warning(
                    "Excel命中多个候选工作表, 将按优先顺序使用 [%s], 其余候选: %s, 文件: %s",
                    selected_sheet_name,
                    matched_sheet_names[1:],
                    path,
                )
            logger.info("Excel解析使用工作表: %s, 文件: %s", selected_sheet_name, path)
            return all_sheets[selected_sheet_name]

        fallback_sheet_name = next(iter(all_sheets.keys()))
        logger.warning(
            "Excel未命中必需列%s, 回退到第一张工作表: %s, 文件: %s",
            sorted(required),
            fallback_sheet_name,
            path,
        )
        return all_sheets[fallback_sheet_name]

    # ---------- 1. 化合物库文件处理 ----------
    def export_chemical_list_to_file(self, output_path: str) -> None:
        """
        功能:
            获取所有化学品并导出到 CSV 文件
        参数:
            output_path: 输出路径
        返回:
            None
        """
        path = Path(output_path)
        chemical_info = self.get_all_chemical_list()
        chemical_list = chemical_info.get("chemical_list", [])

        if not chemical_list:
            logger.warning("化学品列表为空，未写入文件")
            return

        fieldnames = [
            "fid", "name", "sssi", "cas", "element", "state",
            "concentration_str", "chemical_properties", "preparation_method"
        ]
        
        # 确保目录存在
        path.parent.mkdir(parents=True, exist_ok=True)

        with path.open("w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for item in chemical_list:
                writer.writerow(item)
        
        logger.info(f"化学品列表已导出至: {path.resolve()}")

    def sync_chemicals_from_file(self, file_path: str, overwrite: bool = False) -> None:
        """
        功能:
            读取 CSV 文件并通过父类同步化学品到工站
        参数:
            file_path: CSV 文件路径
            overwrite: 是否覆盖更新
        返回:
            None
        """
        path = Path(file_path)
        if not path.exists():
            # 生成模板
            header = ["name", "cas", "element", "state", "concentration_str", "chemical_properties", "preparation_method"]
            with path.open("w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerow(header)
            logger.warning(f"文件不存在，已生成模板: {path}")
            return

        # 读取并清洗数据
        items: List[JsonDict] = []
        with path.open("r", newline="", encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                name = (row.get("name") or "").strip()
                state = (row.get("state") or "").strip()
                if name and state:
                    # 过滤空值键
                    clean_item = {k: v.strip() for k, v in row.items() if v and str(v).strip()}
                    items.append(clean_item)
        
        # 调用父类逻辑处理
        self.sync_chemicals_from_data(items, overwrite=overwrite)

    def check_chemical_library_by_file(self, file_path: str) -> Dict[str, List[str]]:
        """
        功能:
            读取化学品库文件并调用底层校验逻辑，输出校验结果
        参数:
            file_path: str, 化学品库文件路径，支持 Excel/CSV
        返回:
            Dict[str, List[str]], 包含 errors 与 warnings
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {path}")

        # 读取文件后交给控制层做校验
        df = pd.read_excel(path) if path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(path)
        df = df.fillna("")
        rows = df.to_dict(orient="records")
        headers = [str(col).strip() for col in df.columns]

        result = self.check_chemical_library_data(rows, headers)

        for msg in result.get("warnings", []):
            logger.warning(msg)

        if len(result.get("errors", [])) > 0:
            for msg in result["errors"]:
                logger.error(msg)
            raise ValidationError("化学品库完整性检查未通过，请修复错误后重试")

        return result
    
    def deduplicate_chemical_library_by_file(self, file_path: str, output_path: Optional[str] = None) -> List[JsonDict]:
        """
        功能:
            读取化学品库文件，按 substance 自动去重并回写
        参数:
            file_path: str, 输入文件路径，支持 Excel/CSV
            output_path: Optional[str], 输出文件路径，默认覆盖原文件
        返回:
            List[Dict[str, Any]], 去重后的数据
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {path}")

        df = pd.read_excel(path) if path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(path)
        df = df.fillna("")
        headers = [str(c).strip() for c in df.columns]
        rows = df.to_dict(orient="records")

        dedup_rows = self.deduplicate_chemical_library_data(rows, headers)

        target_path = Path(output_path) if output_path else path
        out_df = pd.DataFrame(dedup_rows)
        if target_path.suffix.lower() == ".csv":
            out_df.to_csv(target_path, index=False, encoding="utf-8-sig")
        else:
            safe_excel_write(out_df, target_path, index=False)
            self._beautify_excel_database(target_path)  # 保存后再美化

        logger.info("化合物库去重完成，输出文件: %s", target_path.resolve())
        return dedup_rows
    
    def _beautify_excel_database(self, file_path: Path) -> None:
        """
        功能:
            美化去重后的 Excel: 表头加粗、全居中、列宽自适应、按内容选择中英文字体
        参数:
            file_path: Path, 目标 Excel 路径
        返回:
            None
        """
        wb = load_workbook(file_path)
        ws = wb.active
        MAX_WIDTH = 60  # 列宽上限

        align_center = Alignment(horizontal="center", vertical="center")

        def _is_chinese(text: str) -> bool:
            return re.search(r"[\u4e00-\u9fff]", text) is not None

        # 遍历列计算列宽并设置字体/对齐
        for col_cells in ws.iter_cols():
            max_len = 0
            for idx, cell in enumerate(col_cells):
                val_str = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val_str))

                # 按内容切换字体，表头加粗
                if idx == 0:
                    cell.font = Font(name="微软雅黑", bold=True)
                else:
                    cell.font = Font(name="微软雅黑")

                cell.alignment = align_center

            # 列宽留一点边距，最小 10，最大 40
            col_width = max(10, max_len + 2)
            col_width = min(col_width, MAX_WIDTH)
            ws.column_dimensions[col_cells[0].column_letter].width = col_width

        safe_workbook_save(wb, file_path)

    def align_chemicals_with_file(self, file_path: str, auto_delete: bool = True) -> None:
        """
        功能:
            读取 Excel/CSV 文件，调用父类对齐逻辑，并将结果(fid)写回文件
        参数:
            file_path: 文件路径
            auto_delete: 是否删除不在文件中的工站化学品
        返回:
            None
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品对齐文件: {path}")

        # 读取文件内容为 List[Dict]
        df = pd.read_excel(path) if path.suffix in ['.xlsx', '.xls'] else pd.read_csv(path)
        # 将 NaN 替换为空字符串
        df = df.fillna("")
        rows = df.to_dict(orient='records')
        header = df.columns.tolist()

        # 调用父类进行对齐，父类会修改 rows 中的数据(如回填 chemical_id)
        updated_rows = self.align_chemicals_from_data(rows, auto_delete=auto_delete)

        # 写回文件
        new_df = pd.DataFrame(updated_rows)
        # 保持原有列顺序，如果增加了新列(如 chemical_id 之前没有)，这会包含它
        if path.suffix == '.csv':
            new_df.to_csv(path, index=False, encoding="utf-8-sig")
        else:
            safe_excel_write(new_df, path, index=False)
            self._beautify_excel_database(path)  # 保存后再美化
        
        logger.info(f"化学品对齐完成并回写文件: {path}")

    @staticmethod
    def _has_any_append_core_value(row_data: Dict[str, Any]) -> bool:
        """
        功能:
            判断追加行是否至少包含一个可用于识别化合物的核心字段.
        参数:
            row_data: Dict[str, Any], 准备写入 Excel 的行数据.
        返回:
            bool, True 表示至少包含 CAS, 英文名, 中文名中的一个.
        """
        return any([
            str(row_data.get("cas_number") or "").strip() != "",
            str(row_data.get("substance_english_name") or "").strip() != "",
            str(row_data.get("substance") or "").strip() != "",
        ])

    @staticmethod
    def _format_append_row_summary(row_data: Dict[str, Any]) -> str:
        """
        功能:
            提取 chemical list 关键字段, 生成统一的日志摘要文本.
        参数:
            row_data: Dict[str, Any], 追加到 chemical list 的行数据.
        返回:
            str, 包含 substance, physical_state, physical_form 的摘要文本.
        """
        substance = str(
            row_data.get("substance")
            or row_data.get("substance_chinese_name")
            or ""
        ).strip()
        physical_state = str(row_data.get("physical_state") or "").strip()
        physical_form = str(row_data.get("physical_form") or "").strip()
        return (
            f"substance={substance}, "
            f"physical_state={physical_state}, "
            f"physical_form={physical_form}"
        )

    def _resolve_append_excel_path(self, excel_path: Optional[str]) -> Path:
        """
        功能:
            解析化学品追加流程使用的目标 Excel 路径.
        参数:
            excel_path: Optional[str], 用户指定路径, None 时使用默认库文件.
        返回:
            Path, 已解析的 Excel 路径.
        异常:
            FileNotFoundError: 目标文件不存在时抛出.
        """
        path = Path(excel_path) if excel_path is not None else MODULE_ROOT / "sheet" / "chemical_list.xlsx"
        if path.exists() is False:
            raise FileNotFoundError(f"化学品库文件不存在: {path}")
        return path

    @staticmethod
    def _build_append_header_map(worksheet: Any) -> Dict[str, int]:
        """
        功能:
            从化学品库工作表首行构建表头到列号的映射.
        参数:
            worksheet: Any, openpyxl 工作表对象.
        返回:
            Dict[str, int], 表头名称到列号的映射.
        """
        header_map: Dict[str, int] = {}
        for col_idx in range(1, worksheet.max_column + 1):
            header_val = worksheet.cell(row=1, column=col_idx).value
            if header_val is None:
                continue
            header_map[str(header_val).strip()] = col_idx
        return header_map

    @staticmethod
    def _build_append_row_snapshot(
        worksheet: Any,
        header_map: Dict[str, int],
        row_index: int) -> Dict[str, Any]:
        """
        功能:
            从化学品库工作表中提取单行数据快照, 统一补齐常用别名字段.
        参数:
            worksheet: Any, openpyxl 工作表对象.
            header_map: Dict[str, int], 表头名称到列号映射.
            row_index: int, 目标行号.
        返回:
            Dict[str, Any], 单行字段字典.
        """
        row_data: Dict[str, Any] = {}
        for column_name, column_index in header_map.items():
            row_data[column_name] = worksheet.cell(row=row_index, column=column_index).value

        substance_value = str(row_data.get("substance") or "").strip()
        chinese_name_value = str(row_data.get("substance_chinese_name") or "").strip()
        if substance_value == "" and chinese_name_value != "":
            row_data["substance"] = chinese_name_value
        if chinese_name_value == "" and substance_value != "":
            row_data["substance_chinese_name"] = substance_value
        base_substance = row_data.get("substance") or row_data.get("substance_chinese_name") or ""
        physical_form = str(row_data.get("physical_form") or "").strip().lower()
        if physical_form in {"solution", "beads"} and " (" in str(base_substance):
            base_substance = str(base_substance).split(" (", 1)[0].strip()
        row_data["base_substance"] = base_substance
        return row_data

    def _find_existing_chemical_row(
        self,
        *,
        excel_path: Optional[str] = None,
        cas_number: str = "",
        substance_english_name: str = "",
        substance: str = "") -> Optional[Dict[str, Any]]:
        """
        功能:
            按 CAS, 英文名, 中文名顺序在化学品库中查找已有条目, 优先返回 neat 行.
        参数:
            excel_path: Optional[str], 化学品库文件路径.
            cas_number: str, 候选 CAS 号.
            substance_english_name: str, 候选英文名.
            substance: str, 候选中文名或展示名.
        返回:
            Optional[Dict[str, Any]], 命中时返回包含 row_data 与 row_index 的结果字典.
        """
        path = self._resolve_append_excel_path(excel_path)
        wb = load_workbook(path, data_only=True)
        try:
            ws = wb.active
            header_map = self._build_append_header_map(ws)
            lookup_specs = []
            normalized_cas = str(cas_number or "").strip()
            normalized_english_name = str(substance_english_name or "").strip()
            normalized_substance = str(substance or "").strip()
            if normalized_cas != "":
                lookup_specs.append((("cas_number",), normalized_cas))
            if normalized_english_name != "":
                lookup_specs.append((("substance_english_name",), normalized_english_name))
            if normalized_substance != "":
                lookup_specs.append((("substance", "substance_chinese_name"), normalized_substance))

            for candidate_columns, target_value in lookup_specs:
                matched_rows: List[Dict[str, Any]] = []
                for row_idx in range(2, ws.max_row + 1):
                    for candidate_column in candidate_columns:
                        if candidate_column not in header_map:
                            continue
                        existing_value = str(
                            ws.cell(row=row_idx, column=header_map[candidate_column]).value or ""
                        ).strip()
                        if existing_value != target_value:
                            continue
                        row_data = self._build_append_row_snapshot(ws, header_map, row_idx)
                        matched_rows.append({
                            "row_index": row_idx,
                            "row_data": row_data,
                        })
                        break

                if len(matched_rows) == 0:
                    continue

                neat_rows = [
                    row_item
                    for row_item in matched_rows
                    if str(row_item["row_data"].get("physical_form") or "").strip().lower() == "neat"
                ]
                if len(neat_rows) > 0:
                    return neat_rows[0]
                return matched_rows[0]
        finally:
            wb.close()

        return None

    @staticmethod
    def _parse_positive_float(value: Any, field_name: str) -> float:
        """
        功能:
            将输入值解析为大于 0 的浮点数.
        参数:
            value: Any, 待解析的数值.
            field_name: str, 字段中文名, 用于异常提示.
        返回:
            float, 解析后的正数.
        异常:
            ValidationError: 字段为空, 非数字或不大于 0 时抛出.
        """
        text_value = str(value or "").strip()
        if text_value == "":
            raise ValidationError(f"{field_name}不能为空")

        try:
            numeric_value = float(text_value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"{field_name}必须为数字") from exc

        if numeric_value <= 0:
            raise ValidationError(f"{field_name}必须大于0")
        return numeric_value

    @staticmethod
    def _format_preparation_number(value: float) -> str:
        """
        功能:
            将配液或称量结果格式化为紧凑展示文本.
        参数:
            value: float, 原始数值.
        返回:
            str, 去除多余尾零后的文本.
        """
        return format(float(value), ".6g")

    def _build_solution_recipe(
        self,
        base_row_data: Dict[str, Any],
        concentration_mol_l: float,
        target_volume_ml: float,
        solvent_name: str) -> Dict[str, Any]:
        """
        功能:
            根据母体化合物信息生成溶液配置结果.
        参数:
            base_row_data: Dict[str, Any], 母体化合物行数据.
            concentration_mol_l: float, 目标浓度, 单位 mol/L.
            target_volume_ml: float, 目标定容体积, 单位 mL.
            solvent_name: str, 溶剂名称.
        返回:
            Dict[str, Any], 包含质量, 体积与展示文案的结果字典.
        异常:
            ValidationError: 缺少分子量时抛出.
        """
        molecular_weight = self._parse_positive_float(
            base_row_data.get("molecular_weight"),
            "母体化合物分子量",
        )
        normalized_solvent_name = str(solvent_name or "").strip()
        if normalized_solvent_name == "":
            raise ValidationError("溶剂名称不能为空")

        solute_moles = concentration_mol_l * target_volume_ml / 1000.0
        solute_mass_g = solute_moles * molecular_weight
        instruction_text = (
            f"称取/加入溶质 {self._format_preparation_number(solute_mass_g)} g, "
            f"用 {normalized_solvent_name} 溶解后定容至 "
            f"{self._format_preparation_number(target_volume_ml)} mL"
        )

        solute_volume_ml = None
        density_text = str(base_row_data.get("density (g/mL)") or "").strip()
        physical_state = str(base_row_data.get("physical_state") or "").strip().lower()
        if physical_state == "liquid" and density_text != "":
            try:
                density_value = self._parse_positive_float(density_text, "母体化合物密度")
                solute_volume_ml = solute_mass_g / density_value
            except ValidationError:
                logger.warning("母体液体密度无效, 跳过溶质量取体积估算: density=%s", density_text)

        return {
            "prepared_form": "solution",
            "solute_moles": solute_moles,
            "solute_mass_g": solute_mass_g,
            "solute_volume_ml": solute_volume_ml,
            "target_volume_ml": target_volume_ml,
            "active_content": concentration_mol_l,
            "solvent_name": normalized_solvent_name,
            "instruction_text": instruction_text,
        }

    def _build_beads_recipe(
        self,
        base_row_data: Dict[str, Any],
        wt_percent: float,
        target_active_mmol: float) -> Dict[str, Any]:
        """
        功能:
            根据母体化合物信息生成 beads 称量结果.
        参数:
            base_row_data: Dict[str, Any], 母体化合物行数据.
            wt_percent: float, 有效成分质量分数.
            target_active_mmol: float, 目标活性摩尔数, 单位 mmol.
        返回:
            Dict[str, Any], 包含 beads 质量与展示文案的结果字典.
        异常:
            ValidationError: 缺少分子量时抛出.
        """
        molecular_weight = self._parse_positive_float(
            base_row_data.get("molecular_weight"),
            "母体化合物分子量",
        )
        active_mass_g = target_active_mmol / 1000.0 * molecular_weight
        beads_mass_g = active_mass_g / (wt_percent / 100.0)
        instruction_text = (
            f"称取 beads {self._format_preparation_number(beads_mass_g)} g, "
            f"其中有效成分约为 {self._format_preparation_number(target_active_mmol)} mmol"
        )
        return {
            "prepared_form": "beads",
            "active_mass_g": active_mass_g,
            "beads_mass_g": beads_mass_g,
            "target_active_mmol": target_active_mmol,
            "active_content": wt_percent,
            "instruction_text": instruction_text,
        }

    def _resolve_prepared_base_chemical(
        self,
        identifier: str,
        excel_path: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """
        功能:
            为溶液或 beads 配置流程解析母体化合物, 不存在时自动补录 neat 条目.
        参数:
            identifier: str, CAS 或 SMILES.
            excel_path: Optional[str], 化学品库文件路径.
        返回:
            Optional[Dict[str, Any]], 成功时返回母体行数据与来源信息, 失败时返回 None.
        """
        from ..chem_tools.chemical_lookup import is_cas_number, lookup_chemical_by_smiles

        normalized_identifier = str(identifier or "").strip()
        if normalized_identifier == "":
            raise ValidationError("CAS 或 SMILES 不能为空")

        if is_cas_number(normalized_identifier) is True:
            existing_result = self._find_existing_chemical_row(
                excel_path=excel_path,
                cas_number=normalized_identifier,
            )
            if existing_result is not None:
                existing_result["base_created"] = False
                existing_result["chemicalbook_status"] = ""
                existing_result["chemicalbook_record_path"] = ""
                return existing_result

            append_result = self.lookup_and_append_chemical(normalized_identifier, excel_path)
            if append_result is None:
                return None
            append_result["base_created"] = True
            return append_result

        lookup_info = lookup_chemical_by_smiles(normalized_identifier)
        if lookup_info is None:
            return None

        existing_result = self._find_existing_chemical_row(
            excel_path=excel_path,
            cas_number=str(getattr(lookup_info, "cas_number", "") or "").strip(),
            substance_english_name=str(getattr(lookup_info, "substance_english_name", "") or "").strip(),
            substance=str(getattr(lookup_info, "substance", "") or "").strip(),
        )
        if existing_result is not None:
            existing_result["base_created"] = False
            existing_result["chemicalbook_status"] = ""
            existing_result["chemicalbook_record_path"] = ""
            return existing_result

        append_result = self.lookup_and_append_chemical_by_smiles(normalized_identifier, excel_path)
        if append_result is None:
            return None
        append_result["base_created"] = True
        return append_result

    @staticmethod
    def _find_duplicate_append_row(
        worksheet: Any,
        header_map: Dict[str, int],
        row_data: Dict[str, Any],
    ) -> Optional[Tuple[str, str, str, int]]:
        """
        功能:
            按约定优先级在 Excel 中查找重复化合物.
        参数:
            worksheet: Any, openpyxl 工作表对象.
            header_map: Dict[str, int], 表头名称到列号映射.
            row_data: Dict[str, Any], 准备写入的行数据.
        返回:
            Optional[Tuple[str, str, str, int]], 命中时返回
            (字段标签, 目标值, 表头名, 行号), 否则返回 None.
        """
        duplicate_specs = build_duplicate_check_specs(row_data)
        for candidate_columns, target_value, label_text in duplicate_specs:
            matched_column_name = None
            matched_column_index = None
            for candidate_column in candidate_columns:
                if candidate_column in header_map:
                    matched_column_name = candidate_column
                    matched_column_index = header_map[candidate_column]
                    break

            if matched_column_index is None:
                continue

            for row_idx in range(2, worksheet.max_row + 1):
                existing_value = str(worksheet.cell(row=row_idx, column=matched_column_index).value or "").strip()
                if existing_value == target_value:
                    return label_text, target_value, str(matched_column_name), row_idx
        return None

    def _append_chemical_row_to_excel(
        self,
        row_data: Dict[str, Any],
        excel_path: Optional[str] = None,
    ) -> Optional[int]:
        """
        功能:
            将单条化学品行数据追加到 Excel 末尾, 并执行重复检查.
        参数:
            row_data: Dict[str, Any], 准备写入的行数据.
            excel_path: Optional[str], 目标 Excel 文件路径.
        返回:
            Optional[int], 成功时返回新行行号, 命中重复时返回 None.
        """
        path = self._resolve_append_excel_path(excel_path)

        wb = load_workbook(path)
        try:
            ws = wb.active
            header_map = self._build_append_header_map(ws)

            missing_headers = collect_missing_append_headers(header_map)
            if len(missing_headers) > 0:
                logger.warning("化学品追加时发现缺失表头: %s", ", ".join(missing_headers))

            duplicate_result = self._find_duplicate_append_row(ws, header_map, row_data)
            if duplicate_result is not None:
                label_text, target_value, column_name, row_idx = duplicate_result
                summary_text = self._format_append_row_summary(row_data)
                logger.warning(
                    "化合物已存在, %s=%s, 表头=%s, 行号=%d, %s, 跳过添加",
                    label_text,
                    target_value,
                    column_name,
                    row_idx,
                    summary_text,
                )
                return None

            # 仅写入有值字段, 并设置与已有数据行一致的字体和对齐格式.
            data_font = Font(name="微软雅黑")
            data_alignment = Alignment(horizontal="center", vertical="center")

            new_row = ws.max_row + 1
            for column_name, column_index in header_map.items():
                value = get_excel_write_value(row_data, column_name)
                if value is None:
                    continue
                if isinstance(value, str) is True and value == "":
                    continue
                cell = ws.cell(row=new_row, column=column_index, value=value)
                cell.font = data_font
                cell.alignment = data_alignment

            safe_workbook_save(wb, path)
            return new_row
        finally:
            wb.close()

    def _fetch_chemicalbook_append_artifacts(
        self,
        resolved_cas: str,
    ) -> Tuple[Optional[Dict[str, Any]], str, str]:
        """
        功能:
            根据已解析的 CAS 获取 ChemicalBook 结构化结果及 sidecar 路径.
        参数:
            resolved_cas: str, 已解析出的 CAS 号.
        返回:
            Tuple[Optional[Dict[str, Any]], str, str], 依次为
            chemicalbook_record, chemicalbook_status, chemicalbook_record_path.
        """
        from ..chem_tools.chemicalbook_scraper import fetch_chemicalbook_by_cas

        normalized_cas = str(resolved_cas or "").strip()
        if normalized_cas == "":
            return None, "", ""

        chemicalbook_record = None
        chemicalbook_status = ""
        chemicalbook_record_path = ""
        try:
            chemicalbook_record = fetch_chemicalbook_by_cas(normalized_cas)
            chemicalbook_status = str(chemicalbook_record.get("status") or "")
        except Exception as exc:
            logger.warning("ChemicalBook 结构化抓取异常: CAS=%s, err=%s", normalized_cas, exc)

        if chemicalbook_record is not None:
            try:
                chemicalbook_record_path = save_chemicalbook_record(chemicalbook_record)
            except OSError as exc:
                logger.warning("ChemicalBook sidecar 保存失败: CAS=%s, err=%s", normalized_cas, exc)
                chemicalbook_record_path = ""

        return chemicalbook_record, chemicalbook_status, chemicalbook_record_path

    def lookup_and_append_chemical(
        self, query: str, excel_path: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        功能:
            在线查询化合物信息并追加到化学品库 Excel 文件末尾.
            查询分为两层: 先用多源核心查询获取可用基础信息, 再用 ChemicalBook
            结构化抓取补充全量数据, 并将原始结果保存为 sidecar JSON.
        参数:
            query: str, CAS 号或化合物中英文名称.
            excel_path: Optional[str], 目标 Excel 文件路径, 默认为 sheet/chemical_list.xlsx.
        返回:
            Optional[Dict[str, Any]], 成功返回稳定结果字典, 包含 row_data, row_index,
            chemicalbook_status, chemicalbook_record_path. 查询失败或重复时返回 None.
        """
        from ..chem_tools.chemical_lookup import is_cas_number, lookup_chemical

        normalized_query = str(query or "").strip()
        if normalized_query == "":
            logger.warning("化学品追加失败, 查询参数为空")
            return None

        # 先获取多源核心字段 (PubChem + Common Chemistry)
        info = lookup_chemical(normalized_query)

        resolved_cas = ""
        if info is not None and str(info.cas_number or "").strip() != "":
            resolved_cas = str(info.cas_number).strip()
        elif is_cas_number(normalized_query) is True:
            resolved_cas = normalized_query

        chemicalbook_record, chemicalbook_status, chemicalbook_record_path = self._fetch_chemicalbook_append_artifacts(
            resolved_cas,
        )

        if info is None and chemicalbook_record is None:
            logger.warning("在线查询未找到化合物: %s", normalized_query)
            return None

        row_data = build_append_row_data(
            query=resolved_cas,
            lookup_info=info,
            chemicalbook_record=chemicalbook_record,
        )
        if self._has_any_append_core_value(row_data) is False:
            logger.warning("化学品追加失败, 未获取到可用核心字段: %s", normalized_query)
            return None

        new_row = self._append_chemical_row_to_excel(row_data=row_data, excel_path=excel_path)
        if new_row is None:
            return None

        summary_text = self._format_append_row_summary(row_data)
        logger.info(
            "已追加化合物到 Excel: CAS=%s, 英文名=%s, %s, 行号=%d",
            row_data.get("cas_number"),
            row_data.get("substance_english_name"),
            summary_text,
            new_row,
        )

        return {
            "row_data": row_data,
            "row_index": new_row,
            "chemicalbook_status": chemicalbook_status,
            "chemicalbook_record_path": chemicalbook_record_path,
        }

    def lookup_and_append_chemical_by_smiles(
        self, smiles: str, excel_path: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        功能:
            根据单个完整 SMILES 在线查询化合物信息, 并追加到化学品库 Excel 文件末尾.
            查询顺序固定为 PubChem 结构查询, 拿到 CAS 后再补 Common Chemistry
            与 ChemicalBook.
        参数:
            smiles: str, 单个完整 SMILES 结构式.
            excel_path: Optional[str], 目标 Excel 文件路径, 默认为 sheet/chemical_list.xlsx.
        返回:
            Optional[Dict[str, Any]], 成功返回稳定结果字典, 包含 row_data, row_index,
            chemicalbook_status, chemicalbook_record_path. 查询失败或重复时返回 None.
        """
        from ..chem_tools.chemical_lookup import lookup_chemical_by_smiles

        normalized_smiles = str(smiles or "").strip()
        if normalized_smiles == "":
            logger.warning("SMILES 化学品追加失败, 查询参数为空")
            return None

        info = lookup_chemical_by_smiles(normalized_smiles)
        if info is None:
            logger.warning("SMILES 在线查询未找到化合物: %s", normalized_smiles)
            return None

        resolved_cas = str(info.cas_number or "").strip()
        chemicalbook_record, chemicalbook_status, chemicalbook_record_path = self._fetch_chemicalbook_append_artifacts(
            resolved_cas,
        )

        row_data = build_append_row_data_for_smiles(
            lookup_info=info,
            chemicalbook_record=chemicalbook_record,
        )
        if self._has_any_append_core_value(row_data) is False:
            logger.warning("SMILES 化学品追加失败, 未获取到可用核心字段: %s", normalized_smiles)
            return None

        new_row = self._append_chemical_row_to_excel(row_data=row_data, excel_path=excel_path)
        if new_row is None:
            return None

        summary_text = self._format_append_row_summary(row_data)
        logger.info(
            "已通过 SMILES 追加化合物到 Excel: SMILES=%s, CAS=%s, 英文名=%s, %s, 行号=%d",
            normalized_smiles,
            row_data.get("cas_number"),
            row_data.get("substance_english_name"),
            summary_text,
            new_row,
        )

        return {
            "row_data": row_data,
            "row_index": new_row,
            "chemicalbook_status": chemicalbook_status,
            "chemicalbook_record_path": chemicalbook_record_path,
        }

    def prepare_solution_or_beads(
        self,
        identifier: str,
        prepared_form: str,
        *,
        solvent_name: str = "",
        active_content: Any,
        target_volume_ml: Optional[Any] = None,
        target_active_mmol: Optional[Any] = None,
        excel_path: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        功能:
            根据 CAS 或 SMILES 解析母体化合物, 按指定形态生成溶液或 beads 条目并追加到化学品库.
        参数:
            identifier: str, 母体化合物 CAS 或 SMILES.
            prepared_form: str, 派生形态, 支持 solution 或 beads.
            solvent_name: str, solution 使用的溶剂名称.
            active_content: Any, solution 时表示 mol/L, beads 时表示 wt%.
            target_volume_ml: Optional[Any], solution 目标定容体积, 单位 mL.
            target_active_mmol: Optional[Any], beads 目标活性摩尔数, 单位 mmol.
            excel_path: Optional[str], 目标 Excel 文件路径.
        返回:
            Optional[Dict[str, Any]], 成功返回母体信息, 派生条目信息与配制结果摘要.
            派生条目重复或母体无法解析时返回 None.
        异常:
            ValidationError: 形态或数值参数非法时抛出.
        """
        normalized_form = str(prepared_form or "").strip().lower()
        if normalized_form not in {"solution", "beads"}:
            raise ValidationError("派生形态仅支持 solution 或 beads")

        base_result = self._resolve_prepared_base_chemical(
            identifier=identifier,
            excel_path=excel_path,
        )
        if base_result is None:
            logger.warning("未找到可用于配置的母体化合物: %s", identifier)
            return None

        base_row_data = dict(base_result.get("row_data") or {})
        if base_row_data.get("base_substance") in (None, ""):
            base_row_data["base_substance"] = (
                base_row_data.get("substance")
                or base_row_data.get("substance_chinese_name")
                or base_row_data.get("substance_english_name")
                or ""
            )

        normalized_active_content = self._parse_positive_float(active_content, "活性含量")
        if normalized_form == "solution":
            normalized_target_volume_ml = self._parse_positive_float(target_volume_ml, "目标定容体积")
            derived_row_data = build_prepared_chemical_row_data(
                base_row_data=base_row_data,
                prepared_form="solution",
                active_content=normalized_active_content,
                solvent_name=solvent_name,
            )
            recipe = self._build_solution_recipe(
                base_row_data=base_row_data,
                concentration_mol_l=normalized_active_content,
                target_volume_ml=normalized_target_volume_ml,
                solvent_name=solvent_name,
            )
        else:
            normalized_target_active_mmol = self._parse_positive_float(target_active_mmol, "目标活性 mmol")
            derived_row_data = build_prepared_chemical_row_data(
                base_row_data=base_row_data,
                prepared_form="beads",
                active_content=normalized_active_content,
            )
            recipe = self._build_beads_recipe(
                base_row_data=base_row_data,
                wt_percent=normalized_active_content,
                target_active_mmol=normalized_target_active_mmol,
            )

        derived_row_index = self._append_chemical_row_to_excel(
            row_data=derived_row_data,
            excel_path=excel_path,
        )
        if derived_row_index is None:
            logger.warning(
                "派生条目已存在, 跳过添加: identifier=%s, substance=%s",
                identifier,
                derived_row_data.get("substance"),
            )
            return None

        summary_text = self._format_append_row_summary(derived_row_data)
        logger.info(
            "已完成溶液或 beads 配置: identifier=%s, 母体新建=%s, %s, 行号=%d",
            identifier,
            base_result.get("base_created"),
            summary_text,
            derived_row_index,
        )

        return {
            "base_row_data": base_row_data,
            "base_row_index": base_result.get("row_index"),
            "base_created": bool(base_result.get("base_created")),
            "derived_row_data": derived_row_data,
            "derived_row_index": derived_row_index,
            "recipe": recipe,
            "chemicalbook_status": base_result.get("chemicalbook_status", ""),
            "chemicalbook_record_path": base_result.get("chemicalbook_record_path", ""),
        }

    # ---------- 2. 上料动作 ----------

    def _read_batch_in_records(self, file_path: str) -> List[Dict[str, str]]:
        """
        功能:
            读取上料表格文件(xlsx/csv), 返回标准化记录列表.
        参数:
            file_path: str, 上料文件路径.
        返回:
            List[Dict[str, str]], 包含 position, tray_type, content,
            shelf_position, storage 字段的记录列表.
        异常:
            FileNotFoundError: 文件不存在时自动生成模板并抛出.
        """
        path = Path(file_path)
        if not path.exists():
            logger.warning(f"未找到{file_path}. 自动生成模板文件")
            self._generate_batch_in_tray_template(path.with_suffix(".xlsx"))
            raise FileNotFoundError(f"上料文件不存在: {file_path}")

        if path.suffix == ".xlsx":
            wb = openpyxl.load_workbook(path)
            try:
                ws, header_row, header_map = self._select_batch_in_sheet(wb)
                records = self._iter_batch_in_records(ws, header_row, header_map)
            finally:
                wb.close()
            return records
        else:
            # CSV 回退: 构造与 xlsx 相同的 dict 结构
            df = pd.read_csv(path)
            df = df.fillna("")
            records: List[Dict[str, str]] = []
            for _, row in df.iterrows():
                records.append({
                    "position": str(row[0]).strip(),
                    "tray_type": str(row[1]).strip(),
                    "content": str(row[2]).strip(),
                    "shelf_position": str(row[3]).strip() if len(row) > 3 else "",
                    "storage": str(row[4]).strip() if len(row) > 4 else "",
                })
            return records

    def batch_in_tray_by_file(self, file_path: str) -> JsonDict:
        """
        功能:
            读取上料表格, 转换为中间格式, 调用父类生成 Payload 并执行上料
        参数:
            file_path: 文件路径
        返回:
            Dict: API 响应
        """
        path = Path(file_path)
        if not path.exists():
            logger.warning(f"未找到{file_path}.自动生成模板文件")
            self._generate_batch_in_tray_template(path.with_suffix(".xlsx"))
            return {}

        rows: List[Tuple[str, str, str]] = []

        # 读取文件
        if path.suffix == '.xlsx':
            wb = openpyxl.load_workbook(path)
            try:
                ws, header_row, header_map = self._select_batch_in_sheet(wb)
                batch_records = self._iter_batch_in_records(ws, header_row, header_map)
            finally:
                wb.close()

            for record in batch_records:
                rows.append((record["position"], record["tray_type"], record["content"]))
        else:
            df = pd.read_csv(path)
            df = df.fillna("")
            for _, row in df.iterrows():
                rows.append((str(row[0]), str(row[1]), str(row[2])))

        # 调用父类生成 Payload
        payload = self.build_batch_in_tray_payload(rows)

        if not payload:
            logger.warning("生成的上料数据为空")
            return {}

        # 执行上料
        resp = self.batch_in_tray(payload)

        return resp

    def batch_in_tray_with_agv_transfer(
        self,
        file_path: str = None,
        *,
        block: bool = True,
        chamber_capacity: int = 8,
    ) -> JsonDict:
        """
        功能:
            根据 batch_in_tray.xlsx 中的信息, 分轮次(每轮最多 chamber_capacity 个托盘)
            执行: 开过渡舱门 -> AGV 转运 -> 机器人上料.
            全部轮次结束后 AGV 前往充电站.
            当总托盘数 <= chamber_capacity 时, 行为与旧版本一致(单轮).
        参数:
            file_path: 上料文件路径, 默认为 sheet/batch_in_tray.xlsx
            block: 是否阻塞等待 AGV 转运完成
            chamber_capacity: 过渡舱单次最大容纳托盘数, 默认 8
        返回:
            Dict, 包含多轮次的聚合结果:
                - success: bool, 是否全部轮次成功
                - total_trays: int, 总托盘数
                - transferred_trays: int, 成功转运的托盘数
                - loaded_trays: int, 成功上料的托盘数
                - rounds: List[Dict], 每轮次详情
                - charging_result: Dict, AGV 充电结果
                - errors: List[str], 所有错误信息
                - message: str, 结果摘要
        """
        # 0. 默认文件路径
        if file_path is None:
            file_path = str(MODULE_ROOT / "sheet" / "batch_in_tray.xlsx")

        # 1. 一次性读取全部记录
        try:
            all_records = self._read_batch_in_records(file_path)
        except FileNotFoundError:
            return {
                "success": False,
                "total_trays": 0,
                "transferred_trays": 0,
                "loaded_trays": 0,
                "rounds": [],
                "charging_result": None,
                "errors": ["上料文件不存在"],
                "message": "上料文件不存在",
            }

        if not all_records:
            logger.warning("上料记录为空")
            return {
                "success": False,
                "total_trays": 0,
                "transferred_trays": 0,
                "loaded_trays": 0,
                "rounds": [],
                "charging_result": None,
                "errors": ["上料记录为空"],
                "message": "上料记录为空",
            }

        # 2. 创建 AGV 控制器(仅创建一次)
        import sys
        sys.path.append(str(Path(__file__).resolve().parent.parent.parent))
        from eit_agv.controller.agv_controller import AGVController
        agv_controller = AGVController()

        # 3. 按 chamber_capacity 分轮
        total_records = len(all_records)
        rounds_result: List[JsonDict] = []
        all_errors: List[str] = []
        total_transferred = 0
        total_loaded = 0
        overall_success = True

        for round_start in range(0, total_records, chamber_capacity):
            round_end = min(round_start + chamber_capacity, total_records)
            round_num = round_start // chamber_capacity + 1
            round_records = all_records[round_start:round_end]

            round_validate_errors = self._validate_agv_transfer_round_records(
                round_records,
                round_num=round_num,
            )
            if round_validate_errors:
                all_errors.extend(round_validate_errors)
                logger.error(f"第 {round_num} 轮上料记录校验失败, 终止后续操作")
                rounds_result.append({
                    "round_num": round_num,
                    "success": False,
                    "phase": "validate_round",
                    "errors": round_validate_errors,
                })
                overall_success = False
                break

            logger.info(
                f"===== 第 {round_num} 轮上料开始 "
                f"(记录 {round_start + 1}~{round_end}/{total_records}) ====="
            )

            # 3a. 打开过渡舱外门
            door_errors = self._ensure_outer_door_open()
            if door_errors:
                all_errors.extend(door_errors)
                logger.error(f"第 {round_num} 轮开门失败, 终止后续操作")
                rounds_result.append({
                    "round_num": round_num,
                    "success": False,
                    "phase": "open_door",
                    "errors": door_errors,
                })
                overall_success = False
                break

            # 3b. AGV 转运本轮托盘(内部按 4 个一批)
            round_tasks, build_errors = self._build_agv_transfer_tasks(round_records)
            all_errors.extend(build_errors)
            if build_errors:
                logger.error(f"第 {round_num} 轮 AGV 任务构建失败, 终止后续操作")
                rounds_result.append({
                    "round_num": round_num,
                    "success": False,
                    "phase": "build_agv_tasks",
                    "errors": build_errors,
                })
                overall_success = False
                break

            transferred = 0
            batches: List[JsonDict] = []
            if round_tasks:
                transferred, batches, transfer_errors = (
                    self._execute_agv_transfer_batches(
                        round_tasks, agv_controller, block=block
                    )
                )
                total_transferred += transferred
                all_errors.extend(transfer_errors)

                if transfer_errors:
                    logger.error(f"第 {round_num} 轮 AGV 转运失败, 终止后续操作")
                    rounds_result.append({
                        "round_num": round_num,
                        "success": False,
                        "phase": "agv_transfer",
                        "transferred": transferred,
                        "batches": batches,
                        "errors": transfer_errors,
                    })
                    overall_success = False
                    break

            # 3c. AGV 转运完成后先回充电站等待, 机器人上料期间 AGV 充电
            try:
                charging_result = agv_controller.go_to_charging_station()
                if charging_result is not None:
                    logger.info(f"第 {round_num} 轮 AGV 已返回充电站")
                else:
                    logger.warning(f"第 {round_num} 轮 AGV 返回充电站失败")
            except Exception as e:
                logger.warning(f"第 {round_num} 轮 AGV 返回充电站异常: {e}, 继续执行上料")

            # 3d. 执行本轮上料(机器人将托盘从过渡舱搬到工位)
            round_rows = [
                (r["position"], r["tray_type"], r["content"])
                for r in round_records
            ]
            payload = self.build_batch_in_tray_payload(round_rows)

            in_tray_result = None
            if payload:
                try:
                    in_tray_result = self.batch_in_tray(payload)
                    total_loaded += len(round_rows)
                    logger.info(f"第 {round_num} 轮上料完成")
                except Exception as e:
                    error_msg = f"第 {round_num} 轮上料异常: {e}"
                    logger.error(error_msg)
                    all_errors.append(error_msg)
                    rounds_result.append({
                        "round_num": round_num,
                        "success": False,
                        "phase": "in_tray",
                        "transferred": transferred,
                        "batches": batches,
                        "in_tray_result": None,
                        "errors": [error_msg],
                    })
                    overall_success = False
                    break
            else:
                logger.warning(f"第 {round_num} 轮上料 payload 为空, 跳过上料")

            # 本轮成功
            rounds_result.append({
                "round_num": round_num,
                "success": True,
                "phase": "completed",
                "transferred": transferred,
                "batches": batches,
                "in_tray_result": in_tray_result,
                "loaded_count": len(round_rows),
            })

            logger.info(f"===== 第 {round_num} 轮上料完成 =====")

        # 4. 确保 AGV 在充电站(正常流程中每轮转运后已回充电站, 此处为保底)
        final_charging_result = None
        try:
            final_charging_result = agv_controller.go_to_charging_station()
            if final_charging_result is not None:
                logger.info("AGV 已确认在充电站")
            else:
                logger.warning("AGV 返回充电站失败")
        except Exception as e:
            logger.error(f"AGV 返回充电站时发生异常: {e}")

        # 5. 聚合返回
        overall_success = overall_success and len(all_errors) == 0

        return {
            "success": overall_success,
            "total_trays": total_records,
            "transferred_trays": total_transferred,
            "loaded_trays": total_loaded,
            "rounds": rounds_result,
            "charging_result": final_charging_result,
            "errors": all_errors,
            "message": (
                "全部轮次完成" if overall_success
                else f"执行过程中出现错误, 完成 {len([r for r in rounds_result if r.get('success')])} 轮"
            ),
        }

    def _validate_agv_transfer_round_records(
        self,
        records: List[Dict[str, str]],
        *,
        round_num: int,
    ) -> List[str]:
        """
        功能:
            校验单轮 AGV 上料记录是否满足唯一性要求.
            同一轮次内 position 与 shelf_position 都必须唯一, 且不能为空.
        参数:
            records: List[Dict[str, str]], 单轮上料记录列表.
            round_num: int, 当前轮次编号.
        返回:
            List[str], 校验失败时的错误信息列表.
        """
        errors: List[str] = []
        positions: List[str] = []
        shelf_positions: List[str] = []

        for index, record in enumerate(records, start=1):
            position = str(record.get("position", "")).strip()
            shelf_position = str(record.get("shelf_position", "")).strip()

            if position == "":
                errors.append(f"第 {round_num} 轮第 {index} 条记录缺少 position, 无法执行 AGV 上料")
            else:
                positions.append(position)

            if shelf_position == "":
                errors.append(f"第 {round_num} 轮第 {index} 条记录缺少 shelf_position, 无法执行 AGV 上料")
            else:
                shelf_positions.append(shelf_position)

        duplicate_positions = self._collect_duplicate_texts(positions)
        if duplicate_positions:
            errors.append(
                "第 "
                f"{round_num} 轮存在重复的 position: {', '.join(duplicate_positions)}. "
                "同一轮次不能复用 TB 位, 请检查 batch_in_tray.xlsx 顺序或重新生成上料文件"
            )

        duplicate_shelf_positions = self._collect_duplicate_texts(shelf_positions)
        if duplicate_shelf_positions:
            errors.append(
                "第 "
                f"{round_num} 轮存在重复的 shelf_position: {', '.join(duplicate_shelf_positions)}. "
                "同一轮次不能复用同一个货架位, 请检查 batch_in_tray.xlsx 顺序或重新生成上料文件"
            )

        return errors

    def _generate_batch_in_tray_template(self, file_path: Path) -> None:
        """
        功能:
            生成批量上料Excel模板, 配置上料点位下拉、托盘类型下拉与内容示例
        参数:
            file_path: Path, 模板输出路径
        返回:
            None
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "batch_in_tray"
        ws.append(["position", "tray_type", "content", "shelf_position", "storage"])
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50

        # 位置下拉，包含 TB 列与 W-1-1~W-1-8 货位
        positions_tb = [f"TB-{row}-{col}" for row in (1, 2) for col in range(1, 5)]
        positions_w = [f"W-1-{index}" for index in range(1, 9)]
        positions = positions_tb + positions_w
        dv_pos = DataValidation(type="list", formula1=f"\"{','.join(positions)}\"")
        ws.add_data_validation(dv_pos)
        dv_pos.add("A2:A101")

        # 托盘下拉，耗材显示数量范围，带物质显示点位范围
        consumable_trays = {
            int(ResourceCode.TIP_TRAY_50UL),
            int(ResourceCode.TIP_TRAY_1ML),
            int(ResourceCode.TIP_TRAY_5ML),
            int(ResourceCode.REACTION_SEAL_CAP_TRAY),
            int(ResourceCode.FLASH_FILTER_INNER_BOTTLE_TRAY),
            int(ResourceCode.FLASH_FILTER_OUTER_BOTTLE_TRAY),
            int(ResourceCode.REACTION_TUBE_TRAY_2ML),
            int(ResourceCode.TEST_TUBE_MAGNET_TRAY_2ML),
        }
        tray_display: List[str] = []
        for code, name in TRAY_CODE_DISPLAY_NAME.items():
            base_text = f"{name}({code})"
            try:
                enum_name = ResourceCode(code).name
                spec = getattr(TraySpec, enum_name, None)
            except Exception:
                spec = None

            if spec is None:
                tray_display.append(base_text)
                continue

            col_count, row_count = spec
            if col_count <= 0 or row_count <= 0:
                tray_display.append(base_text)
                continue

            if code in consumable_trays:
                capacity = col_count * row_count
                tray_display.append(f"{base_text} [1-{capacity}]")
            else:
                end_row_char = chr(ord("A") + row_count - 1)
                tray_display.append(f"{base_text} [A1-{end_row_char}{col_count}]")

        # 用隐藏sheet作为数据源，避免下拉字符串过长
        tray_sheet = wb.create_sheet("validation_meta")
        for idx, option in enumerate(tray_display, start=1):
            tray_sheet.cell(row=idx, column=1).value = option
        tray_sheet.sheet_state = "hidden"

        # 定义命名区域, 避免跨 sheet 验证被 Excel 写成 x14 扩展
        options_name = "tray_type_options"
        options_ref  = f"validation_meta!$A$1:$A${len(tray_display)}"
        wb.defined_names.add(DefinedName(options_name, attr_text=options_ref))

        dv_tray = DataValidation(
            type="list",
            formula1=f"={options_name}",
            showInputMessage=True,
        )
        ws.add_data_validation(dv_tray)
        dv_tray.add("B2:B101")

        ws["C1"] = "content(耗材填数量; 物质填: A1|名称|2mL; B2|名称|5mg)"
        ws["D1"] = "shelf_position"
        ws["E1"] = "storage(格式: 物质|位置; 多个用;隔开)"
        safe_workbook_save(wb, file_path)
        logger.info(f"已生成上料模板: {file_path}")

    def _find_header_in_sheet(self, worksheet: Any, header_keyword: str) -> Tuple[Optional[int], Optional[int]]:
        """
        功能:
            在单个工作表的前 50 行和前 50 列中查找目标表头.
        参数:
            worksheet: openpyxl 工作表对象.
            header_keyword: 需要匹配的表头关键词, 例如"实验编号".
        返回:
            Tuple[Optional[int], Optional[int]], 命中时返回(行号, 列号), 未命中返回(None, None).
        """
        max_scan_row = min(worksheet.max_row, 50)
        max_scan_col = min(worksheet.max_column, 50)

        for row_index in range(1, max_scan_row + 1):
            for col_index in range(1, max_scan_col + 1):
                cell_value = worksheet.cell(row_index, col_index).value
                if isinstance(cell_value, str) is False:
                    continue

                # 去除空白字符, 兼容用户在表头中插入空格或换行.
                normalized_text = cell_value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "").strip()
                if header_keyword in normalized_text:
                    return row_index, col_index

        return None, None

    def _select_task_template_sheet(
        self,
        workbook: Workbook,
        header_keyword: str = "实验编号",
    ) -> Tuple[Optional[Any], Optional[int], Optional[int]]:
        """
        功能:
            从任务模板工作簿中选择包含目标表头的工作表.
            选择顺序为: "实验方案设定" -> 当前激活工作表 -> 其余工作表.
        参数:
            workbook: openpyxl Workbook 对象.
            header_keyword: 需要匹配的表头关键词.
        返回:
            Tuple[Optional[Any], Optional[int], Optional[int]], 命中时返回(工作表, 行号, 列号), 未命中返回(None, None, None).
        """
        candidate_sheet_names: List[str] = []
        preferred_sheet_name = "实验方案设定"
        if preferred_sheet_name in workbook.sheetnames:
            candidate_sheet_names.append(preferred_sheet_name)

        active_sheet_name = workbook.active.title
        if active_sheet_name not in candidate_sheet_names:
            candidate_sheet_names.append(active_sheet_name)

        for sheet_name in workbook.sheetnames:
            if sheet_name not in candidate_sheet_names:
                candidate_sheet_names.append(sheet_name)

        for sheet_name in candidate_sheet_names:
            worksheet = workbook[sheet_name]
            header_row, exp_no_col = self._find_header_in_sheet(worksheet, header_keyword)
            if header_row is not None and exp_no_col is not None:
                if sheet_name != active_sheet_name:
                    logger.info(
                        "模板解析使用工作表: %s, 当前激活工作表: %s",
                        sheet_name,
                        active_sheet_name,
                    )
                return worksheet, header_row, exp_no_col

        return None, None, None

    # ---------- 3. 任务生成文件处理 ----------
    def create_task_by_file(self, template_path: str, chemical_db_path: str) -> JsonDict:
        """
        功能:
            读取任务模板和化学品库，解析为中间数据，调用父类生成任务 Payload 并提交
        参数:
            template_path: 实验模板路径
            chemical_db_path: 化学品库路径
        返回:
            Dict: 任务创建结果
        """
        t_path = Path(template_path)
        c_path = Path(chemical_db_path)

        # 1. 检查并生成模板
        if not t_path.exists():
            self._generate_reaction_template(t_path)
            raise FileNotFoundError(f"已生成模板 {t_path}，请填写后重试")

        if not c_path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {c_path}")

        # 2. 读取化学品库 -> Dict
        chem_df = self._read_table_file_with_required_columns(
            c_path,
            required_columns=["substance"],
        )
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row, *keys, default=None):
            for k in keys:
                if k in row and pd.notna(row[k]):
                    return row[k]
            return default

        chemical_db: Dict[str, Dict[str, Any]] = {}
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if not name:
                continue

            # 小写后的列名
            chemical_db[name] = {
                "chemical_id": _pick(row, "chemical_id"),
                "molecular_weight": _pick(row, "molecular_weight", "mw"),
                "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                "density (g/mL)": _pick(row, "density (g/ml)", "density(g/ml)", "density_g_ml", "density", default=None),
                "physical_form": str(_pick(row, "physical_form", default="") or "").strip().lower(),
                "active_content": _pick(row, "active_content","active_content(mmol/ml or wt%)" ,"active_content(mol/l or wt%)", default="" ),
            }

        # 3. 读取任务模板 -> params(Dict), headers(List), data_rows(List[List])
        wb = load_workbook(t_path, data_only=True)
        ws, header_row, exp_no_col = self._select_task_template_sheet(wb, header_keyword="实验编号")
        if ws is None or header_row is None or exp_no_col is None:
            raise ValueError(f"模板中未找到'实验编号'表头, 可用工作表: {wb.sheetnames}")

        # 3.2 提取全局参数（左侧 A/B）
        # - 实验名称：A1是标签，用户通常填在 B1
        params: Dict[str, Any] = {}
        exp_name = ws.cell(1, 2).value  # B1
        if exp_name is not None and str(exp_name).strip() != "":
            params["实验名称"] = str(exp_name).strip()

        # 扫描 A/B（从第2行开始，遇到“注：”不停止也可以；这里仅跳过“注：”本行）
        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 1).value
            val = ws.cell(r, 2).value

            if key is None:
                continue
            key_str = str(key).strip()
            if not key_str:
                continue

            # 跳过注释行（不写入 params；否则会污染）
            if key_str.startswith("注：") or key_str.startswith("注:"):
                continue

            # 分类标题行通常是合并单元格，B 为空；这类不要写入 params
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue

            params[key_str] = val

        # 3.3 生成 headers（从 “实验编号”列开始往右：C..M）
        # 同时把 “试剂_1” -> “试剂名称_1”，让 build_task_payload 能识别
        raw_headers: List[Any] = []
        for c in range(exp_no_col, ws.max_column + 1):
            raw_headers.append(ws.cell(header_row, c).value)

        headers: List[str] = []
        reagent_idx = 0
        for h in raw_headers:
            s = "" if h is None else str(h).strip()

            # 规范化：试剂_1/试剂1 -> 试剂名称_1
            if s.startswith("试剂") and "量" not in s and s != "试剂名称":
                reagent_idx += 1
                headers.append(f"试剂名称_{reagent_idx}")
                continue

            # 规范化：试剂量 -> 试剂量_1/2/...
            if "试剂量" in s:
                # 若前面还没遇到试剂列，给个兜底编号
                idx = reagent_idx if reagent_idx > 0 else (len([x for x in headers if "试剂量" in x]) + 1)
                headers.append(f"试剂量_{idx}")
                continue

            headers.append(s)

        # 3.4 生成 data_rows：从表头下一行开始，按实验编号列读取到最后一列（C..M）
        data_rows: List[List[Any]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            exp_no = ws.cell(r, exp_no_col).value

            # 实验编号为空：认为实验区结束（模板一般后面都是空）
            if exp_no is None or (isinstance(exp_no, str) and exp_no.strip() == ""):
                # 只有在已经读到至少一行实验后才 break，避免中间空行误判
                if data_rows:
                    break
                else:
                    continue

            row_vals: List[Any] = []
            for c in range(exp_no_col, ws.max_column + 1):
                v = ws.cell(r, c).value
                # 这里不要强制 str 化，build_task_payload 内部会 str()；但 None 要变成 ""
                row_vals.append("" if v is None else v)

            data_rows.append(row_vals)

        # 4. 调用父类纯逻辑生成 Payload
        task_payload = self.build_task_payload(params, headers, data_rows, chemical_db)

        # 5. 提交任务信息到工站
        try:
            resp = self._submit_task_payload(task_payload)
        except ApiError as exc:
            if getattr(exc, "code", None) == 409:
                # 自动重命名: 在任务名称后添加当前日期时间(精确到秒)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                task_name = task_payload.get("task_name") or params.get("实验名称")
                new_task_name = f"{task_name}_{timestamp}"

                task_payload["task_name"] = new_task_name
                logger.info(f"任务名称重复, 自动重命名为: {new_task_name}")

                # 重试提交
                try:
                    resp = self._submit_task_payload(task_payload)
                except ApiError as retry_exc:
                    logger.error(f"重命名后任务提交仍失败: {retry_exc}")
                    raise
            else:
                raise

        # 6. 提交任务信息到工站
        task_id = resp.get("task_id")

        # 7. 回写任务ID和任务名称到模板
        try:
            task_id_int = int(task_id)
            id_updated = False
            name_updated = False
            # 获取实际提交的任务名称(可能经过重命名)
            final_task_name = task_payload.get("task_name")

            for r in range(1, ws.max_row + 1):
                key_val = ws.cell(r, 1).value
                if key_val is None:
                    continue
                key_str = str(key_val).strip()
                # 回写实验ID
                if key_str == "实验ID":
                    ws.cell(r, 2, value=task_id_int)
                    id_updated = True
                # 回写任务名称(重命名后同步更新模板)
                if key_str == "实验名称" and final_task_name is not None:
                    ws.cell(r, 2, value=final_task_name)
                    name_updated = True

            if id_updated or name_updated:
                safe_workbook_save(wb, t_path)
                if id_updated:
                    logger.info("已将任务ID写入模板文件: %s", t_path)
                if name_updated:
                    logger.info("已将任务名称同步写入模板文件: %s", final_task_name)
            else:
                logger.warning("未找到'实验ID'位置, 未回写任务ID")
        except Exception as exc:
            logger.warning("任务ID回写失败: %s", exc)

        # 8. 将模板文件拷贝到 data/tasks/<task_id>/ 并重命名为任务ID
        try:
            task_dir = MODULE_ROOT / "data" / "tasks" / str(task_id)
            task_dir.mkdir(parents=True, exist_ok=True)          # 创建任务文件夹(若已存在则忽略)
            dest_path = task_dir / f"{task_id}_experiment_plan{t_path.suffix}"   # 目标路径: <task_id>_experiment_plan.xlsx
            shutil.copy2(t_path, dest_path)                      # 拷贝并保留元数据
            logger.info("已将模板文件拷贝至任务目录: %s", dest_path)
        except Exception as exc:
            logger.warning("模板文件拷贝至任务目录失败: %s", exc)

        return task_id

    def _generate_reaction_template(self, path: Path) -> None:
        """
        生成与 reeaction_template.xlsx 一致的反应模板
        结构：左侧为参数配置区，右侧为实验试剂填报区
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 模板默认字体：等线 11
        base_font = Font(name="Microsoft YaHei", charset=134, family=2, scheme="minor", sz=11)
        title_font = Font(name="Microsoft YaHei", charset=134, family=2, scheme="minor", sz=11, bold=True)
        center = Alignment(horizontal="center", vertical="center")

        # 覆盖默认 Normal 样式，保证空白单元格也用微软雅黑
        for style in getattr(wb, "_named_styles", []):
            if getattr(style, "name", "").lower() == "normal":
                style.font = base_font
                break

        # --- 1. 定义左侧参数配置数据 (行2开始, A列和B列) ---
        left_params = [
            ("实验设定", ""),
            ("实验名称", "Auto_task"),
            ("实验ID", 0),
            ("反应设定", ""),
            ("反应规模(mmol)", "0.2"),
            ("反应器类型", "heat"),
            ("反应时间(min/h)", "8h"),
            ("反应温度(°C)", 40),
            ("转速(rpm)", 500),
            ("搅拌后⽬标温度(°C)", 30),
            ("等待目标温度", "否"),
            ("称量设定", ""),
            ("称量误差(%)", 3),
            ("最大称量误差(mg)", 1),
            ("加料设定", ""),
            ("固定加料顺序", "否"),
            ("自动加磁子", "是"),
            ("内标设定", ""),
            ("内标种类", "1,3,5-三异丙基苯(溶液,1mol/L in MeCN)"),
            ("内标用量(μL/mg)", 100),
            ("加入内标后搅拌时间(min)", 5),
            ("稀释设定", ""),
            ("稀释液种类", "乙腈"),
            ("稀释量(μL)", 500),
            ("闪滤设定", ""),
            ("闪滤液种类", "乙腈"),
            ("闪滤液用量(μL)", 500),
            ("取样量(μL)", 1),
            ("闪滤实验编号", "全部"),   # 空/"全部"=全部实验闪滤; 支持 "1-12,24,28" 格式
            ("", ""),  # 空行
        ]
        left_param_rows = len(left_params)

        # --- 2. 设置第一行表头 (Row 1) ---
        ws.cell(row=1, column=3, value="实验编号").font = base_font
        
        reagent_count = 5
        current_col = 4
        for i in range(1, reagent_count + 1):
            ws.cell(row=1, column=current_col, value=f"试剂").font = base_font
            ws.cell(row=1, column=current_col + 1, value="试剂量").font = base_font
            current_col += 2

        # --- 3. 填充左侧参数区 (Row 2 ~ Row 22) ---
        for idx, (param_name, default_val) in enumerate(left_params):
            row_idx = idx + 1  # 从第2行开始

            # 分类标题：模板是 A:B 合并，只写 A 列，且加粗
            if param_name and default_val == "":
                ws.cell(row=row_idx, column=1, value=param_name).font = title_font
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                continue

            # 空行：保持空
            if param_name == "" and default_val == "":
                continue

            # 普通参数行
            ws.cell(row=row_idx, column=1, value=param_name).font = base_font
            ws.cell(row=row_idx, column=2, value=default_val).font = base_font

        # --- 4. 填充右侧实验编号 (Row 2 ~ Row 25) ---
        for i in range(1, 25):  # 1~24
            row_idx = i + 1
            ws.cell(row=row_idx, column=3, value=i).font = base_font

        # --- 5. 底部注释 (跟随参数行, 预留一行空白) ---
        note_row = left_param_rows + 2
        note_text = "注：试剂量支持单位：(eq,mmol,g,mg,μL,mL）"
        ws.cell(row=note_row, column=1, value=note_text).font = base_font
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
        ws.cell(row=note_row, column=1).alignment = center  # 合并后的单元格居中

        max_template_row = max(note_row, 25)

        # --- 6. 字体铺满 (A1:M*)  ---
        for r in range(1, max_template_row + 1):
            for c in range(1, 14):  # A..M
                cell = ws.cell(r, c)
                # 标题行的粗体不要覆盖
                if cell.font and cell.font.bold:
                    continue
                cell.font = base_font

        # --- 7. 对齐 ---
        # C~L 整块都居中（含空白）
        for r in range(1, max_template_row + 1):
            for c in range(3, 13):  # C..L
                ws.cell(r, c).alignment = center

        # A 列：参数行和注释行居中
        a_rows = []
        b_rows = []
        for idx, (param_name, default_val) in enumerate(left_params):
            row_idx = idx + 1
            if param_name != "":
                a_rows.append(row_idx)
            if param_name != "" and default_val != "":
                b_rows.append(row_idx)
        for r in a_rows + [note_row]:
            ws.cell(r, 1).alignment = center

        # B 列：只有有值的参数行居中（标题行/空白行/合并后的 B 不处理）
        for r in b_rows:
            ws.cell(r, 2).alignment = center

        # M 列：只有表头 M1 居中
        ws.cell(1, 13).alignment = center

        # 表头 A1/C1 也居中（模板如此）
        ws.cell(1, 1).alignment = center
        ws.cell(1, 3).alignment = center

        # --- 8. 列宽： ---
        widths_map = {
            "A": 26.0,
            "B": 38.0,
            "C": 15.0,
            "D": 14.0,
            "E": 14.0,
            "F": 14.0,
            "G": 14.0,
            "H": 14.0,
            "I": 14.0,
            "J": 14.0,
            "K": 14.0,
            "L": 14.0,
            "M": 14.0,
        }
        for col_letter, w in widths_map.items():
            ws.column_dimensions[col_letter].width = w

        safe_workbook_save(wb, path)
        logger.info(f"已生成任务模板: {path}")

    # ---------- 4. 物料核算 ----------
    def check_resource_for_task(self, template_path: str, chemical_db_path: str, auto_generate_batch_file: bool = True) -> JsonDict:
        """
        功能:
            读取实验模板与化学品库, 构建任务 Payload, 获取站内资源并比对是否满足实验需求。
        参数:
            template_path: 实验模板文件路径(xlsx/csv)。
            chemical_db_path: 化学品库文件路径(xlsx/csv)。
            auto_generate_batch_file: 是否自动生成上料文件, 默认为 True。
        返回:
            Dict, analyze_resource_readiness 的结果, 包含需求、库存、缺失与冗余信息。
        """
        t_path = Path(template_path)
        c_path = Path(chemical_db_path)

        if not t_path.exists():
            raise FileNotFoundError(f"未找到实验模板文件: {t_path}")
        if not c_path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {c_path}")

        chem_df = self._read_table_file_with_required_columns(
            c_path,
            required_columns=["substance"],
        )
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row, *keys, default=None):
            for k in keys:
                if k in row and pd.notna(row[k]):
                    return row[k]
            return default

        chemical_db: Dict[str, Dict[str, Any]] = {}
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if not name:
                continue
            chemical_db[name] = {
                "chemical_id": _pick(row, "chemical_id"),
                "molecular_weight": _pick(row, "molecular_weight", "mw"),
                "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                "density (g/mL)": _pick(row, "density (g/ml)", "density(g/ml)", "density_g_ml", "density", default=None),
                "physical_form": str(_pick(row, "physical_form", default="") or "").strip().lower(),
                "active_content": _pick(row, "active_content", "active_content(mmol/ml or wt%)", "active_content(mol/l or wt%)", default=""),
            }

        wb = load_workbook(t_path, data_only=True)
        ws, header_row, exp_no_col = self._select_task_template_sheet(wb, header_keyword="实验编号")
        if ws is None or header_row is None or exp_no_col is None:
            raise ValidationError(f"模板中未找到'实验编号'表头, 可用工作表: {wb.sheetnames}")

        params: Dict[str, Any] = {}
        exp_name = ws.cell(1, 2).value
        if exp_name is not None and str(exp_name).strip() != "":
            params["实验名称"] = str(exp_name).strip()

        task_id = None  # 用于存储实验ID
        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 1).value
            val = ws.cell(r, 2).value
            if key is None:
                continue
            key_str = str(key).strip()
            if key_str == "":
                continue
            if key_str.startswith("注：") or key_str.startswith("注"):
                continue
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            
            # 识别实验ID参数并提取整数值
            if key_str == "实验ID":
                try:
                    task_id = int(val)
                    self._logger.info("从模板中读取到实验ID: %d", task_id)
                except (ValueError, TypeError):
                    self._logger.warning("实验ID格式无效: %s, 将跳过二次校验", val)
            
            params[key_str] = val

        raw_headers: List[Any] = []
        for c in range(exp_no_col, ws.max_column + 1):
            raw_headers.append(ws.cell(header_row, c).value)

        headers: List[str] = []
        reagent_idx = 0
        for h in raw_headers:
            s = "" if h is None else str(h).strip()
            if s.startswith("试剂") and "量" not in s and s != "试剂名称":
                reagent_idx += 1
                headers.append(f"试剂名称_{reagent_idx}")
                continue
            if "试剂量" in s:
                idx = reagent_idx if reagent_idx > 0 else (len([x for x in headers if "试剂量" in x]) + 1)
                headers.append(f"试剂量_{idx}")
                continue
            headers.append(s)

        data_rows: List[List[Any]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            exp_no = ws.cell(r, exp_no_col).value
            if exp_no is None or (isinstance(exp_no, str) and exp_no.strip() == ""):
                if data_rows:
                    break
                else:
                    continue
            row_vals: List[Any] = []
            for c in range(exp_no_col, ws.max_column + 1):
                v = ws.cell(r, c).value
                row_vals.append("" if v is None else v)
            data_rows.append(row_vals)

        task_payload = self.build_task_payload(params, headers, data_rows, chemical_db)
        resource_rows = self.get_resource_info()
        result = self.analyze_resource_readiness(task_payload, resource_rows, chemical_db, task_id=task_id)

        # 自动保存物料核算结果
        if self._data_manager and task_id:
            self._data_manager.save_resource_check(str(task_id), result)

        # 自动生成上料文件
        if auto_generate_batch_file and task_id:
            self.auto_generate_batch_in_tray_from_resource_check(task_id)

        return result

    def auto_generate_batch_in_tray_from_resource_check(self, task_id: Optional[int] = None) -> None:
        """
        功能:
            根据资源核查结果自动修改上料文件, 考虑料盘规格, 优先填满一个料盘再使用下一个
        参数:
            task_id: 任务ID, 如果为None则自动搜索data/tasks中id最大且状态为UNSTARTED的任务
        返回:
            None
        """
        import json
        import math

        # 1. 确定任务ID
        if task_id is None:
            tasks_dir = MODULE_ROOT / "data/tasks"
            if not tasks_dir.exists():
                raise FileNotFoundError(f"任务目录不存在: {tasks_dir}")

            # 获取所有任务文件夹, 按ID降序排列
            task_folders = []
            for folder in tasks_dir.iterdir():
                if folder.is_dir() and folder.name.isdigit():
                    task_folders.append(int(folder.name))

            if not task_folders:
                raise FileNotFoundError("未找到任何任务文件夹")

            task_folders.sort(reverse=True)

            # 查找第一个状态为UNSTARTED的任务
            found_task = None
            for tid in task_folders:
                task_info_path = tasks_dir / str(tid) / "task_info.json"
                if task_info_path.exists():
                    with open(task_info_path, "r", encoding="utf-8") as f:
                        task_info = json.load(f)
                    if task_info.get("status") == "UNSTARTED":
                        found_task = tid
                        break

            if found_task is None:
                raise ValueError("未找到状态为UNSTARTED的任务")

            task_id = found_task
            logger.info(f"自动选择任务ID: {task_id}")

        # 2. 读取resource_check.json
        resource_check_path = MODULE_ROOT / "data" / "tasks" / str(task_id) / "resource_check.json"
        if not resource_check_path.exists():
            raise FileNotFoundError(f"未找到资源核查文件: {resource_check_path}")

        with open(resource_check_path, "r", encoding="utf-8") as f:
            resource_check = json.load(f)

        missing_list = resource_check.get("missing", [])
        if not missing_list:
            logger.info("没有缺失的物资, 无需生成上料文件")
            return

        # 3. 读取chemical_list.xlsx
        chemical_list_path = MODULE_ROOT / "sheet" / "chemical_list.xlsx"
        if not chemical_list_path.exists():
            raise FileNotFoundError(f"未找到化学品列表文件: {chemical_list_path}")

        chem_df = self._read_table_file_with_required_columns(
            chemical_list_path,
            required_columns=["substance", "physical_state", "storage_location"],
        )
        chem_df = chem_df.fillna("")

        # 创建物质名到信息的映射
        chemical_dict = {}
        for _, row in chem_df.iterrows():
            substance = str(row.get("substance", "")).strip()
            if substance:
                chemical_dict[substance] = {
                    "physical_state": str(row.get("physical_state", "")).strip().lower(),
                    "storage_location": str(row.get("storage_location", "")).strip()
                }

        # 4. 准备料盘规格信息和位置管理
        # TB 位共 8 个, 上料完成后空出可复用; 货架共 3 行 × 4 列 = 12 个位置
        position_list = ["TB-2-1", "TB-2-2", "TB-2-3", "TB-2-4",
                         "TB-1-1", "TB-1-2", "TB-1-3", "TB-1-4",
                         "TB-2-1", "TB-2-2", "TB-2-3", "TB-2-4"]
        shelf_position_list = ["3-1", "3-2", "3-3", "3-4",
                               "2-1", "2-2", "2-3", "2-4",
                               "1-1", "1-2", "1-3", "1-4"]

        # 料盘类型到规格的映射
        tray_spec_map = {
            int(ResourceCode.REAGENT_BOTTLE_TRAY_2ML): TraySpec.REAGENT_BOTTLE_TRAY_2ML,
            int(ResourceCode.REAGENT_BOTTLE_TRAY_8ML): TraySpec.REAGENT_BOTTLE_TRAY_8ML,
            int(ResourceCode.REAGENT_BOTTLE_TRAY_40ML): TraySpec.REAGENT_BOTTLE_TRAY_40ML,
            int(ResourceCode.POWDER_BUCKET_TRAY_30ML): TraySpec.POWDER_BUCKET_TRAY_30ML,
            int(ResourceCode.REACTION_TUBE_TRAY_2ML): TraySpec.REACTION_TUBE_TRAY_2ML,
            int(ResourceCode.TEST_TUBE_MAGNET_TRAY_2ML): TraySpec.TEST_TUBE_MAGNET_TRAY_2ML,
            int(ResourceCode.REACTION_SEAL_CAP_TRAY): TraySpec.REACTION_SEAL_CAP_TRAY,
            int(ResourceCode.FLASH_FILTER_INNER_BOTTLE_TRAY): TraySpec.FLASH_FILTER_INNER_BOTTLE_TRAY,
            int(ResourceCode.FLASH_FILTER_OUTER_BOTTLE_TRAY): TraySpec.FLASH_FILTER_OUTER_BOTTLE_TRAY,
            int(ResourceCode.TIP_TRAY_50UL): TraySpec.TIP_TRAY_50UL,
            int(ResourceCode.TIP_TRAY_1ML): TraySpec.TIP_TRAY_1ML,
            int(ResourceCode.TIP_TRAY_5ML): TraySpec.TIP_TRAY_5ML,
        }

        # 跟踪每种料盘类型的使用情况:
        # {tray_type_code: [(position, shelf_position, current_slot_index, max_slots, alloc_idx)]}
        tray_usage = {}
        # 按索引追踪已分配的位置, 避免 TB 位复用时误判
        allocated_position_indices: set = set()

        def _get_slot_name(slot_index: int, cols: int, rows: int) -> str:
            """
            功能:
                根据坑位索引生成坑位名称, 按行优先排列
            参数:
                slot_index: 坑位索引(从0开始)
                cols: 列数
                rows: 行数
            返回:
                坑位名称, 如 "A1", "A2", "B1" 等
            """
            row_idx = slot_index // cols
            col_idx = slot_index % cols
            row_letter = chr(ord('A') + row_idx)
            return f"{row_letter}{col_idx + 1}"

        def _allocate_slot(tray_type_code: int, tray_type_name: str) -> Tuple[str, str, str]:
            """
            功能:
                为指定料盘类型分配一个坑位.
                优先填满已有的未满料盘, 否则从位置列表中分配新位置.
                使用索引追踪已分配位置, 支持 TB 位跨轮次复用.
            参数:
                tray_type_code: 料盘类型代码
                tray_type_name: 料盘类型名称
            返回:
                (position, shelf_position, slot_name) 元组
            """
            # 获取料盘规格
            spec = tray_spec_map.get(tray_type_code)
            if spec is None:
                raise ValueError(f"未找到料盘规格: {tray_type_code}")

            cols, rows = spec
            max_slots = cols * rows

            # 检查是否已有该类型的料盘在使用
            if tray_type_code not in tray_usage:
                tray_usage[tray_type_code] = []

            # 查找是否有未满的料盘
            for tray_info in tray_usage[tray_type_code]:
                position, shelf_position, current_slot, max_slots_in_tray = tray_info[:4]
                if current_slot < max_slots_in_tray:
                    # 找到未满的料盘, 分配下一个坑位
                    slot_name = _get_slot_name(current_slot, cols, rows)
                    tray_info[2] += 1  # 更新当前坑位索引
                    logger.info(f"使用现有料盘 {position}, 坑位 {slot_name}")
                    return position, shelf_position, slot_name

            # 没有未满的料盘, 需要分配新位置(按索引遍历)
            new_idx = None
            for i in range(len(position_list)):
                if i not in allocated_position_indices:
                    new_idx = i
                    break

            if new_idx is None:
                raise ValueError("可用位置已用完, 无法分配新料盘")

            allocated_position_indices.add(new_idx)
            new_position = position_list[new_idx]
            new_shelf_position = shelf_position_list[new_idx]

            # 创建新料盘记录
            slot_name = _get_slot_name(0, cols, rows)
            tray_usage[tray_type_code].append(
                [new_position, new_shelf_position, 1, max_slots, new_idx]
            )
            logger.info(f"分配新料盘 {new_position}(shelf {new_shelf_position}), 类型 {tray_type_name}")
            return new_position, new_shelf_position, slot_name

        def _normalize_consumable_name(raw_name: str) -> str:
            """
            功能:
                将耗材名称归一化, 用于在统一别名表中稳定匹配.
            参数:
                raw_name: str, 原始耗材名称.
            返回:
                str, 归一化后的耗材名称.
            """
            normalized_name = str(raw_name).strip().lower()
            normalized_name = normalized_name.replace("μ", "u").replace("µ", "u")
            normalized_name = normalized_name.replace(" ", "")
            normalized_name = normalized_name.replace("-", "")
            normalized_name = normalized_name.replace("_", "")
            return normalized_name

        # 5. 解析missing列表并生成上料数据
        # 使用字典来按位置分组物资: {position: {"tray_type": ..., "contents": [...], "shelf_position": ..., "storages": [...]}}
        position_groups = {}

        # 分类存储: 先处理固体、液体、耗材，按顺序生成
        solid_items = []
        liquid_items = []
        consumable_items = []

        # 第一遍: 分类
        for missing_item in missing_list:
            # 解析格式: "物质名:数量单位"
            if ":" not in missing_item:
                logger.warning(f"跳过格式错误的缺失项: {missing_item}")
                continue

            substance, amount_str = missing_item.split(":", 1)
            substance = substance.strip()
            amount_str = amount_str.strip()

            # 检查是否为耗材(以"件"结尾)
            is_consumable = amount_str.endswith("件")

            if is_consumable:
                consumable_items.append((substance, amount_str))
            else:
                # 获取物质信息判断固液
                chem_info = chemical_dict.get(substance)
                if not chem_info:
                    logger.warning(f"在化学品列表中未找到物质: {substance}, 跳过")
                    continue

                physical_state = chem_info["physical_state"]
                if physical_state == "solid":
                    solid_items.append((substance, amount_str, chem_info))
                elif physical_state == "liquid":
                    liquid_items.append((substance, amount_str, chem_info))
                else:
                    logger.warning(f"物质状态未知: {physical_state}, 跳过物质 {substance}")

        # 第二遍: 按顺序处理 - 固体
        for substance, amount_str, chem_info in solid_items:
            physical_state = chem_info["physical_state"]
            storage_location = chem_info["storage_location"]

            # 解析数量和单位
            amount_value = 0.0
            unit = ""

            # 提取数字和单位
            import re
            match = re.match(r"([\d.]+)\s*(\w+)", amount_str)
            if match:
                amount_value = float(match.group(1))
                unit = match.group(2).lower()
            else:
                logger.warning(f"无法解析数量: {amount_str}, 跳过")
                continue

            # 固体: 使用粉桶托盘
            tray_type_code = int(ResourceCode.POWDER_BUCKET_TRAY_30ML)
            tray_type_name = f"30 mL粉桶托盘({tray_type_code})"

            # 单位转换为mg
            if unit == "mg":
                final_amount = amount_value
            elif unit == "g":
                final_amount = amount_value * 1000
            else:
                logger.warning(f"固体物质单位不支持: {unit}, 跳过")
                continue

            # 计算上料量: 最小100mg, 大于100mg按实际需要量的两倍取整
            if final_amount <= 100:
                final_amount = 100
            else:
                final_amount = final_amount * 2
                # 向上取整到百位
                final_amount = math.ceil(final_amount / 100) * 100

            final_unit = "mg"

            # 分配坑位
            try:
                position, shelf_position, slot_name = _allocate_slot(tray_type_code, tray_type_name)
            except ValueError as e:
                logger.warning(f"无法分配坑位: {e}, 跳过物资 {substance}")
                break

            # 处理final_amount可能是int或float的情况
            if isinstance(final_amount, int):
                amount_display = final_amount
            elif isinstance(final_amount, float) and final_amount.is_integer():
                amount_display = int(final_amount)
            else:
                amount_display = final_amount

            # 生成单个坑位的内容: "坑位|物质名|数量单位"
            slot_content = f"{slot_name}|{substance}|{amount_display}{final_unit}"

            # 按位置+货架分组(同一 TB 位在不同轮次对应不同货架, 需区分)
            tray_key = f"{position}_{shelf_position}"
            if tray_key not in position_groups:
                position_groups[tray_key] = {
                    "position": position,
                    "tray_type": tray_type_name,
                    "contents": [],
                    "shelf_position": shelf_position,
                    "storages": []
                }

            position_groups[tray_key]["contents"].append(slot_content)
            position_groups[tray_key]["storages"].append(f"{substance}|{storage_location if storage_location else '未知'}")
            logger.info(f"添加固体上料项: {substance} -> {position} {slot_name}, {final_amount}{final_unit}")

        # 第三遍: 处理液体
        for substance, amount_str, chem_info in liquid_items:
            physical_state = chem_info["physical_state"]
            storage_location = chem_info["storage_location"]

            # 解析数量和单位
            amount_value = 0.0
            unit = ""

            # 提取数字和单位
            import re
            match = re.match(r"([\d.]+)\s*(\w+)", amount_str)
            if match:
                amount_value = float(match.group(1))
                unit = match.group(2).lower()
            else:
                logger.warning(f"无法解析数量: {amount_str}, 跳过")
                continue

            # 液体: 根据量选择试剂瓶托盘
            # 单位转换为mL
            if unit == "ml" or unit == "mL":
                final_amount = amount_value
            elif unit == "μl" or unit == "ul":
                final_amount = amount_value / 1000
            elif unit == "l":
                final_amount = amount_value * 1000
            else:
                logger.warning(f"液体物质单位不支持: {unit}, 跳过")
                continue

            # 最少1mL, 向上取整
            final_amount = max(1, math.ceil(final_amount))
            final_unit = "mL"

            # 根据量选择瓶子规格
            if final_amount <= 2:
                tray_type_code = int(ResourceCode.REAGENT_BOTTLE_TRAY_2ML)
                tray_type_name = f"2 mL试剂瓶托盘({tray_type_code})"
            elif final_amount <= 8:
                tray_type_code = int(ResourceCode.REAGENT_BOTTLE_TRAY_8ML)
                tray_type_name = f"8 mL试剂瓶托盘({tray_type_code})"
            elif final_amount <= 40:
                tray_type_code = int(ResourceCode.REAGENT_BOTTLE_TRAY_40ML)
                tray_type_name = f"40 mL试剂瓶托盘({tray_type_code})"
            else:
                logger.warning(f"液体量超过40mL, 不支持: {final_amount}mL, 跳过")
                continue

            # 分配坑位
            try:
                position, shelf_position, slot_name = _allocate_slot(tray_type_code, tray_type_name)
            except ValueError as e:
                logger.warning(f"无法分配坑位: {e}, 跳过物资 {substance}")
                break

            # 处理final_amount可能是int或float的情况
            if isinstance(final_amount, int):
                amount_display = final_amount
            elif isinstance(final_amount, float) and final_amount.is_integer():
                amount_display = int(final_amount)
            else:
                amount_display = final_amount

            # 生成单个坑位的内容: "坑位|物质名|数量单位"
            slot_content = f"{slot_name}|{substance}|{amount_display}{final_unit}"

            # 按位置+货架分组(同一 TB 位在不同轮次对应不同货架, 需区分)
            tray_key = f"{position}_{shelf_position}"
            if tray_key not in position_groups:
                position_groups[tray_key] = {
                    "position": position,
                    "tray_type": tray_type_name,
                    "contents": [],
                    "shelf_position": shelf_position,
                    "storages": []
                }

            position_groups[tray_key]["contents"].append(slot_content)
            position_groups[tray_key]["storages"].append(f"{substance}|{storage_location if storage_location else '未知'}")
            logger.info(f"添加液体上料项: {substance} -> {position} {slot_name}, {final_amount}{final_unit}")

        # 第四遍: 处理耗材
        unrecognized_consumable_names: List[str] = []
        for substance, amount_str in consumable_items:
            # 耗材处理: 解析数量
            import re
            match = re.match(r"([\d.]+)\s*件", amount_str)
            if not match:
                logger.warning(f"无法解析耗材数量: {amount_str}, 跳过")
                continue

            consumable_count = int(float(match.group(1)))

            normalized_substance = _normalize_consumable_name(substance)
            consumable_code = CONSUMABLE_ALIAS_TO_CODE.get(normalized_substance)
            if consumable_code is None:
                logger.error("无法识别耗材类型: %s, 请先在constants.py中注册别名", substance)
                unrecognized_consumable_names.append(substance)
                continue

            tray_type_code = CONSUMABLE_CODE_TO_TRAY_CODE.get(consumable_code)
            if tray_type_code is None:
                logger.error("耗材缺少托盘映射: %s(code=%s), 请检查constants.py", substance, consumable_code)
                unrecognized_consumable_names.append(substance)
                continue

            tray_display_name = TRAY_CODE_DISPLAY_NAME.get(tray_type_code, f"未知托盘")
            tray_type_name = f"{tray_display_name}({tray_type_code})"
            consumable_display_name = CONSUMABLE_CODE_DISPLAY_NAME.get(consumable_code, substance)

            # 获取托盘规格, 计算满盘数量
            spec = tray_spec_map.get(tray_type_code)
            if spec is None:
                logger.warning(f"未找到托盘规格: {tray_type_code}, 跳过")
                continue

            cols, rows = spec
            full_tray_capacity = cols * rows

            # REACTION_TUBE_TRAY_2ML 特殊处理: 按需求数量选择12/24/36/48规格
            if tray_type_code == int(ResourceCode.REACTION_TUBE_TRAY_2ML):
                tray_capacity = 24  # 一盘24个
                available_specs = [12, 24, 36, 48]

                # 选择最小的满足需求的规格, 超出48则取48
                chosen_spec = available_specs[-1]
                for spec_val in available_specs:
                    if consumable_count <= spec_val:
                        chosen_spec = spec_val
                        break

                # 按每盘24个拆分到多个托盘
                remaining = chosen_spec
                tray_allocations = []
                while remaining > 0:
                    tray_amount = min(remaining, tray_capacity)
                    tray_allocations.append(tray_amount)
                    remaining -= tray_amount

                # 为每个托盘分配独立位置
                for tray_amount in tray_allocations:
                    new_idx = None
                    for i in range(len(position_list)):
                        if i not in allocated_position_indices:
                            new_idx = i
                            break

                    if new_idx is None:
                        logger.warning("可用位置已用完, 无法分配新料盘, 跳过耗材 %s", substance)
                        break

                    allocated_position_indices.add(new_idx)
                    position = position_list[new_idx]
                    shelf_position = shelf_position_list[new_idx]

                    # 记录到tray_usage, 标记已用坑位数以避免被复用
                    if tray_type_code not in tray_usage:
                        tray_usage[tray_type_code] = []
                    tray_usage[tray_type_code].append(
                        [position, shelf_position, tray_amount, full_tray_capacity, new_idx]
                    )

                    slot_content = str(tray_amount)

                    tray_key = f"{position}_{shelf_position}"
                    if tray_key not in position_groups:
                        position_groups[tray_key] = {
                            "position": position,
                            "tray_type": tray_type_name,
                            "contents": [],
                            "shelf_position": shelf_position,
                            "storages": []
                        }

                    position_groups[tray_key]["contents"].append(slot_content)
                    position_groups[tray_key]["storages"].append(f"{consumable_display_name}|耗材库")
                    logger.info(
                        "添加耗材上料项(反应试管): %s -> %s, 选定规格 %s, 本盘数量 %s(需求 %s)",
                        substance, position, chosen_spec, tray_amount, consumable_count,
                    )

                continue

            # 其他耗材: 按满盘分配
            try:
                position, shelf_position, slot_name = _allocate_slot(tray_type_code, tray_type_name)
            except ValueError as e:
                logger.warning(f"无法分配坑位: {e}, 跳过耗材 {substance}")
                break

            # 耗材的content格式: 满盘数量
            slot_content = str(full_tray_capacity)

            # 按位置+货架分组(同一 TB 位在不同轮次对应不同货架, 需区分)
            tray_key = f"{position}_{shelf_position}"
            if tray_key not in position_groups:
                position_groups[tray_key] = {
                    "position": position,
                    "tray_type": tray_type_name,
                    "contents": [],
                    "shelf_position": shelf_position,
                    "storages": []
                }

            position_groups[tray_key]["contents"].append(slot_content)
            position_groups[tray_key]["storages"].append(f"{consumable_display_name}|耗材库")
            logger.info(
                "添加耗材上料项: %s -> %s, 标准名=%s, 满盘数量 %s(需求 %s)",
                substance,
                position,
                consumable_display_name,
                full_tray_capacity,
                consumable_count,
            )

        if len(unrecognized_consumable_names) > 0:
            unknown_consumables = sorted(set(unrecognized_consumable_names))
            supported_consumables = sorted(set(CONSUMABLE_CODE_DISPLAY_NAME.values()))
            raise ValidationError(
                "检测到无法识别的耗材类型: "
                + ", ".join(unknown_consumables)
                + ". 请先在constants.py的CONSUMABLE_ALIAS_TO_CODE中注册后再试. "
                + "当前支持的标准耗材: "
                + ", ".join(supported_consumables)
            )

        # 6. 合并同一位置的内容并生成最终数据
        # 保持首次分配顺序, 避免第二轮复用的 TB 位被字符串排序插回第一轮.
        batch_in_data = []
        for group in position_groups.values():
            # 用分号连接同一料盘的所有坑位
            combined_content = ";".join(group["contents"])
            combined_storage = ";".join(group["storages"])
            batch_in_data.append({
                "position": group["position"],
                "tray_type": group["tray_type"],
                "content": combined_content,
                "shelf_position": group["shelf_position"],
                "storage": combined_storage
            })

        # 7. 写入batch_in_tray.xlsx
        if not batch_in_data:
            logger.warning("没有生成任何上料数据")
            return

        batch_in_path = MODULE_ROOT / "sheet" / "batch_in_tray.xlsx"

        # 检查模板是否存在，不存在则生成
        if not batch_in_path.exists():
            logger.info(f"未找到上料模板，正在生成: {batch_in_path}")
            self._generate_batch_in_tray_template(batch_in_path)

        # 读取现有模板并选择目标sheet
        wb = load_workbook(batch_in_path)
        try:
            try:
                ws, _, _ = self._select_batch_in_sheet(wb)
            except ValueError:
                if "batch_in_tray" in wb.sheetnames:
                    ws = wb["batch_in_tray"]
                else:
                    ws = wb.create_sheet("batch_in_tray")
                ws.cell(row=1, column=1, value="position")
                ws.cell(row=1, column=2, value="tray_type")
                ws.cell(row=1, column=3, value="content")
                ws.cell(row=1, column=4, value="shelf_position")
                ws.cell(row=1, column=5, value="storage")

            # 清除现有数据（保留表头，从第2行开始清除）
            max_row = ws.max_row
            if max_row > 1:
                ws.delete_rows(2, max_row - 1)

            # 写入新数据（从第2行开始）
            for idx, item in enumerate(batch_in_data, start=2):
                ws.cell(row=idx, column=1, value=item["position"])
                ws.cell(row=idx, column=2, value=item["tray_type"])
                ws.cell(row=idx, column=3, value=item["content"])
                ws.cell(row=idx, column=4, value=item["shelf_position"])
                ws.cell(row=idx, column=5, value=item["storage"])

            # 保存文件
            safe_workbook_save(wb, batch_in_path)
        finally:
            wb.close()

        logger.info(f"已生成上料文件: {batch_in_path}, 共{len(batch_in_data)}行, 包含{sum(len(g['contents']) for g in position_groups.values())}个物资")
        logger.info(f"请检查文件并根据需要调整")

        # 自动打印试剂名称标签
        try:
            self.print_reagent_labels()
        except Exception as exc:
            logger.warning("自动打印试剂标签失败(不影响上料文件生成): %s", exc)

    # ---------- 4.5 标签打印 ----------

    def _get_label_printer(self):
        """
        功能:
            延迟创建并返回标签打印服务实例.
        返回:
            LabelPrintService 实例.
        """
        if not hasattr(self, "_label_printer") or self._label_printer is None:
            from ..printer import LabelPrintService
            config_path = MODULE_ROOT / "printer" / "25x10x2.yaml"
            self._label_printer = LabelPrintService(str(config_path))
        return self._label_printer

    def print_reagent_labels(self) -> None:
        """
        功能:
            从上料文件(batch_in_tray.xlsx)中提取试剂名称, 打印试剂标签.
            只打印试剂(固体/液体), 不打印耗材.
            通过判断content字段是否包含"|"来区分试剂和耗材.
        """
        batch_in_path = MODULE_ROOT / "sheet" / "batch_in_tray.xlsx"
        if not batch_in_path.exists():
            raise FileNotFoundError(f"未找到上料文件: {batch_in_path}")

        wb = load_workbook(batch_in_path, read_only=True)
        try:
            # 定位 sheet
            ws = None
            for name in wb.sheetnames:
                if "batch_in_tray" in name.lower():
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb.active

            # 解析表头, 找到 content 列
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "content" not in headers:
                raise ValueError("上料文件缺少 content 列")
            content_col_idx = headers.index("content")

            # 提取试剂名称 (含"|"的条目为试剂行, 第二个字段为试剂名)
            reagent_names = []
            seen = set()
            for row in ws.iter_rows(min_row=2):
                cell_value = row[content_col_idx].value
                if cell_value is None:
                    continue
                content_str = str(cell_value).strip()
                # 多个坑位用";"分隔
                for slot_entry in content_str.split(";"):
                    slot_entry = slot_entry.strip()
                    if "|" not in slot_entry:
                        continue  # 耗材行, 跳过
                    parts = slot_entry.split("|")
                    if len(parts) >= 2:
                        substance = parts[1].strip()
                        if substance and substance not in seen:
                            reagent_names.append(substance)
                            seen.add(substance)
        finally:
            wb.close()

        if not reagent_names:
            logger.info("上料文件中没有试剂条目, 无需打印标签")
            return

        logger.info("从上料文件中提取到 %d 种试剂: %s", len(reagent_names), ", ".join(reagent_names))

        # 打印标签
        printer = self._get_label_printer()
        try:
            printer.connect()
            printer.print_batch(reagent_names)
        finally:
            printer.disconnect()

        logger.info("试剂标签打印完成")

    def print_task_number_labels(self, task_id: int) -> None:
        """
        功能:
            根据实验方案文件打印两组编号标签:
            - 反应管标签: R{task_id}-1, R{task_id}-2, ..., R{task_id}-N
            - 检测样品标签: S{task_id}-1, S{task_id}-2, ..., S{task_id}-M
            N 由实验编号最大值决定, M 由闪滤实验编号决定(空=全部).
        参数:
            task_id: int, 任务ID.
        """
        # 1. 查找实验方案文件
        plan_path = MODULE_ROOT / "data" / "tasks" / str(task_id) / f"{task_id}_experiment_plan.xlsx"
        if not plan_path.exists():
            raise FileNotFoundError(f"未找到实验方案文件: {plan_path}")

        wb = load_workbook(plan_path, read_only=True)
        try:
            # 定位实验方案 sheet (第一个 sheet, 不用 wb.active 因为活动页可能是其他 sheet)
            ws = wb.worksheets[0]

            # 2. 读取实验编号 (列C, 从第2行开始)
            experiment_numbers = []
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                val = row[0].value
                if val is not None:
                    try:
                        # 兼容 int/float/str 混合类型
                        num = int(float(str(val)))
                        experiment_numbers.append(num)
                    except (ValueError, TypeError):
                        pass

            if not experiment_numbers:
                raise ValueError(f"任务 {task_id} 的实验方案文件中未找到实验编号")

            max_exp_num = max(experiment_numbers)
            logger.info("任务 %d: 实验编号 1~%d", task_id, max_exp_num)

            # 3. 读取闪滤实验编号 (行29, 列B)
            flash_filter_cell = ws.cell(row=29, column=2).value
            if flash_filter_cell is None or str(flash_filter_cell).strip() == "":
                # 空值: 所有实验都需要闪滤 -> 样品编号 = 全部实验编号
                sample_numbers = list(range(1, max_exp_num + 1))
                logger.info("闪滤实验编号为空, 全部 %d 个实验需要样品标签", max_exp_num)
            else:
                # 解析闪滤编号 (支持 "1,3,5" 或 "1-6" 或混合 "1-3,5,7-9")
                sample_numbers = self._parse_number_range(str(flash_filter_cell).strip())
                logger.info("闪滤实验编号: %s -> 样品标签 %d 张", flash_filter_cell, len(sample_numbers))
        finally:
            wb.close()

        # 4. 生成标签内容
        reaction_labels = [f"R{task_id}-{i}" for i in range(1, max_exp_num + 1)]
        sample_labels = [f"S{task_id}-{i}" for i in sample_numbers]

        printer = self._get_label_printer()
        try:
            printer.connect()

            # 5. 一次确认后连续打印两组标签
            print(f"\n即将打印标签:")
            print(f"  反应管: {len(reaction_labels)} 张 ({reaction_labels[0]} ~ {reaction_labels[-1]})")
            print(f"  检测样品: {len(sample_labels)} 张 ({sample_labels[0]} ~ {sample_labels[-1]})")
            input("请确认打印机就绪, 按回车开始打印...")

            logger.info("正在打印反应管标签: %d 张", len(reaction_labels))
            printer.print_batch(reaction_labels)
            logger.info("反应管标签打印完成")

            logger.info("正在打印检测样品标签: %d 张", len(sample_labels))
            printer.print_batch(sample_labels)
            logger.info("检测样品标签打印完成")
        finally:
            printer.disconnect()

    @staticmethod
    def _parse_number_range(range_str: str) -> List[int]:
        """
        功能:
            解析数字范围字符串, 支持逗号分隔和连字符范围.
            例: "1,3,5" -> [1,3,5], "1-6" -> [1,2,3,4,5,6], "1-3,5,7-9" -> [1,2,3,5,7,8,9]
        参数:
            range_str: str, 数字范围字符串.
        返回:
            List[int], 排序后的编号列表.
        """
        import re
        numbers = set()
        # 按逗号或中文逗号分隔
        parts = re.split(r"[,，]", range_str)
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # 检查是否包含范围连字符
            range_match = re.match(r"(\d+)\s*[-~]\s*(\d+)", part)
            if range_match:
                start = int(range_match.group(1))
                end = int(range_match.group(2))
                numbers.update(range(start, end + 1))
            else:
                try:
                    numbers.add(int(part))
                except ValueError:
                    pass
        return sorted(numbers)

    # ---------- 5. 执行任务 ----------
    def device_init(self, device_id=None, *, poll_interval_s: float = 1.0, timeout_s: float = 600.0):
        return super().device_init(device_id, poll_interval_s=poll_interval_s, timeout_s=timeout_s)

    def start_task(self, task_id: int | None = None, *, check_glovebox_env: bool = True, water_limit_ppm: float = 10.0, oxygen_limit_ppm: float = 10.0):
        return super().start_task(task_id, check_glovebox_env=check_glovebox_env, water_limit_ppm=water_limit_ppm, oxygen_limit_ppm=oxygen_limit_ppm)

    def wait_task_with_ops(self, task_id: int | None = None, *, poll_interval_s: float = 2.0) -> int:
        return super().wait_task_with_ops(task_id, poll_interval_s=poll_interval_s)

    def export_task_report(self, task_id: int, file_type: str = "excel") -> Path:
        return super().export_task_report(task_id, file_type=file_type)

    # ---------- 6. 下料动作 ----------
    def batch_out_task_and_empty_trays(self, task_id: int | None = None, *, poll_interval_s: float = 1.0, ignore_missing: bool = True, timeout_s: float = 900.0, move_type: str = "main_out"):
        return super().batch_out_task_and_empty_trays(task_id, poll_interval_s=poll_interval_s, ignore_missing=ignore_missing, timeout_s=timeout_s, move_type=move_type)

    def batch_out_task_and_chemical_trays(self, task_id: int | None = None, *, poll_interval_s: float = 1.0, ignore_missing: bool = True, timeout_s: float = 900.0, move_type: str = "main_out"):
        return super().batch_out_task_and_chemical_trays(task_id, poll_interval_s=poll_interval_s, ignore_missing=ignore_missing, timeout_s=timeout_s, move_type=move_type)

    def batch_out_task_trays(self, task_id: int | None = None, *, poll_interval_s: float = 1.0, ignore_missing: bool = True, timeout_s: float = 900.0, move_type: str = "main_out"):
        return super().batch_out_task_trays(task_id, poll_interval_s=poll_interval_s, ignore_missing=ignore_missing, timeout_s=timeout_s, move_type=move_type)

    def batch_out_empty_trays(self, *, poll_interval_s: float = 1.0, ignore_missing: bool = True, timeout_s: float = 900.0, move_type: str = "main_out"):
        return super().batch_out_empty_trays(poll_interval_s=poll_interval_s, ignore_missing=ignore_missing, timeout_s=timeout_s, move_type=move_type)

    def batch_out_tray(self, layout_list: list[dict], move_type: str = "main_out", *, task_id: int = None, poll_interval_s: float = 1.0, timeout_s: float = 900.0):
        return super().batch_out_tray(layout_list, move_type=move_type, task_id=task_id, poll_interval_s=poll_interval_s, timeout_s=timeout_s)

    def auto_unload_trays_to_agv(self, batch_out_file: Optional[str] = None, *, block: bool = True, auto_run_analysis: bool = True):
        return super().auto_unload_trays_to_agv(batch_out_file, block=block, auto_run_analysis=auto_run_analysis)

    # ---------- 7. Unilab 接口（待修改） ----------
    def submit_experiment_task(
        self,
        chemical_db_path: str,
        task_name: str = "Unilab_Auto_Job",
        reaction_type: str = "heat",
        duration: str = "8",
        temperature: str = "40",
        stir_speed: str = "500",
        target_temp: str = "30",
        auto_magnet: bool = True,
        fixed_order: bool = False,
        internal_std_name: str = "",
        stir_time_after_std: str = "",
        diluent_name: str = "",
        rows: list = None
    ) -> JsonDict:
        """
        功能:
            提交 Unilab 流程编排任务, 按行数据动态生成表头, 兼容包含“加磁子”的列.
        参数:
            chemical_db_path: str, 化学品库文件路径.
            task_name: str, 任务名称.
            reaction_type: str, 反应类型.
            duration: str, 反应时间, 必须带单位, 如 "8h" 或 "30min".
            temperature: str, 反应温度(°C).
            stir_speed: str, 搅拌速度(rpm).
            target_temp: str, 搅拌后目标温度(°C).
            auto_magnet: bool, 是否自动加磁子.
            fixed_order: bool, 是否固定加料顺序.
            internal_std_name: str, 内标名称.
            stir_time_after_std: str, 内标加入后搅拌时间(min).
            diluent_name: str, 稀释液名称.
            rows: List[List[Any]], 行数据矩阵, 第1列为实验编号, 其余列为试剂或“加磁子”.
        返回:
            Dict[str, Any], 提交成功后返回的任务 ID.
        """
        c_path = Path(chemical_db_path)
        if c_path.exists() is False:
            raise FileNotFoundError(f"未找到化学品库文件: {c_path}")

        chem_df = self._read_table_file_with_required_columns(
            c_path,
            required_columns=["substance"],
        )
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row, *keys, default=None):
            for k in keys:
                if k in row and pd.notna(row[k]):
                    return row[k]
            return default

        chemical_db: Dict[str, Dict[str, Any]] = {}
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if name == "":
                continue
            chemical_db[name] = {
                "chemical_id": _pick(row, "chemical_id"),
                "molecular_weight": _pick(row, "molecular_weight", "mw"),
                "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                "density (g/mL)": _pick(row, "density (g/ml)", "density(g/ml)", "density_g_ml", "density", default=None),
                "physical_form": str(_pick(row, "physical_form", default="") or "").strip().lower(),
                "active_content": _pick(row, "active_content", "active_content(mmol/ml or wt%)", "active_content(mol/l or wt%)", default=""),
            }

        if auto_magnet is True:
            auto_magnet_text = "是"
        else:
            auto_magnet_text = "否"

        if fixed_order is True:
            fixed_order_text = "是"
        else:
            fixed_order_text = "否"

        params = {
            "实验名称": task_name,
            "反应器类型": reaction_type,
            "反应时间(min/h)": duration,
            "反应温度(°C)": temperature,
            "转速(rpm)": stir_speed,
            "搅拌后目标温度(°C)": target_temp,
            "自动加磁子": auto_magnet_text,
            "固定加料顺序": fixed_order_text,
            "内标种类": internal_std_name,
            "加入内标后搅拌时间(min)": stir_time_after_std,
            "稀释液种类": diluent_name,
        }

        if rows is None:
            rows = []

        default_pair_count = 5
        cleaned_rows: List[List[Any]] = []
        magnet_columns: set[int] = set()
        max_col_count = 1

        for row in rows:
            if isinstance(row, (list, tuple)) is False:
                logger.warning("行数据格式需要列表或元组, 已跳过一行")
                continue
            row_values = list(row)
            while len(row_values) > 0:
                tail_text = "" if row_values[-1] is None else str(row_values[-1]).strip()
                if tail_text == "":
                    row_values.pop()
                    continue
                break
            if len(row_values) > max_col_count:
                max_col_count = len(row_values)
            for col_index, cell in enumerate(row_values):
                cell_text = "" if cell is None else str(cell).strip()
                if "加磁子" in cell_text:
                    magnet_columns.add(col_index)
            cleaned_rows.append(row_values)

        if len(cleaned_rows) == 0:
            header_count = 1 + default_pair_count * 2
        else:
            header_count = max_col_count

        headers: List[str] = ["实验编号"]
        for col_index in range(1, header_count):
            if col_index in magnet_columns:
                headers.append("加磁子")
            elif col_index % 2 == 1:
                headers.append("试剂")
            else:
                headers.append("试剂量")

        normalized_rows: List[List[Any]] = []
        for row_values in cleaned_rows:
            padded_values = row_values + [""] * (header_count - len(row_values))
            normalized_rows.append(padded_values)

        try:
            task_payload = self.build_task_payload(params, headers, normalized_rows, chemical_db)
        except AttributeError as exc:
            raise Exception("无法找到 build_task_payload 方法, 请检查 StationController 定义") from exc

        try:
            resp = self._submit_task_payload(task_payload)
        except ApiError as exc:
            if getattr(exc, "code", None) == 409:
                # 自动重命名: 在任务名称后添加当前日期时间(精确到秒)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                task_name_val = task_payload.get("task_name") or params.get("实验名称")
                new_task_name = f"{task_name_val}_{timestamp}"

                task_payload["task_name"] = new_task_name
                logger.info(f"任务名称重复, 自动重命名为: {new_task_name}")

                # 重试提交
                try:
                    resp = self._submit_task_payload(task_payload)
                except ApiError as retry_exc:
                    logger.error(f"重命名后任务提交仍失败: {retry_exc}")
                    raise
            else:
                raise

        task_id = resp.get("task_id")
        return task_id

    # ---------- 分析站对接 ----------
    def run_analysis(self, task_id: Optional[str] = None) -> Dict:
        """
        功能:
            读取指定合成任务(或最新任务)的 xlsx 配置, 自动生成分析任务 CSV
            并通过 AnalysisStationController 提交至对应仪器(当前已实现 GC_MS).
        参数:
            task_id: 合成任务 ID 字符串, 为 None 时自动选取编号最大的最近任务.
        返回:
            Dict: 各仪器提交结果, 格式示例:
                {
                    "gc_ms":      {"success": bool, "return_info": str},
                    "uplc_qtof":  {"success": bool, "return_info": str},
                    "hplc":       {"success": bool, "return_info": str},
                }
        """
        # 延迟绝对导入，避免模块级相对导入越界问题
        # 运行目录为 devices/，eit_analysis_station 可直接作为顶层包访问
        from unilabos.devices.eit_analysis_station.controller.analysis_controller import AnalysisStationController

        # 创建分析站控制器实例，使用其默认配置
        analysis_ctrl = AnalysisStationController()

        logger.info("启动分析任务提交流程, task_id=%s", task_id)
        results = analysis_ctrl.run_analysis(task_id=task_id)
        return results

    def poll_analysis_run(
        self, task_id: Optional[str] = None, poll_interval: float = 30.0
    ) -> Dict:
        """
        功能:
            轮询 GC-MS 分析任务运行状态, 完成后自动触发结果处理(积分+定性+报告).
            内部委托给 AnalysisStationController.poll_analysis_run 执行.
        参数:
            task_id: 合成任务 ID 字符串, 为 None 时自动选取编号最大的最近任务.
            poll_interval: 轮询间隔(秒), 默认 30 秒.
        返回:
            Dict: process_gc_ms_results 的返回值, 包含 success/return_info/report_path.
        """
        from unilabos.devices.eit_analysis_station.controller.analysis_controller import AnalysisStationController

        analysis_ctrl = AnalysisStationController()

        logger.info("启动 GC-MS 轮询流程, task_id=%s, poll_interval=%.0fs", task_id, poll_interval)
        result = analysis_ctrl.poll_analysis_run(task_id=task_id, poll_interval=poll_interval)
        return result
