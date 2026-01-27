# -*- coding: utf-8 -*-
"""
功能:
    调用 DeepSeek API, 通过自然语言一次性生成多组实验行, 并按 reaction_template_3.xlsx 的格式输出 Excel.
    同时读取 chemical_list.xlsx 作为化学品库, 校验所有物质名称是否存在于库的 substance 列, 并输出 storage_location.

协议:
    1) 用户用自然语言描述实验, 可以包含多组条件或多组平行实验.
    2) 脚本调用大模型抽取结构化 JSON:
        - global_params: 左侧参数区设置(默认值与用户输入结合).
        - experiments: 多个实验行, 每行包含 reagents 列表.
    3) 总实验行数必须为 12/24/36/48 之一, 否则提示用户删除或补充并要求模型重写.
    4) 化学品库校验: 所有名称(含稀释液, 闪滤液, 内标等)必须命中 substance 列, 不命中则给相似候选并要求修正.
    5) 写入模板:
        - 单 sheet 最多 24 行, 36/48 行会拆分到 Sheet2.
    6) 输出本次所有实验涉及化学品的 storage_location.
    7) 生成文件后提醒用户检查.

参数:
    无(脚本内通过常量配置路径, 通过环境变量读取 DEEPSEEK_API_KEY).

返回:
    无(在本地输出 Excel 文件, 并在终端输出提示与 storage_location).

依赖:
    pip install requests openpyxl pandas
"""

from __future__ import annotations

import json
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# 日志配置
# -----------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("reaction_plan_generator")


# -----------------------------
# 路径配置(按你的上传文件路径)
# -----------------------------
TEMPLATE_XLSX_PATH = Path(r"D:\Uni-Lab-OS\unilabos\devices\workstation\eit_synthesis_station\reaction_template_3.xlsx")
CHEMICAL_LIBRARY_XLSX_PATH = Path(r"D:\Uni-Lab-OS\unilabos\devices\workstation\eit_synthesis_station\chemical_list.xlsx")


# -----------------------------
# DeepSeek API 配置
# -----------------------------
DEEPSEEK_BASE_URL = "https://api.deepseek.com/v1"
DEEPSEEK_CHAT_COMPLETIONS = f"{DEEPSEEK_BASE_URL}/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"


# -----------------------------
# 数据结构
# -----------------------------
@dataclass
class ReagentItem:
    name: str
    amount_text: str  # 支持 eq, mmol, g, mg, μL, mL


@dataclass
class ExperimentRow:
    global_experiment_no: int  # 1..48
    reagents: List[ReagentItem]


@dataclass
class ReactionPlan:
    # 左侧参数区
    experiment_name: str
    experiment_id: int
    scale_mmol: float
    reactor_type: str  # "light" or "heat"
    time_h: float
    temperature_c: float
    rpm: int
    open_lid_target_temp_c: float
    wait_target_temp: str  # "是" or "否"
    weigh_error_percent: float
    max_weigh_error_mg: float
    fixed_addition_order: str  # "是" or "否"
    auto_add_stir_bar: str  # "是" or "否"

    dilution_solvent: str
    dilution_volume_ul: Optional[float]

    internal_standard_enabled: str  # "是" or "否"
    internal_standard_name: str
    internal_standard_amount_ul_or_mg: Optional[float]
    stir_after_is_min: float

    flash_filter_enabled: str  # "是" or "否"
    flash_filter_solvent: str
    flash_filter_volume_ul: Optional[float]
    sample_volume_ul: float

    # 右侧实验行
    experiments: List[ExperimentRow]


# -----------------------------
# 工具函数
# -----------------------------
def _today_yyyymmdd() -> str:
    """
    功能:
        获取今天日期字符串, 用于默认实验名称.
    参数:
        无.
    返回:
        str, 形如 "2026-01-15".
    """
    return datetime.now().strftime("%Y-%m-%d")


def _extract_json(text: str) -> Dict[str, Any]:
    """
    功能:
        从模型输出中提取 JSON 对象.
    参数:
        text: str, 模型原始输出.
    返回:
        Dict[str, Any], 解析后的 JSON.
    """
    text = text.strip()
    if text.startswith("```") is True:
        text = re.sub(r"^```[a-zA-Z]*\n", "", text)
        text = re.sub(r"\n```$", "", text)

    start_index = text.find("{")
    end_index = text.rfind("}")
    if start_index == -1 or end_index == -1 or end_index <= start_index:
        raise ValueError("未在模型输出中找到 JSON 对象.")

    json_str = text[start_index : end_index + 1]
    return json.loads(json_str)


def _normalize_reactor_type(user_value: str) -> str:
    """
    功能:
        规范化反应器类型到 light/heat.
    参数:
        user_value: str, 输入.
    返回:
        str, "light" or "heat".
    """
    value = (user_value or "").strip().lower()
    if value in ["light", "光照", "照光", "photoredox", "photo"]:
        return "light"
    return "heat"


def _minutes_to_hours(value: float) -> float:
    """
    功能:
        分钟转换为小时.
    参数:
        value: float, 分钟.
    返回:
        float, 小时.
    """
    return value / 60.0


def _safe_float(value: Any) -> Optional[float]:
    """
    功能:
        将输入转为 float, 失败返回 None.
    参数:
        value: Any, 输入.
    返回:
        Optional[float], 成功为 float, 失败为 None.
    """
    try:
        if value is None:
            return None
        return float(value)
    except Exception:
        return None


def _safe_int(value: Any) -> Optional[int]:
    """
    功能:
        将输入转为 int, 失败返回 None.
    参数:
        value: Any, 输入.
    返回:
        Optional[int], 成功为 int, 失败为 None.
    """
    try:
        if value is None:
            return None
        return int(value)
    except Exception:
        return None


def _allowed_experiment_counts() -> List[int]:
    """
    功能:
        返回允许的实验行数规格.
    参数:
        无.
    返回:
        List[int], [12, 24, 36, 48].
    """
    return [12, 24, 36, 48]


# -----------------------------
# 化学品库
# -----------------------------
class ChemicalLibrary:
    """
    功能:
        读取 chemical_list.xlsx, 用 substance 列做名称校验,
        并提供 storage_location 查询与相似名称提示.
    参数:
        xlsx_path: Path, 化学品库路径.
    返回:
        ChemicalLibrary 对象.
    """

    def __init__(self, xlsx_path: Path) -> None:
        self.xlsx_path = xlsx_path
        self.df = pd.read_excel(xlsx_path)
        if "substance" not in self.df.columns:
            raise ValueError("chemical_list.xlsx 缺少 substance 列.")
        if "storage_location" not in self.df.columns:
            raise ValueError("chemical_list.xlsx 缺少 storage_location 列.")

        self.substance_list = [str(x).strip() for x in self.df["substance"].dropna().tolist()]
        self.substance_set = set(self.substance_list)

    def exists(self, name: str) -> bool:
        """
        功能:
            判断物质名称是否存在于库的 substance 列.
        参数:
            name: str, 物质名称.
        返回:
            bool, 是否存在.
        """
        if name is None:
            return False
        return str(name).strip() in self.substance_set

    def suggest(self, name: str, top_k: int = 5) -> List[str]:
        """
        功能:
            给出库中相似名称候选.
        参数:
            name: str, 输入名称.
            top_k: int, 返回候选数量.
        返回:
            List[str], 候选名称列表.
        """
        if name is None:
            return []
        query = str(name).strip()
        candidates = get_close_matches(query, self.substance_list, n=top_k, cutoff=0.72)
        return candidates

    def storage_location(self, name: str) -> Optional[str]:
        """
        功能:
            查询物质 storage_location.
        参数:
            name: str, 物质名称(需在库中).
        返回:
            Optional[str], 找到则返回位置, 否则 None.
        """
        if name is None:
            return None
        query = str(name).strip()
        matched = self.df[self.df["substance"].astype(str).str.strip() == query]
        if matched.empty is True:
            return None
        value = matched.iloc[0]["storage_location"]
        if pd.isna(value) is True:
            return None
        return str(value).strip()


# -----------------------------
# DeepSeek 客户端
# -----------------------------
class DeepSeekClient:
    """
    功能:
        调用 DeepSeek Chat Completions API, 获取模型输出.
    参数:
        api_key: str, DeepSeek API Key.
    返回:
        DeepSeekClient 对象.
    """

    def __init__(self, api_key: str) -> None:
        self.api_key = api_key

    def chat(self, messages: List[Dict[str, str]], temperature: float = 0.2) -> str:
        """
        功能:
            调用 DeepSeek Chat Completions API.
        参数:
            messages: List[Dict[str, str]], messages.
            temperature: float, 生成温度.
        返回:
            str, assistant content.
        """
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": DEEPSEEK_MODEL,
            "messages": messages,
            "temperature": temperature,
        }

        response = requests.post(DEEPSEEK_CHAT_COMPLETIONS, headers=headers, json=payload, timeout=120)
        if response.status_code != 200:
            raise RuntimeError(f"DeepSeek API 调用失败, status_code={response.status_code}, body={response.text}")

        data = response.json()
        content = data["choices"][0]["message"]["content"]
        return content


# -----------------------------
# 模板写入
# -----------------------------
class ReactionTemplateWriter:
    """
    功能:
        读取 reaction_template_3.xlsx 模板, 将 ReactionPlan 写入, 支持 12/24/36/48 行自动拆分到多个 sheet.
    参数:
        template_path: Path, 模板文件路径.
    返回:
        ReactionTemplateWriter 对象.
    """

    def __init__(self, template_path: Path) -> None:
        self.template_path = template_path

    def _build_param_row_map(self, ws: Worksheet) -> Dict[str, int]:
        """
        功能:
            通过 A 列文本构建参数名到行号映射, 避免写死行号.
        参数:
            ws: Worksheet, 工作表.
        返回:
            Dict[str, int], 参数名 -> 行号.
        """
        param_row_map: Dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            if key is None:
                continue
            key_text = str(key).strip()
            if key_text != "":
                param_row_map[key_text] = row
        return param_row_map

    def _set_param(self, ws: Worksheet, param_row_map: Dict[str, int], param_name: str, value: Any) -> None:
        """
        功能:
            向模板左侧参数区写入一个参数值.
        参数:
            ws: Worksheet, 工作表.
            param_row_map: Dict[str, int], 参数行映射.
            param_name: str, 参数名.
            value: Any, 参数值.
        返回:
            None.
        """
        if param_name not in param_row_map:
            logger.warning("模板中未找到参数行: %s, 已跳过写入.", param_name)
            return
        r = param_row_map[param_name]
        ws.cell(row=r, column=2, value=value)

    def _detect_reagent_pairs(self, ws: Worksheet) -> List[Tuple[int, int]]:
        """
        功能:
            按表头自动识别“试剂/试剂量”列对.
        参数:
            ws: Worksheet, 工作表.
        返回:
            List[Tuple[int, int]], [(试剂列, 试剂量列), ...].
        """
        reagent_pairs: List[Tuple[int, int]] = []
        col = 4
        while col <= ws.max_column:
            head_name = ws.cell(row=1, column=col).value
            head_amount = ws.cell(row=1, column=col + 1).value if col + 1 <= ws.max_column else None
            if str(head_name).strip() == "试剂" and str(head_amount).strip() == "试剂量":
                reagent_pairs.append((col, col + 1))
                col = col + 2
                continue
            col = col + 1
        return reagent_pairs

    def _find_row_by_local_experiment_no(self, ws: Worksheet, local_no: int) -> Optional[int]:
        """
        功能:
            在模板中通过 C 列实验编号找到对应行号.
        参数:
            ws: Worksheet, 工作表.
            local_no: int, 1..24.
        返回:
            Optional[int], 找到则返回行号, 否则 None.
        """
        for r in range(2, 26):
            v = ws.cell(row=r, column=3).value
            if v is None:
                continue
            try:
                if int(v) == int(local_no):
                    return r
            except Exception:
                continue
        return None

    def write(self, plan: ReactionPlan, output_path: Path) -> None:
        """
        功能:
            将实验方案写入 Excel 模板并保存, 支持 12/24/36/48 行拆分.
        参数:
            plan: ReactionPlan, 实验方案.
            output_path: Path, 输出文件路径.
        返回:
            None.
        """
        wb = load_workbook(self.template_path)
        ws1 = wb.active

        # 需要的 sheet 数量
        total_rows = len(plan.experiments)
        sheet_count = 1
        if total_rows > 24:
            sheet_count = 2

        # 若需要第二张表, 拷贝模板 sheet, 保留格式
        sheets: List[Worksheet] = [ws1]
        if sheet_count == 2:
            ws2 = wb.copy_worksheet(ws1)
            ws2.title = "Sheet2"
            sheets.append(ws2)

        # 将左侧参数写入每个 sheet, 让每张表的参数一致
        for ws in sheets:
            param_row_map = self._build_param_row_map(ws)

            self._set_param(ws, param_row_map, "实验名称", plan.experiment_name)
            self._set_param(ws, param_row_map, "实验ID", plan.experiment_id)
            self._set_param(ws, param_row_map, "反应规模(mmol)", plan.scale_mmol)
            self._set_param(ws, param_row_map, "反应器类型", plan.reactor_type)
            self._set_param(ws, param_row_map, "反应时间(h)", plan.time_h)
            self._set_param(ws, param_row_map, "反应温度(°C)", plan.temperature_c)
            self._set_param(ws, param_row_map, "转速(rpm)", plan.rpm)
            self._set_param(ws, param_row_map, "搅拌后⽬标温度(°C)", plan.open_lid_target_temp_c)
            self._set_param(ws, param_row_map, "等待目标温度", plan.wait_target_temp)
            self._set_param(ws, param_row_map, "称量误差(%)", plan.weigh_error_percent)
            self._set_param(ws, param_row_map, "最大称量误差(mg)", plan.max_weigh_error_mg)
            self._set_param(ws, param_row_map, "固定加料顺序", plan.fixed_addition_order)
            self._set_param(ws, param_row_map, "自动加磁子", plan.auto_add_stir_bar)

            if plan.dilution_solvent.strip() != "":
                self._set_param(ws, param_row_map, "稀释液种类", plan.dilution_solvent)
                self._set_param(ws, param_row_map, "稀释量(μL)", plan.dilution_volume_ul)
            else:
                self._set_param(ws, param_row_map, "稀释液种类", "")
                self._set_param(ws, param_row_map, "稀释量(μL)", "")

            if plan.internal_standard_enabled == "是":
                self._set_param(ws, param_row_map, "内标种类", plan.internal_standard_name)
                self._set_param(ws, param_row_map, "内标用量(μL/mg)", plan.internal_standard_amount_ul_or_mg)
                self._set_param(ws, param_row_map, "加入内标后搅拌时间(min)", plan.stir_after_is_min)
            else:
                self._set_param(ws, param_row_map, "内标种类", "")
                self._set_param(ws, param_row_map, "内标用量(μL/mg)", "")
                self._set_param(ws, param_row_map, "加入内标后搅拌时间(min)", "")

            if plan.flash_filter_enabled == "是":
                self._set_param(ws, param_row_map, "闪滤液种类", plan.flash_filter_solvent)
                self._set_param(ws, param_row_map, "闪滤液用量(μL)", plan.flash_filter_volume_ul)
                self._set_param(ws, param_row_map, "取样量(μL)", plan.sample_volume_ul)
            else:
                self._set_param(ws, param_row_map, "闪滤液种类", "")
                self._set_param(ws, param_row_map, "闪滤液用量(μL)", "")
                self._set_param(ws, param_row_map, "取样量(μL)", plan.sample_volume_ul)

        # 写入右侧试剂区: 每个实验行对应 global_experiment_no
        for ws in sheets:
            reagent_pairs = self._detect_reagent_pairs(ws)
            max_slots = len(reagent_pairs)

            for exp in plan.experiments:
                # 映射 sheet 与 local_no
                if exp.global_experiment_no <= 24:
                    target_ws = sheets[0]
                    local_no = exp.global_experiment_no
                else:
                    if len(sheets) < 2:
                        continue
                    target_ws = sheets[1]
                    local_no = exp.global_experiment_no - 24

                if target_ws != ws:
                    continue

                target_row = self._find_row_by_local_experiment_no(target_ws, local_no)
                if target_row is None:
                    logger.warning("未找到实验编号对应行: global_no=%s, local_no=%s.", exp.global_experiment_no, local_no)
                    continue

                for idx in range(max_slots):
                    reagent_col, amount_col = reagent_pairs[idx]
                    if idx < len(exp.reagents):
                        target_ws.cell(row=target_row, column=reagent_col, value=exp.reagents[idx].name)
                        target_ws.cell(row=target_row, column=amount_col, value=exp.reagents[idx].amount_text)
                    else:
                        target_ws.cell(row=target_row, column=reagent_col, value="")
                        target_ws.cell(row=target_row, column=amount_col, value="")

        wb.save(output_path)
        logger.info("已生成 Excel 文件: %s", str(output_path))


# -----------------------------
# 方案生成与校验逻辑
# -----------------------------
def _build_system_prompt() -> str:
    """
    功能:
        构造系统提示词, 要求模型输出严格 JSON, 且一次性给出多组实验行.
    参数:
        无.
    返回:
        str, system prompt.
    """
    return (
        "你是一个自动化化学实验助手. 你需要把用户的自然语言实验描述解析成严格 JSON, "
        "用于生成 reaction_template_3.xlsx 的实验模板. 只输出 JSON, 不要输出解释, 不要 Markdown.\n"
        "JSON 结构如下:\n"
        "{\n"
        '  "global_params": {\n'
        '    "experiment_name": "string or empty",\n'
        '    "experiment_id": 0,\n'
        '    "scale_mmol": number or null,\n'
        '    "reactor_type": "light" or "heat" or empty,\n'
        '    "time_value": number or null,\n'
        '    "time_unit": "h" or "min" or empty,\n'
        '    "temperature_c": number or null,\n'
        '    "rpm": number or null,\n'
        '    "preheat": true or false or null,\n'
        '    "weigh_error_percent": number or null,\n'
        '    "max_weigh_error_mg": number or null,\n'
        '    "fixed_addition_order": true or false or null,\n'
        '    "auto_add_stir_bar": true or false or null,\n'
        '    "need_dilution": true or false or null,\n'
        '    "dilution_solvent": "string or empty",\n'
        '    "dilution_volume_ul": number or null,\n'
        '    "internal_standard_specified": true or false or null,\n'
        '    "internal_standard_name": "string or empty",\n'
        '    "internal_standard_amount_ul_or_mg": number or null,\n'
        '    "flash_filter_specified": true or false or null,\n'
        '    "flash_filter_solvent": "string or empty",\n'
        '    "flash_filter_volume_ul": number or null,\n'
        '    "sample_volume_ul": number or null\n'
        "  },\n"
        '  "experiments": [\n'
        "     {\n"
        '       "global_experiment_no": 1,\n'
        '       "reagents": [{"name":"string","amount_text":"string(支持 eq,mmol,g,mg,μL,mL 等)"}]\n'
        "     }\n"
        "  ]\n"
        "}\n"
        "规则:\n"
        "1) reactor_type: 光照/blue LED/照光/photoredox -> light, 其余默认为 heat.\n"
        "2) time_unit 若用户用分钟请填 min.\n"
        "3) experiments 必须一次性列出多组实验行, global_experiment_no 从 1 开始递增.\n"
        "4) experiments 的总长度必须是 12 或 24 或 36 或 48, 如果用户只给了 6 组请补到 12.\n"
        "5) 不确定则填 null 或 empty, 不要编造.\n"
    )


def _plan_from_model_json(model_json: Dict[str, Any]) -> ReactionPlan:
    """
    功能:
        将模型 JSON 转为 ReactionPlan, 并应用默认值与规则修正.
    参数:
        model_json: Dict[str, Any], 模型输出 JSON.
    返回:
        ReactionPlan, 结构化实验方案.
    """
    global_params = model_json.get("global_params") or {}

    experiment_name = str(global_params.get("experiment_name") or "").strip()
    if experiment_name == "":
        experiment_name = _today_yyyymmdd()

    experiment_id = _safe_int(global_params.get("experiment_id"))
    if experiment_id is None:
        experiment_id = 0

    scale_mmol = _safe_float(global_params.get("scale_mmol"))

    reactor_type = _normalize_reactor_type(str(global_params.get("reactor_type") or ""))

    time_value = _safe_float(global_params.get("time_value"))
    time_unit = str(global_params.get("time_unit") or "").strip().lower()
    time_h: Optional[float] = None
    if time_value is not None:
        if time_unit == "min":
            time_h = _minutes_to_hours(time_value)
        else:
            time_h = time_value

    temperature_c = _safe_float(global_params.get("temperature_c"))

    rpm = _safe_int(global_params.get("rpm"))
    if rpm is None:
        rpm = 500

    open_lid_target_temp_c = 30.0
    if temperature_c is not None:
        if temperature_c <= 30:
            open_lid_target_temp_c = float(temperature_c)

    preheat = global_params.get("preheat")
    wait_target_temp = "否"
    if preheat is True:
        wait_target_temp = "是"

    weigh_error_percent = _safe_float(global_params.get("weigh_error_percent"))
    if weigh_error_percent is None:
        weigh_error_percent = 3.0

    max_weigh_error_mg = _safe_float(global_params.get("max_weigh_error_mg"))
    if max_weigh_error_mg is None:
        max_weigh_error_mg = 1.0

    fixed_addition_order_bool = global_params.get("fixed_addition_order")
    fixed_addition_order = "否"
    if fixed_addition_order_bool is True:
        fixed_addition_order = "是"

    auto_add_stir_bar_bool = global_params.get("auto_add_stir_bar")
    auto_add_stir_bar = "是"
    if auto_add_stir_bar_bool is False:
        auto_add_stir_bar = "否"

    need_dilution = global_params.get("need_dilution")
    dilution_solvent = ""
    dilution_volume_ul: Optional[float] = None
    if need_dilution is True:
        dilution_solvent = str(global_params.get("dilution_solvent") or "").strip()
        dilution_volume_ul = _safe_float(global_params.get("dilution_volume_ul"))

    internal_standard_specified = global_params.get("internal_standard_specified")
    internal_standard_enabled = "否"
    internal_standard_name = ""
    internal_standard_amount_ul_or_mg: Optional[float] = None
    stir_after_is_min = 5.0
    if internal_standard_specified is True:
        internal_standard_enabled = "是"
        internal_standard_name = str(global_params.get("internal_standard_name") or "").strip()
        internal_standard_amount_ul_or_mg = _safe_float(global_params.get("internal_standard_amount_ul_or_mg"))

    flash_filter_specified = global_params.get("flash_filter_specified")
    flash_filter_enabled = "否"
    flash_filter_solvent = ""
    flash_filter_volume_ul: Optional[float] = None
    if flash_filter_specified is True:
        flash_filter_enabled = "是"
        flash_filter_solvent = str(global_params.get("flash_filter_solvent") or "").strip()
        flash_filter_volume_ul = _safe_float(global_params.get("flash_filter_volume_ul"))

    sample_volume_ul = _safe_float(global_params.get("sample_volume_ul"))
    if sample_volume_ul is None:
        sample_volume_ul = 1.0

    experiments_json = model_json.get("experiments") or []
    experiments: List[ExperimentRow] = []
    if isinstance(experiments_json, list) is True:
        for exp in experiments_json:
            global_no = _safe_int((exp or {}).get("global_experiment_no"))
            if global_no is None:
                continue

            reagents_json = (exp or {}).get("reagents") or []
            reagents: List[ReagentItem] = []
            if isinstance(reagents_json, list) is True:
                for item in reagents_json:
                    name = str((item or {}).get("name") or "").strip()
                    amount_text = str((item or {}).get("amount_text") or "").strip()
                    if name != "":
                        reagents.append(ReagentItem(name=name, amount_text=amount_text))

            experiments.append(ExperimentRow(global_experiment_no=global_no, reagents=reagents))

    experiments.sort(key=lambda x: x.global_experiment_no)

    return ReactionPlan(
        experiment_name=experiment_name,
        experiment_id=experiment_id,
        scale_mmol=scale_mmol if scale_mmol is not None else -1.0,
        reactor_type=reactor_type,
        time_h=time_h if time_h is not None else -1.0,
        temperature_c=temperature_c if temperature_c is not None else -999.0,
        rpm=rpm,
        open_lid_target_temp_c=open_lid_target_temp_c,
        wait_target_temp=wait_target_temp,
        weigh_error_percent=weigh_error_percent,
        max_weigh_error_mg=max_weigh_error_mg,
        fixed_addition_order=fixed_addition_order,
        auto_add_stir_bar=auto_add_stir_bar,
        dilution_solvent=dilution_solvent,
        dilution_volume_ul=dilution_volume_ul,
        internal_standard_enabled=internal_standard_enabled,
        internal_standard_name=internal_standard_name,
        internal_standard_amount_ul_or_mg=internal_standard_amount_ul_or_mg,
        stir_after_is_min=stir_after_is_min,
        flash_filter_enabled=flash_filter_enabled,
        flash_filter_solvent=flash_filter_solvent,
        flash_filter_volume_ul=flash_filter_volume_ul,
        sample_volume_ul=sample_volume_ul,
        experiments=experiments,
    )


def _validate_plan(plan: ReactionPlan) -> List[str]:
    """
    功能:
        校验实验方案是否满足必填约束, 返回需要追问的问题列表.
    参数:
        plan: ReactionPlan, 当前方案.
    返回:
        List[str], 需要追问的中文问题.
    """
    questions: List[str] = []

    if plan.scale_mmol <= 0:
        questions.append("你这次反应规模是多少 mmol? 需要在 0.05 到 0.4 mmol 之间.")
    else:
        if plan.scale_mmol < 0.05 or plan.scale_mmol > 0.4:
            questions.append("反应规模需要在 0.05 到 0.4 mmol 之间, 你希望设定为多少 mmol?")

    if plan.time_h <= 0:
        questions.append("反应时间是多少? 请给出小时或分钟, 例如 8 h 或 30 min.")

    if plan.temperature_c <= -100:
        questions.append("反应温度是多少 °C?")

    if plan.internal_standard_enabled == "否":
        questions.append("你是否需要加入内标? 如果需要, 请给出内标种类与用量(μL 或 mg).")

    if plan.flash_filter_enabled == "否":
        questions.append("是否需要制备检测样品并进行闪滤? 如果需要, 请给出闪滤液种类与用量(μL).")

    # 新增: 实验行数必须是 12/24/36/48
    total_count = len(plan.experiments)
    if total_count not in _allowed_experiment_counts():
        questions.append(
            f"你这次给出的实验行数为 {total_count}, 必须是 12, 24, 36, 48 之一. 请删除或补充后再生成."
        )

    # 新增: global_experiment_no 必须从 1 连续递增, 且不超过 48
    if total_count > 0:
        expected = list(range(1, total_count + 1))
        actual = [x.global_experiment_no for x in plan.experiments]
        if actual != expected:
            questions.append(
                "实验编号(global_experiment_no)必须从 1 开始连续递增且不重复. 请删除或补充, 例如 1..12 或 1..24."
            )

    if total_count > 48:
        questions.append("模板最多支持 48 行实验(两张表), 请将实验行数控制在 48 以内.")

    return questions


def _collect_all_chemical_names(plan: ReactionPlan) -> List[str]:
    """
    功能:
        汇总方案中出现的所有化学品名称(试剂 + 稀释液 + 内标 + 闪滤液), 覆盖所有实验行.
    参数:
        plan: ReactionPlan, 实验方案.
    返回:
        List[str], 化学品名称列表(去空, 未去重).
    """
    names: List[str] = []

    for exp in plan.experiments:
        for item in exp.reagents:
            if item.name.strip() != "":
                names.append(item.name.strip())

    if plan.dilution_solvent.strip() != "":
        names.append(plan.dilution_solvent.strip())

    if plan.internal_standard_enabled == "是" and plan.internal_standard_name.strip() != "":
        names.append(plan.internal_standard_name.strip())

    if plan.flash_filter_enabled == "是" and plan.flash_filter_solvent.strip() != "":
        names.append(plan.flash_filter_solvent.strip())

    return names


def _check_chemicals_in_library(library: ChemicalLibrary, names: List[str]) -> Tuple[List[str], Dict[str, List[str]]]:
    """
    功能:
        校验化学品是否在库中, 并为不在库者生成相似名称候选.
    参数:
        library: ChemicalLibrary, 化学品库.
        names: List[str], 需要校验的名称.
    返回:
        Tuple[List[str], Dict[str, List[str]]]:
            - missing: List[str], 不在库的名称(去重).
            - suggestions: Dict[str, List[str]], 名称 -> 候选列表.
    """
    missing: List[str] = []
    suggestions: Dict[str, List[str]] = {}

    unique_names: List[str] = []
    for n in names:
        if n not in unique_names:
            unique_names.append(n)

    for name in unique_names:
        if library.exists(name) is False:
            missing.append(name)
            suggestions[name] = library.suggest(name, top_k=5)

    return missing, suggestions


# -----------------------------
# 主流程
# -----------------------------
def main() -> None:
    """
    功能:
        命令行交互入口, 支持一次性生成 12/24/36/48 行实验.
    参数:
        无.
    返回:
        None.
    """
    api_key = 'sk-4c02dad47a564fc0a33f2a58a06330e1'
    if api_key is None or api_key.strip() == "":
        logger.error("未检测到环境变量 DEEPSEEK_API_KEY, 请先设置后再运行.")
        logger.info("示例: export DEEPSEEK_API_KEY='你的key'")
        return

    if TEMPLATE_XLSX_PATH.exists() is False:
        logger.error("未找到模板文件: %s", str(TEMPLATE_XLSX_PATH))
        return

    if CHEMICAL_LIBRARY_XLSX_PATH.exists() is False:
        logger.error("未找到化学品库文件: %s", str(CHEMICAL_LIBRARY_XLSX_PATH))
        return

    library = ChemicalLibrary(CHEMICAL_LIBRARY_XLSX_PATH)
    client = DeepSeekClient(api_key=api_key)
    writer = ReactionTemplateWriter(TEMPLATE_XLSX_PATH)

    system_prompt = _build_system_prompt()
    messages: List[Dict[str, str]] = [{"role": "system", "content": system_prompt}]

    logger.info("请输入你的实验描述(可以包含多组实验). 输入 exit 结束.")
    user_text = input("用户: ").strip()
    if user_text.lower() == "exit":
        return

    messages.append({"role": "user", "content": user_text})

    max_rounds = 10
    plan: Optional[ReactionPlan] = None

    for round_index in range(max_rounds):
        logger.info("正在调用 DeepSeek 生成结构化方案, 第 %s 轮.", str(round_index + 1))
        assistant_text = client.chat(messages=messages, temperature=0.2)

        try:
            model_json = _extract_json(assistant_text)
        except Exception as exc:
            logger.warning("模型输出无法解析为 JSON, 将要求模型重试. 错误: %s", str(exc))
            messages.append({"role": "assistant", "content": assistant_text})
            messages.append({"role": "user", "content": "你刚才没有输出可解析 JSON, 请严格只输出 JSON, 不要解释."})
            continue

        plan = _plan_from_model_json(model_json)

        questions = _validate_plan(plan)
        if len(questions) > 0:
            messages.append({"role": "assistant", "content": assistant_text})
            ask_text = "我还需要你补充或修正以下信息:\n" + "\n".join([f"{i+1}. {q}" for i, q in enumerate(questions)])
            logger.info(ask_text)
            follow = input("用户补充: ").strip()
            if follow.lower() == "exit":
                return
            messages.append({"role": "user", "content": follow})
            continue

        all_names = _collect_all_chemical_names(plan)
        missing, suggestions = _check_chemicals_in_library(library, names=all_names)

        if len(missing) > 0:
            logger.warning("发现以下物质不在 chemical_list.xlsx 的 substance 列中:")
            for name in missing:
                cand = suggestions.get(name) or []
                if len(cand) > 0:
                    logger.warning(" - %s. 你想用的是不是: %s", name, " / ".join(cand))
                else:
                    logger.warning(" - %s. 未找到相似候选, 请确认名称是否需要加入化学品库.", name)

            messages.append({"role": "assistant", "content": assistant_text})
            messages.append(
                {
                    "role": "user",
                    "content": "请根据提示修正不在库中的化学品名称(必须与库的 substance 完全一致), 然后重新输出 JSON.",
                }
            )
            continue

        # 信息完整且化学品通过校验, 结束循环
        break

    if plan is None:
        logger.error("未能生成有效实验方案, 已退出.")
        return

    # 输出 storage_location(覆盖所有实验)
    used_names = _collect_all_chemical_names(plan)
    unique_used: List[str] = []
    for n in used_names:
        if n not in unique_used:
            unique_used.append(n)

    logger.info("本次实验涉及化学品的存储位置(storage_location)如下:")
    for name in unique_used:
        loc = library.storage_location(name)
        if loc is None:
            logger.info(" - %s: 未找到存储位置(请检查化学品库).", name)
        else:
            logger.info(" - %s: %s", name, loc)

    output_name = f"reaction_plan_{plan.experiment_name}_{len(plan.experiments)}rows.xlsx".replace("/", "_").replace("\\", "_")
    output_path = Path.cwd() / output_name
    writer.write(plan=plan, output_path=output_path)

    logger.info("已生成实验 Excel, 请务必打开文件检查各参数与试剂信息是否正确后再执行自动化流程.")


if __name__ == "__main__":
    main()
